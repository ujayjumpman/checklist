import streamlit as st
import requests
import json
import urllib.parse
import urllib3
import certifi
import pandas as pd
from datetime import datetime
import re 
import logging
import os
from dotenv import load_dotenv
import aiohttp
import asyncio
from concurrent.futures import ThreadPoolExecutor, as_completed
import time
import openpyxl
import io
from dotenv import load_dotenv
from uuid import uuid4
import ibm_boto3
from ibm_botocore.client import Config
from tenacity import retry, stop_after_attempt, wait_exponential
import xlsxwriter
from EWS_LIG import *
from dateutil.relativedelta import relativedelta
import traceback



# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Disable SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Load environment variables
load_dotenv()

# IBM COS Configuration
COS_API_KEY = os.getenv("COS_API_KEY")
COS_SERVICE_INSTANCE_ID = os.getenv("COS_SERVICE_INSTANCE_ID")
COS_ENDPOINT = os.getenv("COS_ENDPOINT")
COS_BUCKET = os.getenv("COS_BUCKET")

# WatsonX configuration
WATSONX_API_URL = os.getenv("WATSONX_API_URL_1")
MODEL_ID = os.getenv("MODEL_ID_1")
PROJECT_ID = os.getenv("PROJECT_ID_1")
API_KEY = os.getenv("API_KEY_1")

# API Endpoints
LOGIN_URL = "https://dms.asite.com/apilogin/"
IAM_TOKEN_URL = "https://iam.cloud.ibm.com/identity/token"


def get_missing_cos_config():
    required = {
        "COS_API_KEY": COS_API_KEY,
        "COS_SERVICE_INSTANCE_ID": COS_SERVICE_INSTANCE_ID,
        "COS_ENDPOINT": COS_ENDPOINT,
    }
    return [key for key, value in required.items() if not value]


if "slabreport" not in st.session_state:
    st.session_state.slabreport = {}

# Login Function
async def login_to_asite(email, password):
    headers = {"Accept": "application/json", "Content-Type": "application/x-www-form-urlencoded"}
    payload = {"emailId": email, "password": password}
    response = requests.post(LOGIN_URL, headers=headers, data=payload, verify=certifi.where(), timeout=50)
    if response.status_code == 200:
        try:
            session_id = response.json().get("UserProfile", {}).get("Sessionid")
            logger.info(f"Login successful, Session ID: {session_id}")
            st.session_state.sessionid = session_id
            st.sidebar.success(f"✅ Login successful, Session ID: {session_id}")
            return session_id
        except json.JSONDecodeError:
            logger.error("JSONDecodeError during login")
            st.sidebar.error("❌ Failed to parse login response")
            return None
    logger.error(f"Login failed: {response.status_code} - {response.text}")
    st.sidebar.error(f"❌ Login failed: {response.status_code} - {response.text}")
    return None

# Function to generate access token
@retry(stop=stop_after_attempt(5), wait=wait_exponential(multiplier=2, min=10, max=60))
def get_access_token(API_KEY):
    headers = {"Content-Type": "application/x-www-form-urlencoded", "Accept": "application/json"}
    data = {"grant_type": "urn:ibm:params:oauth:grant-type:apikey", "apikey": API_KEY}
    response = requests.post(IAM_TOKEN_URL, headers=headers, data=data, verify=certifi.where(), timeout=50)
    try:
        if response.status_code == 200:
            token_info = response.json()
            logger.info("Access token generated successfully")
            return token_info['access_token']
        else:
            logger.error(f"Failed to get access token: {response.status_code} - {response.text}")
            st.error(f"❌ Failed to get access token: {response.status_code} - {response.text}")
            raise Exception("Failed to get access token")
    except Exception as e:
        logger.error(f"Exception getting access token: {str(e)}")
        st.error(f"❌ Error getting access token: {str(e)}")
        return None

# Initialize COS client
@st.cache_resource
@retry(stop=stop_after_attempt(5), wait=wait_exponential(multiplier=1, min=4, max=10))
def initialize_cos_client():
    try:
        missing = get_missing_cos_config()
        if missing:
            raise ValueError(f"Missing COS configuration: {', '.join(missing)}")

        logger.info("Attempting to initialize COS client...")
        cos_client = ibm_boto3.client(
            's3',
            ibm_api_key_id=COS_API_KEY,
            ibm_service_instance_id=COS_SERVICE_INSTANCE_ID,
            config=Config(
                signature_version='oauth',
                connect_timeout=180,
                read_timeout=180,
                retries={'max_attempts': 15}
            ),
            endpoint_url=COS_ENDPOINT
        )
        logger.info("COS client initialized successfully")
        return cos_client
    except Exception as e:
        logger.error(f"Error initializing COS client: {str(e)}")
        st.error(f"❌ Error initializing COS client: {str(e)}")
        raise

# Fetch Workspace ID
async def GetWorkspaceID():
    url = "https://dmsak.asite.com/api/workspace/workspacelist"
    headers = {
        'Cookie': f'ASessionID={st.session_state.sessionid}',
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    response = requests.get(url, headers=headers, verify=certifi.where(), timeout=50)
    if response.status_code != 200:
        st.error(f"Failed to fetch workspace list: {response.status_code} - {response.text}")
        raise Exception(f"Failed to fetch workspace list: {response.status_code}")

    response_text = (response.text or "").strip()
    if not response_text:
        st.error("Workspace list API returned an empty response. Please re-login and try again.")
        raise Exception("Workspace list API returned empty response")

    try:
        data = response.json()
    except json.JSONDecodeError:
        preview = response_text[:250].replace("\n", " ")
        st.error(f"Workspace list returned non-JSON response. Preview: {preview}")
        raise Exception("Workspace list API returned non-JSON response")

    try:
        workspace_list = (
            data.get('asiteDataList', {}).get('workspaceVO', [])
            if isinstance(data, dict) else []
        )
        if not workspace_list:
            raise KeyError("asiteDataList.workspaceVO is missing or empty")

        workspace_id = workspace_list[0].get('Workspace_Id')
        if not workspace_id:
            raise KeyError("Workspace_Id missing in workspaceVO[0]")

        st.session_state.workspaceid = workspace_id
        st.write(f"Workspace ID: {st.session_state.workspaceid}")
    except (KeyError, IndexError, TypeError) as e:
        st.error(f"Error parsing workspace ID: {str(e)}")
        raise

# Fetch Project IDs
async def GetProjectId():
    url = f"https://adoddleak.asite.com/commonapi/qaplan/getQualityPlanList;searchCriteria={{'criteria': [{{'field': 'planCreationDate','operator': 7,'values': ['11-Mar-2025']}}], 'projectId': {str(st.session_state.workspaceid)}, 'recordLimit': 1000, 'recordStart': 1}}"
    headers = {
        'Cookie': f'ASessionID={st.session_state.sessionid}',
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    response = requests.get(url, headers=headers)
    if response.status_code != 200:
        st.error(f"Failed to fetch project IDs: {response.status_code} - {response.text}")
        raise Exception(f"Failed to fetch project IDs: {response.status_code}")
    data = response.json()
    if not data.get('data'):
        st.error("No quality plans found for the specified date.")
        raise Exception("No quality plans found")

    possible_name_fields = ['planName', 'name', 'title', 'planTitle', 'description']
    plan_mapping = {}

    for idx, plan in enumerate(data['data']):
        plan_id = plan.get('planId')
        plan_name = None

        for field in possible_name_fields:
            if plan.get(field):
                plan_name = str(plan[field]).strip().upper()
                break

        if not plan_name:
            logger.warning(f"No name field found for EWS/LIG plan at index {idx}. Available fields: {list(plan.keys())}")
            continue

        logger.info(f"EWS/LIG plan {idx}: Name='{plan_name}', ID={plan_id}")

        if 'STRUCTURE' in plan_name and 'EWS' in plan_name and 'LIG' in plan_name:
            plan_mapping['EWS_LIG_structure'] = plan_id
            st.write(f"Found Structure plan: {plan_name}")
        elif 'FINISHING' in plan_name and 'EWS' in plan_name and 'LIG' in plan_name:
            plan_mapping['EWS_LIG_finishing'] = plan_id
            st.write(f"Found Finishing plan: {plan_name}")

    if 'EWS_LIG_structure' not in plan_mapping and data['data']:
        plan_mapping['EWS_LIG_structure'] = data['data'][0]['planId']
        st.warning("Structure plan name not matched exactly. Using the first returned plan as fallback.")

    st.session_state.EWS_LIG_structure = plan_mapping.get('EWS_LIG_structure')
    st.session_state.EWS_LIG_finishing = plan_mapping.get('EWS_LIG_finishing')

    if st.session_state.EWS_LIG_structure:
        st.write(f"EWS_LIG Structure Project ID: {st.session_state.EWS_LIG_structure}")
    else:
        st.error("EWS_LIG Structure plan not found!")

    if st.session_state.EWS_LIG_finishing:
        st.write(f"EWS_LIG Finishing Project ID: {st.session_state.EWS_LIG_finishing}")
    else:
        st.warning("EWS_LIG Finishing plan not found in quality plan list.")

# Asynchronous Fetch Function
async def fetch_data(session, url, headers):
    async with session.get(url, headers=headers) as response:
        if response.status == 200:
            return await response.json()
        elif response.status == 204:
            return None
        else:
            raise Exception(f"Error fetching data: {response.status} - {await response.text()}")

# Fetch All Structure Data
async def GetAllDatas():
    record_limit = 1000
    headers = {'Cookie': f'ASessionID={st.session_state.sessionid}'}
    all_structure_data = []

    async with aiohttp.ClientSession() as session:
        start_record = 1
        st.write("Fetching EWS_LIG  Structure data...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanAssociation/?projectId={st.session_state.workspaceid}&planId={st.session_state.EWS_LIG_structure}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                async with session.get(url, headers=headers) as response:
                    if response.status == 204:
                        st.write("No more EWS_LIG Structure data available (204)")
                        break
                    data = await response.json()
                    if 'associationList' in data and data['associationList']:
                        all_structure_data.extend(data['associationList'])
                    else:
                        all_structure_data.extend(data if isinstance(data, list) else [])
                    st.write(f"Fetched {len(all_structure_data[-record_limit:])} EWS_LIG Structure records (Total: {len(all_structure_data)})")
                    if len(all_structure_data[-record_limit:]) < record_limit:
                        break
                    start_record += record_limit
            except Exception as e:
                st.error(f"❌ Error fetching Structure data: {str(e)}")
                break

    df_structure = pd.DataFrame(all_structure_data)
    
    desired_columns = ['activitySeq', 'qiLocationId']
    if 'statusName' in df_structure.columns:
        desired_columns.append('statusName')
    elif 'statusColor' in df_structure.columns:
        desired_columns.append('statusColor')
        status_mapping = {'#92D050': 'Completed', '#4CB0F0': 'Not Started', '#4C0F0': 'Not Started'}
        df_structure['statusName'] = df_structure['statusColor'].map(status_mapping).fillna('Unknown')
        desired_columns.append('statusName')
    else:
        st.error("❌ Neither statusName nor statusColor found in data!")
        return pd.DataFrame()

    EWS_LIG_structure = df_structure[desired_columns]    

    st.write(f"EWS_LIG STRUCTURE ({', '.join(desired_columns)})")
    st.write(f"Total records: {len(EWS_LIG_structure)}")
    st.write(EWS_LIG_structure)  
    
    return EWS_LIG_structure

# Fetch Activity Data
async def Get_Activity():
    record_limit = 1000
    headers = {
        'Cookie': f'ASessionID={st.session_state.sessionid}',
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    
    all_structure_activity_data = []
    
    async with aiohttp.ClientSession() as session:
        start_record = 1
        st.write("Fetching Activity data for EWS_LIG Structure...")  
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanActivities/?projectId={st.session_state.workspaceid}&planId={st.session_state.EWS_LIG_structure}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more Structure Activity data available (204)")
                    break
                if 'activityList' in data and data['activityList']:
                    all_structure_activity_data.extend(data['activityList'])
                else:
                    all_structure_activity_data.extend(data if isinstance(data, list) else [])
                st.write(f"Fetched {len(all_structure_activity_data[-record_limit:])} Structure Activity records (Total: {len(all_structure_activity_data)})")
                if len(all_structure_activity_data[-record_limit:]) < record_limit:
                    break
                start_record += record_limit
            except Exception as e:
                st.error(f"❌ Error fetching Structure Activity data: {str(e)}")
                break
 
    structure_activity_data = pd.DataFrame(all_structure_activity_data)[['activityName', 'activitySeq', 'formTypeId']]

    st.write("EWS_LIG STRUCTURE ACTIVITY DATA (activityName and activitySeq)")
    st.write(f"Total records: {len(structure_activity_data)}")
    st.write(structure_activity_data)
      
    return structure_activity_data

# Fetch Location/Module Data
async def Get_Location():
    record_limit = 1000
    headers = {
        'Cookie': f'ASessionID={st.session_state.sessionid}',
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    
    all_structure_location_data = []
    
    async with aiohttp.ClientSession() as session:
        start_record = 1
        total_records_fetched = 0
        st.write("Fetching EWS_LIG Structure Location/Module data...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanLocation/?projectId={st.session_state.workspaceid}&planId={st.session_state.EWS_LIG_structure}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more Structure Location data available (204)")
                    break
                if isinstance(data, list):
                    location_data = [{'qiLocationId': item.get('qiLocationId', ''), 'qiParentId': item.get('qiParentId', ''), 'name': item.get('name', '')} 
                                   for item in data if isinstance(item, dict)]
                    all_structure_location_data.extend(location_data)
                    total_records_fetched = len(all_structure_location_data)
                    st.write(f"Fetched {len(location_data)} Structure Location records (Total: {total_records_fetched})")
                elif isinstance(data, dict) and 'locationList' in data and data['locationList']:
                    location_data = [{'qiLocationId': loc.get('qiLocationId', ''), 'qiParentId': loc.get('qiParentId', ''), 'name': loc.get('name', '')} 
                                   for loc in data['locationList']]
                    all_structure_location_data.extend(location_data)
                    total_records_fetched = len(all_structure_location_data)
                    st.write(f"Fetched {len(location_data)} Structure Location records (Total: {total_records_fetched})")
                else:
                    st.warning(f"No 'locationList' in Structure Location data or empty list.")
                    break
                if len(location_data) < record_limit:
                    break
                start_record += record_limit
            except Exception as e:
                st.error(f"❌ Error fetching Structure Location data: {str(e)}")
                break
        
    structure_df = pd.DataFrame(all_structure_location_data)
    
    if 'name' in structure_df.columns and structure_df['name'].isna().all():
        st.error("❌ All 'name' values in Structure Location data are missing or empty!")

    st.write("EWS_LIG STRUCTURE LOCATION/MODULE DATA")
    st.write(f"Total records: {len(structure_df)}")
    st.write(structure_df)
    
    st.session_state.structure_location_data = structure_df
    
    return structure_df


async def GetFinishingDatas():
    if not st.session_state.get('EWS_LIG_finishing'):
        st.warning("EWS_LIG Finishing plan ID is not available.")
        return pd.DataFrame()

    record_limit = 1000
    headers = {'Cookie': f'ASessionID={st.session_state.sessionid}'}
    all_finishing_data = []

    async with aiohttp.ClientSession() as session:
        start_record = 1
        st.write("Fetching EWS_LIG Finishing data...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanAssociation/?projectId={st.session_state.workspaceid}&planId={st.session_state.EWS_LIG_finishing}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                async with session.get(url, headers=headers) as response:
                    if response.status == 204:
                        st.write("No more EWS_LIG Finishing data available (204)")
                        break
                    data = await response.json()
                    if 'associationList' in data and data['associationList']:
                        all_finishing_data.extend(data['associationList'])
                    else:
                        all_finishing_data.extend(data if isinstance(data, list) else [])
                    st.write(f"Fetched {len(all_finishing_data[-record_limit:])} EWS_LIG Finishing records (Total: {len(all_finishing_data)})")
                    if len(all_finishing_data[-record_limit:]) < record_limit:
                        break
                    start_record += record_limit
            except Exception as e:
                st.error(f"âŒ Error fetching Finishing data: {str(e)}")
                break

    df_finishing = pd.DataFrame(all_finishing_data)

    desired_columns = ['activitySeq', 'qiLocationId']
    if 'statusName' in df_finishing.columns:
        desired_columns.append('statusName')
    elif 'statusColor' in df_finishing.columns:
        desired_columns.append('statusColor')
        status_mapping = {'#92D050': 'Completed', '#4CB0F0': 'Not Started', '#4C0F0': 'Not Started'}
        df_finishing['statusName'] = df_finishing['statusColor'].map(status_mapping).fillna('Unknown')
        desired_columns.append('statusName')
    else:
        st.error("âŒ Neither statusName nor statusColor found in finishing data!")
        return pd.DataFrame()

    finishing_df = df_finishing[desired_columns]

    st.write(f"EWS_LIG FINISHING ({', '.join(desired_columns)})")
    st.write(f"Total records: {len(finishing_df)}")
    st.write(finishing_df)

    return finishing_df


async def Get_Finishing_Activity():
    if not st.session_state.get('EWS_LIG_finishing'):
        st.warning("EWS_LIG Finishing plan ID is not available.")
        return pd.DataFrame()

    record_limit = 1000
    headers = {
        'Cookie': f'ASessionID={st.session_state.sessionid}',
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }

    all_finishing_activity_data = []

    async with aiohttp.ClientSession() as session:
        start_record = 1
        st.write("Fetching Activity data for EWS_LIG Finishing...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanActivities/?projectId={st.session_state.workspaceid}&planId={st.session_state.EWS_LIG_finishing}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more Finishing Activity data available (204)")
                    break
                if 'activityList' in data and data['activityList']:
                    all_finishing_activity_data.extend(data['activityList'])
                else:
                    all_finishing_activity_data.extend(data if isinstance(data, list) else [])
                st.write(f"Fetched {len(all_finishing_activity_data[-record_limit:])} Finishing Activity records (Total: {len(all_finishing_activity_data)})")
                if len(all_finishing_activity_data[-record_limit:]) < record_limit:
                    break
                start_record += record_limit
            except Exception as e:
                st.error(f"âŒ Error fetching Finishing Activity data: {str(e)}")
                break

    finishing_activity_data = pd.DataFrame(all_finishing_activity_data)[['activityName', 'activitySeq', 'formTypeId']]

    st.write("EWS_LIG FINISHING ACTIVITY DATA (activityName and activitySeq)")
    st.write(f"Total records: {len(finishing_activity_data)}")
    st.write(finishing_activity_data)

    return finishing_activity_data


async def Get_Finishing_Location():
    if not st.session_state.get('EWS_LIG_finishing'):
        st.warning("EWS_LIG Finishing plan ID is not available.")
        return pd.DataFrame()

    record_limit = 1000
    headers = {
        'Cookie': f'ASessionID={st.session_state.sessionid}',
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }

    all_finishing_location_data = []

    async with aiohttp.ClientSession() as session:
        start_record = 1
        total_records_fetched = 0
        st.write("Fetching EWS_LIG Finishing Location/Module data...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanLocation/?projectId={st.session_state.workspaceid}&planId={st.session_state.EWS_LIG_finishing}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more Finishing Location data available (204)")
                    break
                if isinstance(data, list):
                    location_data = [{'qiLocationId': item.get('qiLocationId', ''), 'qiParentId': item.get('qiParentId', ''), 'name': item.get('name', '')}
                                   for item in data if isinstance(item, dict)]
                    all_finishing_location_data.extend(location_data)
                    total_records_fetched = len(all_finishing_location_data)
                    st.write(f"Fetched {len(location_data)} Finishing Location records (Total: {total_records_fetched})")
                elif isinstance(data, dict) and 'locationList' in data and data['locationList']:
                    location_data = [{'qiLocationId': loc.get('qiLocationId', ''), 'qiParentId': loc.get('qiParentId', ''), 'name': loc.get('name', '')}
                                   for loc in data['locationList']]
                    all_finishing_location_data.extend(location_data)
                    total_records_fetched = len(all_finishing_location_data)
                    st.write(f"Fetched {len(location_data)} Finishing Location records (Total: {total_records_fetched})")
                else:
                    st.warning("No 'locationList' in Finishing Location data or empty list.")
                    break
                if len(location_data) < record_limit:
                    break
                start_record += record_limit
            except Exception as e:
                st.error(f"âŒ Error fetching Finishing Location data: {str(e)}")
                break

    finishing_df = pd.DataFrame(all_finishing_location_data)

    if 'name' in finishing_df.columns and finishing_df['name'].isna().all():
        st.error("âŒ All 'name' values in Finishing Location data are missing or empty!")

    st.write("EWS_LIG FINISHING LOCATION/MODULE DATA")
    st.write(f"Total records: {len(finishing_df)}")
    st.write(finishing_df)

    st.session_state.finishing_location_data = finishing_df

    return finishing_df


async def GetAllDatas_EWSLIG_Style():
    record_limit = 1000
    headers = {'Cookie': f'ASessionID={st.session_state.sessionid}'}
    all_finishing_data = []
    all_structure_data = []

    async with aiohttp.ClientSession() as session:
        if st.session_state.get('EWS_LIG_finishing'):
            start_record = 1
            st.write("Fetching EWS_LIG Finishing data...")
            while True:
                url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanAssociation/?projectId={st.session_state.workspaceid}&planId={st.session_state.EWS_LIG_finishing}&recordStart={start_record}&recordLimit={record_limit}"
                try:
                    data = await fetch_data(session, url, headers)
                    if data is None:
                        st.write("No more EWS_LIG Finishing data available (204)")
                        break
                    if 'associationList' in data and data['associationList']:
                        all_finishing_data.extend(data['associationList'])
                    else:
                        all_finishing_data.extend(data if isinstance(data, list) else [])
                    st.write(f"Fetched {len(all_finishing_data[-record_limit:])} EWS_LIG Finishing records (Total: {len(all_finishing_data)})")
                    if len(all_finishing_data[-record_limit:]) < record_limit:
                        break
                    start_record += record_limit
                except Exception as e:
                    st.error(f"Error fetching Finishing data: {str(e)}")
                    break

        start_record = 1
        st.write("Fetching EWS_LIG Structure data...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanAssociation/?projectId={st.session_state.workspaceid}&planId={st.session_state.EWS_LIG_structure}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more EWS_LIG Structure data available (204)")
                    break
                if 'associationList' in data and data['associationList']:
                    all_structure_data.extend(data['associationList'])
                else:
                    all_structure_data.extend(data if isinstance(data, list) else [])
                st.write(f"Fetched {len(all_structure_data[-record_limit:])} EWS_LIG Structure records (Total: {len(all_structure_data)})")
                if len(all_structure_data[-record_limit:]) < record_limit:
                    break
                start_record += record_limit
            except Exception as e:
                st.error(f"Error fetching Structure data: {str(e)}")
                break

    def build_association_df(raw_data, dataset_label):
        df = pd.DataFrame(raw_data)
        desired_columns = ['activitySeq', 'qiLocationId']
        if 'statusName' in df.columns:
            desired_columns.append('statusName')
        elif 'statusColor' in df.columns:
            desired_columns.append('statusColor')
            status_mapping = {'#92D050': 'Completed', '#4CB0F0': 'Not Started', '#4C0F0': 'Not Started'}
            df['statusName'] = df['statusColor'].map(status_mapping).fillna('Unknown')
            desired_columns.append('statusName')
        elif df.empty:
            return pd.DataFrame(columns=['activitySeq', 'qiLocationId', 'statusName'])
        else:
            st.error(f"Neither statusName nor statusColor found in {dataset_label} data!")
            return pd.DataFrame()

        result_df = df[desired_columns]
        st.write(f"{dataset_label.upper()} ({', '.join(desired_columns)})")
        st.write(f"Total records: {len(result_df)}")
        st.write(result_df)
        return result_df

    finishing_df = build_association_df(all_finishing_data, "EWS_LIG Finishing")
    structure_df = build_association_df(all_structure_data, "EWS_LIG Structure")

    return finishing_df, structure_df


async def Get_Activity_EWSLIG_Style():
    record_limit = 1000
    headers = {
        'Cookie': f'ASessionID={st.session_state.sessionid}',
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }

    all_finishing_activity_data = []
    all_structure_activity_data = []

    async with aiohttp.ClientSession() as session:
        if st.session_state.get('EWS_LIG_finishing'):
            start_record = 1
            st.write("Fetching Activity data for EWS_LIG Finishing...")
            while True:
                url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanActivities/?projectId={st.session_state.workspaceid}&planId={st.session_state.EWS_LIG_finishing}&recordStart={start_record}&recordLimit={record_limit}"
                try:
                    data = await fetch_data(session, url, headers)
                    if data is None:
                        st.write("No more Finishing Activity data available (204)")
                        break
                    if 'activityList' in data and data['activityList']:
                        all_finishing_activity_data.extend(data['activityList'])
                    else:
                        all_finishing_activity_data.extend(data if isinstance(data, list) else [])
                    st.write(f"Fetched {len(all_finishing_activity_data[-record_limit:])} Finishing Activity records (Total: {len(all_finishing_activity_data)})")
                    if len(all_finishing_activity_data[-record_limit:]) < record_limit:
                        break
                    start_record += record_limit
                except Exception as e:
                    st.error(f"Error fetching Finishing Activity data: {str(e)}")
                    break

        start_record = 1
        st.write("Fetching Activity data for EWS_LIG Structure...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanActivities/?projectId={st.session_state.workspaceid}&planId={st.session_state.EWS_LIG_structure}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more Structure Activity data available (204)")
                    break
                if 'activityList' in data and data['activityList']:
                    all_structure_activity_data.extend(data['activityList'])
                else:
                    all_structure_activity_data.extend(data if isinstance(data, list) else [])
                st.write(f"Fetched {len(all_structure_activity_data[-record_limit:])} Structure Activity records (Total: {len(all_structure_activity_data)})")
                if len(all_structure_activity_data[-record_limit:]) < record_limit:
                    break
                start_record += record_limit
            except Exception as e:
                st.error(f"Error fetching Structure Activity data: {str(e)}")
                break

    def safe_select(df, cols):
        if df.empty:
            return pd.DataFrame(columns=cols)
        missing = [col for col in cols if col not in df.columns]
        for col in missing:
            df[col] = None
        return df[cols]

    finishing_activity_data = safe_select(pd.DataFrame(all_finishing_activity_data), ['activityName', 'activitySeq', 'formTypeId'])
    structure_activity_data = safe_select(pd.DataFrame(all_structure_activity_data), ['activityName', 'activitySeq', 'formTypeId'])

    st.write("EWS_LIG FINISHING ACTIVITY DATA (activityName, activitySeq, formTypeId)")
    st.write(f"Total records: {len(finishing_activity_data)}")
    st.write(finishing_activity_data)
    st.write("EWS_LIG STRUCTURE ACTIVITY DATA (activityName, activitySeq, formTypeId)")
    st.write(f"Total records: {len(structure_activity_data)}")
    st.write(structure_activity_data)

    return finishing_activity_data, structure_activity_data


async def Get_Location_EWSLIG_Style():
    record_limit = 1000
    headers = {
        'Cookie': f'ASessionID={st.session_state.sessionid}',
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }

    all_finishing_location_data = []
    all_structure_location_data = []

    async with aiohttp.ClientSession() as session:
        if st.session_state.get('EWS_LIG_finishing'):
            start_record = 1
            total_records_fetched = 0
            st.write("Fetching EWS_LIG Finishing Location/Module data...")
            while True:
                url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanLocation/?projectId={st.session_state.workspaceid}&planId={st.session_state.EWS_LIG_finishing}&recordStart={start_record}&recordLimit={record_limit}"
                try:
                    data = await fetch_data(session, url, headers)
                    if data is None:
                        st.write("No more Finishing Location data available (204)")
                        break
                    if isinstance(data, list):
                        location_data = [{'qiLocationId': item.get('qiLocationId', ''), 'qiParentId': item.get('qiParentId', ''), 'name': item.get('name', '')}
                                       for item in data if isinstance(item, dict)]
                        all_finishing_location_data.extend(location_data)
                        total_records_fetched = len(all_finishing_location_data)
                        st.write(f"Fetched {len(location_data)} Finishing Location records (Total: {total_records_fetched})")
                    elif isinstance(data, dict) and 'locationList' in data and data['locationList']:
                        location_data = [{'qiLocationId': loc.get('qiLocationId', ''), 'qiParentId': loc.get('qiParentId', ''), 'name': loc.get('name', '')}
                                       for loc in data['locationList']]
                        all_finishing_location_data.extend(location_data)
                        total_records_fetched = len(all_finishing_location_data)
                        st.write(f"Fetched {len(location_data)} Finishing Location records (Total: {total_records_fetched})")
                    else:
                        st.warning("No 'locationList' in Finishing Location data or empty list.")
                        break
                    if len(location_data) < record_limit:
                        break
                    start_record += record_limit
                except Exception as e:
                    st.error(f"Error fetching Finishing Location data: {str(e)}")
                    break

        start_record = 1
        total_records_fetched = 0
        st.write("Fetching EWS_LIG Structure Location/Module data...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanLocation/?projectId={st.session_state.workspaceid}&planId={st.session_state.EWS_LIG_structure}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more Structure Location data available (204)")
                    break
                if isinstance(data, list):
                    location_data = [{'qiLocationId': item.get('qiLocationId', ''), 'qiParentId': item.get('qiParentId', ''), 'name': item.get('name', '')}
                                   for item in data if isinstance(item, dict)]
                    all_structure_location_data.extend(location_data)
                    total_records_fetched = len(all_structure_location_data)
                    st.write(f"Fetched {len(location_data)} Structure Location records (Total: {total_records_fetched})")
                elif isinstance(data, dict) and 'locationList' in data and data['locationList']:
                    location_data = [{'qiLocationId': loc.get('qiLocationId', ''), 'qiParentId': loc.get('qiParentId', ''), 'name': loc.get('name', '')}
                                   for loc in data['locationList']]
                    all_structure_location_data.extend(location_data)
                    total_records_fetched = len(all_structure_location_data)
                    st.write(f"Fetched {len(location_data)} Structure Location records (Total: {total_records_fetched})")
                else:
                    st.warning("No 'locationList' in Structure Location data or empty list.")
                    break
                if len(location_data) < record_limit:
                    break
                start_record += record_limit
            except Exception as e:
                st.error(f"Error fetching Structure Location data: {str(e)}")
                break

    finishing_df = pd.DataFrame(all_finishing_location_data)
    structure_df = pd.DataFrame(all_structure_location_data)

    if 'name' in finishing_df.columns and finishing_df['name'].isna().all():
        st.error("All 'name' values in Finishing Location data are missing or empty!")
    if 'name' in structure_df.columns and structure_df['name'].isna().all():
        st.error("All 'name' values in Structure Location data are missing or empty!")

    st.write("EWS_LIG FINISHING LOCATION/MODULE DATA")
    st.write(f"Total records: {len(finishing_df)}")
    st.write(finishing_df)
    st.write("EWS_LIG STRUCTURE LOCATION/MODULE DATA")
    st.write(f"Total records: {len(structure_df)}")
    st.write(structure_df)

    st.session_state.finishing_location_data = finishing_df
    st.session_state.structure_location_data = structure_df

    return finishing_df, structure_df

# Process individual chunk
def process_chunk(chunk, chunk_idx, dataset_name, location_df):
    logger.info(f"Starting thread for {dataset_name} Chunk {chunk_idx + 1}")
    generated_text = format_chunk_locally(chunk, chunk_idx, len(chunk), dataset_name, location_df)
    logger.info(f"Completed thread for {dataset_name} Chunk {chunk_idx + 1}")
    return generated_text, chunk_idx

# Process data with manual counting
def process_manually(analysis_df, total, dataset_name, chunk_size=1000, max_workers=4):
    if analysis_df.empty:
        st.warning(f"No completed activities found for {dataset_name}.")
        return "No completed activities found."

    unique_activities = analysis_df['activityName'].unique()
    logger.info(f"Unique activities in {dataset_name} dataset: {list(unique_activities)}")
    logger.info(f"Total records in {dataset_name} dataset: {len(analysis_df)}")

    st.write(f"Saved EWS_LIG {dataset_name} data to EWS_LIG_{dataset_name.lower()}_data.json")
    chunks = [analysis_df[i:i + chunk_size] for i in range(0, len(analysis_df), chunk_size)]

    location_df = st.session_state.structure_location_data

    chunk_results = {}
    progress_bar = st.progress(0)
    status_text = st.empty()

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_chunk = {
            executor.submit(process_chunk, chunk, idx, dataset_name, location_df): idx 
            for idx, chunk in enumerate(chunks)
        }

        completed_chunks = 0
        for future in as_completed(future_to_chunk):
            chunk_idx = future_to_chunk[future]
            try:
                generated_text, idx = future.result()
                chunk_results[idx] = generated_text
                completed_chunks += 1
                progress_percent = completed_chunks / len(chunks)
                progress_bar.progress(progress_percent)
                status_text.text(f"Processed chunk {completed_chunks} of {len(chunks)} ({progress_percent:.1%} complete)")
            except Exception as e:
                logger.error(f"Error processing chunk {chunk_idx + 1} for {dataset_name}: {str(e)}")
                st.error(f"❌ Error processing chunk {chunk_idx + 1}: {str(e)}")

    parsed_data = {}
    for chunk_idx in sorted(chunk_results.keys()):
        generated_text = chunk_results[chunk_idx]
        if not generated_text:
            continue

        current_tower = None
        tower_activities = []
        lines = generated_text.split("\n")
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            if line.startswith("Tower:"):
                try:
                    tower_parts = line.split("Tower:", 1)
                    if len(tower_parts) > 1:
                        if current_tower and tower_activities:
                            if current_tower not in parsed_data:
                                parsed_data[current_tower] = []
                            parsed_data[current_tower].extend(tower_activities)
                            tower_activities = []
                        current_tower = tower_parts[1].strip()
                except Exception as e:
                    logger.warning(f"Error parsing Tower line: {line}, error: {str(e)}")
                    if not current_tower:
                        current_tower = f"Unknown Tower {chunk_idx}"
                    
            elif line.startswith("Total Completed Activities:"):
                continue
            elif not line.strip().startswith("activityName"):
                try:
                    parts = re.split(r'\s{2,}', line.strip())
                    if len(parts) >= 2:
                        activity_name = ' '.join(parts[:-1]).strip()
                        count_str = parts[-1].strip()
                        count_match = re.search(r'\d+', count_str)
                        if count_match:
                            count = int(count_match.group())
                            if current_tower:
                                tower_activities.append({
                                    "activityName": activity_name,
                                    "completedCount": count
                                })
                    else:
                        match = re.match(r'^\s*(.+?)\s+(\d+)$', line)
                        if match and current_tower:
                            activity_name = match.group(1).strip()
                            count = int(match.group(2).strip())
                            tower_activities.append({
                                "activityName": activity_name,
                                "completedCount": count
                            })
                except (ValueError, IndexError) as e:
                    logger.warning(f"Skipping malformed activity line: {line}, error: {str(e)}")

        if current_tower and tower_activities:
            if current_tower not in parsed_data:
                parsed_data[current_tower] = []
            parsed_data[current_tower].extend(tower_activities)

    aggregated_data = {}
    for tower_name, activities in parsed_data.items():
        tower_short_name = tower_name.split('/')[1] if '/' in tower_name else tower_name
        if tower_short_name not in aggregated_data:
            aggregated_data[tower_short_name] = {}
        
        for activity in activities:
            name = activity.get("activityName", "Unknown")
            count = activity.get("completedCount", 0)
            if name in aggregated_data[tower_short_name]:
                aggregated_data[tower_short_name][name] += count
            else:
                aggregated_data[tower_short_name][name] = count

    combined_output_lines = []
    sorted_towers = sorted(aggregated_data.keys())
    
    for i, tower_short_name in enumerate(sorted_towers):
        combined_output_lines.append(f"{tower_short_name:<11} activityName            CompletedCount")
        activity_dict = aggregated_data[tower_short_name]
        tower_total = 0
        for name, count in sorted(activity_dict.items()):
            combined_output_lines.append(f"{'':<11} {name:<23} {count:>14}")
            tower_total += count
        combined_output_lines.append(f"{'':<11} Total for {tower_short_name:<11}: {tower_total:>14}")
        if i < len(sorted_towers) - 1:
            combined_output_lines.append("")
    
    combined_output = "\n".join(combined_output_lines)
    return combined_output

# Local formatting function for manual counting
def format_chunk_locally(chunk, chunk_idx, chunk_size, dataset_name, location_df):
    towers_data = {}
    
    for _, row in chunk.iterrows():
        tower_name = row['tower_name']
        activity_name = row['activityName']
        count = int(row['CompletedCount'])
        
        if tower_name not in towers_data:
            towers_data[tower_name] = []
            
        towers_data[tower_name].append({
            "activityName": activity_name,
            "completedCount": count
        })
    
    output = ""
    total_activities = 0
    
    for tower_name, activities in sorted(towers_data.items()):
        output += f"Tower: {tower_name}\n"
        output += "activityName            CompletedCount\n"
        activity_dict = {}
        for activity in activities:
            name = activity['activityName']
            count = activity['completedCount']
            activity_dict[name] = activity_dict.get(name, 0) + count
        for name, count in sorted(activity_dict.items()):
            output += f"{name:<30} {count}\n"
            total_activities += count
    
    output += f"Total Completed Activities: {total_activities}"
    return output

def get_tower_name(full_path):
    parts = full_path.split('/')
    if len(parts) < 2:
        logger.warning(f"Invalid full_path format: {full_path}. Returning as-is.")
        return full_path

    logger.info(f"Processing path: {full_path}")
    logger.info(f"Path parts: {parts}")

    path_lower = full_path.lower()
    is_ews = "ews" in path_lower
    is_lig = "lig" in path_lower

    tower = None
    tower_number = None
    found_part = None
    
    for i, part in enumerate(parts):
        logger.info(f"Checking part {i}: '{part}'")
        
        # Look for "Tower X" or "Pour X" format
        part_lower = part.lower()
        if part_lower.startswith("tower"):
            tower = part
            found_part = f"Tower format: {part}"
            tower_parts = part.split()
            if len(tower_parts) > 1 and tower_parts[1].isdigit():
                tower_number = tower_parts[1]
            else:
                logger.warning(f"Could not extract tower number from: {part}")
                tower_number = "Unknown"
            break
        elif part_lower.startswith("pour"):
            found_part = f"Pour format: {part}"
            pour_parts = part.split()
            if len(pour_parts) > 1 and pour_parts[1].isdigit():
                tower_number = pour_parts[1]
                tower = f"Pour {tower_number}"
            else:
                logger.warning(f"Could not extract pour number from: {part}")
                tower_number = "Unknown"
                tower = part
            break
        elif "tower" in part_lower or "pour" in part_lower:
            found_part = f"Contains 'tower' or 'pour': {part}"
            logger.info(f"Found part containing 'tower' or 'pour': {part}")
            number_match = re.search(r'\d+', part)
            if number_match:
                tower_number = number_match.group()
                tower = f"Tower {tower_number}" if "tower" in part_lower else f"Pour {tower_number}"
            else:
                logger.warning(f"Could not extract number from: {part}")
                tower_number = "Unknown"
                tower = part
            break
    
    logger.info(f"Found part: {found_part}")
    logger.info(f"Tower: {tower}, Tower number: {tower_number}")
    
    if not tower:
        logger.warning(f"Tower/Pour name not found in path: {full_path}. Returning as-is.")
        return full_path

    if is_ews and not is_lig:
        prefix = "EWS"
    elif is_lig and not is_ews:
        prefix = "LIG"
    elif is_ews and is_lig:
        # Handle cases where both EWS and LIG appear in the path
        ews_idx = path_lower.find("ews")
        lig_idx = path_lower.find("lig")
        prefix = "EWS" if ews_idx < lig_idx else "LIG"
        logger.info(f"Both EWS and LIG in path, using prefix: {prefix}")
    else:
        logger.warning(f"Could not classify EWS/LIG for path: {full_path}. Defaulting to 'Unknown' prefix.")
        prefix = "Unknown"

    tower_name = f"{prefix} {tower}"
    logger.info(f"Final tower name: {tower_name}")
    return tower_name

def get_full_path(location_id, parent_child_dict, name_dict):
    path = []
    current_id = location_id
    max_depth = 15
    depth = 0
    visited_ids = set()
    
    while current_id and depth < max_depth:
        if current_id in visited_ids:
            logger.warning(f"Circular reference detected for location_id {location_id} at {current_id}. Path so far: {path}")
            break
        visited_ids.add(current_id)
        
        if current_id not in parent_child_dict or current_id not in name_dict:
            logger.warning(f"Location ID {current_id} not found in parent_child_dict or name_dict. Path so far: {path}")
            break
        
        parent_id = parent_child_dict.get(current_id)
        name = name_dict.get(current_id, "Unknown")
        path.append(name)
        
        if not parent_id:
            break
        
        current_id = parent_id
        depth += 1
    
    if depth >= max_depth:
        logger.warning(f"Max depth reached while computing path for location_id {location_id}. Possible deep hierarchy or error. Path: {path}")
    
    if not path:
        logger.warning(f"No path constructed for location_id {location_id}. Using 'Unknown'.")
        return "Unknown"
    
    full_path = '/'.join(reversed(path))
    logger.debug(f"Full path for location_id {location_id}: {full_path}")
    return full_path

def is_roof_slab_only(full_path):
    parts = full_path.split('/')
    last_part = parts[-1].lower()
    is_slab = any(keyword in last_part for keyword in ['roof slab', 'slab', 'roofslab', 'slab level'])
    logger.debug(f"Checking roof slab for path: {full_path}, result: {is_slab}")
    return is_slab

def process_data(df, activity_df, location_df, dataset_name, use_module_hierarchy_for_finishing=False):
    completed = df[df['statusName'] == 'Completed'].copy()
    
    asite_activities = [
        "Wall Conducting", "Plumbing Works", "POP & Gypsum Plaster", "Wiring & Switch Socket",
        "Slab Conducting", "Electrical Cable", "Door/Window Frame", "Waterproofing - Sunken",
        "Wall Tile", "Floor Tile", "Door/Window Shutter", "Shuttering", "Reinforcement",
        "Sewer Line", "Rain Water/Storm Line", "Granular Sub-base", "WMM",
        "Saucer drain/Paver block", "Kerb Stone", "Concreting"
    ]
    
    count_table = pd.DataFrame({'Count': [0] * len(asite_activities)}, index=asite_activities)
    
    if completed.empty:
        logger.warning(f"No completed activities found in {dataset_name} data.")
        return pd.DataFrame(), 0, count_table

    completed = completed.merge(location_df[['qiLocationId', 'name']], on='qiLocationId', how='left')
    completed = completed.merge(activity_df[['activitySeq', 'activityName']], on='activitySeq', how='left')

    if 'qiActivityId' not in completed.columns:
        completed['qiActivityId'] = completed['qiLocationId'].astype(str) + '$$' + completed['activitySeq'].astype(str)

    if completed['name'].isna().all():
        logger.error(f"All 'name' values are missing in {dataset_name} data after merge!")
        st.error(f"All 'name' values are missing in {dataset_name} data after merge! Check location data.")
        completed['name'] = 'Unknown'
    else:
        completed['name'] = completed['name'].fillna('Unknown')

    def normalize_activity_name(name):
        typo_corrections = {
            "Wall Conduting": "Wall Conducting",
            "Slab conduting": "Slab Conducting",
            "WallTile": "Wall Tile",
            "FloorTile": "Floor Tile",
            "Wall Tiling": "Wall Tile",
            "Floor Tiling": "Floor Tile",
            "wall tile": "Wall Tile",
            "floor tile": "Floor Tile",
            "wall tiling": "Wall Tile",
            "floor tiling": "Floor Tile",
            "DoorWindowFrame": "Door/Window Frame",
            "DoorWindowShutter": "Door/Window Shutter",
            "Second Roof Slab": "Roof Slab",
            "First Roof Slab": "Roof Slab",
            "Roof slab": "Roof Slab",
            "Beam": "Beam",
            "Column": "Column",
            "Reinforcement": "Reinforcement",
            "Shuttering": "Shuttering",
            "Concreting": "Concreting",
            "DeShuttering": "De-Shuttering"
        }
        for typo, correct in typo_corrections.items():
            if name.lower() == typo.lower():
                return correct
        return name

    completed['activityName'] = completed['activityName'].apply(normalize_activity_name).fillna('Unknown')

    parent_child_dict = dict(zip(location_df['qiLocationId'], location_df['qiParentId']))
    name_dict = dict(zip(location_df['qiLocationId'], location_df['name']))

    completed['full_path'] = completed['qiLocationId'].apply(
        lambda x: get_full_path(x, parent_child_dict, name_dict)
    )

    logger.debug(f"All unique full_path values in {dataset_name} dataset BEFORE filtering:")
    full_path_counts = completed['full_path'].value_counts()
    for path, count in full_path_counts.items():
        logger.debug(f"  Path: {path}, Count: {count}")

    completed['temp_tower_name'] = completed['full_path'].apply(
        lambda x: x.split('/')[1] if len(x.split('/')) > 1 and ('Tower' in x.split('/')[1] or 'Pour' in x.split('/')[1]) else 'Unknown'
    )
    tower_counts_before = completed['temp_tower_name'].value_counts()
    logger.debug(f"Tower distribution BEFORE filtering in {dataset_name}:")
    for tower, count in tower_counts_before.items():
        logger.debug(f"  {tower}: {count} records")

    def has_flat_number(full_path):
        parts = full_path.split('/')
        last_part = parts[-1]
        match = re.match(r'^\d+(?:(?:\s*\(LL\))|(?:\s*\(UL\))|(?:\s*LL)|(?:\s*UL))?$', last_part)
        return bool(match)

    def is_roof_slab_only(full_path):
        parts = full_path.split('/')
        last_part = parts[-1].lower()
        is_slab = any(keyword in last_part for keyword in ['roof slab', 'slab', 'roofslab', 'slab level'])
        logger.debug(f"Checking roof slab for path: {full_path}, result: {is_slab}")
        return is_slab

    if dataset_name.lower() == 'structure':
        logger.debug(f"Applying roof slab filtering for {dataset_name} dataset")
        completed_before_filter = len(completed)
        
        logger.debug(f"All unique paths before roof slab filtering:")
        for path, count in full_path_counts.items():
            logger.debug(f"  Path: {path}, Count: {count}")
        
        logger.debug("Paths that WILL be filtered out by is_roof_slab_only:")
        paths_to_be_filtered = completed[~completed['full_path'].apply(is_roof_slab_only)]['full_path'].unique()
        for path in sorted(paths_to_be_filtered):
            logger.debug(f"  Path: {path}")
        
        completed = completed[completed['full_path'].apply(is_roof_slab_only)]
        completed_after_filter = len(completed)
        logger.debug(f"Roof slab filtering: {completed_before_filter} -> {completed_after_filter} records")
        
        if not completed.empty:
            logger.debug(f"Paths that passed roof slab filtering:")
            full_path_counts_after = completed['full_path'].value_counts()
            for path, count in full_path_counts_after.items():
                logger.debug(f"  Path: {path}, Count: {count}")
        else:
            logger.warning(f"No paths contain 'roof slab', 'slab', 'roofslab', or 'slab level' in {dataset_name} dataset")
    
    else:
        completed = completed[completed['full_path'].apply(has_flat_number)]
        if completed.empty:
            logger.warning(f"No completed activities with flat numbers found in {dataset_name} data after filtering.")
            return pd.DataFrame(), 0, count_table

    completed['temp_tower_name'] = completed['full_path'].apply(
        lambda x: x.split('/')[1] if len(x.split('/')) > 1 and ('Tower' in x.split('/')[1] or 'Pour' in x.split('/')[1]) else 'Unknown'
    )
    tower_counts_after = completed['temp_tower_name'].value_counts()
    logger.debug(f"Tower distribution AFTER filtering in {dataset_name}:")
    for tower, count in tower_counts_after.items():
        logger.debug(f"  {tower}: {count} records")
    completed = completed.drop(columns=['temp_tower_name'])

    completed['tower_name'] = completed['full_path'].apply(get_tower_name)

    logger.debug(f"All tower_name values after get_tower_name in {dataset_name}:")
    tower_name_counts = completed['tower_name'].value_counts()
    for tower_name, count in tower_name_counts.items():
        logger.debug(f"  {tower_name}: {count} records")

    logger.debug(f"Sample full_path to tower_name mapping in {dataset_name}:")
    for idx, row in completed[['full_path', 'tower_name']].head(20).iterrows():
        logger.debug(f"  full_path: {row['full_path']} -> tower_name: {row['tower_name']}")

    # Changed: This now represents CLOSED checklists from Asite
    analysis = completed.groupby(['tower_name', 'activityName'])['qiLocationId'].nunique().reset_index(name='ClosedChecklistCount')
    analysis = analysis.sort_values(by=['tower_name', 'activityName'], ascending=True)
    total_closed = analysis['ClosedChecklistCount'].sum()

    activity_counts = completed.groupby('activityName')['qiLocationId'].nunique().reset_index(name='Count')
    for activity in asite_activities:
        if activity in activity_counts['activityName'].values:
            count_table.loc[activity, 'Count'] = activity_counts[activity_counts['activityName'] == activity]['Count'].iloc[0]

    logger.info(f"Total closed checklists for {dataset_name}: {total_closed}")
    logger.info(f"Count table for {dataset_name}:\n{count_table.to_string()}")
    
    logger.debug(f"Final analysis results for {dataset_name} by tower:")
    for tower in sorted(analysis['tower_name'].unique()):
        tower_data = analysis[analysis['tower_name'] == tower]
        tower_total = tower_data['ClosedChecklistCount'].sum()
        logger.debug(f"  {tower}: {tower_total} total closed checklists")
    
    return analysis, total_closed, count_table



# Main analysis function for EWSLIG  Structure
def AnalyzeStatusManually(email=None, password=None):
    start_time = time.time()

    if 'sessionid' not in st.session_state:
        st.error("❌ Please log in first!")
        logger.error("AnalyzeStatusManually failed: No sessionid in st.session_state")
        return

    required_data = [
        'EWSLIG_structure',
        'structure_activity_data',
        'structure_location_data'
    ]
    
    for data_key in required_data:
        if data_key not in st.session_state:
            st.error(f"❌ Please fetch required data first! Missing: {data_key}")
            logger.error(f"AnalyzeStatusManually failed: Missing {data_key} in st.session_state")
            return
        if not isinstance(st.session_state[data_key], pd.DataFrame):
            st.error(f"❌ {data_key} is not a DataFrame! Found type: {type(st.session_state[data_key])}")
            logger.error(f"AnalyzeStatusManually failed: {data_key} is not a DataFrame, found type {type(st.session_state[data_key])}")
            return
        if st.session_state[data_key].empty:
            st.error(f"❌ {data_key} is an empty DataFrame!")
            logger.error(f"AnalyzeStatusManually failed: {data_key} is an empty DataFrame")
            return

    structure_data = st.session_state.EWSLIG_structure
    structure_activity = st.session_state.structure_activity_data
    structure_locations = st.session_state.structure_location_data
    
    for df, name in [(structure_data, "Structure")]:
        required_columns = ['statusName', 'qiLocationId', 'activitySeq']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            st.error(f"❌ Missing columns {missing_columns} in {name} data!")
            logger.error(f"AnalyzeStatusManually failed: Missing columns {missing_columns} in {name} data")
            return

    for df, name in [(structure_locations, "Structure Location")]:
        required_columns = ['qiLocationId', 'name']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            st.error(f"❌ Missing columns {missing_columns} in {name} data!")
            logger.error(f"AnalyzeStatusManually failed: Missing columns {missing_columns} in {name} data")
            return

    for df, name in [(structure_activity, "Structure Activity")]:
        required_columns = ['activitySeq', 'activityName']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            st.error(f"❌ Missing columns {missing_columns} in {name} data!")
            logger.error(f"AnalyzeStatusManually failed: Missing columns {missing_columns} in {name} data")
            return

    try:
        logger.info("Starting structure data processing...")
        structure_analysis, structure_total, _ = process_data(structure_data, structure_activity, structure_locations, "Structure")
        logger.info(f"Structure data processed. Total closed checklists: {structure_total}")
    except Exception as e:
        st.error(f"❌ Error processing structure data: {str(e)}")
        logger.error(f"AnalyzeStatusManually failed: Error processing structure data: {str(e)}")
        logger.error(f"Stack trace:\n{traceback.format_exc()}")
        return

    st.session_state.structure_analysis = structure_analysis
    st.session_state.structure_total = structure_total
    logger.info("Structure analysis stored in st.session_state")

    if 'tower_name' not in structure_analysis.columns:
        st.error("❌ Tower names not found in structure analysis. Check location data and tower extraction logic.")
        logger.error("AnalyzeStatusManually failed: tower_name column not found in structure_analysis")
        return

    unique_towers = structure_analysis['tower_name'].unique()
    logger.debug(f"Unique towers in structure_analysis: {list(unique_towers)}")

    if len(unique_towers) <= 1:
        logger.warning(f"Only {len(unique_towers)} tower(s) found: {unique_towers}. Expected multiple towers for EWS LIG.")
        st.warning(f"Only {len(unique_towers)} tower(s) found. This may indicate an issue with location data separation.")

    def sort_key(tower_name):
        prefix = tower_name.split()[0]  # EWS or LIG
        tower_num = int(tower_name.split()[-1]) if tower_name.split()[-1].isdigit() else 0
        return (prefix, tower_num)

    sorted_towers = sorted(unique_towers, key=sort_key)

    st.write("### EWS_LIG Structure Quality Analysis (Closed Checklists from Asite):")
    
    for tower in sorted_towers:
        tower_data = structure_analysis[structure_analysis['tower_name'] == tower]
        if tower_data.empty:
            st.write(f"**{tower}:** No closed checklists found.")
            logger.info(f"No closed checklists for {tower}")
            continue

        st.write(f"**{tower}:**")
        output_lines = []
        output_lines.append("activityName ClosedChecklistCount")
        tower_total = 0
        for _, row in tower_data.iterrows():
            # CHANGED: Use ClosedChecklistCount instead of CompletedCount
            output_lines.append(f"{row['activityName']:<30} {row['ClosedChecklistCount']}")
            tower_total += row['ClosedChecklistCount']
        output_lines.append(f"{'Total for ' + tower:<30} {tower_total}")
        st.text("\n".join(output_lines))
        logger.info(f"Displayed output for {tower}")

    st.write(f"**Total Closed Checklists Across All Towers:** {structure_total}")

    end_time = time.time()
    st.write(f"Total execution time: {end_time - start_time:.2f} seconds")
    logger.info(f"AnalyzeStatusManually completed in {end_time - start_time:.2f} seconds")

def get_cos_files():
    try:
        # Initialize COS client
        cos_client = initialize_cos_client()
        if not cos_client:
            st.error("❌ Failed to initialize COS client. Check credentials or configuration.")
            logger.error("Failed to initialize COS client")
            return None

        # Step 1: List all objects in the bucket to inspect structure
        st.write(f"Listing all objects in bucket '{COS_BUCKET}' (no prefix)")
        response = cos_client.list_objects_v2(Bucket=COS_BUCKET)
        if 'Contents' not in response:
            st.error(f"❌ No objects found in bucket '{COS_BUCKET}'. Verify bucket name and permissions.")
            logger.error(f"No objects found in bucket {COS_BUCKET}")
            return None

        all_files = [obj['Key'] for obj in response.get('Contents', [])]
        st.write("**All files in bucket:**")
        if all_files:
            st.write("\n".join(all_files))
        else:
            st.write("No files found in the bucket.")
            logger.warning(f"Bucket {COS_BUCKET} is empty")
            return None

        # Extract folder names (prefixes)
        folders = set()
        for file in all_files:
            if '/' in file:
                folder = file.split('/')[0] + '/'
                folders.add(folder)
        st.write("**Available folders in bucket:**")
        st.write("\n".join(folders) if folders else "No folders found.")

        # Step 2: Focus on the EWS LIG P4 folder with variations
        possible_prefixes = [
            "EWS LIG P4/",  # Exact match
            "EWS LIG P4",   # Without trailing slash
            "ews lig p4/",  # Lowercase
            "EWS LIG P4 /", # Extra space
            "EWS_LIG_P4/",  # Underscores instead of spaces
            "EWS-LIG-P4/",  # Hyphens instead of spaces
        ]

        target_files = []
        for prefix in possible_prefixes:
            st.write(f"\nListing objects in bucket '{COS_BUCKET}' with prefix '{prefix}'")
            response = cos_client.list_objects_v2(Bucket=COS_BUCKET, Prefix=prefix)
            
            if 'Contents' not in response:
                st.write(f"No files found in '{prefix}' folder.")
                logger.info(f"No objects found in {prefix} folder")
                continue

            prefix_files = [obj['Key'] for obj in response.get('Contents', [])]
            st.write(f"**Files in {prefix} folder:**")
            if prefix_files:
                st.write("\n".join(prefix_files))
            else:
                st.write("No files found.")
                logger.info(f"{prefix} folder is empty")
                continue

            # Updated regex pattern to match "Structure Work Tracker" instead of "Checklist Report"
            pattern = re.compile(
                r"(?i)EWS\s*LIG\s*P4/.*?Structure\s*Work\s*Tracker.*?[\(\s]*(.*?)(?:[\)\s]*\.xlsx)$"
            )

            # Supported date formats for parsing
            date_formats = [
                "%d-%m-%Y", "%Y-%m-%d", "%d-%m-%y",
                "%d/%m/%Y", "%d.%m.%Y", "%Y%m%d",
                "%d%m%Y", "%Y.%m.%d"
            ]

            for key in prefix_files:
                match = pattern.match(key)
                if match:
                    date_str = match.group(1).strip('()').strip()
                    parsed_date = None
                    for fmt in date_formats:
                        try:
                            parsed_date = datetime.strptime(date_str, fmt)
                            break
                        except ValueError:
                            continue
                    if parsed_date:
                        target_files.append({'key': key, 'date': parsed_date})
                    else:
                        logger.warning(f"Could not parse date in filename: {key}")
                        st.warning(f"Skipping file with unparseable date: {key}")
                else:
                    st.write(f"File '{key}' does not match the expected pattern.")

        if not target_files:
            st.error(f"❌ No Excel files matched the expected pattern in any of the folders: {', '.join(possible_prefixes)}")
            logger.error("No files matched the expected pattern")
            return None

        # Find the latest file based on the parsed date
        latest_file = max(target_files, key=lambda x: x['date'])
        file_key = latest_file['key']
        st.success(f"Found matching file: {file_key}")
        return file_key

    except Exception as e:
        st.error(f"❌ Error fetching COS files: {str(e)}")
        logger.error(f"Error fetching COS files: {str(e)}")
        return None

if 'cos_df_Revised_Baseline_45daysNGT_Rai' not in st.session_state:
    st.session_state.cos_df_Revised_Baseline_45daysNGT_Rai = None

if 'cos_finishing_towers' not in st.session_state:
    st.session_state.cos_finishing_towers = {}

if 'cos_finishing_files' not in st.session_state:
    st.session_state.cos_finishing_files = []

if 'ai_response' not in st.session_state:
    st.session_state.ai_response = {} 


def make_streamlit_safe_df(df):
    safe_df = df.copy()
    safe_df.columns = [str(col) for col in safe_df.columns]
    for col in safe_df.columns:
        if safe_df[col].dtype == object:
            safe_df[col] = safe_df[col].apply(
                lambda value: value.isoformat(sep=' ') if isinstance(value, datetime) else (str(value) if value is not None else '')
            )
        else:
            # Convert all non-object types to string to avoid Arrow serialization issues
            safe_df[col] = safe_df[col].astype(str)
    return safe_df

# Process Excel files for EWSLIG blocks with updated sheet names and expected_columns
def process_file(file_stream, filename):
    try:
        workbook = openpyxl.load_workbook(file_stream)
        available_sheets = workbook.sheetnames
        st.write(f"Available sheets in {filename}: {', '.join(available_sheets)}")

        target_sheets = ["Revised Baseline 45daysNGT+Rai"]
        results = []

        for sheet_name in target_sheets:
            if sheet_name not in available_sheets:
                st.warning(f"Sheet '{sheet_name}' not found in file: {filename}")
                continue

            file_stream.seek(0)

            try:
                # Read the first few rows to inspect the data
                df_preview = pd.read_excel(file_stream, sheet_name=sheet_name, nrows=10)
                st.write(f"Preview of first 10 rows in {sheet_name}:")
                st.write(make_streamlit_safe_df(df_preview))

                # Try different header rows
                header_found = False
                actual_finish_col = None
                for header_row in [4, 5, 6, 3, 2]:
                    file_stream.seek(0)
                    df = pd.read_excel(file_stream, sheet_name=sheet_name, header=header_row)
                    st.write(f"Testing header row {header_row} in {sheet_name}. Raw columns: {list(df.columns)}")

                    df.columns = [col.strip() if isinstance(col, str) else col for col in df.columns]

                    # Check for 'Floors' or floor identifiers
                    if 'Floors' in df.columns or any('Floor' in str(col) for col in df.columns):
                        header_found = True
                    elif not df.empty and any(str(df.iloc[i, 0]).strip() in ['GF', '1F', '2F', '3F', '4F', '5F'] for i in range(min(5, len(df)))):
                        if df.columns[0] != 'Floors':
                            df.rename(columns={df.columns[0]: 'Floors'}, inplace=True)
                        header_found = True

                    # Check for 'Actual Finish' or variants
                    for col in df.columns:
                        if str(col).lower() in ['actual finish', 'actual_finish', 'finish date', 'completion date']:
                            actual_finish_col = col
                            break

                    if header_found and actual_finish_col:
                        break

                if not header_found:
                    st.error(f"No valid header row found in {sheet_name}. Expected to find 'Floors' column or floor identifiers.")
                    continue

                # Clean up the dataframe
                df = df.dropna(subset=[df.columns[0]])
                df = df[~df.iloc[:, 0].astype(str).str.contains('Floor|Pour|Baseline|Days', case=False, na=False)]
                
                floor_pattern = r'^(GF|\d{1,2}F)$'
                df = df[df.iloc[:, 0].astype(str).str.match(floor_pattern, na=False)]

                df.rename(columns={df.columns[0]: 'Activity Name'}, inplace=True)

                # Rename 'Actual Finish' if found
                if actual_finish_col:
                    df.rename(columns={actual_finish_col: 'Actual Finish'}, inplace=True)
                else:
                    st.write(f"No 'Actual Finish' column found in {sheet_name}. Using empty values.")
                    df['Actual Finish'] = pd.NA
                    logger.info(f"No 'Actual Finish' column in {sheet_name}")

                target_columns = ['Activity Name', 'Actual Finish']
                available_columns = [col for col in target_columns if col in df.columns]
                for col in df.columns:
                    if col not in target_columns:
                        available_columns.append(col)

                if len(available_columns) <= 1:
                    st.error(f"Only 'Activity Name' found in {sheet_name}. No additional columns to process.")
                    continue

                df = df[available_columns]
                df['Activity Name'] = df['Activity Name'].astype(str).str.strip()

                if 'Actual Finish' in df.columns:
                    df['Actual_Finish_Original'] = df['Actual Finish'].astype(str)
                    df['Actual Finish'] = pd.to_datetime(df['Actual Finish'], errors='coerce', dayfirst=True)
                    has_na_mask = (
                        pd.isna(df['Actual Finish']) |
                        (df['Actual_Finish_Original'].str.upper() == 'NAT') |
                        (df['Actual_Finish_Original'].str.lower().isin(['nan', 'na', 'n/a', 'none', '']))
                    )
                    st.write(f"Sample of rows with NA or invalid values in Actual Finish for {sheet_name}:")
                    na_rows = df[has_na_mask][['Activity Name', 'Actual Finish']]
                    if not na_rows.empty:
                        st.write(na_rows.head(10))
                    else:
                        st.write("No NA or invalid values found in Actual Finish")
                    df.drop('Actual_Finish_Original', axis=1, inplace=True)

                st.write(f"Unique Activity Names (Floor identifiers) in {sheet_name}:")
                st.write(make_streamlit_safe_df(df[['Activity Name']].drop_duplicates()))

                st.write(f"Final processed dataframe shape: {df.shape}")
                st.write(f"Final columns: {list(df.columns)}")
                st.write("Sample of processed data:")
                st.write(make_streamlit_safe_df(df.head()))

                results.append((df, sheet_name))

            except Exception as e:
                st.error(f"Error processing sheet {sheet_name}: {str(e)}")
                logger.error(f"Error processing sheet {sheet_name}: {str(e)}")
                continue

        if not results:
            st.error(f"No valid sheets ({', '.join(target_sheets)}) found in file: {filename}")
            return [(None, None)]

        return results

    except Exception as e:
        st.error(f"Error loading Excel file: {str(e)}")
        logger.error(f"Error loading Excel file: {str(e)}")
        return [(None, None)]


def get_cos_tracker_files():
    try:
        cos_client = initialize_cos_client()
        if not cos_client:
            st.error("❌ Failed to initialize COS client. Check credentials or configuration.")
            logger.error("Failed to initialize COS client")
            return []

        possible_prefixes = [
            "EWS LIG P4/",
            "EWS LIG P4",
            "ews lig p4/",
            "EWS LIG P4 /",
            "EWS_LIG_P4/",
            "EWS-LIG-P4/",
        ]

        date_formats = [
            "%d-%m-%Y", "%Y-%m-%d", "%d-%m-%y",
            "%d/%m/%Y", "%d.%m.%Y", "%Y%m%d",
            "%d%m%Y", "%Y.%m.%d"
        ]

        expected_finishing_towers = [
            "EWS_1", "EWS_2", "EWS_3",
            "LIG_1", "LIG_2", "LIG_3",
        ]
        structure_candidates = []
        finishing_candidates = {}
        scanned_files = set()

        structure_pattern = re.compile(
            r"(?i)Structure\s*Work\s*Tracker[\(\s]*(.*?)(?:[\)\s]*\.xlsx)$"
        )
        finishing_pattern = re.compile(
            r"(?i)\b(EWS|LIG)[\s_-]*(?:Tower|T)[\s_-]*([123])\b[\s_-]*Finishing[\s_-]*Tracker[\(\s_-]*(.*?)(?:[\)\s]*\.xlsx)$"
        )

        def parse_tracker_date(date_str, fallback_date=None):
            cleaned_date = str(date_str or "").strip().strip("()[]")
            if cleaned_date:
                for fmt in date_formats:
                    try:
                        return datetime.strptime(cleaned_date, fmt)
                    except ValueError:
                        continue
            return fallback_date

        def build_finishing_display_name(segment, tower_num, parsed_date):
            formatted_date = parsed_date.strftime("%d-%m-%Y") if parsed_date else "Unknown Date"
            return f"{segment} Tower {tower_num} Finishing Tracker ({formatted_date})"

        for prefix in possible_prefixes:
            response = cos_client.list_objects_v2(Bucket=COS_BUCKET, Prefix=prefix)
            if 'Contents' not in response:
                continue

            for obj in response.get('Contents', []):
                key = obj['Key']
                if key in scanned_files or not key.lower().endswith('.xlsx'):
                    continue
                scanned_files.add(key)

                filename = key.split('/')[-1]
                last_modified = obj.get('LastModified')

                structure_match = structure_pattern.search(filename)
                if structure_match:
                    date_str = structure_match.group(1).strip('()').strip()
                    parsed_date = parse_tracker_date(date_str, fallback_date=last_modified)
                    if parsed_date:
                        structure_candidates.append({
                            'key': key,
                            'date': parsed_date,
                            'type': 'structure'
                        })
                    continue

                finishing_match = finishing_pattern.search(filename)
                if finishing_match:
                    segment = finishing_match.group(1).upper()
                    tower_num = finishing_match.group(2)
                    date_str = finishing_match.group(3).strip('()').strip()
                    parsed_date = parse_tracker_date(date_str, fallback_date=last_modified)
                    if not parsed_date:
                        logger.warning(f"Could not parse date in finishing tracker filename: {key}")
                        continue

                    tower_id = f"{segment}_{tower_num}"
                    current = finishing_candidates.get(tower_id)
                    if current is None or parsed_date > current['date']:
                        finishing_candidates[tower_id] = {
                            'key': key,
                            'date': parsed_date,
                            'type': 'finishing',
                            'tower_id': tower_id,
                            'tower_name': f"{segment} Tower {tower_num}",
                            'display_name': build_finishing_display_name(segment, tower_num, parsed_date)
                        }

        selected_files = []
        if structure_candidates:
            selected_files.append(max(structure_candidates, key=lambda x: x['date']))
        selected_files.extend(
            [finishing_candidates[tower_id] for tower_id in expected_finishing_towers if tower_id in finishing_candidates]
        )

        if selected_files:
            st.write("Selected COS tracker files:")
            for info in selected_files:
                label = info.get('display_name') or info.get('tower_name', info['type'])
                st.write(f"- {label}: {info['key']}")
        else:
            st.warning("No matching structure or finishing tracker files were found in COS.")

        missing_towers = [tower_id for tower_id in expected_finishing_towers if tower_id not in finishing_candidates]
        if missing_towers:
            missing_labels = [tower_id.replace("_", " Tower ") for tower_id in missing_towers]
            st.warning(f"Missing finishing tracker(s) in IBM COS bucket: {', '.join(missing_labels)}")

        return selected_files

    except Exception as e:
        st.error(f"❌ Error fetching tracker files from COS: {str(e)}")
        logger.error(f"Error fetching tracker files from COS: {str(e)}")
        return []


def process_finishing_tracker_file(file_stream, filename, tower_name=None):
    try:
        workbook = openpyxl.load_workbook(file_stream)
        available_sheets = workbook.sheetnames
        st.write(f"Available sheets in {filename}: {', '.join(available_sheets)}")
        candidate_sheets = available_sheets

        inferred_tower_name = tower_name
        filename_only = filename.split('/')[-1]
        if not inferred_tower_name:
            match = re.search(r'(?i)\b(EWS|LIG)[\s_-]*(?:Tower|T)[\s_-]*([123])\b', filename_only)
            if match:
                inferred_tower_name = f"{match.group(1).upper()} Tower {match.group(2)}"

        if not inferred_tower_name:
            st.error(f"Cannot determine tower name from filename: {filename}")
            return (None, None)

        tower_match = re.search(r'(?i)\b(EWS|LIG)[\s_-]*(?:Tower|T)[\s_-]*([123])\b', inferred_tower_name)
        segment = tower_match.group(1).upper() if tower_match else None
        tower_num = tower_match.group(2) if tower_match else None

        target_sheet_name = f"{segment} T{tower_num} FINISHING." if segment and tower_num else None
        if target_sheet_name:
            matching_sheets = [
                sheet for sheet in available_sheets
                if normalize_activity_label(sheet) == normalize_activity_label(target_sheet_name)
            ]
            if matching_sheets:
                candidate_sheets = matching_sheets
                st.write(f"Using finishing sheet for {inferred_tower_name}: {', '.join(candidate_sheets)}")
            else:
                st.warning(
                    f"Expected finishing sheet '{target_sheet_name}' not found in {filename}. "
                    "Checking all sheets as fallback."
                )

        possible_sheet_names = []
        if segment and tower_num:
            possible_sheet_names = [
                target_sheet_name,
                f"{segment} TOWER {tower_num} FINISHING.",
                f"{segment} TOWER {tower_num}.",
                f"{segment} {tower_num} FINISHING.",
                f"{segment} {tower_num}.",
                f"TOWER {tower_num} FINISHING.",
                f"TOWER {tower_num}",
                f"{segment}TOWER{tower_num}FINISHING",
                f"{segment}TOWER{tower_num}",
                f"{segment}-{tower_num}",
                "Finish",
                "Finishing",
            ]

        def parse_excel_date_series(series):
            cleaned = series.replace(['', ' ', 'nan', 'NaN', 'NAT', 'nat', None], pd.NA)
            parsed = pd.Series(pd.NaT, index=series.index, dtype='datetime64[ns]')

            datetime_mask = cleaned.apply(lambda value: isinstance(value, (datetime, pd.Timestamp)))
            if datetime_mask.any():
                parsed.loc[datetime_mask] = pd.to_datetime(cleaned.loc[datetime_mask], errors='coerce')

            numeric_values = pd.to_numeric(cleaned, errors='coerce')
            numeric_mask = numeric_values.notna() & parsed.isna()
            if numeric_mask.any():
                numeric_dates = pd.to_datetime(
                    numeric_values.loc[numeric_mask],
                    unit='D',
                    origin='1899-12-30',
                    errors='coerce'
                )
                plausible_numeric = numeric_dates.between(pd.Timestamp('2020-01-01'), pd.Timestamp('2035-12-31'))
                parsed.loc[numeric_dates.index[plausible_numeric]] = numeric_dates.loc[plausible_numeric]

            text_mask = cleaned.notna() & parsed.isna()
            text_values = cleaned.loc[text_mask].astype(str).str.strip()
            date_formats = [
                "%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y",
                "%Y-%m-%d", "%Y/%m/%d",
                "%d-%m-%Y %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%d.%m.%Y %H:%M:%S",
                "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S",
                "%d-%b-%Y", "%d %b %Y", "%d-%B-%Y", "%d %B %Y",
                "%d-%m-%y", "%d/%m/%y", "%d.%m.%y",
            ]
            for date_format in date_formats:
                remaining = text_values.index[parsed.loc[text_values.index].isna()]
                if len(remaining) == 0:
                    break
                format_parsed = pd.to_datetime(text_values.loc[remaining], format=date_format, errors='coerce')
                parsed.loc[remaining] = parsed.loc[remaining].fillna(format_parsed)

            return parsed

        def sheet_priority(sheet_name):
            upper_name = sheet_name.strip().upper()
            for idx, candidate in enumerate(possible_sheet_names):
                if upper_name == candidate.upper():
                    return idx
            if segment and tower_num and segment in upper_name and tower_num in upper_name and "FINISH" in upper_name:
                return 100
            if "FINISH" in upper_name:
                return 200
            return 999

        ordered_sheets = sorted(candidate_sheets, key=sheet_priority)
        if not ordered_sheets:
            st.error(f"No usable finishing sheets found in file: {filename}")
            logger.error(f"No usable finishing sheets found in {filename}")
            return (None, None)

        combined_frames = []
        processed_sheet_names = []

        for sheet_name in ordered_sheets:
            file_stream.seek(0)
            raw_df = pd.read_excel(sheet_name=sheet_name, io=file_stream, header=None)

            if raw_df.empty or raw_df.shape[1] <= 11:
                logger.info(f"{filename} - sheet {sheet_name} skipped because it has only {raw_df.shape[1]} columns.")
                continue

            raw_df = raw_df.copy()
            raw_df["Actual Finish"] = parse_excel_date_series(raw_df.iloc[:, 11])

            logger.info(
                f"{filename} - sheet {sheet_name}: using activity column F and Actual Finish column L"
            )

            sheet_df = pd.DataFrame({
                "Module": None,
                "Floor": None,
                "Flat": None,
                "Activity ID": None,
                "Activity Name": raw_df.iloc[:, 5],
                "Actual Finish": raw_df["Actual Finish"],
                "Source Sheet": sheet_name,
            })
            sheet_df = sheet_df.dropna(subset=['Activity Name'])
            sheet_df['Activity Name'] = sheet_df['Activity Name'].astype(str).str.strip()
            sheet_df = sheet_df[sheet_df['Activity Name'].str.lower().ne('nan')]
            sheet_df = sheet_df[sheet_df['Actual Finish'].notna()]

            if sheet_df.empty:
                logger.info(f"{filename} - sheet {sheet_name} produced no usable finishing rows.")
                continue

            combined_frames.append(sheet_df)
            processed_sheet_names.append(sheet_name)

        if not combined_frames:
            st.error(f"No usable finishing data found in any analyzed sheet of file: {filename}")
            logger.error(f"No usable finishing data found in any analyzed sheet of {filename}")
            return (None, None)

        df = pd.concat(combined_frames, ignore_index=True)
        st.write(
            f"Processed finishing tracker for {inferred_tower_name}: {len(df)} rows "
            f"from sheets: {', '.join(processed_sheet_names)}"
        )
        st.write(make_streamlit_safe_df(df.head()))
        return (df, inferred_tower_name)

    except Exception as e:
        st.error(f"Error loading finishing tracker {filename}: {str(e)}")
        logger.exception(f"Error loading finishing tracker {filename}")
        return (None, None)

# Function to get access token for WatsonX API
def get_access_token(api_key):
    try:
        headers = {"Content-Type": "application/x-www-form-urlencoded"}
        data = {
            "grant_type": "urn:ibm:params:oauth:grant-type:apikey",
            "apikey": api_key
        }
        response = requests.post("https://iam.cloud.ibm.com/identity/token", headers=headers, data=data)
        if response.status_code == 200:
            return response.json().get("access_token")
        else:
            logger.error(f"Failed to get access token: {response.status_code} - {response.text}")
            return None
    except Exception as e:
        logger.error(f"Error getting access token: {str(e)}")
        return None

#Slab code
def GetSlabReport():
    st.write("📊 Fetching latest EWS LIG Structure Work Tracker...")
    try:
        missing = get_missing_cos_config()
        if missing:
            message = f"Missing COS configuration: {', '.join(missing)}"
            st.error(f"ERROR: {message}. Please set these env values in .env.")
            logger.error(message)
            return "No Data Found"

        logger.info("Initializing IBM COS client...")
        cos_client = initialize_cos_client()
        bucket_name = COS_BUCKET or "projectreportnew"

        # ✅ List all Excel files containing “Structure Work Tracker”
        response = cos_client.list_objects_v2(Bucket=bucket_name)
        files = [
            obj['Key'] for obj in response.get('Contents', [])
            if obj['Key'].endswith('.xlsx') and "Structure Work Tracker" in obj['Key']
        ]

        if not files:
            st.error("❌ No tracker Excel files found in bucket.")
            return "No Data Found"

        # ✅ Try to extract a date from each filename
        pattern = re.compile(r"(?i)Structure\s*Work\s*Tracker.*?(\d{1,2}[-_/]\d{1,2}[-_/]\d{2,4})")
        date_formats = ["%d-%m-%Y", "%d_%m_%Y", "%d-%m-%y", "%Y-%m-%d"]

        file_dates = []
        for f in files:
            match = pattern.search(f)
            if match:
                date_str = match.group(1)
                for fmt in date_formats:
                    try:
                        date_parsed = datetime.strptime(date_str, fmt)
                        file_dates.append((f, date_parsed))
                        break
                    except ValueError:
                        continue

        # ✅ If no valid date found, fall back to last modified date
        if not file_dates:
            response = cos_client.list_objects_v2(Bucket=bucket_name)
            file_dates = [
                (obj['Key'], obj['LastModified'])
                for obj in response.get('Contents', [])
                if obj['Key'] in files
            ]

        # ✅ Pick the latest file
        latest_file, latest_date = max(file_dates, key=lambda x: x[1])
        st.success(f"✅ Using latest tracker file: {latest_file} (Date: {latest_date.date()})")
        logger.info(f"Using latest tracker file: {latest_file}")

        # ✅ Read and process latest file
        response = cos_client.get_object(Bucket=bucket_name, Key=latest_file)
        tracker_bytes = io.BytesIO(response['Body'].read())

        if st.session_state.ignore_year and st.session_state.ignore_month:
            st.session_state.slabreport = ProcessEWS_LIG(tracker_bytes, st.session_state.ignore_year, st.session_state.ignore_month)
        else:
            st.session_state.slabreport = ProcessEWS_LIG(tracker_bytes, st.session_state.ignore_year, st.session_state.ignore_month)

        return st.session_state.slabreport

    except Exception as e:
        st.error(f"❌ Error fetching latest tracker: {e}")
        logger.exception(f"Error in GetSlabReport: {e}")
        return "No Data Found"



# WatsonX Prompt Generation
def generatePrompt(json_datas, tower_name):
    try:
        logger.info(f"Using deterministic fallback totals for {tower_name}")
        return generate_fallback_totals(json_datas)
    
    except Exception as e:
        logger.error(f"Error in WatsonX API call for {tower_name}: {str(e)}")
        return generate_fallback_totals(json_datas)

# Fallback Total Calculation
def generate_fallback_totals(count_table):
    try:
        if not isinstance(count_table, pd.DataFrame):
            logger.error("Fallback method received invalid input: not a DataFrame")
            return json.dumps([
                {"Category": "Civil Works", "Activities": [
                    {"Activity Name": "Concreting", "Total": 0},
                    {"Activity Name": "Shuttering", "Total": 0},
                    {"Activity Name": "Reinforcement", "Total": 0},
                    {"Activity Name": "De-Shuttering", "Total": 0}
                ]},
                {"Category": "MEP Works", "Activities": [
                    {"Activity Name": "Plumbing Works", "Total": 0},
                    {"Activity Name": "Min. count of UP-First Fix and CP-First Fix", "Total": 0},
                    {"Activity Name": "Slab Conducting", "Total": 0},
                    {"Activity Name": "Wall Conducting", "Total": 0},
                    {"Activity Name": "Wiring & Switch Socket", "Total": 0}
                ]},
                {"Category": "Interior Finishing Works", "Activities": [
                    {"Activity Name": "Floor Tile", "Total": 0},
                    {"Activity Name": "POP & Gypsum Plaster", "Total": 0},
                    {"Activity Name": "Wall Tile", "Total": 0},
                    {"Activity Name": "Waterproofing - Sunken", "Total": 0}
                ]},
                {"Category": "External Development Activities", "Activities": [
                    {"Activity Name": "Granular Sub-base", "Total": 0},
                    {"Activity Name": "Kerb Stone", "Total": 0},
                    {"Activity Name": "Rain Water/Storm Line", "Total": 0},
                    {"Activity Name": "Saucer drain/Paver block", "Total": 0},
                    {"Activity Name": "Sewer Line", "Total": 0},
                    {"Activity Name": "Stamp Concrete", "Total": 0},
                    {"Activity Name": "Storm Line", "Total": 0},
                    {"Activity Name": "WMM", "Total": 0}
                ]}
            ], indent=2)

        categories = {
            "Civil Works": [
                "Concreting", "Shuttering", "Reinforcement", "De-Shuttering"
            ],
            "MEP Works": [
                "Plumbing Works", "Min. count of UP-First Fix and CP-First Fix",
                "Slab Conducting", "Wall Conducting", "Wiring & Switch Socket"
            ],
            "Interior Finishing Works": [
                "Floor Tiling", "POP & Gypsum Plaster", "Wall Tiling", "Waterproofing"
            ],
        }

        result = []
        for category, activities in categories.items():
            category_data = {"Category": category, "Activities": []}
            
            if category == "MEP Works":
                for activity in activities:
                    if activity == "Min. count of UP-First Fix and CP-First Fix":
                        combined_count = count_table.loc["UP-First Fix and CP-First Fix", "Count"] if "UP-First Fix and CP-First Fix" in count_table.index else 0
                        total = combined_count
                    else:
                        total = count_table.loc[activity, "Count"] if activity in count_table.index else 0
                    category_data["Activities"].append({
                        "Activity Name": activity,
                        "Total": int(total) if pd.notna(total) else 0
                    })
            else:
                for activity in activities:
                    total = count_table.loc[activity, "Count"] if activity in count_table.index else 0
                    category_data["Activities"].append({
                        "Activity Name": activity,
                        "Total": int(total) if pd.notna(total) else 0
                    })
            
            result.append(category_data)

        return json.dumps(result, indent=2)
    except Exception as e:
        logger.error(f"Error in fallback total calculation: {str(e)}")
        st.error(f"Error in fallback total calculation: {str(e)}")
        return json.dumps([
            {"Category": "Civil Works", "Activities": [
                {"Activity Name": "Concreting", "Total": 0},
                {"Activity Name": "Shuttering", "Total": 0},
                {"Activity Name": "Reinforcement", "Total": 0},
                {"Activity Name": "De-Shuttering", "Total": 0}
            ]},
            {"Category": "MEP Works", "Activities": [
                {"Activity Name": "Plumbing Works", "Total": 0},
                {"Activity Name": "Min. count of UP-First Fix and CP-First Fix", "Total": 0},
                {"Activity Name": "Slab Conducting", "Total": 0},
                {"Activity Name": "Wall Conducting", "Total": 0},
                {"Activity Name": "Wiring & Switch Socket", "Total": 0}
            ]},
            {"Category": "Interior Finishing Works", "Activities": [
                {"Activity Name": "Floor Tile", "Total": 0},
                {"Activity Name": "POP & Gypsum Plaster", "Total": 0},
                {"Activity Name": "Wall Tile", "Total": 0},
                {"Activity Name": "Waterproofing - Sunken", "Total": 0}
            ]},
            {"Category": "External Development Activities", "Activities": [
                {"Activity Name": "Granular Sub-base", "Total": 0},
                {"Activity Name": "Kerb Stone", "Total": 0},
                {"Activity Name": "Rain Water/Storm Line", "Total": 0},
                {"Activity Name": "Saucer drain/Paver block", "Total": 0},
                {"Activity Name": "Sewer Line", "Total": 0},
                {"Activity Name": "Stamp Concrete", "Total": 0},
                {"Activity Name": "Storm Line", "Total": 0},
                {"Activity Name": "WMM", "Total": 0}
            ]}
        ], indent=2)

def build_categorized_activity_response(activity_counts):
    # Keys must match Asite canonical names used in extract_cos_activity_counts
    categories = {
        "Civil Works": [
            "Concreting", "Shuttering", "Reinforcement", "De-Shuttering"
        ],
        "MEP Works": [
            "Plumbing Works", "UP-First Fix and CP-First Fix",
            "Slab Conducting", "Wall Conducting", "Wiring & Switch Socket"
        ],
        "Interior Finishing Works": [
            "Floor Tile", "POP & Gypsum Plaster", "Wall Tile", "Waterproofing - Sunken"
        ],
    }

    result = []
    for category, activities in categories.items():
        category_data = {"Category": category, "Activities": []}
        for activity in activities:
            display_name = activity
            if activity == "UP-First Fix and CP-First Fix":
                display_name = "Min. count of UP-First Fix and CP-First Fix"
            category_data["Activities"].append({
                "Activity Name": display_name,
                "Total": int(activity_counts.get(activity, 0) or 0),
            })
        result.append(category_data)
    return result

def getTotal(ai_data):
    st.write(ai_data)
    try:
        if isinstance(ai_data, str):
            ai_data = json.loads(ai_data)
            
        if not isinstance(ai_data, list):
            logger.error(f"AI data is not a list: {ai_data}")
            return {}

        totals = {}
        for category_data in ai_data:
            if isinstance(category_data, dict) and 'Activities' in category_data:
                for activity in category_data['Activities']:
                    if isinstance(activity, dict) and 'Total' in activity:
                        activity_name = activity['Activity Name']
                        total = activity['Total']
                        totals[activity_name] = int(total) if isinstance(total, (int, float)) and pd.notna(total) else 0
                    else:
                        logger.warning(f"Invalid activity format: {activity}")
            else:
                logger.warning(f"Invalid category format: {category_data}")
        return totals
    except Exception as e:
        logger.error(f"Error parsing AI data: {str(e)}")
        st.error(f"Error parsing AI data: {str(e)}")
        return {}


def normalize_activity_label(label):
    return re.sub(r'[^a-z0-9]+', '', str(label or '').strip().lower())


def normalize_report_activity_name(activity_name):
    normalized = normalize_activity_label(activity_name)
    
    report_aliases = {
        # Civil
        "concreting":                           "Concreting",
        "shuttering":                           "Shuttering",
        "reinforcement":                        "Reinforcement",
        "deshuttering":                         "De-Shuttering",
        # MEP
        "plumbingworks":                        "Plumbing Works",
        "slabconducting":                       "Slab Conducting",
        "wallconducting":                       "Wall Conducting",
        "wallconduting":                        "Wall Conducting",
        "elfirstfix":                           "Wall Conducting",
        "el1stfix":                             "Wall Conducting",
        "wiringswitchsocket":                   "Wiring & Switch Socket",
        "wiringandswitchsocket":                "Wiring & Switch Socket",
        "elsecondfix":                          "Wiring & Switch Socket",
        "el2ndfix":                             "Wiring & Switch Socket",
        "eisecondfix":                          "Wiring & Switch Socket",
        "ei2ndfix":                             "Wiring & Switch Socket",
        "upfirstfixandcpfirstfix":              "Plumbing Works",
        "mincountofupfirstfixandcpfirstfix":    "Plumbing Works",
        # Finishing — ONLY exact normalized matches allowed for tiles
        "floortile":                            "Floor Tile",
        "floortiling":                          "Floor Tile",
        "walltile":                             "Wall Tile",
        "walltiling":                           "Wall Tile",
        "popgypsumplaster":                     "POP & Gypsum Plaster",
        "popandgypsumplaster":                  "POP & Gypsum Plaster",
        "gypsumpoppunning":                     "POP & Gypsum Plaster",
        "gypsumandpoppunning":                  "POP & Gypsum Plaster",
        "gypsumanpoppunning":                   "POP & Gypsum Plaster",
        "waterproofingsunken":                  "Waterproofing - Sunken",
        "waterproofingworks":                   "Waterproofing - Sunken",
        "waterproofingwork":                    "Waterproofing - Sunken",
        "waterproofworks":                      "Waterproofing - Sunken",
        "waterproofing":                        "Waterproofing - Sunken",
        "waterproofingworksunken":              "Waterproofing - Sunken",
        "waterproofingworksforsunken":          "Waterproofing - Sunken",
        "waterproofingsunkenarea":              "Waterproofing - Sunken",
        "waterproofingforsunken":               "Waterproofing - Sunken",
        # Door
        "installationofdoors":                  "Door/Window Frame",
        "installationdoors":                    "Door/Window Frame",
        "doorwindowframe":                      "Door/Window Frame",
        "doorwindowshutter":                    "Door/Window Shutter",
    }

    if normalized in report_aliases:
        return report_aliases[normalized]

    # Limited fuzzy fallbacks — NO fuzzy for Wall Tile or Floor Tile
    if re.fullmatch(r'waterproof\w*', normalized):
        return "Waterproofing - Sunken"
    if "gypsum" in normalized:
        return "POP & Gypsum Plaster"
    if re.search(r'switch|socket|wiring', normalized):
        return "Wiring & Switch Socket"
    if re.fullmatch(r'wallconduct\w*', normalized):
        return "Wall Conducting"
    if re.fullmatch(r'slabconduct\w*', normalized):
        return "Slab Conducting"
    if "plumbing" in normalized:
        return "Plumbing Works"

    # Unrecognized — return as-is, caller will skip it
    return activity_name


def extract_cos_activity_counts(tower_df, tower_name):
    try:
        expected_activities = [
            "Concreting", "Shuttering", "Reinforcement", "De-Shuttering",
            "Slab Conducting", "Wall Conducting", "Wiring & Switch Socket",
            "Plumbing Works", "Door/Window Frame",
            "Floor Tile", "Wall Tile",
            "POP & Gypsum Plaster", "Waterproofing - Sunken",
            "UP-First Fix and CP-First Fix",
        ]
        counts = {activity: 0 for activity in expected_activities}

        if tower_df is None or not isinstance(tower_df, pd.DataFrame) or tower_df.empty:
            logger.warning(f"{tower_name}: COS finishing DataFrame is empty or invalid.")
            return counts

        working_df = tower_df.copy()
        if 'Actual Finish' in working_df.columns:
            working_df['Actual Finish'] = pd.to_datetime(working_df['Actual Finish'], errors='coerce')
            working_df = working_df[working_df['Actual Finish'].notna()].copy()

        if working_df.empty:
            logger.warning(f"{tower_name}: No completed COS rows found after Actual Finish filtering.")
            return counts

        activity_occurrences = {}

        for _, row in working_df.iterrows():
            raw_activity = str(row.get('Activity Name') or '').strip()
            mapped_activity = normalize_report_activity_name(raw_activity)

            # Log every Wall Tile / Floor Tile hit for debugging
            if mapped_activity in ("Floor Tile", "Wall Tile"):
                logger.info(f"{tower_name}: raw='{raw_activity}' → mapped='{mapped_activity}'")

            # Unrecognized — mapped_activity equals raw_activity, skip unless UP/CP split
            if mapped_activity == raw_activity:
                activity_key = normalize_activity_label(raw_activity)
                if activity_key == "upfirstfix":
                    activity_occurrences["UP_FIRST_FIX"] = activity_occurrences.get("UP_FIRST_FIX", 0) + 1
                elif activity_key == "cpfirstfix":
                    activity_occurrences["CP_FIRST_FIX"] = activity_occurrences.get("CP_FIRST_FIX", 0) + 1
                # All other unrecognized activities are skipped entirely
                continue

            activity_occurrences[mapped_activity] = activity_occurrences.get(mapped_activity, 0) + 1

        # Write recognized Asite-name activities directly
        for activity in [
            "Concreting", "Shuttering", "Reinforcement", "De-Shuttering",
            "Slab Conducting", "Wall Conducting", "Wiring & Switch Socket",
            "Plumbing Works", "Door/Window Frame",
            "Floor Tile", "Wall Tile",
            "POP & Gypsum Plaster", "Waterproofing - Sunken",
        ]:
            counts[activity] = activity_occurrences.get(activity, 0)

        # UP/CP combined logic
        up_count       = activity_occurrences.get("UP_FIRST_FIX", 0)
        cp_count       = activity_occurrences.get("CP_FIRST_FIX", 0)
        combined_count = activity_occurrences.get("Plumbing Works", 0)

        if up_count and cp_count:
            counts["UP-First Fix and CP-First Fix"] = min(up_count, cp_count)
        elif combined_count:
            counts["UP-First Fix and CP-First Fix"] = combined_count
        else:
            counts["UP-First Fix and CP-First Fix"] = max(up_count, cp_count)

        if counts["Plumbing Works"] == 0 and counts["UP-First Fix and CP-First Fix"] > 0:
            counts["Plumbing Works"] = counts["UP-First Fix and CP-First Fix"]

        logger.info(f"{tower_name}: Final COS activity counts: {counts}")
        return counts

    except Exception as e:
        logger.error(f"{tower_name}: Error extracting COS activity counts: {str(e)}")
        st.error(f"{tower_name}: Error extracting COS activity counts: {str(e)}")
        return {}


def display_activity_count():
    try:
        st.write("Starting display_activity_count function")

        # All Asite canonical names
        specific_activities = [
            "Concreting", "Shuttering", "Reinforcement", "De-Shuttering",
            "Plumbing Works", "Slab Conducting", "Wall Conducting", "Wiring & Switch Socket",
            "Floor Tile", "Wall Tile", "POP & Gypsum Plaster", "Waterproofing - Sunken",
        ]
        all_activities = specific_activities + ["UP-First Fix and CP-First Fix"]

        category_mapping = {
            "Concreting":               "Civil Works",
            "Shuttering":               "Civil Works",
            "Reinforcement":            "Civil Works",
            "De-Shuttering":            "Civil Works",
            "Plumbing Works":           "MEP Works",
            "Slab Conducting":          "MEP Works",
            "Wall Conducting":          "MEP Works",
            "Wiring & Switch Socket":   "MEP Works",
            "UP-First Fix and CP-First Fix": "MEP Works",
            "Floor Tile":               "Interior Finishing Works",
            "Wall Tile":                "Interior Finishing Works",
            "POP & Gypsum Plaster":     "Interior Finishing Works",
            "Waterproofing - Sunken":   "Interior Finishing Works",
        }

        structure_analysis = st.session_state.get('structure_analysis')
        if structure_analysis is None or not isinstance(structure_analysis, pd.DataFrame) or structure_analysis.empty:
            st.error("structure_analysis is missing or empty.")
            return

        required_columns = ['tower_name', 'activityName', 'ClosedChecklistCount']
        missing_columns = [col for col in required_columns if col not in structure_analysis.columns]
        if missing_columns:
            st.error(f"Missing required columns in structure_analysis: {missing_columns}")
            return

        # Extract Asite closed checklist counts using Asite activity names
        asite_closed_counts = {}
        for tower in structure_analysis['tower_name'].unique():
            tower_data = structure_analysis[structure_analysis['tower_name'] == tower]
            tower_counts = {}
            for activity in specific_activities:
                activity_data = tower_data[tower_data['activityName'] == activity]
                tower_counts[activity] = int(activity_data['ClosedChecklistCount'].iloc[0]) if not activity_data.empty else 0
            asite_closed_counts[tower] = tower_counts

        if 'ai_response' not in st.session_state or not isinstance(st.session_state.ai_response, dict):
            st.session_state.ai_response = {}
        st.session_state.cos_activity_counts = {}

        count_tables = {}
        cos_finishing_towers = st.session_state.get('cos_finishing_towers', {})
        target_towers = ['EWS Tower 1', 'EWS Tower 2', 'EWS Tower 3',
                         'LIG Tower 1', 'LIG Tower 2', 'LIG Tower 3']

        for tower in target_towers:
            tower_segment = tower.split()[0]
            tower_num = tower.split()[-1]
            finishing_key = f"{tower_segment}_{tower_num}"
            finishing_info = cos_finishing_towers.get(finishing_key)

            cos_counts = {}
            if finishing_info and isinstance(finishing_info, dict):
                cos_df_tower = finishing_info.get('df')
                cos_counts = extract_cos_activity_counts(cos_df_tower, tower)
                st.write(f"COS counts for {tower}: {cos_counts}")

            if not cos_counts and tower in asite_closed_counts:
                st.warning(f"No COS finishing counts for {tower}. Falling back to Asite counts.")
                cos_counts = {a: asite_closed_counts[tower].get(a, 0) for a in all_activities}

            if cos_counts:
                st.session_state.cos_activity_counts[tower] = cos_counts
                count_table = pd.DataFrame({
                    'Closed_Checklist_Unfiltered': [cos_counts.get(a, 0) for a in all_activities],
                    'Closed_Checklist_Filtered':   [cos_counts.get(a, 0) for a in all_activities],
                }, index=all_activities)
                count_tables[tower] = count_table

        if not count_tables:
            st.error("No count tables generated for any towers.")
            return

        for tname, count_table in count_tables.items():
            try:
                source_counts = st.session_state.cos_activity_counts.get(tname, {})
                ai_data = build_categorized_activity_response(source_counts)
                st.session_state.ai_response[tname] = ai_data

                totals_mapping = getTotal(ai_data)

                display_df = count_table.reset_index().rename(columns={'index': 'Activity Name'})
                display_df['Total'] = display_df['Activity Name'].map(
                    lambda x: totals_mapping.get(x, 0)
                )
                display_df['Category'] = display_df['Activity Name'].map(category_mapping)
                display_df['Asite_Closed_Count'] = display_df['Activity Name'].map(
                    lambda x: asite_closed_counts.get(tname, {}).get(x, 0)
                )
                st.write(f"Activity Count with Totals for {tname}:")
                st.dataframe(display_df[['Activity Name', 'Closed_Checklist_Filtered',
                                         'Total', 'Asite_Closed_Count', 'Category']])
            except Exception as tower_error:
                st.error(f"Error processing {tname}: {str(tower_error)}")
                st.code(traceback.format_exc())

        st.write("display_activity_count completed successfully")

    except Exception as main_error:
        st.error(f"MAIN ERROR in display_activity_count: {str(main_error)}")
        st.code(traceback.format_exc())


# Combined function for Initialize and Fetch Data
async def initialize_and_fetch_data(email, password):
    # Quick check: if data is already loaded, offer to skip
    if all(key in st.session_state for key in ['structure_location_data', 'structure_activity_data', 'cos_finishing_towers']):
        st.info("ℹ️ Data already loaded.")
        return True
    
    with st.spinner("Starting initialization and data fetching process..."):
        # Step 1: Login
        if not email or not password:
            st.sidebar.error("Please provide both email and password!")
            return False
        try:
            st.sidebar.write("Logging in...")
            session_id = await login_to_asite(email, password)
            if not session_id:
                st.sidebar.error("Login failed!")
                return False
            st.sidebar.success("Login successful!")
        except Exception as e:
            st.sidebar.error(f"Login failed: {str(e)}")
            return False

        # Step 2: Get Workspace ID
        try:
            st.sidebar.write("Fetching Workspace ID...")
            await GetWorkspaceID()
            st.sidebar.success("Workspace ID fetched successfully!")
        except Exception as e:
            st.sidebar.error(f"Failed to fetch Workspace ID: {str(e)}")
            return False

        # Step 3: Get Project IDs
        try:
            st.sidebar.write("Fetching Project IDs...")
            await GetProjectId()
            st.sidebar.success("Project IDs fetched successfully!")
        except Exception as e:
            st.sidebar.error(f"Failed to fetch Project IDs: {str(e)}")
            return False

        # Step 4: Get All Data (Structure only)
        try:
            st.sidebar.write("Fetching All Data...")
            finishing_data, EWSLIG_structure = await GetAllDatas_EWSLIG_Style()
            st.session_state.ews_lig_finishing = finishing_data
            st.session_state.EWSLIG_structure = EWSLIG_structure
            st.sidebar.success("All Data fetched successfully!")
        except Exception as e:
            st.sidebar.error(f"Failed to fetch All Data: {str(e)}")
            return False

        # Step 5: Get Activity Data
        try:
            st.sidebar.write("Fetching Activity Data...")
            finishing_activity_data, structure_activity_data = await Get_Activity_EWSLIG_Style()
            st.session_state.finishing_activity_data = finishing_activity_data
            st.session_state.structure_activity_data = structure_activity_data
            st.sidebar.success("Activity Data fetched successfully!")
        except Exception as e:
            st.sidebar.error(f"Failed to fetch Activity Data: {str(e)}")
            return False

        # Step 6: Get Location/Module Data
        try:
            st.sidebar.write("Fetching Location/Module Data...")
            finishing_location_data, structure_location_data = await Get_Location_EWSLIG_Style()
            st.session_state.finishing_location_data = finishing_location_data
            st.session_state.structure_location_data = structure_location_data
            st.sidebar.success("Location/Module Data fetched successfully!")
        except Exception as e:
            st.sidebar.error(f"Failed to fetch Location/Module Data: {str(e)}")
            return False

        # Step 7: Fetch COS Files
        try:
            st.sidebar.write("Fetching COS files from EWS-LIG folder...")
            tracker_files = get_cos_tracker_files()
            st.session_state.file_key = tracker_files
            st.session_state.cos_finishing_towers = {}
            st.session_state.cos_finishing_files = tracker_files
            if tracker_files:
                st.success(f"Found {len(tracker_files)} tracker file(s) in COS storage")
                try:
                    cos_client = initialize_cos_client()
                    if not cos_client:
                        st.error("Failed to initialize COS client during file fetch")
                        logger.error("COS client initialization failed during file fetch")
                        return False

                    for tracker_info in tracker_files:
                        file_key = tracker_info['key']
                        st.write(f"Processing file: {file_key}")
                        response = cos_client.get_object(Bucket=COS_BUCKET, Key=file_key)
                        file_bytes = io.BytesIO(response['Body'].read())
                        st.write("File fetched successfully. Processing sheets...")

                        if tracker_info.get('type') == 'structure':
                            results = process_file(file_bytes, file_key)
                            st.write(f"Processing results: {len(results)} sheets processed")
                            for df, sheet_name in results:
                                if df is not None and sheet_name == "Revised Baseline 45daysNGT+Rai":
                                    st.session_state.cos_df_Revised_Baseline_45daysNGT_Rai = df
                                    st.session_state.cos_tname_Revised_Baseline_45daysNGT_Rai = "Revised Baseline 45daysNGT+Rai"
                                    st.write(f"Processed Data for {sheet_name} - {len(df)} rows:")
                                    st.write(make_streamlit_safe_df(df.head()))
                        elif tracker_info.get('type') == 'finishing':
                            df, tower_name = process_finishing_tracker_file(
                                file_bytes,
                                file_key,
                                tower_name=tracker_info.get('tower_name')
                            )
                            if df is not None and tower_name:
                                tower_id = tracker_info.get('tower_id', tower_name)
                                st.session_state.cos_finishing_towers[tower_id] = {
                                    'df': df,
                                    'tname': tower_name,
                                    'rows': len(df),
                                    'file': file_key
                                }
                                st.write(f"Processed finishing tracker for {tower_name} - {len(df)} rows:")
                                st.write(make_streamlit_safe_df(df.head()))
                            else:
                                st.warning(f"No data processed for finishing tracker {file_key}.")
                except Exception as e:
                    st.error(f"Error loading tracker files from cloud storage: {str(e)}")
                    logger.error(f"Error loading tracker files: {str(e)}")
                    return False
            else:
                st.warning("No expected tracker files available in the EWS LIG P4 folder of the COS bucket.")
                return False
        except Exception as e:
            st.sidebar.error(f"Failed to fetch COS files: {str(e)}")
            logger.error(f"Failed to fetch COS files: {str(e)}")
            return False

    st.sidebar.success("All steps completed successfully!")
    return True


def generate_consolidated_Checklist_excel(structure_analysis=None, activity_counts=None):
    try:
        if structure_analysis is None:
            structure_analysis = st.session_state.get('structure_analysis')
        if activity_counts is None:
            activity_counts = st.session_state.get('ai_response', {})

        if structure_analysis is None or not isinstance(structure_analysis, pd.DataFrame):
            st.error("No valid structure_analysis available.")
            return None
        if not activity_counts:
            st.error("No activity counts available.")
            return None

        expected_columns = ['tower_name', 'activityName', 'ClosedChecklistCount']
        missing = [c for c in expected_columns if c not in structure_analysis.columns]
        if missing:
            st.error(f"Missing columns in structure_analysis: {missing}")
            return None

        # Flatten ai_response — activity names already Asite canonical after build_categorized_activity_response
        transformed_activity_counts = []
        if isinstance(activity_counts, dict):
            for tower, categories_data in activity_counts.items():
                for category_data in categories_data:
                    for activity_data in category_data.get("Activities", []):
                        raw_name = activity_data.get("Activity Name", "")
                        # "Min. count of UP-First Fix..." → "Plumbing Works"
                        canonical = normalize_report_activity_name(raw_name)
                        transformed_activity_counts.append({
                            "tower":           tower,
                            "activity":        canonical,
                            "completed_count": activity_data.get("Total", 0),
                        })

        # Categories use Asite names throughout
        categories = {
            "Civil Works": [
                "Concreting", "Shuttering", "Reinforcement", "De-Shuttering"
            ],
            "MEP Works": [
                "Plumbing Works", "Slab Conducting", "Wall Conducting", "Wiring & Switch Socket"
            ],
            "Interior Finishing Works": [
                "Floor Tile", "POP & Gypsum Plaster", "Wall Tile", "Waterproofing - Sunken"
            ],
        }

        # cos_to_asite_mapping is now identity — both sides are Asite names
        cos_to_asite_mapping = {a: a for cat in categories.values() for a in cat}

        towers = ["EWS Tower 1", "EWS Tower 2", "EWS Tower 3",
                  "LIG Tower 1", "LIG Tower 2", "LIG Tower 3"]

        if "slabreport" not in st.session_state or not st.session_state.slabreport:
            GetSlabReport()

        try:
            if isinstance(st.session_state.slabreport, str) and st.session_state.slabreport == "No Data Found":
                st.error("No slab report data found.")
                return None
            slab_data = (json.loads(st.session_state.slabreport)
                         if isinstance(st.session_state.slabreport, str)
                         else st.session_state.slabreport)
            if not isinstance(slab_data, list):
                st.error("Invalid slab report format.")
                return None
        except Exception as e:
            st.error(f"Error parsing slab report: {e}")
            return None

        expected_towers = ["EWST1", "EWST2", "EWST3", "LIGT1", "LIGT2", "LIGT3"]
        tower_mapping = {
            "EWST1": "EWS Tower 1", "EWST2": "EWS Tower 2", "EWST3": "EWS Tower 3",
            "LIGT1": "LIG Tower 1", "LIGT2": "LIG Tower 2", "LIGT3": "LIG Tower 3",
        }
        tower_counts_raw = {t: 0 for t in expected_towers}
        for entry in slab_data:
            t = entry.get("Tower")
            c = entry.get("Slab Count", 0)
            if t in tower_counts_raw:
                tower_counts_raw[t] = int(c) if isinstance(c, (int, float)) and not pd.isna(c) else 0

        tracker_completed_mapping = {
            tower_mapping[t]: cnt for t, cnt in tower_counts_raw.items()
        }
        st.write("DEBUG - Tracker Completed Mapping:", tracker_completed_mapping)

        civil_activities = {"Concreting", "Shuttering", "Reinforcement", "De-Shuttering"}

        consolidated_rows = []
        for tower in towers:
            tower_completed_from_tracker = tracker_completed_mapping.get(tower, 0)

            for category, activities in categories.items():
                for activity in activities:
                    asite_activity = cos_to_asite_mapping.get(activity, activity)

                    if activity in civil_activities or activity == "Slab Conducting":
                        completed_work = tower_completed_from_tracker
                    else:
                        completed_work = 0
                        for item in transformed_activity_counts:
                            if item["tower"] == tower and item["activity"] == activity:
                                completed_work = item["completed_count"]
                                break

                    tower_asite = structure_analysis[structure_analysis['tower_name'] == tower]
                    activity_row = tower_asite[tower_asite['activityName'] == asite_activity]
                    closed_checklist = int(activity_row['ClosedChecklistCount'].iloc[0]) if not activity_row.empty else 0

                    open_missing = max(0, completed_work - closed_checklist) if completed_work > 0 else 0

                    consolidated_rows.append({
                        "Tower":                        tower,
                        "Category":                     category,
                        "Activity Name":                asite_activity,
                        "Completed Work*(Count of Flat)": completed_work,
                        "In progress":                  0,
                        "Closed checklist":             closed_checklist,
                        "Open/Missing check list":      open_missing,
                    })

        df = pd.DataFrame(consolidated_rows)
        if df.empty:
            st.warning("No data to generate checklist.")
            return None
        df.sort_values(by=["Tower", "Category"], inplace=True)

        # --- Excel writing (unchanged from original) ---
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet("Consolidated Checklist")
        header_format = workbook.add_format({'bold': True, 'bg_color': '#D3D3D3', 'border': 1,
                                              'align': 'center', 'valign': 'vcenter'})
        total_format  = workbook.add_format({'bold': True, 'bg_color': '#FFDAB9', 'border': 1, 'align': 'center'})
        cell_format   = workbook.add_format({'border': 1})
        headers_row   = ["Activity Name", "Completed", "In progress",
                         "Closed checklist", "Open/Missing check list"]

        current_row = 0
        for tower, tower_group in df.groupby('Tower'):
            for category, cat_group in tower_group.groupby('Category'):
                worksheet.merge_range(current_row, 0, current_row, 4,
                                      f"{tower} {category} Checklist Status", header_format)
                for j, h in enumerate(headers_row):
                    worksheet.write(current_row + 1, j, h, header_format)
                data_start = current_row + 2
                total_pending = 0
                for idx, (_, row) in enumerate(cat_group.iterrows()):
                    r = data_start + idx
                    worksheet.write(r, 0, row["Activity Name"],                     cell_format)
                    worksheet.write(r, 1, row["Completed Work*(Count of Flat)"],    cell_format)
                    worksheet.write(r, 2, row["In progress"],                       cell_format)
                    worksheet.write(r, 3, row["Closed checklist"],                  cell_format)
                    worksheet.write(r, 4, row["Open/Missing check list"],           cell_format)
                    total_pending += row["Open/Missing check list"]
                min_rows = 5
                for empty in range(len(cat_group), min_rows):
                    for col in range(5):
                        worksheet.write(data_start + empty, col, "", cell_format)
                total_row = data_start + min_rows
                worksheet.merge_range(total_row, 0, total_row, 3, "Total pending check list", total_format)
                worksheet.write(total_row, 4, total_pending, total_format)
                current_row += min_rows + 3  # header(1) + subheader(1) + rows(5) + total(1) + gap(1)

        for col in range(5):
            worksheet.set_column(col, col, 22)

        # Sheet 2
        current_month = datetime.now().strftime("%B")
        ws2 = workbook.add_worksheet(f"Checklist {current_month}")
        sheet2_headers = ["Site",
                          "Total of Missing & Open Checklist-Civil",
                          "Total of Missing & Open Checklist-MEP",
                          "Total of Missing & Open Checklist-Interior Finishing",
                          "TOTAL"]
        for col, h in enumerate(sheet2_headers):
            ws2.write(0, col, h, header_format)

        category_type_map = {
            "Civil Works": "Civil", "MEP Works": "MEP",
            "Interior Finishing Works": "Interior",
        }
        summary_data = {}
        for _, row in df.iterrows():
            tower, category, open_missing = row["Tower"], row["Category"], row["Open/Missing check list"]
            tower_type, tower_num = tower.split(" Tower ")
            site_name = f"EWSLIG-{tower_type} Tower {tower_num.zfill(2)}"
            t = category_type_map.get(category, "Civil")
            if site_name not in summary_data:
                summary_data[site_name] = {"Civil": 0, "MEP": 0, "Interior": 0}
            summary_data[site_name][t] += open_missing

        for i, (site_name, counts) in enumerate(sorted(summary_data.items()), start=1):
            total = sum(counts.values())
            ws2.write(i, 0, site_name,          cell_format)
            ws2.write(i, 1, counts["Civil"],     cell_format)
            ws2.write(i, 2, counts["MEP"],       cell_format)
            ws2.write(i, 3, counts["Interior"],  cell_format)
            ws2.write(i, 4, total,               cell_format)

        for col in range(5):
            ws2.set_column(col, col, 30)

        workbook.close()
        output.seek(0)
        return output

    except Exception as e:
        logger.error(f"Error generating consolidated Excel: {traceback.format_exc()}")
        st.error(f"Error generating Excel file: {str(e)}")
        return None

# Combined function to handle analysis and display
def run_analysis_and_display():
    try:
        # Step 1: Run status analysis
        st.write("Running status analysis...")
        AnalyzeStatusManually()
        
        # Check if structure_analysis was populated
        if 'structure_analysis' not in st.session_state or st.session_state.structure_analysis is None:
            st.error("❌ Status analysis failed to generate structure_analysis. Please check the logs and ensure data fetching was successful.")
            logger.error("run_analysis_and_display failed: structure_analysis not populated after AnalyzeStatusManually")
            return
        st.success("Status analysis completed successfully!")

        # Step 2: Initialize AI response if needed
        if 'ai_response' not in st.session_state or not isinstance(st.session_state.ai_response, dict):
            st.session_state.ai_response = {}
            logger.info("Initialized st.session_state.ai_response in run_analysis_and_display")

        # Step 3: Display activity counts
        st.write("Displaying activity counts and generating AI data...")
        logger.debug("COS DataFrame columns: {}".format(
            list(st.session_state.get('cos_df_Revised_Baseline_45daysNGT_Rai', pd.DataFrame()).columns)
        ))
        display_activity_count()
        st.success("Activity counts displayed successfully!")

        
        
        # Check structure_analysis
        structure_analysis = st.session_state.get('structure_analysis')
        if structure_analysis is None:
            st.error("❌ No structure analysis data available.")
            logger.error("No structure_analysis in st.session_state")
            return
        
        # Check ai_response (activity_counts)
        if not st.session_state.ai_response:
            st.error("❌ No AI data available in st.session_state.ai_response. Attempting to regenerate.")
            logger.error("No AI data in st.session_state.ai_response after display_activity_count")
            
            logger.debug("Retrying COS DataFrame columns: {}".format(
                list(st.session_state.get('cos_df_Revised_Baseline_45daysNGT_Rai', pd.DataFrame()).columns)
            ))
            display_activity_count()
            if not st.session_state.ai_response:
                st.error("❌ Failed to regenerate AI data. Please check data fetching and try again.")
                logger.error("Failed to regenerate AI data")
                return

     
        # Step 6: Generate Excel file
        st.write("Generating consolidated checklist Excel file...")
        
        with st.spinner("Generating Excel file... This may take a moment."):
            excel_file = generate_consolidated_Checklist_excel(
                structure_analysis=structure_analysis, 
                activity_counts=st.session_state.ai_response
            )
        
        # Step 7: Handle download
        if excel_file:
            timestamp = pd.Timestamp.now(tz='Asia/Kolkata').strftime('%Y%m%d_%H%M')
            file_name = f"Consolidated_Checklist_EWSLIG_{timestamp}.xlsx"
            
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                st.download_button(
                    label="📥 Download Checklist Excel",
                    data=excel_file,
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_excel_button",
                    help="Click to download the consolidated checklist in Excel format."
                )
            st.success("✅ Excel file generated successfully!")
        else:
            st.error("Error generating Excel file. Check logs for details.")
            logger.error("Failed to generate Excel file - function returned None")

    except Exception as e:
        error_msg = str(e)
        st.error(f"❌ Error during analysis, display, or Excel generation: {error_msg}")
        logger.error(f"Error during analysis, display, or Excel generation: {error_msg}")
        logger.error(f"Stack trace:\n{traceback.format_exc()}")
            
# Streamlit UI
st.markdown(
    """
    <h1 style='font-family: "Arial Black", Gadget, sans-serif; 
               color: red; 
               font-size: 48px; 
               text-align: center;'>
        CheckList - Report
    </h1>
    """,
    unsafe_allow_html=True
)

# Initialize and Fetch Data
st.sidebar.title("🔒 Asite Initialization")
email = st.sidebar.text_input("Email", "impwatson@gadieltechnologies.com", key="email_input")
password = st.sidebar.text_input("Password", "Wave2026@123$", type="password", key="password_input")

if st.sidebar.button("Initialize and Fetch Data"):
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        success = loop.run_until_complete(initialize_and_fetch_data(email, password))
        if success:
            st.sidebar.success("Initialization and data fetching completed successfully!")
        else:
            st.sidebar.error("Initialization and data fetching failed!")
    except Exception as e:
        st.sidebar.error(f"Initialization and data fetching failed: {str(e)}")
    finally:
        loop.close()

# Analyze and Display
st.sidebar.title("📊 Status Analysis")
if st.sidebar.button("Analyze and Display Activity Counts"):
    try:
        run_analysis_and_display()  # This function already handles the full workflow
    except Exception as e:
        logging.error(f"Error during analysis and display: {str(e)}")
        logging.error(f"Stack trace:\n{traceback.format_exc()}")
        st.error(f"Error occurred: {str(e)}\nCheck logs for details.")

st.sidebar.title("📊 Slab Cycle")
st.session_state.ignore_year = datetime.now().year
st.session_state.ignore_month = datetime.now().month
