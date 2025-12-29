# import streamlit as st
# import requests
# import json
# import urllib.parse
# import urllib3
# import certifi
# import pandas as pd
# from datetime import datetime
# import re
# import logging
# import os
# from dotenv import load_dotenv
# import aiohttp
# import asyncio
# from concurrent.futures import ThreadPoolExecutor, as_completed
# import time
# import openpyxl
# import io
# from dotenv import load_dotenv
# from uuid import uuid4
# import ibm_boto3
# from ibm_botocore.client import Config
# from tenacity import retry, stop_after_attempt, wait_exponential
# import xlsxwriter

# # Set up logging
# logging.basicConfig(level=logging.INFO)
# logger = logging.getLogger(__name__)

# # Configure logging
# logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
# logger = logging.getLogger(__name__)

# # Disable SSL warnings
# urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# # Load environment variables
# load_dotenv()

# # IBM COS Configuration
# COS_API_KEY = os.getenv("COS_API_KEY")
# COS_SERVICE_INSTANCE_ID = os.getenv("COS_SERVICE_INSTANCE_ID")
# COS_ENDPOINT = os.getenv("COS_ENDPOINT")
# COS_BUCKET = os.getenv("COS_BUCKET")

# # WatsonX configuration
# WATSONX_API_URL = os.getenv("WATSONX_API_URL_1")
# MODEL_ID = os.getenv("MODEL_ID_1")
# PROJECT_ID = os.getenv("PROJECT_ID_1")
# API_KEY = os.getenv("API_KEY_1")

# # API Endpoints
# LOGIN_URL = "https://dms.asite.com/apilogin/"
# IAM_TOKEN_URL = "https://iam.cloud.ibm.com/identity/token"

# # Login Function
# async def login_to_asite(email, password):
#     headers = {"Accept": "application/json", "Content-Type": "application/x-www-form-urlencoded"}
#     payload = {"emailId": email, "password": password}
#     response = requests.post(LOGIN_URL, headers=headers, data=payload, verify=certifi.where(), timeout=50)
#     if response.status_code == 200:
#         try:
#             session_id = response.json().get("UserProfile", {}).get("Sessionid")
#             logger.info(f"Login successful, Session ID: {session_id}")
#             st.session_state.sessionid = session_id
#             st.sidebar.success(f"✅ Login successful, Session ID: {session_id}")
#             return session_id
#         except json.JSONDecodeError:
#             logger.error("JSONDecodeError during login")
#             st.sidebar.error("❌ Failed to parse login response")
#             return None
#     logger.error(f"Login failed: {response.status_code} - {response.text}")
#     st.sidebar.error(f"❌ Login failed: {response.status_code} - {response.text}")
#     return None

# # Function to generate access token
# @retry(stop=stop_after_attempt(5), wait=wait_exponential(multiplier=2, min=10, max=60))
# def get_access_token(API_KEY):
#     headers = {"Content-Type": "application/x-www-form-urlencoded", "Accept": "application/json"}
#     data = {"grant_type": "urn:ibm:params:oauth:grant-type:apikey", "apikey": API_KEY}
#     response = requests.post(IAM_TOKEN_URL, headers=headers, data=data, verify=certifi.where(), timeout=50)
#     try:
#         if response.status_code == 200:
#             token_info = response.json()
#             logger.info("Access token generated successfully")
#             return token_info['access_token']
#         else:
#             logger.error(f"Failed to get access token: {response.status_code} - {response.text}")
#             st.error(f"❌ Failed to get access token: {response.status_code} - {response.text}")
#             raise Exception("Failed to get access token")
#     except Exception as e:
#         logger.error(f"Exception getting access token: {str(e)}")
#         st.error(f"❌ Error getting access token: {str(e)}")
#         return None

# # Initialize COS client
# @retry(stop=stop_after_attempt(5), wait=wait_exponential(multiplier=1, min=4, max=10))
# def initialize_cos_client():
#     try:
#         logger.info("Attempting to initialize COS client...")
#         cos_client = ibm_boto3.client(
#             's3',
#             ibm_api_key_id=COS_API_KEY,
#             ibm_service_instance_id=COS_SERVICE_INSTANCE_ID,
#             config=Config(
#                 signature_version='oauth',
#                 connect_timeout=180,
#                 read_timeout=180,
#                 retries={'max_attempts': 15}
#             ),
#             endpoint_url=COS_ENDPOINT
#         )
#         logger.info("COS client initialized successfully")
#         return cos_client
#     except Exception as e:
#         logger.error(f"Error initializing COS client: {str(e)}")
#         st.error(f"❌ Error initializing COS client: {str(e)}")
#         raise

# # Fetch Workspace ID
# async def GetWorkspaceID():
#     url = "https://dmsak.asite.com/api/workspace/workspacelist"
#     headers = {
#         'Cookie': f'ASessionID={st.session_state.sessionid}',
#         "Accept": "application/json",
#         "Content-Type": "application/x-www-form-urlencoded",
#     }
#     response = requests.get(url, headers=headers)
#     if response.status_code != 200:
#         st.error(f"Failed to fetch workspace list: {response.status_code} - {response.text}")
#         raise Exception(f"Failed to fetch workspace list: {response.status_code}")
#     try:
#         data = response.json()
#         st.write(f"Workspace list response: {data}")
#         st.session_state.workspaceid = data['asiteDataList']['workspaceVO'][2]['Workspace_Id']
#         st.write(f"Workspace ID: {st.session_state.workspaceid}")
#     except (KeyError, IndexError) as e:
#         st.error(f"Error parsing workspace ID: {str(e)}")
#         raise

# # Fetch Project IDs
# async def GetProjectId():
#     url = f"https://adoddleak.asite.com/commonapi/qaplan/getQualityPlanList;searchCriteria={{'criteria': [{{'field': 'planCreationDate','operator': 6,'values': ['11-Mar-2025']}}], 'projectId': {str(st.session_state.workspaceid)}, 'recordLimit': 1000, 'recordStart': 1}}"
#     headers = {
#         'Cookie': f'ASessionID={st.session_state.sessionid}',
#         "Accept": "application/json",
#         "Content-Type": "application/x-www-form-urlencoded",
#     }
#     response = requests.get(url, headers=headers)
#     if response.status_code != 200:
#         st.error(f"Failed to fetch project IDs: {response.status_code} - {response.text}")
#         raise Exception(f"Failed to fetch project IDs: {response.status_code}")
#     data = response.json()
#     st.write(f"Project ID response: {data}")
#     if not data.get('data'):
#         st.error("No quality plans found for the specified date.")
#         raise Exception("No quality plans found")
#     st.session_state.Eden_structure = data['data'][0]['planId']
#     st.write(f"Eden Structure Project ID: {st.session_state.Eden_structure}")

# # Asynchronous Fetch Function
# async def fetch_data(session, url, headers):
#     async with session.get(url, headers=headers) as response:
#         if response.status == 200:
#             return await response.json()
#         elif response.status == 204:
#             return None
#         else:
#             raise Exception(f"Error fetching data: {response.status} - {await response.text()}")

# # Fetch All Structure Data
# async def GetAllDatas():
#     record_limit = 1000
#     headers = {'Cookie': f'ASessionID={st.session_state.sessionid}'}
#     all_structure_data = []

#     async with aiohttp.ClientSession() as session:
#         start_record = 1
#         st.write("Fetching Eden Structure data...")
#         while True:
#             url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanAssociation/?projectId={st.session_state.workspaceid}&planId={st.session_state.Eden_structure}&recordStart={start_record}&recordLimit={record_limit}"
#             try:
#                 async with session.get(url, headers=headers) as response:
#                     if response.status == 204:
#                         st.write("No more Structure data available (204)")
#                         break
#                     data = await response.json()
#                     if 'associationList' in data and data['associationList']:
#                         all_structure_data.extend(data['associationList'])
#                     else:
#                         all_structure_data.extend(data if isinstance(data, list) else [])
#                     st.write(f"Fetched {len(all_structure_data[-record_limit:])} Structure records (Total: {len(all_structure_data)})")
#                     if len(all_structure_data[-record_limit:]) < record_limit:
#                         break
#                     start_record += record_limit
#             except Exception as e:
#                 st.error(f"❌ Error fetching Structure data: {str(e)}")
#                 break

#     df_structure = pd.DataFrame(all_structure_data)
    
#     desired_columns = ['activitySeq', 'qiLocationId']
#     if 'statusName' in df_structure.columns:
#         desired_columns.append('statusName')
#     elif 'statusColor' in df_structure.columns:
#         desired_columns.append('statusColor')
#         status_mapping = {'#4CAF50': 'Completed', '#4CB0F0': 'Not Started', '#4C0F0': 'Not Started'}
#         df_structure['statusName'] = df_structure['statusColor'].map(status_mapping).fillna('Unknown')
#         desired_columns.append('statusName')
#     else:
#         st.error("❌ Neither statusName nor statusColor found in data!")
#         return pd.DataFrame()

#     eden_structure = df_structure[desired_columns]

#     st.write(f"EDEN STRUCTURE ({', '.join(desired_columns)})")
#     st.write(f"Total records: {len(eden_structure)}")
#     st.write(eden_structure)
    
#     return eden_structure

# # Fetch Activity Data
# async def Get_Activity():
#     record_limit = 1000
#     headers = {
#         'Cookie': f'ASessionID={st.session_state.sessionid}',
#         "Accept": "application/json",
#         "Content-Type": "application/x-www-form-urlencoded",
#     }
    
#     all_structure_activity_data = []
    
#     async with aiohttp.ClientSession() as session:
#         start_record = 1
#         st.write("Fetching Activity data for Eden Structure...")
#         while True:
#             url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanActivities/?projectId={st.session_state.workspaceid}&planId={st.session_state.Eden_structure}&recordStart={start_record}&recordLimit={record_limit}"
#             try:
#                 data = await fetch_data(session, url, headers)
#                 if data is None:
#                     st.write("No more Structure Activity data available (204)")
#                     break
#                 if 'activityList' in data and data['activityList']:
#                     all_structure_activity_data.extend(data['activityList'])
#                 else:
#                     all_structure_activity_data.extend(data if isinstance(data, list) else [])
#                 st.write(f"Fetched {len(all_structure_activity_data[-record_limit:])} Structure Activity records (Total: {len(all_structure_activity_data)})")
#                 if len(all_structure_activity_data[-record_limit:]) < record_limit:
#                     break
#                 start_record += record_limit
#             except Exception as e:
#                 st.error(f"❌ Error fetching Structure Activity data: {str(e)}")
#                 break
 
#     structure_activity_data = pd.DataFrame(all_structure_activity_data)[['activityName', 'activitySeq', 'formTypeId']]

#     st.write("EDEN STRUCTURE ACTIVITY DATA (activityName and activitySeq)")
#     st.write(f"Total records: {len(structure_activity_data)}")
#     st.write(structure_activity_data)
      
#     return structure_activity_data

# # Fetch Location/Module Data
# async def Get_Location():
#     record_limit = 1000
#     headers = {
#         'Cookie': f'ASessionID={st.session_state.sessionid}',
#         "Accept": "application/json",
#         "Content-Type": "application/x-www-form-urlencoded",
#     }
    
#     all_structure_location_data = []
    
#     async with aiohttp.ClientSession() as session:
#         start_record = 1
#         total_records_fetched = 0
#         st.write("Fetching Eden Structure Location/Module data...")
#         while True:
#             url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanLocation/?projectId={st.session_state.workspaceid}&planId={st.session_state.Eden_structure}&recordStart={start_record}&recordLimit={record_limit}"
#             try:
#                 data = await fetch_data(session, url, headers)
#                 if data is None:
#                     st.write("No more Structure Location data available (204)")
#                     break
#                 if isinstance(data, list):
#                     location_data = [{'qiLocationId': item.get('qiLocationId', ''), 'qiParentId': item.get('qiParentId', ''), 'name': item.get('name', '')} 
#                                    for item in data if isinstance(item, dict)]
#                     all_structure_location_data.extend(location_data)
#                     total_records_fetched = len(all_structure_location_data)
#                     st.write(f"Fetched {len(location_data)} Structure Location records (Total: {total_records_fetched})")
#                 elif isinstance(data, dict) and 'locationList' in data and data['locationList']:
#                     location_data = [{'qiLocationId': loc.get('qiLocationId', ''), 'qiParentId': loc.get('qiParentId', ''), 'name': loc.get('name', '')} 
#                                    for loc in data['locationList']]
#                     all_structure_location_data.extend(location_data)
#                     total_records_fetched = len(all_structure_location_data)
#                     st.write(f"Fetched {len(location_data)} Structure Location records (Total: {total_records_fetched})")
#                 else:
#                     st.warning(f"No 'locationList' in Structure Location data or empty list.")
#                     break
#                 if len(location_data) < record_limit:
#                     break
#                 start_record += record_limit
#             except Exception as e:
#                 st.error(f"❌ Error fetching Structure Location data: {str(e)}")
#                 break
        
#     structure_df = pd.DataFrame(all_structure_location_data)
    
#     if 'name' in structure_df.columns and structure_df['name'].isna().all():
#         st.error("❌ All 'name' values in Structure Location data are missing or empty!")

#     st.write("EDEN STRUCTURE LOCATION/MODULE DATA")
#     st.write(f"Total records: {len(structure_df)}")
#     st.write(structure_df)
    
#     st.session_state.structure_location_data = structure_df
    
#     return structure_df

# # Process individual chunk
# def process_chunk(chunk, chunk_idx, dataset_name, location_df):
#     logger.info(f"Starting thread for {dataset_name} Chunk {chunk_idx + 1}")
#     generated_text = format_chunk_locally(chunk, chunk_idx, len(chunk), dataset_name, location_df)
#     logger.info(f"Completed thread for {dataset_name} Chunk {chunk_idx + 1}")
#     return generated_text, chunk_idx

# # Process data with manual counting
# def process_manually(analysis_df, total, dataset_name, chunk_size=1000, max_workers=4):
#     if analysis_df.empty:
#         st.warning(f"No completed activities found for {dataset_name}.")
#         return "No completed activities found."

#     unique_activities = analysis_df['activityName'].unique()
#     logger.info(f"Unique activities in {dataset_name} dataset: {list(unique_activities)}")
#     logger.info(f"Total records in {dataset_name} dataset: {len(analysis_df)}")

#     st.write(f"Saved Eden {dataset_name} data to eden_{dataset_name.lower()}_data.json")
#     chunks = [analysis_df[i:i + chunk_size] for i in range(0, len(analysis_df), chunk_size)]

#     location_df = st.session_state.structure_location_data

#     chunk_results = {}
#     progress_bar = st.progress(0)
#     status_text = st.empty()

#     with ThreadPoolExecutor(max_workers=max_workers) as executor:
#         future_to_chunk = {
#             executor.submit(process_chunk, chunk, idx, dataset_name, location_df): idx 
#             for idx, chunk in enumerate(chunks)
#         }

#         completed_chunks = 0
#         for future in as_completed(future_to_chunk):
#             chunk_idx = future_to_chunk[future]
#             try:
#                 generated_text, idx = future.result()
#                 chunk_results[idx] = generated_text
#                 completed_chunks += 1
#                 progress_percent = completed_chunks / len(chunks)
#                 progress_bar.progress(progress_percent)
#                 status_text.text(f"Processed chunk {completed_chunks} of {len(chunks)} ({progress_percent:.1%} complete)")
#             except Exception as e:
#                 logger.error(f"Error processing chunk {chunk_idx + 1} for {dataset_name}: {str(e)}")
#                 st.error(f"❌ Error processing chunk {chunk_idx + 1}: {str(e)}")

#     parsed_data = {}
#     for chunk_idx in sorted(chunk_results.keys()):
#         generated_text = chunk_results[chunk_idx]
#         if not generated_text:
#             continue

#         current_tower = None
#         tower_activities = []
#         lines = generated_text.split("\n")
        
#         for line in lines:
#             line = line.strip()
#             if not line:
#                 continue
            
#             if line.startswith("Tower:"):
#                 try:
#                     tower_parts = line.split("Tower:", 1)
#                     if len(tower_parts) > 1:
#                         if current_tower and tower_activities:
#                             if current_tower not in parsed_data:
#                                 parsed_data[current_tower] = []
#                             parsed_data[current_tower].extend(tower_activities)
#                             tower_activities = []
#                         current_tower = tower_parts[1].strip()
#                 except Exception as e:
#                     logger.warning(f"Error parsing Tower line: {line}, error: {str(e)}")
#                     if not current_tower:
#                         current_tower = f"Unknown Tower {chunk_idx}"
                    
#             elif line.startswith("Total Completed Activities:"):
#                 continue
#             elif not line.strip().startswith("activityName"):
#                 try:
#                     parts = re.split(r'\s{2,}', line.strip())
#                     if len(parts) >= 2:
#                         activity_name = ' '.join(parts[:-1]).strip()
#                         count_str = parts[-1].strip()
#                         count_match = re.search(r'\d+', count_str)
#                         if count_match:
#                             count = int(count_match.group())
#                             if current_tower:
#                                 tower_activities.append({
#                                     "activityName": activity_name,
#                                     "completedCount": count
#                                 })
#                     else:
#                         match = re.match(r'^\s*(.+?)\s+(\d+)$', line)
#                         if match and current_tower:
#                             activity_name = match.group(1).strip()
#                             count = int(match.group(2).strip())
#                             tower_activities.append({
#                                 "activityName": activity_name,
#                                 "completedCount": count
#                             })
#                 except (ValueError, IndexError) as e:
#                     logger.warning(f"Skipping malformed activity line: {line}, error: {str(e)}")

#         if current_tower and tower_activities:
#             if current_tower not in parsed_data:
#                 parsed_data[current_tower] = []
#             parsed_data[current_tower].extend(tower_activities)

#     aggregated_data = {}
#     for tower_name, activities in parsed_data.items():
#         tower_short_name = tower_name.split('/')[1] if '/' in tower_name else tower_name
#         if tower_short_name not in aggregated_data:
#             aggregated_data[tower_short_name] = {}
        
#         for activity in activities:
#             name = activity.get("activityName", "Unknown")
#             count = activity.get("completedCount", 0)
#             if name in aggregated_data[tower_short_name]:
#                 aggregated_data[tower_short_name][name] += count
#             else:
#                 aggregated_data[tower_short_name][name] = count

#     combined_output_lines = []
#     sorted_towers = sorted(aggregated_data.keys())
    
#     for i, tower_short_name in enumerate(sorted_towers):
#         combined_output_lines.append(f"{tower_short_name:<11} activityName            CompletedCount")
#         activity_dict = aggregated_data[tower_short_name]
#         tower_total = 0
#         for name, count in sorted(activity_dict.items()):
#             combined_output_lines.append(f"{'':<11} {name:<23} {count:>14}")
#             tower_total += count
#         combined_output_lines.append(f"{'':<11} Total for {tower_short_name:<11}: {tower_total:>14}")
#         if i < len(sorted_towers) - 1:
#             combined_output_lines.append("")
    
#     combined_output = "\n".join(combined_output_lines)
#     return combined_output

# # Local formatting function for manual counting
# def format_chunk_locally(chunk, chunk_idx, chunk_size, dataset_name, location_df):
#     towers_data = {}
    
#     for _, row in chunk.iterrows():
#         tower_name = row['tower_name']
#         activity_name = row['activityName']
#         count = int(row['CompletedCount'])
        
#         if tower_name not in towers_data:
#             towers_data[tower_name] = []
            
#         towers_data[tower_name].append({
#             "activityName": activity_name,
#             "completedCount": count
#         })
    
#     output = ""
#     total_activities = 0
    
#     for tower_name, activities in sorted(towers_data.items()):
#         output += f"Tower: {tower_name}\n"
#         output += "activityName            CompletedCount\n"
#         activity_dict = {}
#         for activity in activities:
#             name = activity['activityName']
#             count = activity['completedCount']
#             activity_dict[name] = activity_dict.get(name, 0) + count
#         for name, count in sorted(activity_dict.items()):
#             output += f"{name:<30} {count}\n"
#             total_activities += count
    
#     output += f"Total Completed Activities: {total_activities}"
#     return output

# def process_data(df, activity_df, location_df, dataset_name):
#     completed = df[df['statusName'] == 'Completed']
#     if completed.empty:
#         logger.warning(f"No completed activities found in {dataset_name} data.")
#         return pd.DataFrame(), 0

#     completed = completed.merge(location_df[['qiLocationId', 'name']], on='qiLocationId', how='left')
#     completed = completed.merge(activity_df[['activitySeq', 'activityName']], on='activitySeq', how='left')

#     if 'qiActivityId' not in completed.columns:
#         completed['qiActivityId'] = completed['qiLocationId'].astype(str) + '$$' + completed['activitySeq'].astype(str)

#     if completed['name'].isna().all():
#         logger.error(f"All 'name' values are missing in {dataset_name} data after merge!")
#         st.error(f"❌ All 'name' values are missing in {dataset_name} data after merge! Check location data.")
#         completed['name'] = 'Unknown'
#     else:
#         completed['name'] = completed['name'].fillna('Unknown')

#     completed['activityName'] = completed['activityName'].fillna('Unknown')

#     parent_child_dict = dict(zip(location_df['qiLocationId'], location_df['qiParentId']))
#     name_dict = dict(zip(location_df['qiLocationId'], location_df['name']))

#     def get_full_path(location_id):
#         path = []
#         current_id = location_id
#         max_depth = 10
#         depth = 0
        
#         while current_id and depth < max_depth:
#             if current_id not in parent_child_dict or current_id not in name_dict:
#                 logger.warning(f"Location ID {current_id} not found in parent_child_dict or name_dict. Path so far: {path}")
#                 break
            
#             parent_id = parent_child_dict.get(current_id)
#             name = name_dict.get(current_id, "Unknown")
            
#             if not parent_id:
#                 if name != "Quality":
#                     path.append(name)
#                     path.append("Quality")
#                 else:
#                     path.append(name)
#                 break
            
#             path.append(name)
#             current_id = parent_id
#             depth += 1
        
#         if depth >= max_depth:
#             logger.warning(f"Max depth reached while computing path for location_id {location_id}. Possible circular reference. Path: {path}")
        
#         if not path:
#             logger.warning(f"No path constructed for location_id {location_id}. Using 'Unknown'.")
#             return "Unknown"
        
#         full_path = '/'.join(reversed(path))
#         logger.debug(f"Full path for location_id {location_id}: {full_path}")
#         return full_path

#     completed['full_path'] = completed['qiLocationId'].apply(get_full_path)

#     # **NEW: Filter to only include records with "/Pour" in the full path**
#     logger.info(f"Total completed records before /Pour filter: {len(completed)}")
#     completed = completed[completed['full_path'].str.contains('/Pour', case=False, na=False)]
#     logger.info(f"Total completed records after /Pour filter: {len(completed)}")
    
#     if completed.empty:
#         logger.warning(f"No completed activities with /Pour in path found in {dataset_name} data after filtering.")
#         st.warning(f"No completed activities with /Pour in path found for {dataset_name}")
#         return pd.DataFrame(), 0

#     def get_tower_name(full_path):
#         parts = full_path.split('/')
#         if len(parts) < 2:
#             return full_path
        
#         tower = parts[1]
#         if tower == "Tower 4" and len(parts) > 2:
#             module = parts[2]
#             module_number = module.replace("Module ", "").strip()
#             try:
#                 module_num = int(module_number)
#                 if 1 <= module_num <= 4:
#                     return "Tower 4(B)"
#                 elif 5 <= module_num <= 8:
#                     return "Tower 4(A)"
#             except ValueError:
#                 logger.warning(f"Could not parse module number from {module} in path {full_path}")
        
#         return tower

#     completed['tower_name'] = completed['full_path'].apply(get_tower_name)

#     # **NEW: Extract tower number for better matching**
#     def extract_tower_key(tower_name):
#         """Extract tower key like 'T4', 'T5', 'T6', 'T7' from tower name"""
#         if 'Tower 4' in tower_name or 'T4' in tower_name:
#             return 'T4'
#         elif 'Tower 5' in tower_name or 'T5' in tower_name:
#             return 'T5'
#         elif 'Tower 6' in tower_name or 'T6' in tower_name:
#             return 'T6'
#         elif 'Tower 7' in tower_name or 'T7' in tower_name:
#             return 'T7'
#         else:
#             # Try to extract number from tower name
#             match = re.search(r'Tower\s*(\d+)', tower_name, re.IGNORECASE)
#             if match:
#                 return f'T{match.group(1)}'
#             return tower_name

#     completed['tower_key'] = completed['tower_name'].apply(extract_tower_key)

#     # Group by tower_key and activityName to get counts
#     analysis = completed.groupby(['tower_key', 'activityName'])['qiLocationId'].nunique().reset_index(name='CompletedCount')
    
#     # Also keep tower_name for display purposes
#     tower_mapping = completed[['tower_key', 'tower_name']].drop_duplicates()
#     analysis = analysis.merge(tower_mapping, on='tower_key', how='left')
    
#     analysis = analysis.sort_values(by=['tower_key', 'activityName'], ascending=True)
#     total_completed = analysis['CompletedCount'].sum()

#     logger.info(f"Total completed activities for {dataset_name} after processing with /Pour filter: {total_completed}")
    
#     # Log some sample paths to verify filtering
#     if not completed.empty:
#         sample_paths = completed['full_path'].head(10).tolist()
#         logger.info(f"Sample full paths after /Pour filter: {sample_paths}")
    
#     return analysis, total_completed

# # Main analysis function for Eden Structure
# def AnalyzeStatusManually(email=None, password=None):
#     start_time = time.time()

#     if 'sessionid' not in st.session_state:
#         st.error("❌ Please log in first!")
#         return

#     required_data = [
#         'eden_structure',
#         'structure_activity_data',
#         'structure_location_data'
#     ]
    
#     for data_key in required_data:
#         if data_key not in st.session_state:
#             st.error(f"❌ Please fetch required data first! Missing: {data_key}")
#             return
#         if not isinstance(st.session_state[data_key], pd.DataFrame):
#             st.error(f"❌ {data_key} is not a DataFrame! Found type: {type(st.session_state[data_key])}")
#             return

#     structure_data = st.session_state.eden_structure
#     structure_activity = st.session_state.structure_activity_data
#     structure_locations = st.session_state.structure_location_data
    
#     for df, name in [(structure_data, "Structure")]:
#         if 'statusName' not in df.columns:
#             st.error(f"❌ statusName column not found in {name} data!")
#             return
#         if 'qiLocationId' not in df.columns:
#             st.error(f"❌ qiLocationId column not found in {name} data!")
#             return
#         if 'activitySeq' not in df.columns:
#             st.error(f"❌ activitySeq column not found in {name} data!")
#             return

#     for df, name in [(structure_locations, "Structure Location")]:
#         if 'qiLocationId' not in df.columns or 'name' not in df.columns:
#             st.error(f"❌ qiLocationId or name column not found in {name} data!")
#             return

#     for df, name in [(structure_activity, "Structure Activity")]:
#         if 'activitySeq' not in df.columns or 'activityName' not in df.columns:
#             st.error(f"❌ activitySeq or activityName column not found in {name} data!")
#             return

#     structure_analysis, structure_total = process_data(structure_data, structure_activity, structure_locations, "Structure")

#     st.write("### Eden Structure Quality Analysis (Completed Activities):")
#     st.write("**Full Output (Structure):**")
#     structure_output = process_manually(structure_analysis, structure_total, "Structure")
#     if structure_output:
#         st.text(structure_output)

#     # Store structure analysis in session state for Excel generation
#     st.session_state.structure_analysis = structure_analysis

#     end_time = time.time()
#     st.write(f"Total execution time: {end_time - start_time:.2f} seconds")

# def get_cos_files():
#     try:
#         cos_client = initialize_cos_client()
#         if not cos_client:
#             st.error("❌ Failed to initialize COS client.")
#             return None

#         st.write(f"Attempting to list objects in bucket '{COS_BUCKET}' with prefix 'Eden/'")
#         response = cos_client.list_objects_v2(Bucket=COS_BUCKET, Prefix="Eden/")
#         if 'Contents' not in response:
#             st.error(f"❌ No files found in the 'Eden' folder of bucket '{COS_BUCKET}'.")
#             logger.error("No objects found in Eden folder")
#             return None

#         all_files = [obj['Key'] for obj in response.get('Contents', [])]
#         st.write("**All files in Eden folder:**")
#         if all_files:
#             st.write("\n".join(all_files))
#         else:
#             st.write("No files found.")
#             logger.warning("Eden folder is empty")
#             return None

#         pattern = re.compile(
#             r"Eden/Structure\s*Work\s*Tracker[\(\s]*(.*?)(?:[\)\s]*\.xlsx)$",
#             re.IGNORECASE
#         )
        
#         date_formats = ["%d-%m-%Y", "%d-%m-%y", "%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y"]

#         file_info = []
#         for obj in response.get('Contents', []):
#             key = obj['Key']
#             match = pattern.match(key)
#             if match:
#                 date_str = match.group(1).strip('()').strip()
#                 parsed_date = None
#                 for fmt in date_formats:
#                     try:
#                         parsed_date = datetime.strptime(date_str, fmt)
#                         break
#                     except ValueError:
#                         continue
#                 if parsed_date:
#                     file_info.append({'key': key, 'date': parsed_date})
#                 else:
#                     logger.warning(f"Could not parse date in filename: {key}")
#                     st.warning(f"Skipping file with unparseable date: {key}")
#             else:
#                 st.write(f"File '{key}' does not match the expected pattern 'Eden/Structure Work Tracker (DD-MM-YYYY).xlsx'")

#         if not file_info:
#             st.error("❌ No Excel files matched the expected pattern in the 'Eden' folder.")
#             logger.error("No files matched the expected pattern")
#             return None

#         latest_file = max(file_info, key=lambda x: x['date']) if file_info else None
#         if not latest_file:
#             st.error("❌ No valid Excel files found for Structure Work Tracker.")
#             logger.error("No valid files after date parsing")
#             return None

#         file_key = latest_file['key']
#         st.success(f"Found matching file: {file_key}")
#         return file_key
#     except Exception as e:
#         st.error(f"❌ Error fetching COS files: {str(e)}")
#         logger.error(f"Error fetching COS files: {str(e)}")
#         return None

# # Initialize session state variables
# if 'cos_df_tower4' not in st.session_state:
#     st.session_state.cos_df_tower4 = None
# if 'cos_df_tower5' not in st.session_state:
#     st.session_state.cos_df_tower5 = None
# if 'cos_df_tower6' not in st.session_state:
#     st.session_state.cos_df_tower6 = None
# if 'cos_df_tower7' not in st.session_state:
#     st.session_state.cos_df_tower7 = None
# if 'cos_tname_tower4' not in st.session_state:
#     st.session_state.cos_tname_tower4 = None
# if 'cos_tname_tower5' not in st.session_state:
#     st.session_state.cos_tname_tower5 = None
# if 'cos_tname_tower6' not in st.session_state:
#     st.session_state.cos_tname_tower6 = None
# if 'cos_tname_tower7' not in st.session_state:
#     st.session_state.cos_tname_tower7 = None
# if 'ai_response' not in st.session_state:
#     st.session_state.ai_response = {}  # Initialize as empty dictionary

# # Process Excel files for Towers 4, 5, 6, 7
# def process_file(file_stream, filename):
#     try:
#         workbook = openpyxl.load_workbook(file_stream)
#         available_sheets = workbook.sheetnames
#         st.write(f"Available sheets in {filename}: {', '.join(available_sheets)}")

#         target_sheets = ["Tower 4", "Tower 5", "Tower 6", "Tower 7"]
#         results = []

#         expected_columns = [
#             'Tower', 'Activity No.', 'Monthly Lookahead ID', 'Task Name', 
#             'Actual Start', 'Actual Finish', '% Complete-MSP', 'Duration', 
#             'Start', 'Finish', 'Baseline Duration', 'Baseline Start', 
#             'Baseline Finish', 'Week1', 'Week2', 'Week3', 'Week4', 
#             'Total for the month', 'Total for the tower'
#         ]

#         for sheet_name in target_sheets:
#             if sheet_name not in available_sheets:
#                 st.warning(f"Sheet '{sheet_name}' not found in file: {filename}")
#                 continue

#             file_stream.seek(0)

#             try:
#                 df = pd.read_excel(file_stream, sheet_name=sheet_name, header=1)
#                 st.write(f"Raw columns in {sheet_name}: {list(df.columns)}")

#                 if len(df.columns) != len(expected_columns):
#                     st.error(f"Sheet {sheet_name} has {len(df.columns)} columns, but {len(expected_columns)} were expected: {list(df.columns)}")
#                     continue

#                 df.columns = expected_columns

#                 target_columns = ["Task Name", "Actual Start", "Actual Finish"]
#                 available_columns = [col for col in target_columns if col in df.columns]

#                 if len(available_columns) < len(target_columns):
#                     missing_cols = [col for col in target_columns if col not in available_columns]
#                     st.warning(f"Missing columns in sheet {sheet_name}: {', '.join(missing_cols)}")
#                     for col in missing_cols:
#                         df[col] = None

#                 df_original = df.copy()
#                 df = df[target_columns]
#                 df = df.dropna(subset=['Task Name'])
#                 df['Task Name'] = df['Task Name'].astype(str).str.strip()

#                 if 'Actual Finish' in df.columns:
#                     df['Actual_Finish_Original'] = df['Actual Finish'].astype(str)
#                     df['Actual Finish'] = pd.to_datetime(df['Actual Finish'], errors='coerce')
#                     has_na_mask = (
#                         pd.isna(df['Actual Finish']) |
#                         (df['Actual_Finish_Original'].str.upper() == 'NAT') |
#                         (df['Actual_Finish_Original'].str.lower().isin(['nan', 'na', 'n/a', 'none', '']))
#                     )
#                     st.write(f"Sample of rows with NA or invalid values in Actual Finish for {sheet_name}:")
#                     na_rows = df[has_na_mask][['Task Name', 'Actual Finish']]
#                     if not na_rows.empty:
#                         st.write(na_rows.head(10))
#                     else:
#                         st.write("No NA or invalid values found in Actual Finish")
#                     df.drop('Actual_Finish_Original', axis=1, inplace=True)

#                 st.write(f"Unique Task Names in {sheet_name}:")
#                 unique_tasks = df[['Task Name']].drop_duplicates()
#                 st.write(unique_tasks)

#                 results.append((df, sheet_name))

#             except Exception as e:
#                 st.error(f"Error processing sheet {sheet_name}: {str(e)}")
#                 continue

#         if not results:
#             st.error(f"No valid sheets ({', '.join(target_sheets)}) found in file: {filename}")
#             return [(None, None)]

#         return results

#     except Exception as e:
#         st.error(f"Error loading Excel file: {str(e)}")
#         return [(None, None)]

# # Function to get access token for WatsonX API
# def get_access_token(api_key):
#     try:
#         headers = {"Content-Type": "application/x-www-form-urlencoded"}
#         data = {
#             "grant_type": "urn:ibm:params:oauth:grant-type:apikey",
#             "apikey": api_key
#         }
#         response = requests.post("https://iam.cloud.ibm.com/identity/token", headers=headers, data=data)
#         if response.status_code == 200:
#             return response.json().get("access_token")
#         else:
#             logger.error(f"Failed to get access token: {response.status_code} - {response.text}")
#             return None
#     except Exception as e:
#         logger.error(f"Error getting access token: {str(e)}")
#         return None

# # NEW FUNCTION: Count concreting activities with special logic
# def count_concreting_from_cos(tower_df, tower_name):
#     """
#     Count concreting/casting activities from Structure Work Tracker that have a valid date in column F (Actual Finish).
#     Only activities with the word "casting" in the Task Name and a valid date in Actual Finish are counted.
    
#     Args:
#         tower_df: DataFrame for the specific tower sheet
#         tower_name: Name of the tower (e.g., "Tower 4")
    
#     Returns:
#         int: Count of completed concreting activities
#     """
#     try:
#         if tower_df is None or tower_df.empty:
#             logger.warning(f"No data available for {tower_name}")
#             return 0
        
#         concreting_count = 0
        
#         logger.info(f"\n{'='*80}")
#         logger.info(f"Starting concreting count for {tower_name}")
#         logger.info(f"{'='*80}")
        
#         # Iterate through all rows in the tower sheet
#         for idx, row in tower_df.iterrows():
#             task_name = str(row.get('Task Name', '')).strip()
#             task_name_lower = task_name.lower()
#             actual_finish = row.get('Actual Finish')
            
#             # Check: Does the task name contain the word "casting"?
#             if 'casting' not in task_name_lower:
#                 continue  # Skip if it doesn't have "casting" in the name
            
#             # Check: Does column F (Actual Finish) have a valid date?
#             has_valid_date = False
#             date_str = "No date"
            
#             if pd.notna(actual_finish):
#                 # Check if it's already a datetime object
#                 if isinstance(actual_finish, (pd.Timestamp, datetime)):
#                     has_valid_date = True
#                     date_str = actual_finish.strftime('%Y-%m-%d')
#                 # Check if it's a string that's not NA
#                 elif isinstance(actual_finish, str):
#                     actual_finish_cleaned = actual_finish.strip().upper()
#                     # Exclude common NA values
#                     if actual_finish_cleaned not in ['NA', 'N/A', 'NAT', 'NONE', '', 'NAN']:
#                         # Try to parse it as a date to ensure it's valid
#                         try:
#                             parsed_date = pd.to_datetime(actual_finish)
#                             has_valid_date = True
#                             date_str = parsed_date.strftime('%Y-%m-%d')
#                         except:
#                             has_valid_date = False
#                             logger.warning(f"{tower_name} (row {idx}): Casting activity '{task_name}' has invalid date format: '{actual_finish}'")
#                 else:
#                     # Handle numeric dates (Excel serial dates)
#                     try:
#                         converted_date = pd.to_datetime(actual_finish, errors='coerce')
#                         if pd.notna(converted_date):
#                             has_valid_date = True
#                             date_str = converted_date.strftime('%Y-%m-%d')
#                     except:
#                         has_valid_date = False
            
#             # Log and count if it has "casting" and a valid date
#             if has_valid_date:
#                 concreting_count += 1
#                 logger.info(f"✓ {tower_name} (row {idx}): COUNTED - '{task_name}' | Actual Finish: {date_str}")
#             else:
#                 logger.info(f"✗ {tower_name} (row {idx}): SKIPPED - '{task_name}' | Actual Finish: {actual_finish} (No valid date)")
        
#         logger.info(f"{'='*80}")
#         logger.info(f"Total concreting count for {tower_name}: {concreting_count}")
#         logger.info(f"{'='*80}\n")
#         return concreting_count
        
#     except Exception as e:
#         logger.error(f"Error counting concreting for {tower_name}: {str(e)}")
#         import traceback
#         logger.error(f"Full traceback: {traceback.format_exc()}")
#         return 0

# # WatsonX Prompt Generation (Updated with new categories)
# def generatePrompt(json_datas):
#     try:
#         if isinstance(json_datas, pd.DataFrame):
#             json_str = json_datas.reset_index().to_json(orient='records', indent=2)
#         else:
#             json_str = str(json_datas)

#         body = {
#             "input": f"""
#             Read the table data provided below and categorize the activities into the following categories: Civil Works, MEP Works, Interior Finishing Works, and External Development Activities. Compute the total count of each activity within its respective category and return the results as a JSON array, following the example format provided. For activities like "UP-First Fix" and "CP-First Fix", combine them as "Plumbing Works". If an activity is not found in the data, include it with a count of 0. Ensure the counts are accurate, the output is grouped by category, and the JSON structure is valid with no nested or repeated keys.

#             Table Data:
#             {json_str}

#             Categories and Activities:
#             - Civil Works: Concreting, Shuttering, Reinforcement, De-Shuttering
#             - MEP Works: Plumbing Works, Slab conduting, Wall Conduiting, Wiring & Switch Socket
#             - Interior Finishing Works: Floor Tiling, POP & Gypsum Plaster, Wall Tiling, Waterproofing – Sunken
#             - External Development Activities: Granular Sub-Base, Kerb Stone, Rain Water / Storm Line, Saucer Drain / Paver Block, Sewer Line, Stamp Concrete, Storm Line, WMM

#             Example JSON format needed:
#             [
#               {{
#                 "Category": "Civil Works",
#                 "Activities": [
#                   {{"Activity Name": "Concreting", "Total": 0}},
#                   {{"Activity Name": "Shuttering", "Total": 0}},
#                   {{"Activity Name": "Reinforcement", "Total": 0}},
#                   {{"Activity Name": "De-Shuttering", "Total": 0}}
#                 ]
#               }},
#               {{
#                 "Category": "MEP Works",
#                 "Activities": [
#                   {{"Activity Name": "Plumbing Works", "Total": 0}},
#                   {{"Activity Name": "Slab conduting", "Total": 0}},
#                   {{"Activity Name": "Wall Conduiting", "Total": 0}},
#                   {{"Activity Name": "Wiring & Switch Socket", "Total": 0}}
#                 ]
#               }},
#               {{
#                 "Category": "Interior Finishing Works",
#                 "Activities": [
#                   {{"Activity Name": "Floor Tiling", "Total": 0}},
#                   {{"Activity Name": "POP & Gypsum Plaster", "Total": 0}},
#                   {{"Activity Name": "Wall Tiling", "Total": 0}},
#                   {{"Activity Name": "Waterproofing – Sunken", "Total": 0}}
#                 ]
#               }},
#               {{
#                 "Category": "External Development Activities",
#                 "Activities": [
#                   {{"Activity Name": "Granular Sub-Base", "Total": 0}},
#                   {{"Activity Name": "Kerb Stone", "Total": 0}},
#                   {{"Activity Name": "Rain Water / Storm Line", "Total": 0}},
#                   {{"Activity Name": "Saucer Drain / Paver Block", "Total": 0}},
#                   {{"Activity Name": "Sewer Line", "Total": 0}},
#                   {{"Activity Name": "Stamp Concrete", "Total": 0}},
#                   {{"Activity Name": "Storm Line", "Total": 0}},
#                   {{"Activity Name": "WMM", "Total": 0}}
#                 ]
#               }}
#             ]

#             Return only the JSON array, no additional text, explanations, or code. Ensure the counts are accurate, activities are correctly categorized, and the JSON structure is valid.
#             """,
#             "parameters": {
#                 "decoding_method": "greedy",
#                 "max_new_tokens": 8100,
#                 "min_new_tokens": 0,
#                 "stop_sequences": [";"],
#                 "repetition_penalty": 1.05,
#                 "temperature": 0.5
#             },
#             "model_id": os.getenv("MODEL_ID_1"),
#             "project_id": os.getenv("PROJECT_ID_1")
#         }
        
#         access_token = get_access_token(os.getenv("API_KEY_1"))
#         if not access_token:
#             logger.error("Failed to obtain access token for WatsonX API")
#             return generate_fallback_totals(json_datas)
            
#         headers = {
#             "Accept": "application/json",
#             "Content-Type": "application/json",
#             "Authorization": f"Bearer {access_token}"
#         }
        
#         logger.info("Sending request to WatsonX API")
#         response = requests.post(os.getenv("WATSONX_API_URL_1"), headers=headers, json=body, timeout=60)
        
#         logger.info(f"WatsonX API response status: {response.status_code}")
#         logger.debug(f"WatsonX API response text: {response.text[:1000]}...")  # Log first 1000 chars
        
#         if response.status_code != 200:
#             logger.error(f"WatsonX API call failed: {response.status_code} - {response.text}")
#             st.warning(f"WatsonX API failed with status {response.status_code}: {response.text}. Using fallback method to calculate totals.")
#             return generate_fallback_totals(json_datas)
            
#         response_data = response.json()
#         logger.debug(f"WatsonX API response data: {response_data}")
        
#         if 'results' not in response_data or not response_data['results']:
#             logger.error("WatsonX API response does not contain 'results' key")
#             st.warning("WatsonX API response invalid. Using fallback method to calculate totals.")
#             return generate_fallback_totals(json_datas)

#         generated_text = response_data['results'][0].get('generated_text', '').strip()
#         logger.debug(f"Generated text from WatsonX: {generated_text[:1000]}...")  # Log first 1000 chars
        
#         if not generated_text:
#             logger.error("WatsonX API returned empty generated text")
#             st.warning("WatsonX API returned empty response. Using fallback method to calculate totals.")
#             return generate_fallback_totals(json_datas)

#         if not (generated_text.startswith('[') and generated_text.endswith(']')):
#             start_idx = generated_text.find('[')
#             end_idx = generated_text.rfind(']')
#             if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
#                 generated_text = generated_text[start_idx:end_idx+1]
#                 logger.info("Extracted JSON array from response")
#             else:
#                 logger.error(f"Could not extract valid JSON array from response: {generated_text[:1000]}...")
#                 return generate_fallback_totals(json_datas)
        
#         try:
#             parsed_json = json.loads(generated_text)
#             if not all(isinstance(item, dict) and 'Category' in item and 'Activities' in item for item in parsed_json):
#                 logger.warning("JSON structure doesn't match expected format")
#                 return generate_fallback_totals(json_datas)
#             logger.info("Successfully parsed WatsonX API response")
#             return generated_text
#         except json.JSONDecodeError as e:
#             logger.error(f"WatsonX API returned invalid JSON: {e}")
#             st.warning(f"WatsonX API returned invalid JSON. Error: {str(e)}. Using fallback method to calculate totals.")
#             error_position = int(str(e).split('(char ')[1].split(')')[0]) if '(char ' in str(e) else 0
#             context_start = max(0, error_position - 50)
#             context_end = min(len(generated_text), error_position + 50)
#             logger.error(f"JSON error context: ...{generated_text[context_start:error_position]}[ERROR HERE]{generated_text[error_position:context_end]}...")
#             return generate_fallback_totals(json_datas)
    
#     except Exception as e:
#         logger.error(f"Error in WatsonX API call: {str(e)}")
#         st.warning(f"Error in WatsonX API call: {str(e)}. Using fallback method to calculate totals.")
#         return generate_fallback_totals(json_datas)

# # Fallback Total Calculation (Updated with new categories)
# def generate_fallback_totals(count_table):
#     try:
#         if not isinstance(count_table, pd.DataFrame):
#             logger.error("Fallback method received invalid input: not a DataFrame")
#             return json.dumps([
#                 {"Category": "Civil Works", "Activities": [
#                     {"Activity Name": "Concreting", "Total": 0},
#                     {"Activity Name": "Shuttering", "Total": 0},
#                     {"Activity Name": "Reinforcement", "Total": 0},
#                     {"Activity Name": "De-Shuttering", "Total": 0}
#                 ]},
#                 {"Category": "MEP Works", "Activities": [
#                     {"Activity Name": "Plumbing Works", "Total": 0},
#                     {"Activity Name": "Slab conduting", "Total": 0},
#                     {"Activity Name": "Wall Conduiting", "Total": 0},
#                     {"Activity Name": "Wiring & Switch Socket", "Total": 0}
#                 ]},
#                 {"Category": "Interior Finishing Works", "Activities": [
#                     {"Activity Name": "Floor Tiling", "Total": 0},
#                     {"Activity Name": "POP & Gypsum Plaster", "Total": 0},
#                     {"Activity Name": "Wall Tiling", "Total": 0},
#                     {"Activity Name": "Waterproofing – Sunken", "Total": 0}
#                 ]},
#                 {"Category": "External Development Activities", "Activities": [
#                     {"Activity Name": "Granular Sub-Base", "Total": 0},
#                     {"Activity Name": "Kerb Stone", "Total": 0},
#                     {"Activity Name": "Rain Water / Storm Line", "Total": 0},
#                     {"Activity Name": "Saucer Drain / Paver Block", "Total": 0},
#                     {"Activity Name": "Sewer Line", "Total": 0},
#                     {"Activity Name": "Stamp Concrete", "Total": 0},
#                     {"Activity Name": "Storm Line", "Total": 0},
#                     {"Activity Name": "WMM", "Total": 0}
#                 ]}
#             ], indent=2)

#         categories = {
#             "Civil Works": [
#                 "Concreting", "Shuttering", "Reinforcement", "De-Shuttering"
#             ],
#             "MEP Works": [
#                 "Plumbing Works", "Slab conduting", "Wall Conduiting", "Wiring & Switch Socket"
#             ],
#             "Interior Finishing Works": [
#                 "Floor Tiling", "POP & Gypsum Plaster", "Wall Tiling", "Waterproofing – Sunken"
#             ],
#             "External Development Activities": [
#                 "Granular Sub-Base", "Kerb Stone", "Rain Water / Storm Line", "Saucer Drain / Paver Block", 
#                 "Sewer Line", "Stamp Concrete", "Storm Line", "WMM"
#             ]
#         }

#         result = []
#         for category, activities in categories.items():
#             category_data = {"Category": category, "Activities": []}
            
#             for activity in activities:
#                 if activity == "Plumbing Works":
#                     combined_count = count_table.loc["UP-First Fix and CP-First Fix", "Count"] if "UP-First Fix and CP-First Fix" in count_table.index else 0
#                     total = combined_count
#                 else:
#                     total = count_table.loc[activity, "Count"] if activity in count_table.index else 0
#                 category_data["Activities"].append({
#                     "Activity Name": activity,
#                     "Total": int(total) if pd.notna(total) else 0
#                 })
            
#             result.append(category_data)

#         return json.dumps(result, indent=2)
#     except Exception as e:
#         logger.error(f"Error in fallback total calculation: {str(e)}")
#         st.error(f"Error in fallback total calculation: {str(e)}")
#         return json.dumps([
#             {"Category": "Civil Works", "Activities": [
#                 {"Activity Name": "Concreting", "Total": 0},
#                 {"Activity Name": "Shuttering", "Total": 0},
#                 {"Activity Name": "Reinforcement", "Total": 0},
#                 {"Activity Name": "De-Shuttering", "Total": 0}
#             ]},
#             {"Category": "MEP Works", "Activities": [
#                 {"Activity Name": "Plumbing Works", "Total": 0},
#                 {"Activity Name": "Slab conduting", "Total": 0},
#                 {"Activity Name": "Wall Conduiting", "Total": 0},
#                 {"Activity Name": "Wiring & Switch Socket", "Total": 0}
#             ]},
#             {"Category": "Interior Finishing Works", "Activities": [
#                 {"Activity Name": "Floor Tiling", "Total": 0},
#                 {"Activity Name": "POP & Gypsum Plaster", "Total": 0},
#                 {"Activity Name": "Wall Tiling", "Total": 0},
#                 {"Activity Name": "Waterproofing – Sunken", "Total": 0}
#             ]},
#             {"Category": "External Development Activities", "Activities": [
#                 {"Activity Name": "Granular Sub-Base", "Total": 0},
#                 {"Activity Name": "Kerb Stone", "Total": 0},
#                 {"Activity Name": "Rain Water / Storm Line", "Total": 0},
#                 {"Activity Name": "Saucer Drain / Paver Block", "Total": 0},
#                 {"Activity Name": "Sewer Line", "Total": 0},
#                 {"Activity Name": "Stamp Concrete", "Total": 0},
#                 {"Activity Name": "Storm Line", "Total": 0},
#                 {"Activity Name": "WMM", "Total": 0}
#             ]}
#         ], indent=2)

# # Extract Totals from AI Data
# def getTotal(ai_data):
#     try:
#         if isinstance(ai_data, str):
#             ai_data = json.loads(ai_data)
            
#         if not isinstance(ai_data, list):
#             logger.error(f"AI data is not a list: {ai_data}")
#             return [0] * len(st.session_state.get('sheduledf', pd.DataFrame()).index)

#         share = []
#         for category_data in ai_data:
#             if isinstance(category_data, dict) and 'Activities' in category_data:
#                 for activity in category_data['Activities']:
#                     if isinstance(activity, dict) and 'Total' in activity:
#                         total = activity['Total']
#                         share.append(int(total) if isinstance(total, (int, float)) and pd.notna(total) else 0)
#                     else:
#                         share.append(0)
#             else:
#                 share.append(0)
#         return share
#     except Exception as e:
#         logger.error(f"Error parsing AI data: {str(e)}")
#         st.error(f"Error parsing AI data: {str(e)}")
#         return [0] * len(st.session_state.get('sheduledf', pd.DataFrame()).index)

# # UPDATED: Function to handle activity count display
# def display_activity_count():
#     # Updated specific activities according to new categorization
#     specific_activities = [
#         "Concreting", "Shuttering", "Reinforcement", "De-Shuttering",  # Civil Works
#         "Slab conduting", "Wall Conduiting", "Wiring & Switch Socket",  # MEP Works (changed from "Slab Conduiting")
#         "Floor Tiling", "POP & Gypsum Plaster", "Wall Tiling", "Waterproofing – Sunken",  # Interior Finishing Works
#         "Granular Sub-Base", "Kerb Stone", "Rain Water / Storm Line", "Saucer Drain / Paver Block",
#         "Sewer Line", "Stamp Concrete", "Storm Line", "WMM"  # External Development Activities
#     ]
#     all_activities = specific_activities + ["UP-First Fix and CP-First Fix"]

#     # Updated category mapping
#     category_mapping = {
#         "Concreting": "Civil Works",
#         "Shuttering": "Civil Works", 
#         "Reinforcement": "Civil Works",
#         "De-Shuttering": "Civil Works",
#         "UP-First Fix and CP-First Fix": "MEP Works",
#         "Slab conduting": "MEP Works",  # Changed from "Slab Conduiting"
#         "Wall Conduiting": "MEP Works", 
#         "Wiring & Switch Socket": "MEP Works",
#         "Floor Tiling": "Interior Finishing Works",
#         "POP & Gypsum Plaster": "Interior Finishing Works",
#         "Wall Tiling": "Interior Finishing Works",
#         "Waterproofing – Sunken": "Interior Finishing Works",
#         "Granular Sub-Base": "External Development Activities",
#         "Kerb Stone": "External Development Activities",
#         "Rain Water / Storm Line": "External Development Activities",
#         "Saucer Drain / Paver Block": "External Development Activities",
#         "Sewer Line": "External Development Activities",
#         "Stamp Concrete": "External Development Activities",
#         "Storm Line": "External Development Activities",
#         "WMM": "External Development Activities"
#     }

#     count_tables = {}
#     if 'ai_response' not in st.session_state or not isinstance(st.session_state.ai_response, dict):
#         st.session_state.ai_response = {}
#         logger.info("Re-initialized st.session_state.ai_response as empty dictionary")

#     def process_tower_data(tower_data, tname):
#         if tower_data is None or tower_data.empty:
#             logger.warning(f"No data available for {tname}")
#             return tname, None

#         tower_data = tower_data.copy()
        
#         st.write(f"Debug - First few rows from {tname}:")
#         st.write(tower_data.head(3))
        
#         # **NEW: Special handling for Concreting count**
#         concreting_count_special = count_concreting_from_cos(tower_data, tname)
#         st.write(f"**Special Concreting Count for {tname}: {concreting_count_special}**")
        
#         count_table = pd.DataFrame({
#             'Count_Unfiltered': [0] * len(all_activities),
#             'Count_Filtered': [0] * len(all_activities)
#         }, index=all_activities)
        
#         # Filter tower data by Actual Finish date
#         tower_data_filtered = tower_data.copy()
#         if 'Actual Finish' in tower_data.columns:
#             tower_data['Actual_Finish_Original'] = tower_data['Actual Finish'].astype(str)
#             tower_data['Actual Finish'] = pd.to_datetime(tower_data['Actual Finish'], errors='coerce')
#             has_na_mask = (
#                 pd.isna(tower_data['Actual Finish']) | 
#                 (tower_data['Actual_Finish_Original'].str.upper() == 'NAT') |
#                 (tower_data['Actual_Finish_Original'].str.lower().isin(['nan', 'na', 'n/a', 'none', '']))
#             )
#             tower_data_filtered = tower_data[~has_na_mask].copy()
#             tower_data.drop('Actual_Finish_Original', axis=1, inplace=True)
        
#         # For non-concreting activities, use standard counting
#         for activity in specific_activities:
#             if activity == "Concreting":
#                 # Use the special count for concreting
#                 count_table.loc[activity, 'Count_Filtered'] = concreting_count_special
#                 count_table.loc[activity, 'Count_Unfiltered'] = concreting_count_special
#                 continue
            
#             # Standard counting for other activities
#             exact_matches = tower_data[tower_data['Task Name'] == activity]
#             if len(exact_matches) > 0:
#                 count_table.loc[activity, 'Count_Unfiltered'] = len(exact_matches)
#             else:
#                 case_insensitive_matches = tower_data[tower_data['Task Name'].str.lower() == activity.lower()]
#                 count_table.loc[activity, 'Count_Unfiltered'] = len(case_insensitive_matches)
            
#             exact_matches_filtered = tower_data_filtered[tower_data_filtered['Task Name'] == activity]
#             if len(exact_matches_filtered) > 0:
#                 count_table.loc[activity, 'Count_Filtered'] = len(exact_matches_filtered)
#             else:
#                 case_insensitive_matches_filtered = tower_data_filtered[tower_data_filtered['Task Name'].str.lower() == activity.lower()]
#                 count_table.loc[activity, 'Count_Filtered'] = len(case_insensitive_matches_filtered)
        
#         # Handle UP-First Fix and CP-First Fix combination
#         up_first_fix_matches = tower_data[tower_data['Task Name'].str.lower() == "up-first fix".lower()]
#         cp_first_fix_matches = tower_data[tower_data['Task Name'].str.lower() == "cp-first fix".lower()]
#         up_first_fix_count = len(up_first_fix_matches)
#         cp_first_fix_count = len(cp_first_fix_matches)
#         count_table.loc["UP-First Fix and CP-First Fix", "Count_Unfiltered"] = up_first_fix_count + cp_first_fix_count
        
#         up_first_fix_matches_filtered = tower_data_filtered[tower_data_filtered['Task Name'].str.lower() == "up-first fix".lower()]
#         cp_first_fix_matches_filtered = tower_data_filtered[tower_data_filtered['Task Name'].str.lower() == "cp-first fix".lower()]
#         up_first_fix_count_filtered = len(up_first_fix_matches_filtered)
#         cp_first_fix_count_filtered = len(cp_first_fix_matches_filtered)
#         count_table.loc["UP-First Fix and CP-First Fix", "Count_Filtered"] = up_first_fix_count_filtered + cp_first_fix_count_filtered
        
#         count_table['Count_Unfiltered'] = count_table['Count_Unfiltered'].astype(int)
#         count_table['Count_Filtered'] = count_table['Count_Filtered'].astype(int)
        
#         return tname, count_table

#     # Process each tower's data
#     for tower, tname_key in [
#         (st.session_state.cos_df_tower4, 'cos_tname_tower4'),
#         (st.session_state.cos_df_tower5, 'cos_tname_tower5'),
#         (st.session_state.cos_df_tower6, 'cos_tname_tower6'),
#         (st.session_state.cos_df_tower7, 'cos_tname_tower7')
#     ]:
#         if tower is not None:
#             tname = st.session_state.get(tname_key)
#             tname, count_table = process_tower_data(tower, tname)
#             if count_table is not None:
#                 count_tables[tname] = count_table

#     if not count_tables:
#         st.error("No processed COS data available. Please click 'Fetch COS' first.")
#         st.stop()

#     for tname, count_table in count_tables.items():
#         with st.spinner(f"Processing activity counts for {tname}..."):
#             try:
#                 st.write(f"Activity Count for {tname} (Unfiltered vs Filtered):")
#                 st.write(count_table)
                
#                 count_table_filtered = count_table[['Count_Filtered']].rename(columns={'Count_Filtered': 'Count'})
#                 ai_response = generatePrompt(count_table_filtered)
                
#                 try:
#                     ai_data = json.loads(ai_response)
#                     if not all(isinstance(item, dict) and 'Category' in item and 'Activities' in item for item in ai_data):
#                         logger.warning(f"Invalid AI data structure for {tname}: {ai_data}")
#                         ai_response = generate_fallback_totals(count_table_filtered)
#                         ai_data = json.loads(ai_response)
                    
#                     # Store AI response
#                     st.session_state.ai_response[tname] = ai_data
#                     logger.info(f"Stored AI response for {tname}: {ai_data}")
                    
#                     totals_mapping = {}
#                     for category_data in ai_data:
#                         for activity in category_data['Activities']:
#                             totals_mapping[activity['Activity Name']] = activity['Total']
                    
#                     display_df = count_table.reset_index()
#                     display_df.rename(columns={'index': 'Activity Name'}, inplace=True)
                    
#                     display_df['Total'] = display_df['Activity Name'].map(
#                         lambda x: totals_mapping.get(x, display_df.loc[display_df['Activity Name'] == x, 'Count_Filtered'].iloc[0])
#                     )
                    
#                     display_df['Category'] = display_df['Activity Name'].map(lambda x: category_mapping.get(x, "Other"))
                    
#                     display_df = display_df.sort_values(['Category', 'Activity Name'])
                    
#                     st.write(f"Activity Count with Totals for {tname}:")
#                     st.write(display_df[['Activity Name', 'Count_Unfiltered', 'Total', 'Category']])
                    
#                     st.write(f"Activity Counts by Category for {tname}:")
#                     # Updated category display order
#                     for category in ['Civil Works', 'MEP Works', 'Interior Finishing Works', 'External Development Activities']:
#                         category_df = display_df[display_df['Category'] == category]
#                         if not category_df.empty:
#                             st.write(f"**{category}**")
#                             st.write(category_df[['Activity Name', 'Count_Filtered', 'Total']])
                    
#                 except Exception as e:
#                     logger.error(f"Error processing WatsonX for {tname}: {str(e)}")
#                     st.warning(f"Failed to process AI-generated totals for {tname}. Using fallback method.")
                    
#                     ai_response = generate_fallback_totals(count_table_filtered)
#                     ai_data = json.loads(ai_response)
#                     st.session_state.ai_response[tname] = ai_data
#                     logger.info(f"Stored fallback AI response for {tname}: {ai_data}")
                    
#                     display_df = count_table.reset_index()
#                     display_df.rename(columns={'index': 'Activity Name'}, inplace=True)
#                     display_df['Category'] = display_df['Activity Name'].map(lambda x: category_mapping.get(x, "Other"))
#                     display_df['Total'] = display_df['Count_Filtered']
#                     display_df = display_df.sort_values(['Category', 'Activity Name'])
                    
#                     st.write(f"Activity Counts by Category for {tname} (using raw counts):")
#                     # Updated category display order
#                     for category in ['Civil Works', 'MEP Works', 'Interior Finishing Works', 'External Development Activities']:
#                         category_df = display_df[display_df['Category'] == category]
#                         if not category_df.empty:
#                             st.write(f"**{category}**")
#                             st.write(category_df[['Activity Name', 'Count_Filtered', 'Total']])
                
#             except Exception as e:
#                 logger.error(f"Error displaying activity count for {tname}: {str(e)}")
#                 st.error(f"Error displaying activity count for {tname}: {str(e)}")
#                 # Ensure ai_response has an entry even in case of error
#                 count_table_filtered = count_table[['Count_Filtered']].rename(columns={'Count_Filtered': 'Count'})
#                 ai_response = generate_fallback_totals(count_table_filtered)
#                 ai_data = json.loads(ai_response)
#                 st.session_state.ai_response[tname] = ai_data
#                 logger.info(f"Stored fallback AI response for {tname} on error: {ai_data}")

#     # Create consolidated External Development Activities table
#     st.write("### External Development Activities (Consolidated)")
#     external_dev_activities = [
#         "Granular Sub-Base", "Kerb Stone", "Rain Water / Storm Line", "Saucer Drain / Paver Block",
#         "Sewer Line", "Stamp Concrete", "Storm Line", "WMM"
#     ]
    
#     external_dev_data = []
#     for activity in external_dev_activities:
#         total_count = 0
#         for tname, count_table in count_tables.items():
#             if activity in count_table.index:
#                 total_count += count_table.loc[activity, 'Count_Filtered']
#         external_dev_data.append({
#             "Activity Name": activity,
#             "Total Count": total_count
#         })
    
#     external_dev_df = pd.DataFrame(external_dev_data)
#     st.write(external_dev_df)


# # Combined function for Initialize and Fetch Data
# async def initialize_and_fetch_data(email, password):
#     with st.spinner("Starting initialization and data fetching process..."):
#         # Step 1: Login
#         if not email or not password:
#             st.sidebar.error("Please provide both email and password!")
#             return False
#         try:
#             st.sidebar.write("Logging in...")
#             session_id = await login_to_asite(email, password)
#             if not session_id:
#                 st.sidebar.error("Login failed!")
#                 return False
#             st.sidebar.success("Login successful!")
#         except Exception as e:
#             st.sidebar.error(f"Login failed: {str(e)}")
#             return False

#         # Step 2: Get Workspace ID
#         try:
#             st.sidebar.write("Fetching Workspace ID...")
#             await GetWorkspaceID()
#             st.sidebar.success("Workspace ID fetched successfully!")
#         except Exception as e:
#             st.sidebar.error(f"Failed to fetch Workspace ID: {str(e)}")
#             return False

#         # Step 3: Get Project IDs
#         try:
#             st.sidebar.write("Fetching Project IDs...")
#             await GetProjectId()
#             st.sidebar.success("Project IDs fetched successfully!")
#         except Exception as e:
#             st.sidebar.error(f"Failed to fetch Project IDs: {str(e)}")
#             return False

#         # Step 4: Get All Data (Structure only)
#         try:
#             st.sidebar.write("Fetching All Data...")
#             Edenstructure = await GetAllDatas()
#             st.session_state.eden_structure = Edenstructure
#             st.sidebar.success("All Data fetched successfully!")
#         except Exception as e:
#             st.sidebar.error(f"Failed to fetch All Data: {str(e)}")
#             return False

#         # Step 5: Get Activity Data
#         try:
#             st.sidebar.write("Fetching Activity Data...")
#             structure_activity_data = await Get_Activity()
#             st.session_state.structure_activity_data = structure_activity_data
#             st.sidebar.success("Activity Data fetched successfully!")
#         except Exception as e:
#             st.sidebar.error(f"Failed to fetch Activity Data: {str(e)}")
#             return False

#         # Step 6: Get Location/Module Data
#         try:
#             st.sidebar.write("Fetching Location/Module Data...")
#             structure_location_data = await Get_Location()
#             st.session_state.structure_location_data = structure_location_data 
#             st.sidebar.success("Location/Module Data fetched successfully!")
#         except Exception as e:
#             st.sidebar.error(f"Failed to fetch Location/Module Data: {str(e)}")
#             return False

#         # Step 7: Fetch COS Files
#         try:
#             st.sidebar.write("Fetching COS files from Eden folder...")
#             file_key = get_cos_files()
#             st.session_state.file_key = file_key
#             if file_key:
#                 st.success(f"Found 1 file in COS storage: {file_key}")
#                 try:
#                     st.write(f"Processing file: {file_key}")
#                     cos_client = initialize_cos_client()
#                     if not cos_client:
#                         st.error("Failed to initialize COS client during file fetch")
#                         logger.error("COS client initialization failed during file fetch")
#                         return False
#                     st.write("Fetching file from COS...")
#                     response = cos_client.get_object(Bucket=COS_BUCKET, Key=file_key)
#                     file_bytes = io.BytesIO(response['Body'].read())
#                     st.write("File fetched successfully. Processing sheets...")
#                     results = process_file(file_bytes, file_key)
#                     st.write(f"Processing results: {len(results)} sheets processed")
#                     for df, tname in results:
#                         if df is not None:
#                             if "Tower 4" in tname:
#                                 st.session_state.cos_df_tower4 = df
#                                 st.session_state.cos_tname_tower4 = tname
#                                 st.write(f"Processed Data for {tname} - {len(df)} rows:")
#                                 st.write(df.head())
#                             elif "Tower 5" in tname:
#                                 st.session_state.cos_df_tower5 = df
#                                 st.session_state.cos_tname_tower5 = tname
#                                 st.write(f"Processed Data for {tname} - {len(df)} rows:")
#                                 st.write(df.head())
#                             elif "Tower 6" in tname:
#                                 st.session_state.cos_df_tower6 = df
#                                 st.session_state.cos_tname_tower6 = tname
#                                 st.write(f"Processed Data for {tname} - {len(df)} rows:")
#                                 st.write(df.head())
#                             elif "Tower 7" in tname:
#                                 st.session_state.cos_df_tower7 = df
#                                 st.session_state.cos_tname_tower7 = tname
#                                 st.write(f"Processed Data for {tname} - {len(df)} rows:")
#                                 st.write(df.head())
#                         else:
#                             st.warning(f"No data processed for {tname} in {file_key}.")
#                 except Exception as e:
#                     st.error(f"Error loading {file_key} from cloud storage: {str(e)}")
#                     logger.error(f"Error loading {file_key}: {str(e)}")
#                     return False
#             else:
#                 st.warning("No expected Excel files available in the 'Eden' folder of the COS bucket.")
#                 return False
#         except Exception as e:
#             st.sidebar.error(f"Failed to fetch COS files: {str(e)}")
#             logger.error(f"Failed to fetch COS files: {str(e)}")
#             return False

#     st.sidebar.success("All steps completed successfully!")
#     return True


# # Updated generate_consolidated_Checklist_excel function with new categories
# # Updated generate_consolidated_Checklist_excel function with multiplier for Civil Works
# def generate_consolidated_Checklist_excel(structure_analysis, activity_counts):
#     try:
#         # Debug: Print activity_counts structure
#         logger.info(f"Activity counts keys: {list(activity_counts.keys()) if activity_counts else 'None'}")
#         logger.info(f"Activity counts type: {type(activity_counts)}")
        
#         # Also log structure_analysis to see what we have
#         if structure_analysis is not None and not structure_analysis.empty:
#             logger.info(f"Structure analysis columns: {structure_analysis.columns.tolist()}")
#             logger.info(f"Structure analysis sample:\n{structure_analysis.head()}")
#             logger.info(f"Unique tower_keys in structure_analysis: {structure_analysis['tower_key'].unique().tolist()}")
#             logger.info(f"Unique activityNames in structure_analysis: {structure_analysis['activityName'].unique().tolist()}")
        
#         # Define categories and activities according to new structure
#         categories = {
#             "Civil Works": ["Concreting", "Shuttering", "Reinforcement", "De-Shuttering"],
#             "MEP Works": ["Plumbing Works", "Slab Conduting", "Wall Conduiting", "Wiring & Switch Socket"],
#             "Interior Finishing Works": ["Floor Tiling", "POP & Gypsum Plaster", "Wall Tiling", "Waterproofing – Sunken"]
#         }

#         # Define the COS to Asite activity name mapping (updated)
#         cos_to_asite_mapping = {
#             "Concreting": "Concreting",
#             "Shuttering": "Shuttering", 
#             "Reinforcement": "Reinforcement",
#             "De-Shuttering": "De-Shuttering",
#             "Plumbing Works": ["UP-First Fix", "CP-First Fix"],  # Will sum both
#             "Slab Conduting": "Slab conduting",  # lowercase 'c' to match Asite
#             "Wall Conduiting": "Wall Conducting",
#             "Wiring & Switch Socket": "Wiring & Switch Socket",
#             "Floor Tiling": "Floor Tiling",
#             "POP & Gypsum Plaster": "POP & Gypsum Plaster",
#             "Wall Tiling": "Wall Tile",
#             "Waterproofing – Sunken": "Waterproofing - Sunken"
#         }

#         # Activities that should have the same completed count as Concreting
#         activities_same_as_concreting = ["Shuttering", "Reinforcement", "De-Shuttering"]

#         # Include all towers dynamically
#         all_towers = ["Tower 4", "Tower 5", "Tower 6", "Tower 7"]
        
#         logger.info(f"Processing all towers: {all_towers}")

#         # Initialize list to store consolidated data
#         consolidated_rows = []

#         # Process data for each tower and category
#         for tower in all_towers:
#             tower_key = f"T{tower.replace('Tower ', '')}"  # e.g., "Tower 4" -> "T4"
#             tower_num = tower.replace('Tower ', '')
            
#             logger.info(f"\n{'='*60}")
#             logger.info(f"Processing {tower} (key: {tower_key})")
#             logger.info(f"{'='*60}")
            
#             for category, activities in categories.items():
#                 # Process each activity in the category
#                 for activity in activities:
#                     # SPECIAL HANDLING FOR SLAB CONDUTING (Final Concreting - 5)
#                     if activity == "Slab Conduting":
#                         # Get the concreting count from activity_counts (COS data)
#                         concreting_count = 0
#                         tower_name_variations = [tower_key, tower, f"Tower {tower_num}"]
                        
#                         found_tower = None
#                         for tower_var in tower_name_variations:
#                             if activity_counts and tower_var in activity_counts:
#                                 found_tower = tower_var
#                                 break
                        
#                         if found_tower:
#                             ai_data = activity_counts[found_tower]
#                             # Extract the Concreting total from AI response data
#                             if isinstance(ai_data, list):
#                                 for category_data in ai_data:
#                                     if isinstance(category_data, dict) and 'Category' in category_data:
#                                         if category_data['Category'] == 'Civil Works':
#                                             for activity_data in category_data['Activities']:
#                                                 if isinstance(activity_data, dict) and activity_data.get('Activity Name') == 'Concreting':
#                                                     concreting_count = activity_data.get('Total', 0)
#                                                     logger.info(f"Found raw Concreting count for {found_tower}: {concreting_count}")
#                                                     break
#                         else:
#                             logger.info(f"No COS data found for {tower}, concreting_count = 0")
                        
#                         # Apply Civil Works formula to concreting count first: (count*2)-1
#                         final_concreting = (concreting_count * 2) - 1
#                         logger.info(f"Applied Civil Works formula to Concreting: ({concreting_count}*2)-1 = {final_concreting}")
                        
#                         # Calculate Slab Conduting as Final Concreting - 5
#                         completed_flats = max(0, final_concreting - 5)  # Ensure non-negative
#                         logger.info(f"Calculated Slab Conduting for {tower}: {final_concreting} - 5 = {completed_flats}")
                        
#                         # Get closed checklist from Asite (normal processing)
#                         asite_activity = cos_to_asite_mapping.get(activity, activity)
#                         if isinstance(asite_activity, list):
#                             asite_activities = asite_activity
#                         else:
#                             asite_activities = [asite_activity]

#                         closed_checklist = 0
#                         if structure_analysis is not None and not structure_analysis.empty:
#                             for asite_act in asite_activities:
#                                 # Try case-insensitive match
#                                 matching_rows = structure_analysis[
#                                     (structure_analysis['tower_key'] == tower_key) &
#                                     (structure_analysis['activityName'].str.lower() == asite_act.lower())
#                                 ]
                                
#                                 if not matching_rows.empty:
#                                     closed_checklist += matching_rows['CompletedCount'].sum()
#                                     logger.info(f"Found {tower_key} - {asite_act}: {matching_rows['CompletedCount'].sum()} completed")
#                                 else:
#                                     logger.info(f"No match found for {tower_key} - {asite_act}")

#                         # Calculate Open/Missing
#                         in_progress = 0
#                         if completed_flats == 0 or closed_checklist > completed_flats:
#                             open_missing = 0
#                         else:
#                             open_missing = abs(completed_flats - closed_checklist)

#                         display_activity = asite_activities[0] if isinstance(asite_activity, list) else asite_activity

#                         consolidated_rows.append({
#                             "Tower": tower,
#                             "Category": category,
#                             "Activity Name": display_activity,
#                             "Completed Work*(Count of Flat)": completed_flats,
#                             "In progress": in_progress,
#                             "Closed checklist": closed_checklist,
#                             "Open/Missing check list": open_missing
#                         })
                        
#                         continue  # Skip to next activity
                    
#                     # SPECIAL HANDLING FOR ACTIVITIES THAT MATCH CONCRETING
#                     # Shuttering, Reinforcement, De-Shuttering have same completed count as Concreting
#                     if activity in activities_same_as_concreting:
#                         # Get the concreting count from activity_counts (COS data)
#                         completed_flats = 0
#                         tower_name_variations = [tower_key, tower, f"Tower {tower_num}"]
                        
#                         found_tower = None
#                         for tower_var in tower_name_variations:
#                             if activity_counts and tower_var in activity_counts:
#                                 found_tower = tower_var
#                                 break
                        
#                         if found_tower:
#                             ai_data = activity_counts[found_tower]
#                             # Extract the Concreting total from AI response data
#                             if isinstance(ai_data, list):
#                                 for category_data in ai_data:
#                                     if isinstance(category_data, dict) and 'Category' in category_data:
#                                         if category_data['Category'] == 'Civil Works':
#                                             for activity_data in category_data['Activities']:
#                                                 if isinstance(activity_data, dict) and activity_data.get('Activity Name') == 'Concreting':
#                                                     completed_flats = activity_data.get('Total', 0)
#                                                     logger.info(f"Found Concreting count for {found_tower}: {completed_flats}, using for {activity}")
#                                                     break
#                         else:
#                             logger.info(f"No COS data found for {tower}, {activity} completed_flats = 0")
                        
#                         # *** APPLY FORMULA (count*2)-1 FOR CIVIL WORKS ***
#                         if category == "Civil Works":
#                             completed_flats = (completed_flats * 2) - 1
#                             logger.info(f"Applied Civil Works formula ((count*2)-1) for {activity}: {completed_flats}")
                        
#                         # Get closed checklist from Asite (normal processing)
#                         asite_activity = cos_to_asite_mapping.get(activity, activity)
#                         if isinstance(asite_activity, list):
#                             asite_activities = asite_activity
#                         else:
#                             asite_activities = [asite_activity]

#                         closed_checklist = 0
#                         if structure_analysis is not None and not structure_analysis.empty:
#                             for asite_act in asite_activities:
#                                 # Try case-insensitive match
#                                 matching_rows = structure_analysis[
#                                     (structure_analysis['tower_key'] == tower_key) &
#                                     (structure_analysis['activityName'].str.lower() == asite_act.lower())
#                                 ]
                                
#                                 if not matching_rows.empty:
#                                     closed_checklist += matching_rows['CompletedCount'].sum()
#                                     logger.info(f"Found {tower_key} - {asite_act}: {matching_rows['CompletedCount'].sum()} completed")
#                                 else:
#                                     logger.info(f"No match found for {tower_key} - {asite_act}")

#                         # Calculate Open/Missing
#                         in_progress = 0
#                         if completed_flats == 0 or closed_checklist > completed_flats:
#                             open_missing = 0
#                         else:
#                             open_missing = abs(completed_flats - closed_checklist)

#                         display_activity = asite_activities[0] if isinstance(asite_activity, list) else asite_activity

#                         consolidated_rows.append({
#                             "Tower": tower,
#                             "Category": category,
#                             "Activity Name": display_activity,
#                             "Completed Work*(Count of Flat)": completed_flats,
#                             "In progress": in_progress,
#                             "Closed checklist": closed_checklist,
#                             "Open/Missing check list": open_missing
#                         })
                        
#                         continue  # Skip to next activity
                    
#                     # REGULAR PROCESSING FOR OTHER ACTIVITIES (Concreting, Plumbing Works, etc.)
#                     # Map COS activity name to Asite name(s)
#                     asite_activity = cos_to_asite_mapping.get(activity, activity)
#                     if isinstance(asite_activity, list):
#                         asite_activities = asite_activity
#                     else:
#                         asite_activities = [asite_activity]

#                     # Get completed count from structure_analysis (Asite data with /Pour filter)
#                     closed_checklist = 0
#                     if structure_analysis is not None and not structure_analysis.empty:
#                         for asite_act in asite_activities:
#                             # Try case-insensitive match
#                             matching_rows = structure_analysis[
#                                 (structure_analysis['tower_key'] == tower_key) &
#                                 (structure_analysis['activityName'].str.lower() == asite_act.lower())
#                             ]
                            
#                             if not matching_rows.empty:
#                                 closed_checklist += matching_rows['CompletedCount'].sum()
#                                 logger.info(f"Found {tower_key} - {asite_act}: {matching_rows['CompletedCount'].sum()} completed")
#                             else:
#                                 logger.info(f"No match found for {tower_key} - {asite_act}")
#                     else:
#                         logger.info(f"No Asite data available for {tower_key} - {activity}")

#                     # Get completed flats count from activity_counts (COS data)
#                     completed_flats = 0
#                     tower_name_variations = [tower_key, tower, f"Tower {tower_num}"]
                    
#                     found_tower = None
#                     for tower_var in tower_name_variations:
#                         if activity_counts and tower_var in activity_counts:
#                             found_tower = tower_var
#                             break
                    
#                     if found_tower:
#                         ai_data = activity_counts[found_tower]
#                         # Extract the total from AI response data
#                         if isinstance(ai_data, list):
#                             for category_data in ai_data:
#                                 if isinstance(category_data, dict) and 'Activities' in category_data:
#                                     for activity_data in category_data['Activities']:
#                                         if isinstance(activity_data, dict) and activity_data.get('Activity Name') == activity:
#                                             completed_flats = activity_data.get('Total', 0)
#                                             logger.info(f"Found COS data for {found_tower} - {activity}: {completed_flats}")
#                                             break
#                     else:
#                         logger.info(f"No COS data found for {tower} - {activity}, completed_flats = 0")

#                     # *** APPLY FORMULA (count*2)-1 FOR CIVIL WORKS ***
#                     if category == "Civil Works":
#                         completed_flats = (completed_flats * 2) - 1
#                         logger.info(f"Applied Civil Works formula ((count*2)-1) for {activity}: {completed_flats}")

#                     # Calculate Open/Missing check list per clarified requirements
#                     in_progress = 0  # Not calculated in the current code
#                     if completed_flats == 0 or closed_checklist > completed_flats:
#                         open_missing = 0
#                     else:
#                         open_missing = abs(completed_flats - closed_checklist)

#                     # Use the first Asite activity name for display
#                     display_activity = asite_activities[0] if isinstance(asite_activity, list) else asite_activity

#                     consolidated_rows.append({
#                         "Tower": tower,
#                         "Category": category,
#                         "Activity Name": display_activity,
#                         "Completed Work*(Count of Flat)": completed_flats,
#                         "In progress": in_progress,
#                         "Closed checklist": closed_checklist,
#                         "Open/Missing check list": open_missing
#                     })

#         # Create DataFrame
#         df = pd.DataFrame(consolidated_rows)
#         if df.empty:
#             st.warning("No data available to generate consolidated checklist.")
#             return None

#         # Sort by Tower and Category for consistency
#         df.sort_values(by=["Tower", "Category"], inplace=True)
        
#         logger.info(f"\n{'='*60}")
#         logger.info(f"Final consolidated data summary:")
#         logger.info(f"Total rows: {len(df)}")
#         logger.info(f"Towers included: {df['Tower'].unique().tolist()}")
#         logger.info(f"{'='*60}\n")

#         # Create a BytesIO buffer for the Excel file
#         output = io.BytesIO()
#         workbook = xlsxwriter.Workbook(output)
#         worksheet = workbook.add_worksheet("Consolidated Checklist")

#         # Define styles
#         header_format = workbook.add_format({'bold': True, 'bg_color': '#D3D3D3'})
#         total_format = workbook.add_format({'bold': True, 'bg_color': '#FFDAB9'})
#         cell_format = workbook.add_format({'border': 1})

#         # Column headers
#         headers = ["Activity Name", "Completed", "In progress", "Closed checklist", "Open/Missing check list"]

#         # Starting positions for each section
#         col_start = 1  # Start from column B
#         row_start = 0

#         # Group by Tower
#         grouped_by_tower = df.groupby('Tower', sort=False)

#         for tower, tower_group in grouped_by_tower:
#             # Reset column position for each tower
#             col_pos = col_start

#             # Group Categories within this Tower
#             grouped_by_category = tower_group.groupby('Category', sort=False)

#             # Process each Category side by side
#             for category, cat_group in grouped_by_category:
#                 # Write category header
#                 worksheet.merge_range(row_start, col_pos, row_start, col_pos + 4, f"{tower} {category} Checklist Status", header_format)
#                 row_pos = row_start + 1

#                 # Write column headers
#                 for i, header in enumerate(headers, start=0):
#                     worksheet.write(row_pos, col_pos + i, header, header_format)
#                 row_pos += 1

#                 # Write activity data
#                 for _, row in cat_group.iterrows():
#                     worksheet.write(row_pos, col_pos, row["Activity Name"], cell_format)
#                     worksheet.write(row_pos, col_pos + 1, row["Completed Work*(Count of Flat)"], cell_format)
#                     worksheet.write(row_pos, col_pos + 2, row["In progress"], cell_format)
#                     worksheet.write(row_pos, col_pos + 3, row["Closed checklist"], cell_format)
#                     worksheet.write(row_pos, col_pos + 4, row["Open/Missing check list"], cell_format)
#                     row_pos += 1

#                 # Write total pending checklist
#                 total_pending = cat_group["Open/Missing check list"].sum()
#                 worksheet.merge_range(row_pos, col_pos, row_pos, col_pos + 3, "Total pending check list", total_format)
#                 worksheet.write(row_pos, col_pos + 4, total_pending, total_format)
#                 row_pos += 2

#                 # Move to the next column position (side-by-side sections)
#                 col_pos += 6

#             # Move to the next tower (below the current sections)
#             row_start = row_pos

#         # Auto-adjust column widths for Sheet 1
#         for col in range(col_start, col_pos):
#             worksheet.set_column(col, col, 20)

#         # Create Sheet 2: Checklist with current month
#         current_month = datetime.now().strftime("%B")  # Gets full month name (e.g., "January", "February")
#         sheet2_name = f"Checklist {current_month}"
#         worksheet2 = workbook.add_worksheet(sheet2_name)
#         current_row = 0

#         # Write title
#         worksheet2.write(current_row, 0, f"Checklist: {current_month}", header_format)
#         current_row += 1

#         # Write headers
#         headers = [
#             "Site",
#             "Total of Missing & Open Checklist-Civil",
#             "Total of Missing & Open Checklist-MEP",
#             "Total of Missing & Open Checklist-Interior Finishing",
#             "TOTAL"
#         ]
#         for col, header in enumerate(headers, start=0):
#             worksheet2.write(current_row, col, header, header_format)
#         current_row += 1

#         # Categorize categories into Civil, MEP, and Interior Finishing
#         def map_category_to_type(category):
#             if category in ["Civil Works"]:
#                 return "Civil"
#             elif category in ["MEP Works"]:
#                 return "MEP"
#             elif category in ["Interior Finishing Works"]:
#                 return "Interior"
#             else:
#                 return "Civil"  # Default to Civil

#         # Aggregate open/missing counts by tower and type (Civil/MEP) for ALL towers
#         summary_data = {}
#         for _, row in df.iterrows():
#             tower = row["Tower"]
#             category = row["Category"]
#             open_missing = row["Open/Missing check list"]
            
#             # Convert tower name to display format
#             tower_num = tower.replace("Tower ", "")
#             if len(tower_num) == 1:
#                 tower_num = f"0{tower_num}"
#             site_name = f"Eden-Tower {tower_num}"

#             type_ = map_category_to_type(category)
            
#             if site_name not in summary_data:
#                 summary_data[site_name] = {"Civil": 0, "MEP": 0, "Interior": 0}
            
#             summary_data[site_name][type_] += open_missing

#         logger.info(f"Summary data for Sheet 2: {summary_data}")

#         # Write summary data to Sheet 2 (all towers, even if zero)
#         for site_name, counts in sorted(summary_data.items()):
#             civil_count = counts["Civil"]
#             mep_count = counts["MEP"]
#             interior_count = counts["Interior"]
#             total_count = civil_count + mep_count + interior_count
            
#             worksheet2.write(current_row, 0, site_name, cell_format)
#             worksheet2.write(current_row, 1, civil_count, cell_format)
#             worksheet2.write(current_row, 2, mep_count, cell_format)
#             worksheet2.write(current_row, 3, interior_count, cell_format)
#             worksheet2.write(current_row, 4, total_count, cell_format)
            
#             current_row += 1

#         # Create Sheet 3: External Development Activities (Always create this sheet)
#         worksheet3 = workbook.add_worksheet("External Development Activities")
#         current_row = 0

#         # Write title
#         worksheet3.write(current_row, 0, "External Development Activities - Consolidated Report", header_format)
#         current_row += 2

#         # External Development Activities - Always include all activities
#         external_dev_activities = [
#             "Granular Sub-Base", 
#             "Kerb Stone", 
#             "Rain Water / Storm Line", 
#             "Saucer Drain / Paver Block",
#             "Sewer Line", 
#             "Stamp Concrete", 
#             "Storm Line", 
#             "WMM"
#         ]

#         # Write headers for External Development
#         ext_headers = ["Activity Name", "Total Count Across All Towers"]
#         for col, header in enumerate(ext_headers, start=0):
#             worksheet3.write(current_row, col, header, header_format)
#         current_row += 1

#         # Always process and display all external development activities
#         logger.info("Creating External Development Activities table with all activities...")
#         grand_total = 0
        
#         for activity in external_dev_activities:
#             total_count = 0
            
#             # Check all towers for this activity
#             if activity_counts:
#                 tower_keys = list(activity_counts.keys())
#                 logger.info(f"Processing activity: {activity} across towers: {tower_keys}")
                
#                 for tower_key in tower_keys:
#                     ai_data = activity_counts[tower_key]
                    
#                     # Extract the total from AI response data
#                     if isinstance(ai_data, list):
#                         for category_data in ai_data:
#                             if isinstance(category_data, dict) and 'Category' in category_data:
#                                 if category_data['Category'] == 'External Development Activities':
#                                     for activity_data in category_data['Activities']:
#                                         if isinstance(activity_data, dict) and activity_data.get('Activity Name') == activity:
#                                             activity_total = activity_data.get('Total', 0)
#                                             total_count += activity_total
#                                             logger.info(f"Found {activity} in {tower_key}: {activity_total}")
#                                             break
            
#             # Always write the activity to the sheet, even if count is 0
#             worksheet3.write(current_row, 0, activity, cell_format)
#             worksheet3.write(current_row, 1, total_count, cell_format)
#             grand_total += total_count
#             current_row += 1
#             logger.info(f"Added {activity} to External Development sheet with total: {total_count}")

#         # Add a summary row with grand total
#         current_row += 1
#         worksheet3.write(current_row, 0, "GRAND TOTAL", total_format)
#         worksheet3.write(current_row, 1, grand_total, total_format)
        
#         # Add note about future updates
#         current_row += 2
#         worksheet3.write(current_row, 0, "Note: This table will reflect updated counts when data source is refreshed", 
#                         workbook.add_format({'italic': True, 'font_size': 10}))
        
#         logger.info(f"External Development Activities sheet created with {len(external_dev_activities)} activities and grand total: {grand_total}")

#         # Auto-adjust column widths for all sheets
#         for col in range(5):
#             worksheet2.set_column(col, col, 25)
#         for col in range(2):
#             worksheet3.set_column(col, col, 25)

#         # Close the workbook
#         workbook.close()
#         output.seek(0)
#         return output

#     except Exception as e:
#         logger.error(f"Error generating consolidated Excel: {str(e)}")
#         st.error(f"❌ Error generating Excel file: {str(e)}")
#         import traceback
#         logger.error(f"Full traceback: {traceback.format_exc()}")
#         return None
    

# # Combined function to handle analysis and display
# def run_analysis_and_display():
#     try:
#         st.write("Running status analysis...")
#         AnalyzeStatusManually()
#         st.success("Status analysis completed successfully!")

#         # Ensure ai_response is initialized
#         if 'ai_response' not in st.session_state or not isinstance(st.session_state.ai_response, dict):
#             st.session_state.ai_response = {}
#             logger.info("Initialized st.session_state.ai_response in run_analysis_and_display")

#         # Step 1: Display activity counts and generate AI data
#         st.write("Displaying activity counts and generating AI data...")
#         display_activity_count()
#         st.success("Activity counts displayed successfully!")

#         # Step 2: Check AI data totals
#         st.write("Checking AI data totals...")
#         logger.info(f"st.session_state.ai_response contents: {st.session_state.ai_response}")
#         if not st.session_state.ai_response:
#             st.error("❌ No AI data available in st.session_state.ai_response. Attempting to regenerate.")
#             logger.error("No AI data in st.session_state.ai_response after display_activity_count")
#             # Re-run display_activity_count to attempt recovery
#             display_activity_count()
#             if not st.session_state.ai_response:
#                 st.error("❌ Failed to regenerate AI data. Please check data fetching and try again.")
#                 logger.error("Failed to regenerate AI data")
#                 return

#         # Step 3: Generate consolidated checklist Excel file
#         st.write("Generating consolidated checklist Excel file...")
#         structure_analysis = st.session_state.get('structure_analysis', None)
#         if structure_analysis is None:
#             st.error("❌ No structure analysis data available. Please ensure analysis ran successfully.")
#             logger.error("No structure_analysis in st.session_state")
#             return

#         with st.spinner("Generating Excel file... This may take a moment."):
#             excel_file = generate_consolidated_Checklist_excel(structure_analysis, st.session_state.ai_response)
        
#         if excel_file:
#             timestamp = pd.Timestamp.now(tz='Asia/Kolkata').strftime('%Y%m%d_%H%M')
#             file_name = f"Consolidated_Checklist_Eden_{timestamp}.xlsx"
            
#             col1, col2, col3 = st.columns([1, 2, 1])
#             with col2:
#                 st.sidebar.download_button(
#                     label="📥 Download Checklist Excel",
#                     data=excel_file,
#                     file_name=file_name,
#                     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#                     key="download_excel_button",
#                     help="Click to download the consolidated checklist in Excel format."
#                 )
#             st.success("Excel file generated successfully! Click the button above to download.")
#         else:
#             st.error("Failed to generate Excel file. Please check the logs for details.")
#             logger.error("Failed to generate Excel file")

#     except Exception as e:
#         st.error(f"Error during analysis, display, or Excel generation: {str(e)}")
#         logger.error(f"Error during analysis, display, or Excel generation: {str(e)}")

# # Streamlit UI
# st.markdown(
#     """
#     <h1 style='font-family: "Arial Black", Gadget, sans-serif; 
#                color: red; 
#                font-size: 48px; 
#                text-align: center;'>
#         Eden CheckList - Report
#     </h1>
#     """,
#     unsafe_allow_html=True
# )

# # Initialize and Fetch Data
# st.sidebar.title("🔐 Asite Initialization")
# email = st.sidebar.text_input("Email", "impwatson@gadieltechnologies.com", key="email_input")
# password = st.sidebar.text_input("Password", "Srihari@790$", type="password", key="password_input")

# if st.sidebar.button("Initialize and Fetch Data"):
#     loop = asyncio.new_event_loop()
#     asyncio.set_event_loop(loop)
#     try:
#         success = loop.run_until_complete(initialize_and_fetch_data(email, password))
#         if success:
#             st.sidebar.success("Initialization and data fetching completed successfully!")
#         else:
#             st.sidebar.error("Initialization and data fetching failed!")
#     except Exception as e:
#         st.sidebar.error(f"Initialization and data fetching failed: {str(e)}")
#     finally:
#         loop.close()

# # Analyze and Display
# st.sidebar.title("📊 Status Analysis")
# if st.sidebar.button("Analyze and Display Activity Counts"):
#     with st.spinner("Running analysis and displaying activity counts..."):
#         run_analysis_and_display()
































































import streamlit as st
import requests
import json
import urllib.parse
import urllib3
import certifi
import pandas as pd
from datetime import datetime
import re
import logging
import os
from dotenv import load_dotenv
import aiohttp
import asyncio
from concurrent.futures import ThreadPoolExecutor, as_completed
import time
import openpyxl
import io
from dotenv import load_dotenv
from uuid import uuid4
import ibm_boto3
from ibm_botocore.client import Config
from tenacity import retry, stop_after_attempt, wait_exponential
import xlsxwriter

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Disable SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Load environment variables
load_dotenv()

# IBM COS Configuration
COS_API_KEY = os.getenv("COS_API_KEY")
COS_SERVICE_INSTANCE_ID = os.getenv("COS_SERVICE_INSTANCE_ID")
COS_ENDPOINT = os.getenv("COS_ENDPOINT")
COS_BUCKET = os.getenv("COS_BUCKET")

# WatsonX configuration
WATSONX_API_URL = os.getenv("WATSONX_API_URL_1")
MODEL_ID = os.getenv("MODEL_ID_1")
PROJECT_ID = os.getenv("PROJECT_ID_1")
API_KEY = os.getenv("API_KEY_1")

# API Endpoints
LOGIN_URL = "https://dms.asite.com/apilogin/"
IAM_TOKEN_URL = "https://iam.cloud.ibm.com/identity/token"

# Login Function
async def login_to_asite(email, password):
    headers = {"Accept": "application/json", "Content-Type": "application/x-www-form-urlencoded"}
    payload = {"emailId": email, "password": password}
    response = requests.post(LOGIN_URL, headers=headers, data=payload, verify=certifi.where(), timeout=50)
    if response.status_code == 200:
        try:
            session_id = response.json().get("UserProfile", {}).get("Sessionid")
            logger.info(f"Login successful, Session ID: {session_id}")
            st.session_state.sessionid = session_id
            st.sidebar.success(f"✅ Login successful, Session ID: {session_id}")
            return session_id
        except json.JSONDecodeError:
            logger.error("JSONDecodeError during login")
            st.sidebar.error("❌ Failed to parse login response")
            return None
    logger.error(f"Login failed: {response.status_code} - {response.text}")
    st.sidebar.error(f"❌ Login failed: {response.status_code} - {response.text}")
    return None

# Function to generate access token
@retry(stop=stop_after_attempt(5), wait=wait_exponential(multiplier=2, min=10, max=60))
def get_access_token(API_KEY):
    headers = {"Content-Type": "application/x-www-form-urlencoded", "Accept": "application/json"}
    data = {"grant_type": "urn:ibm:params:oauth:grant-type:apikey", "apikey": API_KEY}
    response = requests.post(IAM_TOKEN_URL, headers=headers, data=data, verify=certifi.where(), timeout=50)
    try:
        if response.status_code == 200:
            token_info = response.json()
            logger.info("Access token generated successfully")
            return token_info['access_token']
        else:
            logger.error(f"Failed to get access token: {response.status_code} - {response.text}")
            st.error(f"❌ Failed to get access token: {response.status_code} - {response.text}")
            raise Exception("Failed to get access token")
    except Exception as e:
        logger.error(f"Exception getting access token: {str(e)}")
        st.error(f"❌ Error getting access token: {str(e)}")
        return None

# Initialize COS client
@retry(stop=stop_after_attempt(5), wait=wait_exponential(multiplier=1, min=4, max=10))
def initialize_cos_client():
    try:
        logger.info("Attempting to initialize COS client...")
        cos_client = ibm_boto3.client(
            's3',
            ibm_api_key_id=COS_API_KEY,
            ibm_service_instance_id=COS_SERVICE_INSTANCE_ID,
            config=Config(
                signature_version='oauth',
                connect_timeout=180,
                read_timeout=180,
                retries={'max_attempts': 15}
            ),
            endpoint_url=COS_ENDPOINT
        )
        logger.info("COS client initialized successfully")
        return cos_client
    except Exception as e:
        logger.error(f"Error initializing COS client: {str(e)}")
        st.error(f"❌ Error initializing COS client: {str(e)}")
        raise

# Fetch Workspace ID
async def GetWorkspaceID():
    url = "https://dmsak.asite.com/api/workspace/workspacelist"
    headers = {
        'Cookie': f'ASessionID={st.session_state.sessionid}',
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    response = requests.get(url, headers=headers)
    if response.status_code != 200:
        st.error(f"Failed to fetch workspace list: {response.status_code} - {response.text}")
        raise Exception(f"Failed to fetch workspace list: {response.status_code}")
    try:
        data = response.json()
        st.write(f"Workspace list response: {data}")
        st.session_state.workspaceid = data['asiteDataList']['workspaceVO'][2]['Workspace_Id']
        st.write(f"Workspace ID: {st.session_state.workspaceid}")
    except (KeyError, IndexError) as e:
        st.error(f"Error parsing workspace ID: {str(e)}")
        raise

# Fetch Project IDs
async def GetProjectId():
    url = f"https://adoddleak.asite.com/commonapi/qaplan/getQualityPlanList;searchCriteria={{'criteria': [{{'field': 'planCreationDate','operator': 6,'values': ['11-Mar-2025']}}], 'projectId': {str(st.session_state.workspaceid)}, 'recordLimit': 1000, 'recordStart': 1}}"
    headers = {
        'Cookie': f'ASessionID={st.session_state.sessionid}',
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    response = requests.get(url, headers=headers)
    if response.status_code != 200:
        st.error(f"Failed to fetch project IDs: {response.status_code} - {response.text}")
        raise Exception(f"Failed to fetch project IDs: {response.status_code}")
    data = response.json()
    st.write(f"Project ID response: {data}")
    if not data.get('data'):
        st.error("No quality plans found for the specified date.")
        raise Exception("No quality plans found")
    st.session_state.Eden_structure = data['data'][0]['planId']
    st.write(f"Eden Structure Project ID: {st.session_state.Eden_structure}")

# Asynchronous Fetch Function
async def fetch_data(session, url, headers):
    async with session.get(url, headers=headers) as response:
        if response.status == 200:
            return await response.json()
        elif response.status == 204:
            return None
        else:
            raise Exception(f"Error fetching data: {response.status} - {await response.text()}")

# Fetch All Structure Data
async def GetAllDatas():
    record_limit = 1000
    headers = {'Cookie': f'ASessionID={st.session_state.sessionid}'}
    all_structure_data = []

    async with aiohttp.ClientSession() as session:
        start_record = 1
        st.write("Fetching Eden Structure data...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanAssociation/?projectId={st.session_state.workspaceid}&planId={st.session_state.Eden_structure}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                async with session.get(url, headers=headers) as response:
                    if response.status == 204:
                        st.write("No more Structure data available (204)")
                        break
                    data = await response.json()
                    if 'associationList' in data and data['associationList']:
                        all_structure_data.extend(data['associationList'])
                    else:
                        all_structure_data.extend(data if isinstance(data, list) else [])
                    st.write(f"Fetched {len(all_structure_data[-record_limit:])} Structure records (Total: {len(all_structure_data)})")
                    if len(all_structure_data[-record_limit:]) < record_limit:
                        break
                    start_record += record_limit
            except Exception as e:
                st.error(f"❌ Error fetching Structure data: {str(e)}")
                break

    df_structure = pd.DataFrame(all_structure_data)
    
    desired_columns = ['activitySeq', 'qiLocationId']
    if 'statusName' in df_structure.columns:
        desired_columns.append('statusName')
    elif 'statusColor' in df_structure.columns:
        desired_columns.append('statusColor')
        status_mapping = {'#4CAF50': 'Completed', '#4CB0F0': 'Not Started', '#4C0F0': 'Not Started'}
        df_structure['statusName'] = df_structure['statusColor'].map(status_mapping).fillna('Unknown')
        desired_columns.append('statusName')
    else:
        st.error("❌ Neither statusName nor statusColor found in data!")
        return pd.DataFrame()

    eden_structure = df_structure[desired_columns]

    st.write(f"EDEN STRUCTURE ({', '.join(desired_columns)})")
    st.write(f"Total records: {len(eden_structure)}")
    st.write(eden_structure)
    
    return eden_structure

# Fetch Activity Data
async def Get_Activity():
    record_limit = 1000
    headers = {
        'Cookie': f'ASessionID={st.session_state.sessionid}',
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    
    all_structure_activity_data = []
    
    async with aiohttp.ClientSession() as session:
        start_record = 1
        st.write("Fetching Activity data for Eden Structure...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanActivities/?projectId={st.session_state.workspaceid}&planId={st.session_state.Eden_structure}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more Structure Activity data available (204)")
                    break
                if 'activityList' in data and data['activityList']:
                    all_structure_activity_data.extend(data['activityList'])
                else:
                    all_structure_activity_data.extend(data if isinstance(data, list) else [])
                st.write(f"Fetched {len(all_structure_activity_data[-record_limit:])} Structure Activity records (Total: {len(all_structure_activity_data)})")
                if len(all_structure_activity_data[-record_limit:]) < record_limit:
                    break
                start_record += record_limit
            except Exception as e:
                st.error(f"❌ Error fetching Structure Activity data: {str(e)}")
                break
 
    structure_activity_data = pd.DataFrame(all_structure_activity_data)[['activityName', 'activitySeq', 'formTypeId']]

    st.write("EDEN STRUCTURE ACTIVITY DATA (activityName and activitySeq)")
    st.write(f"Total records: {len(structure_activity_data)}")
    st.write(structure_activity_data)
      
    return structure_activity_data

# Fetch Location/Module Data
async def Get_Location():
    record_limit = 1000
    headers = {
        'Cookie': f'ASessionID={st.session_state.sessionid}',
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    
    all_structure_location_data = []
    
    async with aiohttp.ClientSession() as session:
        start_record = 1
        total_records_fetched = 0
        st.write("Fetching Eden Structure Location/Module data...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanLocation/?projectId={st.session_state.workspaceid}&planId={st.session_state.Eden_structure}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more Structure Location data available (204)")
                    break
                if isinstance(data, list):
                    location_data = [{'qiLocationId': item.get('qiLocationId', ''), 'qiParentId': item.get('qiParentId', ''), 'name': item.get('name', '')} 
                                   for item in data if isinstance(item, dict)]
                    all_structure_location_data.extend(location_data)
                    total_records_fetched = len(all_structure_location_data)
                    st.write(f"Fetched {len(location_data)} Structure Location records (Total: {total_records_fetched})")
                elif isinstance(data, dict) and 'locationList' in data and data['locationList']:
                    location_data = [{'qiLocationId': loc.get('qiLocationId', ''), 'qiParentId': loc.get('qiParentId', ''), 'name': loc.get('name', '')} 
                                   for loc in data['locationList']]
                    all_structure_location_data.extend(location_data)
                    total_records_fetched = len(all_structure_location_data)
                    st.write(f"Fetched {len(location_data)} Structure Location records (Total: {total_records_fetched})")
                else:
                    st.warning(f"No 'locationList' in Structure Location data or empty list.")
                    break
                if len(location_data) < record_limit:
                    break
                start_record += record_limit
            except Exception as e:
                st.error(f"❌ Error fetching Structure Location data: {str(e)}")
                break
        
    structure_df = pd.DataFrame(all_structure_location_data)
    
    if 'name' in structure_df.columns and structure_df['name'].isna().all():
        st.error("❌ All 'name' values in Structure Location data are missing or empty!")

    st.write("EDEN STRUCTURE LOCATION/MODULE DATA")
    st.write(f"Total records: {len(structure_df)}")
    st.write(structure_df)
    
    st.session_state.structure_location_data = structure_df
    
    return structure_df

# Process individual chunk
def process_chunk(chunk, chunk_idx, dataset_name, location_df):
    logger.info(f"Starting thread for {dataset_name} Chunk {chunk_idx + 1}")
    generated_text = format_chunk_locally(chunk, chunk_idx, len(chunk), dataset_name, location_df)
    logger.info(f"Completed thread for {dataset_name} Chunk {chunk_idx + 1}")
    return generated_text, chunk_idx

# Process data with manual counting
def process_manually(analysis_df, total, dataset_name, chunk_size=1000, max_workers=4):
    if analysis_df.empty:
        st.warning(f"No completed activities found for {dataset_name}.")
        return "No completed activities found."

    unique_activities = analysis_df['activityName'].unique()
    logger.info(f"Unique activities in {dataset_name} dataset: {list(unique_activities)}")
    logger.info(f"Total records in {dataset_name} dataset: {len(analysis_df)}")

    st.write(f"Saved Eden {dataset_name} data to eden_{dataset_name.lower()}_data.json")
    chunks = [analysis_df[i:i + chunk_size] for i in range(0, len(analysis_df), chunk_size)]

    location_df = st.session_state.structure_location_data

    chunk_results = {}
    progress_bar = st.progress(0)
    status_text = st.empty()

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_chunk = {
            executor.submit(process_chunk, chunk, idx, dataset_name, location_df): idx 
            for idx, chunk in enumerate(chunks)
        }

        completed_chunks = 0
        for future in as_completed(future_to_chunk):
            chunk_idx = future_to_chunk[future]
            try:
                generated_text, idx = future.result()
                chunk_results[idx] = generated_text
                completed_chunks += 1
                progress_percent = completed_chunks / len(chunks)
                progress_bar.progress(progress_percent)
                status_text.text(f"Processed chunk {completed_chunks} of {len(chunks)} ({progress_percent:.1%} complete)")
            except Exception as e:
                logger.error(f"Error processing chunk {chunk_idx + 1} for {dataset_name}: {str(e)}")
                st.error(f"❌ Error processing chunk {chunk_idx + 1}: {str(e)}")

    parsed_data = {}
    for chunk_idx in sorted(chunk_results.keys()):
        generated_text = chunk_results[chunk_idx]
        if not generated_text:
            continue

        current_tower = None
        tower_activities = []
        lines = generated_text.split("\n")
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            if line.startswith("Tower:"):
                try:
                    tower_parts = line.split("Tower:", 1)
                    if len(tower_parts) > 1:
                        if current_tower and tower_activities:
                            if current_tower not in parsed_data:
                                parsed_data[current_tower] = []
                            parsed_data[current_tower].extend(tower_activities)
                            tower_activities = []
                        current_tower = tower_parts[1].strip()
                except Exception as e:
                    logger.warning(f"Error parsing Tower line: {line}, error: {str(e)}")
                    if not current_tower:
                        current_tower = f"Unknown Tower {chunk_idx}"
                    
            elif line.startswith("Total Completed Activities:"):
                continue
            elif not line.strip().startswith("activityName"):
                try:
                    parts = re.split(r'\s{2,}', line.strip())
                    if len(parts) >= 2:
                        activity_name = ' '.join(parts[:-1]).strip()
                        count_str = parts[-1].strip()
                        count_match = re.search(r'\d+', count_str)
                        if count_match:
                            count = int(count_match.group())
                            if current_tower:
                                tower_activities.append({
                                    "activityName": activity_name,
                                    "completedCount": count
                                })
                    else:
                        match = re.match(r'^\s*(.+?)\s+(\d+)$', line)
                        if match and current_tower:
                            activity_name = match.group(1).strip()
                            count = int(match.group(2).strip())
                            tower_activities.append({
                                "activityName": activity_name,
                                "completedCount": count
                            })
                except (ValueError, IndexError) as e:
                    logger.warning(f"Skipping malformed activity line: {line}, error: {str(e)}")

        if current_tower and tower_activities:
            if current_tower not in parsed_data:
                parsed_data[current_tower] = []
            parsed_data[current_tower].extend(tower_activities)

    aggregated_data = {}
    for tower_name, activities in parsed_data.items():
        tower_short_name = tower_name.split('/')[1] if '/' in tower_name else tower_name
        if tower_short_name not in aggregated_data:
            aggregated_data[tower_short_name] = {}
        
        for activity in activities:
            name = activity.get("activityName", "Unknown")
            count = activity.get("completedCount", 0)
            if name in aggregated_data[tower_short_name]:
                aggregated_data[tower_short_name][name] += count
            else:
                aggregated_data[tower_short_name][name] = count

    combined_output_lines = []
    sorted_towers = sorted(aggregated_data.keys())
    
    for i, tower_short_name in enumerate(sorted_towers):
        combined_output_lines.append(f"{tower_short_name:<11} activityName            CompletedCount")
        activity_dict = aggregated_data[tower_short_name]
        tower_total = 0
        for name, count in sorted(activity_dict.items()):
            combined_output_lines.append(f"{'':<11} {name:<23} {count:>14}")
            tower_total += count
        combined_output_lines.append(f"{'':<11} Total for {tower_short_name:<11}: {tower_total:>14}")
        if i < len(sorted_towers) - 1:
            combined_output_lines.append("")
    
    combined_output = "\n".join(combined_output_lines)
    return combined_output

# Local formatting function for manual counting
def format_chunk_locally(chunk, chunk_idx, chunk_size, dataset_name, location_df):
    towers_data = {}
    
    for _, row in chunk.iterrows():
        tower_name = row['tower_name']
        activity_name = row['activityName']
        count = int(row['CompletedCount'])
        
        if tower_name not in towers_data:
            towers_data[tower_name] = []
            
        towers_data[tower_name].append({
            "activityName": activity_name,
            "completedCount": count
        })
    
    output = ""
    total_activities = 0
    
    for tower_name, activities in sorted(towers_data.items()):
        output += f"Tower: {tower_name}\n"
        output += "activityName            CompletedCount\n"
        activity_dict = {}
        for activity in activities:
            name = activity['activityName']
            count = activity['completedCount']
            activity_dict[name] = activity_dict.get(name, 0) + count
        for name, count in sorted(activity_dict.items()):
            output += f"{name:<30} {count}\n"
            total_activities += count
    
    output += f"Total Completed Activities: {total_activities}"
    return output

def process_data(df, activity_df, location_df, dataset_name):
    completed = df[df['statusName'] == 'Completed']
    if completed.empty:
        logger.warning(f"No completed activities found in {dataset_name} data.")
        return pd.DataFrame(), 0

    completed = completed.merge(location_df[['qiLocationId', 'name']], on='qiLocationId', how='left')
    completed = completed.merge(activity_df[['activitySeq', 'activityName']], on='activitySeq', how='left')

    if 'qiActivityId' not in completed.columns:
        completed['qiActivityId'] = completed['qiLocationId'].astype(str) + '$$' + completed['activitySeq'].astype(str)

    if completed['name'].isna().all():
        logger.error(f"All 'name' values are missing in {dataset_name} data after merge!")
        st.error(f"❌ All 'name' values are missing in {dataset_name} data after merge! Check location data.")
        completed['name'] = 'Unknown'
    else:
        completed['name'] = completed['name'].fillna('Unknown')

    completed['activityName'] = completed['activityName'].fillna('Unknown')

    parent_child_dict = dict(zip(location_df['qiLocationId'], location_df['qiParentId']))
    name_dict = dict(zip(location_df['qiLocationId'], location_df['name']))

    def get_full_path(location_id):
        path = []
        current_id = location_id
        max_depth = 10
        depth = 0
        
        while current_id and depth < max_depth:
            if current_id not in parent_child_dict or current_id not in name_dict:
                logger.warning(f"Location ID {current_id} not found in parent_child_dict or name_dict. Path so far: {path}")
                break
            
            parent_id = parent_child_dict.get(current_id)
            name = name_dict.get(current_id, "Unknown")
            
            if not parent_id:
                if name != "Quality":
                    path.append(name)
                    path.append("Quality")
                else:
                    path.append(name)
                break
            
            path.append(name)
            current_id = parent_id
            depth += 1
        
        if depth >= max_depth:
            logger.warning(f"Max depth reached while computing path for location_id {location_id}. Possible circular reference. Path: {path}")
        
        if not path:
            logger.warning(f"No path constructed for location_id {location_id}. Using 'Unknown'.")
            return "Unknown"
        
        full_path = '/'.join(reversed(path))
        logger.debug(f"Full path for location_id {location_id}: {full_path}")
        return full_path

    completed['full_path'] = completed['qiLocationId'].apply(get_full_path)

    # **NEW: Filter to only include records with "/Pour" in the full path**
    logger.info(f"Total completed records before /Pour filter: {len(completed)}")
    completed = completed[completed['full_path'].str.contains('/Pour', case=False, na=False)]
    logger.info(f"Total completed records after /Pour filter: {len(completed)}")
    
    if completed.empty:
        logger.warning(f"No completed activities with /Pour in path found in {dataset_name} data after filtering.")
        st.warning(f"No completed activities with /Pour in path found for {dataset_name}")
        return pd.DataFrame(), 0

    def get_tower_name(full_path):
        parts = full_path.split('/')
        if len(parts) < 2:
            return full_path
        
        tower = parts[1]
        if tower == "Tower 4" and len(parts) > 2:
            module = parts[2]
            module_number = module.replace("Module ", "").strip()
            try:
                module_num = int(module_number)
                if 1 <= module_num <= 4:
                    return "Tower 4(B)"
                elif 5 <= module_num <= 8:
                    return "Tower 4(A)"
            except ValueError:
                logger.warning(f"Could not parse module number from {module} in path {full_path}")
        
        return tower

    completed['tower_name'] = completed['full_path'].apply(get_tower_name)

    # **NEW: Extract tower number for better matching**
    def extract_tower_key(tower_name):
        """Extract tower key like 'T4', 'T5', 'T6', 'T7' from tower name"""
        if 'Tower 4' in tower_name or 'T4' in tower_name:
            return 'T4'
        elif 'Tower 5' in tower_name or 'T5' in tower_name:
            return 'T5'
        elif 'Tower 6' in tower_name or 'T6' in tower_name:
            return 'T6'
        elif 'Tower 7' in tower_name or 'T7' in tower_name:
            return 'T7'
        else:
            # Try to extract number from tower name
            match = re.search(r'Tower\s*(\d+)', tower_name, re.IGNORECASE)
            if match:
                return f'T{match.group(1)}'
            return tower_name

    completed['tower_key'] = completed['tower_name'].apply(extract_tower_key)

    # Group by tower_key and activityName to get counts
    analysis = completed.groupby(['tower_key', 'activityName'])['qiLocationId'].nunique().reset_index(name='CompletedCount')
    
    # Also keep tower_name for display purposes
    tower_mapping = completed[['tower_key', 'tower_name']].drop_duplicates()
    analysis = analysis.merge(tower_mapping, on='tower_key', how='left')
    
    analysis = analysis.sort_values(by=['tower_key', 'activityName'], ascending=True)
    total_completed = analysis['CompletedCount'].sum()

    logger.info(f"Total completed activities for {dataset_name} after processing with /Pour filter: {total_completed}")
    
    # Log some sample paths to verify filtering
    if not completed.empty:
        sample_paths = completed['full_path'].head(10).tolist()
        logger.info(f"Sample full paths after /Pour filter: {sample_paths}")
    
    return analysis, total_completed

# Main analysis function for Eden Structure
def AnalyzeStatusManually(email=None, password=None):
    start_time = time.time()

    if 'sessionid' not in st.session_state:
        st.error("❌ Please log in first!")
        return

    required_data = [
        'eden_structure',
        'structure_activity_data',
        'structure_location_data'
    ]
    
    for data_key in required_data:
        if data_key not in st.session_state:
            st.error(f"❌ Please fetch required data first! Missing: {data_key}")
            return
        if not isinstance(st.session_state[data_key], pd.DataFrame):
            st.error(f"❌ {data_key} is not a DataFrame! Found type: {type(st.session_state[data_key])}")
            return

    structure_data = st.session_state.eden_structure
    structure_activity = st.session_state.structure_activity_data
    structure_locations = st.session_state.structure_location_data
    
    for df, name in [(structure_data, "Structure")]:
        if 'statusName' not in df.columns:
            st.error(f"❌ statusName column not found in {name} data!")
            return
        if 'qiLocationId' not in df.columns:
            st.error(f"❌ qiLocationId column not found in {name} data!")
            return
        if 'activitySeq' not in df.columns:
            st.error(f"❌ activitySeq column not found in {name} data!")
            return

    for df, name in [(structure_locations, "Structure Location")]:
        if 'qiLocationId' not in df.columns or 'name' not in df.columns:
            st.error(f"❌ qiLocationId or name column not found in {name} data!")
            return

    for df, name in [(structure_activity, "Structure Activity")]:
        if 'activitySeq' not in df.columns or 'activityName' not in df.columns:
            st.error(f"❌ activitySeq or activityName column not found in {name} data!")
            return

    structure_analysis, structure_total = process_data(structure_data, structure_activity, structure_locations, "Structure")

    st.write("### Eden Structure Quality Analysis (Completed Activities):")
    st.write("**Full Output (Structure):**")
    structure_output = process_manually(structure_analysis, structure_total, "Structure")
    if structure_output:
        st.text(structure_output)

    # Store structure analysis in session state for Excel generation
    st.session_state.structure_analysis = structure_analysis

    end_time = time.time()
    st.write(f"Total execution time: {end_time - start_time:.2f} seconds")

def get_cos_files():
    try:
        cos_client = initialize_cos_client()
        if not cos_client:
            st.error("❌ Failed to initialize COS client.")
            return None

        st.write(f"Attempting to list objects in bucket '{COS_BUCKET}' with prefix 'Eden/'")
        response = cos_client.list_objects_v2(Bucket=COS_BUCKET, Prefix="Eden/")
        if 'Contents' not in response:
            st.error(f"❌ No files found in the 'Eden' folder of bucket '{COS_BUCKET}'.")
            logger.error("No objects found in Eden folder")
            return None

        all_files = [obj['Key'] for obj in response.get('Contents', [])]
        st.write("**All files in Eden folder:**")
        if all_files:
            st.write("\n".join(all_files))
        else:
            st.write("No files found.")
            logger.warning("Eden folder is empty")
            return None

        pattern = re.compile(
            r"Eden/Structure\s*Work\s*Tracker[\(\s]*(.*?)(?:[\)\s]*\.xlsx)$",
            re.IGNORECASE
        )
        
        date_formats = ["%d-%m-%Y", "%d-%m-%y", "%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y"]

        file_info = []
        for obj in response.get('Contents', []):
            key = obj['Key']
            match = pattern.match(key)
            if match:
                date_str = match.group(1).strip('()').strip()
                parsed_date = None
                for fmt in date_formats:
                    try:
                        parsed_date = datetime.strptime(date_str, fmt)
                        break
                    except ValueError:
                        continue
                if parsed_date:
                    file_info.append({'key': key, 'date': parsed_date})
                else:
                    logger.warning(f"Could not parse date in filename: {key}")
                    st.warning(f"Skipping file with unparseable date: {key}")
            else:
                st.write(f"File '{key}' does not match the expected pattern 'Eden/Structure Work Tracker (DD-MM-YYYY).xlsx'")

        if not file_info:
            st.error("❌ No Excel files matched the expected pattern in the 'Eden' folder.")
            logger.error("No files matched the expected pattern")
            return None

        latest_file = max(file_info, key=lambda x: x['date']) if file_info else None
        if not latest_file:
            st.error("❌ No valid Excel files found for Structure Work Tracker.")
            logger.error("No valid files after date parsing")
            return None

        file_key = latest_file['key']
        st.success(f"Found matching file: {file_key}")
        return file_key
    except Exception as e:
        st.error(f"❌ Error fetching COS files: {str(e)}")
        logger.error(f"Error fetching COS files: {str(e)}")
        return None

# Initialize session state variables
if 'cos_df_tower4' not in st.session_state:
    st.session_state.cos_df_tower4 = None
if 'cos_df_tower5' not in st.session_state:
    st.session_state.cos_df_tower5 = None
if 'cos_df_tower6' not in st.session_state:
    st.session_state.cos_df_tower6 = None
if 'cos_df_tower7' not in st.session_state:
    st.session_state.cos_df_tower7 = None
if 'cos_tname_tower4' not in st.session_state:
    st.session_state.cos_tname_tower4 = None
if 'cos_tname_tower5' not in st.session_state:
    st.session_state.cos_tname_tower5 = None
if 'cos_tname_tower6' not in st.session_state:
    st.session_state.cos_tname_tower6 = None
if 'cos_tname_tower7' not in st.session_state:
    st.session_state.cos_tname_tower7 = None
if 'ai_response' not in st.session_state:
    st.session_state.ai_response = {}  # Initialize as empty dictionary

# Process Excel files for Towers 4, 5, 6, 7
def process_file(file_stream, filename):
    try:
        workbook = openpyxl.load_workbook(file_stream)
        available_sheets = workbook.sheetnames
        st.write(f"Available sheets in {filename}: {', '.join(available_sheets)}")

        target_sheets = ["Tower 4", "Tower 5", "Tower 6", "Tower 7"]
        results = []

        expected_columns = [
            'Tower', 'Activity No.', 'Monthly Lookahead ID', 'Task Name', 
            'Actual Start', 'Actual Finish', '% Complete-MSP', 'Duration', 
            'Start', 'Finish', 'Baseline Duration', 'Baseline Start', 
            'Baseline Finish', 'Week1', 'Week2', 'Week3', 'Week4', 
            'Total for the month', 'Total for the tower', 'Unique ID',
            'Extra Col 1', 'Extra Col 2', 'Extra Col 3', 'Extra Col 4', 
            'Extra Col 5', 'Slab Cycle'
        ]

        for sheet_name in target_sheets:
            if sheet_name not in available_sheets:
                st.warning(f"Sheet '{sheet_name}' not found in file: {filename}")
                continue

            file_stream.seek(0)

            try:
                df = pd.read_excel(file_stream, sheet_name=sheet_name, header=1)
                st.write(f"Raw columns in {sheet_name}: {list(df.columns)}")

                if len(df.columns) != len(expected_columns):
                    st.error(f"Sheet {sheet_name} has {len(df.columns)} columns, but {len(expected_columns)} were expected: {list(df.columns)}")
                    continue

                df.columns = expected_columns

                target_columns = ["Task Name", "Actual Start", "Actual Finish"]
                available_columns = [col for col in target_columns if col in df.columns]

                if len(available_columns) < len(target_columns):
                    missing_cols = [col for col in target_columns if col not in available_columns]
                    st.warning(f"Missing columns in sheet {sheet_name}: {', '.join(missing_cols)}")
                    for col in missing_cols:
                        df[col] = None

                df_original = df.copy()
                df = df[target_columns]
                df = df.dropna(subset=['Task Name'])
                df['Task Name'] = df['Task Name'].astype(str).str.strip()

                if 'Actual Finish' in df.columns:
                    df['Actual_Finish_Original'] = df['Actual Finish'].astype(str)
                    df['Actual Finish'] = pd.to_datetime(df['Actual Finish'], errors='coerce')
                    has_na_mask = (
                        pd.isna(df['Actual Finish']) |
                        (df['Actual_Finish_Original'].str.upper() == 'NAT') |
                        (df['Actual_Finish_Original'].str.lower().isin(['nan', 'na', 'n/a', 'none', '']))
                    )
                    st.write(f"Sample of rows with NA or invalid values in Actual Finish for {sheet_name}:")
                    na_rows = df[has_na_mask][['Task Name', 'Actual Finish']]
                    if not na_rows.empty:
                        st.write(na_rows.head(10))
                    else:
                        st.write("No NA or invalid values found in Actual Finish")
                    df.drop('Actual_Finish_Original', axis=1, inplace=True)

                st.write(f"Unique Task Names in {sheet_name}:")
                unique_tasks = df[['Task Name']].drop_duplicates()
                st.write(unique_tasks)

                results.append((df, sheet_name))

            except Exception as e:
                st.error(f"Error processing sheet {sheet_name}: {str(e)}")
                continue

        if not results:
            st.error(f"No valid sheets ({', '.join(target_sheets)}) found in file: {filename}")
            return [(None, None)]

        return results

    except Exception as e:
        st.error(f"Error loading Excel file: {str(e)}")
        return [(None, None)]

# Function to get access token for WatsonX API
def get_access_token(api_key):
    try:
        headers = {"Content-Type": "application/x-www-form-urlencoded"}
        data = {
            "grant_type": "urn:ibm:params:oauth:grant-type:apikey",
            "apikey": api_key
        }
        response = requests.post("https://iam.cloud.ibm.com/identity/token", headers=headers, data=data)
        if response.status_code == 200:
            return response.json().get("access_token")
        else:
            logger.error(f"Failed to get access token: {response.status_code} - {response.text}")
            return None
    except Exception as e:
        logger.error(f"Error getting access token: {str(e)}")
        return None

# NEW FUNCTION: Count concreting activities with special logic
def count_concreting_from_cos(tower_df, tower_name):
    """
    Count concreting/casting activities from Structure Work Tracker that have a valid date in column F (Actual Finish).
    Only activities with the word "casting" in the Task Name and a valid date in Actual Finish are counted.
    
    Args:
        tower_df: DataFrame for the specific tower sheet
        tower_name: Name of the tower (e.g., "Tower 4")
    
    Returns:
        int: Count of completed concreting activities
    """
    try:
        if tower_df is None or tower_df.empty:
            logger.warning(f"No data available for {tower_name}")
            return 0
        
        concreting_count = 0
        
        logger.info(f"\n{'='*80}")
        logger.info(f"Starting concreting count for {tower_name}")
        logger.info(f"{'='*80}")
        
        # Iterate through all rows in the tower sheet
        for idx, row in tower_df.iterrows():
            task_name = str(row.get('Task Name', '')).strip()
            task_name_lower = task_name.lower()
            actual_finish = row.get('Actual Finish')
            
            # Check: Does the task name contain the word "casting"?
            if 'casting' not in task_name_lower:
                continue  # Skip if it doesn't have "casting" in the name
            
            # Check: Does column F (Actual Finish) have a valid date?
            has_valid_date = False
            date_str = "No date"
            
            if pd.notna(actual_finish):
                # Check if it's already a datetime object
                if isinstance(actual_finish, (pd.Timestamp, datetime)):
                    has_valid_date = True
                    date_str = actual_finish.strftime('%Y-%m-%d')
                # Check if it's a string that's not NA
                elif isinstance(actual_finish, str):
                    actual_finish_cleaned = actual_finish.strip().upper()
                    # Exclude common NA values
                    if actual_finish_cleaned not in ['NA', 'N/A', 'NAT', 'NONE', '', 'NAN']:
                        # Try to parse it as a date to ensure it's valid
                        try:
                            parsed_date = pd.to_datetime(actual_finish)
                            has_valid_date = True
                            date_str = parsed_date.strftime('%Y-%m-%d')
                        except:
                            has_valid_date = False
                            logger.warning(f"{tower_name} (row {idx}): Casting activity '{task_name}' has invalid date format: '{actual_finish}'")
                else:
                    # Handle numeric dates (Excel serial dates)
                    try:
                        converted_date = pd.to_datetime(actual_finish, errors='coerce')
                        if pd.notna(converted_date):
                            has_valid_date = True
                            date_str = converted_date.strftime('%Y-%m-%d')
                    except:
                        has_valid_date = False
            
            # Log and count if it has "casting" and a valid date
            if has_valid_date:
                concreting_count += 1
                logger.info(f"✓ {tower_name} (row {idx}): COUNTED - '{task_name}' | Actual Finish: {date_str}")
            else:
                logger.info(f"✗ {tower_name} (row {idx}): SKIPPED - '{task_name}' | Actual Finish: {actual_finish} (No valid date)")
        
        logger.info(f"{'='*80}")
        logger.info(f"Total concreting count for {tower_name}: {concreting_count}")
        logger.info(f"{'='*80}\n")
        return concreting_count
        
    except Exception as e:
        logger.error(f"Error counting concreting for {tower_name}: {str(e)}")
        import traceback
        logger.error(f"Full traceback: {traceback.format_exc()}")
        return 0

# WatsonX Prompt Generation (Updated with new categories)
def generatePrompt(json_datas):
    try:
        if isinstance(json_datas, pd.DataFrame):
            json_str = json_datas.reset_index().to_json(orient='records', indent=2)
        else:
            json_str = str(json_datas)

        body = {
            "input": f"""
            Read the table data provided below and categorize the activities into the following categories: Civil Works, MEP Works, Interior Finishing Works, and External Development Activities. Compute the total count of each activity within its respective category and return the results as a JSON array, following the example format provided. For activities like "UP-First Fix" and "CP-First Fix", combine them as "Plumbing Works". If an activity is not found in the data, include it with a count of 0. Ensure the counts are accurate, the output is grouped by category, and the JSON structure is valid with no nested or repeated keys.

            Table Data:
            {json_str}

            Categories and Activities:
            - Civil Works: Concreting, Shuttering, Reinforcement, De-Shuttering
            - MEP Works: Plumbing Works, Slab conduting, Wall Conduiting, Wiring & Switch Socket
            - Interior Finishing Works: Floor Tiling, POP & Gypsum Plaster, Wall Tiling, Waterproofing – Sunken
            - External Development Activities: Granular Sub-Base, Kerb Stone, Rain Water / Storm Line, Saucer Drain / Paver Block, Sewer Line, Stamp Concrete, Storm Line, WMM

            Example JSON format needed:
            [
              {{
                "Category": "Civil Works",
                "Activities": [
                  {{"Activity Name": "Concreting", "Total": 0}},
                  {{"Activity Name": "Shuttering", "Total": 0}},
                  {{"Activity Name": "Reinforcement", "Total": 0}},
                  {{"Activity Name": "De-Shuttering", "Total": 0}}
                ]
              }},
              {{
                "Category": "MEP Works",
                "Activities": [
                  {{"Activity Name": "Plumbing Works", "Total": 0}},
                  {{"Activity Name": "Slab conduting", "Total": 0}},
                  {{"Activity Name": "Wall Conduiting", "Total": 0}},
                  {{"Activity Name": "Wiring & Switch Socket", "Total": 0}}
                ]
              }},
              {{
                "Category": "Interior Finishing Works",
                "Activities": [
                  {{"Activity Name": "Floor Tiling", "Total": 0}},
                  {{"Activity Name": "POP & Gypsum Plaster", "Total": 0}},
                  {{"Activity Name": "Wall Tiling", "Total": 0}},
                  {{"Activity Name": "Waterproofing – Sunken", "Total": 0}}
                ]
              }},
              {{
                "Category": "External Development Activities",
                "Activities": [
                  {{"Activity Name": "Granular Sub-Base", "Total": 0}},
                  {{"Activity Name": "Kerb Stone", "Total": 0}},
                  {{"Activity Name": "Rain Water / Storm Line", "Total": 0}},
                  {{"Activity Name": "Saucer Drain / Paver Block", "Total": 0}},
                  {{"Activity Name": "Sewer Line", "Total": 0}},
                  {{"Activity Name": "Stamp Concrete", "Total": 0}},
                  {{"Activity Name": "Storm Line", "Total": 0}},
                  {{"Activity Name": "WMM", "Total": 0}}
                ]
              }}
            ]

            Return only the JSON array, no additional text, explanations, or code. Ensure the counts are accurate, activities are correctly categorized, and the JSON structure is valid.
            """,
            "parameters": {
                "decoding_method": "greedy",
                "max_new_tokens": 8100,
                "min_new_tokens": 0,
                "stop_sequences": [";"],
                "repetition_penalty": 1.05,
                "temperature": 0.5
            },
            "model_id": os.getenv("MODEL_ID_1"),
            "project_id": os.getenv("PROJECT_ID_1")
        }
        
        access_token = get_access_token(os.getenv("API_KEY_1"))
        if not access_token:
            logger.error("Failed to obtain access token for WatsonX API")
            return generate_fallback_totals(json_datas)
            
        headers = {
            "Accept": "application/json",
            "Content-Type": "application/json",
            "Authorization": f"Bearer {access_token}"
        }
        
        logger.info("Sending request to WatsonX API")
        response = requests.post(os.getenv("WATSONX_API_URL_1"), headers=headers, json=body, timeout=60)
        
        logger.info(f"WatsonX API response status: {response.status_code}")
        logger.debug(f"WatsonX API response text: {response.text[:1000]}...")  # Log first 1000 chars
        
        if response.status_code != 200:
            logger.error(f"WatsonX API call failed: {response.status_code} - {response.text}")
            st.warning(f"WatsonX API failed with status {response.status_code}: {response.text}. Using fallback method to calculate totals.")
            return generate_fallback_totals(json_datas)
            
        response_data = response.json()
        logger.debug(f"WatsonX API response data: {response_data}")
        
        if 'results' not in response_data or not response_data['results']:
            logger.error("WatsonX API response does not contain 'results' key")
            st.warning("WatsonX API response invalid. Using fallback method to calculate totals.")
            return generate_fallback_totals(json_datas)

        generated_text = response_data['results'][0].get('generated_text', '').strip()
        logger.debug(f"Generated text from WatsonX: {generated_text[:1000]}...")  # Log first 1000 chars
        
        if not generated_text:
            logger.error("WatsonX API returned empty generated text")
            st.warning("WatsonX API returned empty response. Using fallback method to calculate totals.")
            return generate_fallback_totals(json_datas)

        if not (generated_text.startswith('[') and generated_text.endswith(']')):
            start_idx = generated_text.find('[')
            end_idx = generated_text.rfind(']')
            if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
                generated_text = generated_text[start_idx:end_idx+1]
                logger.info("Extracted JSON array from response")
            else:
                logger.error(f"Could not extract valid JSON array from response: {generated_text[:1000]}...")
                return generate_fallback_totals(json_datas)
        
        try:
            parsed_json = json.loads(generated_text)
            if not all(isinstance(item, dict) and 'Category' in item and 'Activities' in item for item in parsed_json):
                logger.warning("JSON structure doesn't match expected format")
                return generate_fallback_totals(json_datas)
            logger.info("Successfully parsed WatsonX API response")
            return generated_text
        except json.JSONDecodeError as e:
            logger.error(f"WatsonX API returned invalid JSON: {e}")
            st.warning(f"WatsonX API returned invalid JSON. Error: {str(e)}. Using fallback method to calculate totals.")
            error_position = int(str(e).split('(char ')[1].split(')')[0]) if '(char ' in str(e) else 0
            context_start = max(0, error_position - 50)
            context_end = min(len(generated_text), error_position + 50)
            logger.error(f"JSON error context: ...{generated_text[context_start:error_position]}[ERROR HERE]{generated_text[error_position:context_end]}...")
            return generate_fallback_totals(json_datas)
    
    except Exception as e:
        logger.error(f"Error in WatsonX API call: {str(e)}")
        st.warning(f"Error in WatsonX API call: {str(e)}. Using fallback method to calculate totals.")
        return generate_fallback_totals(json_datas)

# Fallback Total Calculation (Updated with new categories)
def generate_fallback_totals(count_table):
    try:
        if not isinstance(count_table, pd.DataFrame):
            logger.error("Fallback method received invalid input: not a DataFrame")
            return json.dumps([
                {"Category": "Civil Works", "Activities": [
                    {"Activity Name": "Concreting", "Total": 0},
                    {"Activity Name": "Shuttering", "Total": 0},
                    {"Activity Name": "Reinforcement", "Total": 0},
                    {"Activity Name": "De-Shuttering", "Total": 0}
                ]},
                {"Category": "MEP Works", "Activities": [
                    {"Activity Name": "Plumbing Works", "Total": 0},
                    {"Activity Name": "Slab conduting", "Total": 0},
                    {"Activity Name": "Wall Conduiting", "Total": 0},
                    {"Activity Name": "Wiring & Switch Socket", "Total": 0}
                ]},
                {"Category": "Interior Finishing Works", "Activities": [
                    {"Activity Name": "Floor Tiling", "Total": 0},
                    {"Activity Name": "POP & Gypsum Plaster", "Total": 0},
                    {"Activity Name": "Wall Tiling", "Total": 0},
                    {"Activity Name": "Waterproofing – Sunken", "Total": 0}
                ]},
                {"Category": "External Development Activities", "Activities": [
                    {"Activity Name": "Granular Sub-Base", "Total": 0},
                    {"Activity Name": "Kerb Stone", "Total": 0},
                    {"Activity Name": "Rain Water / Storm Line", "Total": 0},
                    {"Activity Name": "Saucer Drain / Paver Block", "Total": 0},
                    {"Activity Name": "Sewer Line", "Total": 0},
                    {"Activity Name": "Stamp Concrete", "Total": 0},
                    {"Activity Name": "Storm Line", "Total": 0},
                    {"Activity Name": "WMM", "Total": 0}
                ]}
            ], indent=2)

        categories = {
            "Civil Works": [
                "Concreting", "Shuttering", "Reinforcement", "De-Shuttering"
            ],
            "MEP Works": [
                "Plumbing Works", "Slab conduting", "Wall Conduiting", "Wiring & Switch Socket"
            ],
            "Interior Finishing Works": [
                "Floor Tiling", "POP & Gypsum Plaster", "Wall Tiling", "Waterproofing – Sunken"
            ],
            "External Development Activities": [
                "Granular Sub-Base", "Kerb Stone", "Rain Water / Storm Line", "Saucer Drain / Paver Block", 
                "Sewer Line", "Stamp Concrete", "Storm Line", "WMM"
            ]
        }

        result = []
        for category, activities in categories.items():
            category_data = {"Category": category, "Activities": []}
            
            for activity in activities:
                if activity == "Plumbing Works":
                    combined_count = count_table.loc["UP-First Fix and CP-First Fix", "Count"] if "UP-First Fix and CP-First Fix" in count_table.index else 0
                    total = combined_count
                else:
                    total = count_table.loc[activity, "Count"] if activity in count_table.index else 0
                category_data["Activities"].append({
                    "Activity Name": activity,
                    "Total": int(total) if pd.notna(total) else 0
                })
            
            result.append(category_data)

        return json.dumps(result, indent=2)
    except Exception as e:
        logger.error(f"Error in fallback total calculation: {str(e)}")
        st.error(f"Error in fallback total calculation: {str(e)}")
        return json.dumps([
            {"Category": "Civil Works", "Activities": [
                {"Activity Name": "Concreting", "Total": 0},
                {"Activity Name": "Shuttering", "Total": 0},
                {"Activity Name": "Reinforcement", "Total": 0},
                {"Activity Name": "De-Shuttering", "Total": 0}
            ]},
            {"Category": "MEP Works", "Activities": [
                {"Activity Name": "Plumbing Works", "Total": 0},
                {"Activity Name": "Slab conduting", "Total": 0},
                {"Activity Name": "Wall Conduiting", "Total": 0},
                {"Activity Name": "Wiring & Switch Socket", "Total": 0}
            ]},
            {"Category": "Interior Finishing Works", "Activities": [
                {"Activity Name": "Floor Tiling", "Total": 0},
                {"Activity Name": "POP & Gypsum Plaster", "Total": 0},
                {"Activity Name": "Wall Tiling", "Total": 0},
                {"Activity Name": "Waterproofing – Sunken", "Total": 0}
            ]},
            {"Category": "External Development Activities", "Activities": [
                {"Activity Name": "Granular Sub-Base", "Total": 0},
                {"Activity Name": "Kerb Stone", "Total": 0},
                {"Activity Name": "Rain Water / Storm Line", "Total": 0},
                {"Activity Name": "Saucer Drain / Paver Block", "Total": 0},
                {"Activity Name": "Sewer Line", "Total": 0},
                {"Activity Name": "Stamp Concrete", "Total": 0},
                {"Activity Name": "Storm Line", "Total": 0},
                {"Activity Name": "WMM", "Total": 0}
            ]}
        ], indent=2)

# Extract Totals from AI Data
def getTotal(ai_data):
    try:
        if isinstance(ai_data, str):
            ai_data = json.loads(ai_data)
            
        if not isinstance(ai_data, list):
            logger.error(f"AI data is not a list: {ai_data}")
            return [0] * len(st.session_state.get('sheduledf', pd.DataFrame()).index)

        share = []
        for category_data in ai_data:
            if isinstance(category_data, dict) and 'Activities' in category_data:
                for activity in category_data['Activities']:
                    if isinstance(activity, dict) and 'Total' in activity:
                        total = activity['Total']
                        share.append(int(total) if isinstance(total, (int, float)) and pd.notna(total) else 0)
                    else:
                        share.append(0)
            else:
                share.append(0)
        return share
    except Exception as e:
        logger.error(f"Error parsing AI data: {str(e)}")
        st.error(f"Error parsing AI data: {str(e)}")
        return [0] * len(st.session_state.get('sheduledf', pd.DataFrame()).index)

# UPDATED: Function to handle activity count display
def display_activity_count():
    # Updated specific activities according to new categorization
    specific_activities = [
        "Concreting", "Shuttering", "Reinforcement", "De-Shuttering",  # Civil Works
        "Slab conduting", "Wall Conduiting", "Wiring & Switch Socket",  # MEP Works (changed from "Slab Conduiting")
        "Floor Tiling", "POP & Gypsum Plaster", "Wall Tiling", "Waterproofing – Sunken",  # Interior Finishing Works
        "Granular Sub-Base", "Kerb Stone", "Rain Water / Storm Line", "Saucer Drain / Paver Block",
        "Sewer Line", "Stamp Concrete", "Storm Line", "WMM"  # External Development Activities
    ]
    all_activities = specific_activities + ["UP-First Fix and CP-First Fix"]

    # Updated category mapping
    category_mapping = {
        "Concreting": "Civil Works",
        "Shuttering": "Civil Works", 
        "Reinforcement": "Civil Works",
        "De-Shuttering": "Civil Works",
        "UP-First Fix and CP-First Fix": "MEP Works",
        "Slab conduting": "MEP Works",  # Changed from "Slab Conduiting"
        "Wall Conduiting": "MEP Works", 
        "Wiring & Switch Socket": "MEP Works",
        "Floor Tiling": "Interior Finishing Works",
        "POP & Gypsum Plaster": "Interior Finishing Works",
        "Wall Tiling": "Interior Finishing Works",
        "Waterproofing – Sunken": "Interior Finishing Works",
        "Granular Sub-Base": "External Development Activities",
        "Kerb Stone": "External Development Activities",
        "Rain Water / Storm Line": "External Development Activities",
        "Saucer Drain / Paver Block": "External Development Activities",
        "Sewer Line": "External Development Activities",
        "Stamp Concrete": "External Development Activities",
        "Storm Line": "External Development Activities",
        "WMM": "External Development Activities"
    }

    count_tables = {}
    if 'ai_response' not in st.session_state or not isinstance(st.session_state.ai_response, dict):
        st.session_state.ai_response = {}
        logger.info("Re-initialized st.session_state.ai_response as empty dictionary")

    def process_tower_data(tower_data, tname):
        if tower_data is None or tower_data.empty:
            logger.warning(f"No data available for {tname}")
            return tname, None

        tower_data = tower_data.copy()
        
        st.write(f"Debug - First few rows from {tname}:")
        st.write(tower_data.head(3))
        
        # **NEW: Special handling for Concreting count**
        concreting_count_special = count_concreting_from_cos(tower_data, tname)
        st.write(f"**Special Concreting Count for {tname}: {concreting_count_special}**")
        
        count_table = pd.DataFrame({
            'Count_Unfiltered': [0] * len(all_activities),
            'Count_Filtered': [0] * len(all_activities)
        }, index=all_activities)
        
        # Filter tower data by Actual Finish date
        tower_data_filtered = tower_data.copy()
        if 'Actual Finish' in tower_data.columns:
            tower_data['Actual_Finish_Original'] = tower_data['Actual Finish'].astype(str)
            tower_data['Actual Finish'] = pd.to_datetime(tower_data['Actual Finish'], errors='coerce')
            has_na_mask = (
                pd.isna(tower_data['Actual Finish']) | 
                (tower_data['Actual_Finish_Original'].str.upper() == 'NAT') |
                (tower_data['Actual_Finish_Original'].str.lower().isin(['nan', 'na', 'n/a', 'none', '']))
            )
            tower_data_filtered = tower_data[~has_na_mask].copy()
            tower_data.drop('Actual_Finish_Original', axis=1, inplace=True)
        
        # For non-concreting activities, use standard counting
        for activity in specific_activities:
            if activity == "Concreting":
                # Use the special count for concreting
                count_table.loc[activity, 'Count_Filtered'] = concreting_count_special
                count_table.loc[activity, 'Count_Unfiltered'] = concreting_count_special
                continue
            
            # Standard counting for other activities
            exact_matches = tower_data[tower_data['Task Name'] == activity]
            if len(exact_matches) > 0:
                count_table.loc[activity, 'Count_Unfiltered'] = len(exact_matches)
            else:
                case_insensitive_matches = tower_data[tower_data['Task Name'].str.lower() == activity.lower()]
                count_table.loc[activity, 'Count_Unfiltered'] = len(case_insensitive_matches)
            
            exact_matches_filtered = tower_data_filtered[tower_data_filtered['Task Name'] == activity]
            if len(exact_matches_filtered) > 0:
                count_table.loc[activity, 'Count_Filtered'] = len(exact_matches_filtered)
            else:
                case_insensitive_matches_filtered = tower_data_filtered[tower_data_filtered['Task Name'].str.lower() == activity.lower()]
                count_table.loc[activity, 'Count_Filtered'] = len(case_insensitive_matches_filtered)
        
        # Handle UP-First Fix and CP-First Fix combination
        up_first_fix_matches = tower_data[tower_data['Task Name'].str.lower() == "up-first fix".lower()]
        cp_first_fix_matches = tower_data[tower_data['Task Name'].str.lower() == "cp-first fix".lower()]
        up_first_fix_count = len(up_first_fix_matches)
        cp_first_fix_count = len(cp_first_fix_matches)
        count_table.loc["UP-First Fix and CP-First Fix", "Count_Unfiltered"] = up_first_fix_count + cp_first_fix_count
        
        up_first_fix_matches_filtered = tower_data_filtered[tower_data_filtered['Task Name'].str.lower() == "up-first fix".lower()]
        cp_first_fix_matches_filtered = tower_data_filtered[tower_data_filtered['Task Name'].str.lower() == "cp-first fix".lower()]
        up_first_fix_count_filtered = len(up_first_fix_matches_filtered)
        cp_first_fix_count_filtered = len(cp_first_fix_matches_filtered)
        count_table.loc["UP-First Fix and CP-First Fix", "Count_Filtered"] = up_first_fix_count_filtered + cp_first_fix_count_filtered
        
        count_table['Count_Unfiltered'] = count_table['Count_Unfiltered'].astype(int)
        count_table['Count_Filtered'] = count_table['Count_Filtered'].astype(int)
        
        return tname, count_table

    # Process each tower's data
    for tower, tname_key in [
        (st.session_state.cos_df_tower4, 'cos_tname_tower4'),
        (st.session_state.cos_df_tower5, 'cos_tname_tower5'),
        (st.session_state.cos_df_tower6, 'cos_tname_tower6'),
        (st.session_state.cos_df_tower7, 'cos_tname_tower7')
    ]:
        if tower is not None:
            tname = st.session_state.get(tname_key)
            tname, count_table = process_tower_data(tower, tname)
            if count_table is not None:
                count_tables[tname] = count_table

    if not count_tables:
        st.error("No processed COS data available. Please click 'Fetch COS' first.")
        st.stop()

    for tname, count_table in count_tables.items():
        with st.spinner(f"Processing activity counts for {tname}..."):
            try:
                st.write(f"Activity Count for {tname} (Unfiltered vs Filtered):")
                st.write(count_table)
                
                count_table_filtered = count_table[['Count_Filtered']].rename(columns={'Count_Filtered': 'Count'})
                ai_response = generatePrompt(count_table_filtered)
                
                try:
                    ai_data = json.loads(ai_response)
                    if not all(isinstance(item, dict) and 'Category' in item and 'Activities' in item for item in ai_data):
                        logger.warning(f"Invalid AI data structure for {tname}: {ai_data}")
                        ai_response = generate_fallback_totals(count_table_filtered)
                        ai_data = json.loads(ai_response)
                    
                    # Store AI response
                    st.session_state.ai_response[tname] = ai_data
                    logger.info(f"Stored AI response for {tname}: {ai_data}")
                    
                    totals_mapping = {}
                    for category_data in ai_data:
                        for activity in category_data['Activities']:
                            totals_mapping[activity['Activity Name']] = activity['Total']
                    
                    display_df = count_table.reset_index()
                    display_df.rename(columns={'index': 'Activity Name'}, inplace=True)
                    
                    display_df['Total'] = display_df['Activity Name'].map(
                        lambda x: totals_mapping.get(x, display_df.loc[display_df['Activity Name'] == x, 'Count_Filtered'].iloc[0])
                    )
                    
                    display_df['Category'] = display_df['Activity Name'].map(lambda x: category_mapping.get(x, "Other"))
                    
                    display_df = display_df.sort_values(['Category', 'Activity Name'])
                    
                    st.write(f"Activity Count with Totals for {tname}:")
                    st.write(display_df[['Activity Name', 'Count_Unfiltered', 'Total', 'Category']])
                    
                    st.write(f"Activity Counts by Category for {tname}:")
                    # Updated category display order
                    for category in ['Civil Works', 'MEP Works', 'Interior Finishing Works', 'External Development Activities']:
                        category_df = display_df[display_df['Category'] == category]
                        if not category_df.empty:
                            st.write(f"**{category}**")
                            st.write(category_df[['Activity Name', 'Count_Filtered', 'Total']])
                    
                except Exception as e:
                    logger.error(f"Error processing WatsonX for {tname}: {str(e)}")
                    st.warning(f"Failed to process AI-generated totals for {tname}. Using fallback method.")
                    
                    ai_response = generate_fallback_totals(count_table_filtered)
                    ai_data = json.loads(ai_response)
                    st.session_state.ai_response[tname] = ai_data
                    logger.info(f"Stored fallback AI response for {tname}: {ai_data}")
                    
                    display_df = count_table.reset_index()
                    display_df.rename(columns={'index': 'Activity Name'}, inplace=True)
                    display_df['Category'] = display_df['Activity Name'].map(lambda x: category_mapping.get(x, "Other"))
                    display_df['Total'] = display_df['Count_Filtered']
                    display_df = display_df.sort_values(['Category', 'Activity Name'])
                    
                    st.write(f"Activity Counts by Category for {tname} (using raw counts):")
                    # Updated category display order
                    for category in ['Civil Works', 'MEP Works', 'Interior Finishing Works', 'External Development Activities']:
                        category_df = display_df[display_df['Category'] == category]
                        if not category_df.empty:
                            st.write(f"**{category}**")
                            st.write(category_df[['Activity Name', 'Count_Filtered', 'Total']])
                
            except Exception as e:
                logger.error(f"Error displaying activity count for {tname}: {str(e)}")
                st.error(f"Error displaying activity count for {tname}: {str(e)}")
                # Ensure ai_response has an entry even in case of error
                count_table_filtered = count_table[['Count_Filtered']].rename(columns={'Count_Filtered': 'Count'})
                ai_response = generate_fallback_totals(count_table_filtered)
                ai_data = json.loads(ai_response)
                st.session_state.ai_response[tname] = ai_data
                logger.info(f"Stored fallback AI response for {tname} on error: {ai_data}")

    # Create consolidated External Development Activities table
    st.write("### External Development Activities (Consolidated)")
    external_dev_activities = [
        "Granular Sub-Base", "Kerb Stone", "Rain Water / Storm Line", "Saucer Drain / Paver Block",
        "Sewer Line", "Stamp Concrete", "Storm Line", "WMM"
    ]
    
    external_dev_data = []
    for activity in external_dev_activities:
        total_count = 0
        for tname, count_table in count_tables.items():
            if activity in count_table.index:
                total_count += count_table.loc[activity, 'Count_Filtered']
        external_dev_data.append({
            "Activity Name": activity,
            "Total Count": total_count
        })
    
    external_dev_df = pd.DataFrame(external_dev_data)
    st.write(external_dev_df)


# Combined function for Initialize and Fetch Data
async def initialize_and_fetch_data(email, password):
    with st.spinner("Starting initialization and data fetching process..."):
        # Step 1: Login
        if not email or not password:
            st.sidebar.error("Please provide both email and password!")
            return False
        try:
            st.sidebar.write("Logging in...")
            session_id = await login_to_asite(email, password)
            if not session_id:
                st.sidebar.error("Login failed!")
                return False
            st.sidebar.success("Login successful!")
        except Exception as e:
            st.sidebar.error(f"Login failed: {str(e)}")
            return False

        # Step 2: Get Workspace ID
        try:
            st.sidebar.write("Fetching Workspace ID...")
            await GetWorkspaceID()
            st.sidebar.success("Workspace ID fetched successfully!")
        except Exception as e:
            st.sidebar.error(f"Failed to fetch Workspace ID: {str(e)}")
            return False

        # Step 3: Get Project IDs
        try:
            st.sidebar.write("Fetching Project IDs...")
            await GetProjectId()
            st.sidebar.success("Project IDs fetched successfully!")
        except Exception as e:
            st.sidebar.error(f"Failed to fetch Project IDs: {str(e)}")
            return False

        # Step 4: Get All Data (Structure only)
        try:
            st.sidebar.write("Fetching All Data...")
            Edenstructure = await GetAllDatas()
            st.session_state.eden_structure = Edenstructure
            st.sidebar.success("All Data fetched successfully!")
        except Exception as e:
            st.sidebar.error(f"Failed to fetch All Data: {str(e)}")
            return False

        # Step 5: Get Activity Data
        try:
            st.sidebar.write("Fetching Activity Data...")
            structure_activity_data = await Get_Activity()
            st.session_state.structure_activity_data = structure_activity_data
            st.sidebar.success("Activity Data fetched successfully!")
        except Exception as e:
            st.sidebar.error(f"Failed to fetch Activity Data: {str(e)}")
            return False

        # Step 6: Get Location/Module Data
        try:
            st.sidebar.write("Fetching Location/Module Data...")
            structure_location_data = await Get_Location()
            st.session_state.structure_location_data = structure_location_data 
            st.sidebar.success("Location/Module Data fetched successfully!")
        except Exception as e:
            st.sidebar.error(f"Failed to fetch Location/Module Data: {str(e)}")
            return False

        # Step 7: Fetch COS Files
        try:
            st.sidebar.write("Fetching COS files from Eden folder...")
            file_key = get_cos_files()
            st.session_state.file_key = file_key
            if file_key:
                st.success(f"Found 1 file in COS storage: {file_key}")
                try:
                    st.write(f"Processing file: {file_key}")
                    cos_client = initialize_cos_client()
                    if not cos_client:
                        st.error("Failed to initialize COS client during file fetch")
                        logger.error("COS client initialization failed during file fetch")
                        return False
                    st.write("Fetching file from COS...")
                    response = cos_client.get_object(Bucket=COS_BUCKET, Key=file_key)
                    file_bytes = io.BytesIO(response['Body'].read())
                    st.write("File fetched successfully. Processing sheets...")
                    results = process_file(file_bytes, file_key)
                    st.write(f"Processing results: {len(results)} sheets processed")
                    for df, tname in results:
                        if df is not None:
                            if "Tower 4" in tname:
                                st.session_state.cos_df_tower4 = df
                                st.session_state.cos_tname_tower4 = tname
                                st.write(f"Processed Data for {tname} - {len(df)} rows:")
                                st.write(df.head())
                            elif "Tower 5" in tname:
                                st.session_state.cos_df_tower5 = df
                                st.session_state.cos_tname_tower5 = tname
                                st.write(f"Processed Data for {tname} - {len(df)} rows:")
                                st.write(df.head())
                            elif "Tower 6" in tname:
                                st.session_state.cos_df_tower6 = df
                                st.session_state.cos_tname_tower6 = tname
                                st.write(f"Processed Data for {tname} - {len(df)} rows:")
                                st.write(df.head())
                            elif "Tower 7" in tname:
                                st.session_state.cos_df_tower7 = df
                                st.session_state.cos_tname_tower7 = tname
                                st.write(f"Processed Data for {tname} - {len(df)} rows:")
                                st.write(df.head())
                        else:
                            st.warning(f"No data processed for {tname} in {file_key}.")
                except Exception as e:
                    st.error(f"Error loading {file_key} from cloud storage: {str(e)}")
                    logger.error(f"Error loading {file_key}: {str(e)}")
                    return False
            else:
                st.warning("No expected Excel files available in the 'Eden' folder of the COS bucket.")
                return False
        except Exception as e:
            st.sidebar.error(f"Failed to fetch COS files: {str(e)}")
            logger.error(f"Failed to fetch COS files: {str(e)}")
            return False

    st.sidebar.success("All steps completed successfully!")
    return True


# Updated generate_consolidated_Checklist_excel function with new categories
# Updated generate_consolidated_Checklist_excel function with multiplier for Civil Works
def generate_consolidated_Checklist_excel(structure_analysis, activity_counts):
    try:
        # Debug: Print activity_counts structure
        logger.info(f"Activity counts keys: {list(activity_counts.keys()) if activity_counts else 'None'}")
        logger.info(f"Activity counts type: {type(activity_counts)}")
        
        # Also log structure_analysis to see what we have
        if structure_analysis is not None and not structure_analysis.empty:
            logger.info(f"Structure analysis columns: {structure_analysis.columns.tolist()}")
            logger.info(f"Structure analysis sample:\n{structure_analysis.head()}")
            logger.info(f"Unique tower_keys in structure_analysis: {structure_analysis['tower_key'].unique().tolist()}")
            logger.info(f"Unique activityNames in structure_analysis: {structure_analysis['activityName'].unique().tolist()}")
        
        # Define categories and activities according to new structure
        categories = {
            "Civil Works": ["Concreting", "Shuttering", "Reinforcement", "De-Shuttering"],
            "MEP Works": ["Plumbing Works", "Slab Conduting", "Wall Conduiting", "Wiring & Switch Socket"],
            "Interior Finishing Works": ["Floor Tiling", "POP & Gypsum Plaster", "Wall Tiling", "Waterproofing – Sunken"]
        }

        # Define the COS to Asite activity name mapping (updated)
        cos_to_asite_mapping = {
            "Concreting": "Concreting",
            "Shuttering": "Shuttering", 
            "Reinforcement": "Reinforcement",
            "De-Shuttering": "De-Shuttering",
            "Plumbing Works": ["UP-First Fix", "CP-First Fix"],  # Will sum both
            "Slab Conduting": "Slab conduting",  # lowercase 'c' to match Asite
            "Wall Conduiting": "Wall Conducting",
            "Wiring & Switch Socket": "Wiring & Switch Socket",
            "Floor Tiling": "Floor Tiling",
            "POP & Gypsum Plaster": "POP & Gypsum Plaster",
            "Wall Tiling": "Wall Tile",
            "Waterproofing – Sunken": "Waterproofing - Sunken"
        }

        # Activities that should have the same completed count as Concreting
        activities_same_as_concreting = ["Shuttering", "Reinforcement", "De-Shuttering"]

        # Include all towers dynamically
        all_towers = ["Tower 4", "Tower 5", "Tower 6", "Tower 7"]
        
        logger.info(f"Processing all towers: {all_towers}")

        # Initialize list to store consolidated data
        consolidated_rows = []

        # Process data for each tower and category
        for tower in all_towers:
            tower_key = f"T{tower.replace('Tower ', '')}"  # e.g., "Tower 4" -> "T4"
            tower_num = tower.replace('Tower ', '')
            
            logger.info(f"\n{'='*60}")
            logger.info(f"Processing {tower} (key: {tower_key})")
            logger.info(f"{'='*60}")
            
            for category, activities in categories.items():
                # Process each activity in the category
                for activity in activities:
                    # SPECIAL HANDLING FOR SLAB CONDUTING (Final Concreting - 5)
                    if activity == "Slab Conduting":
                        # Get the concreting count from activity_counts (COS data)
                        concreting_count = 0
                        tower_name_variations = [tower_key, tower, f"Tower {tower_num}"]
                        
                        found_tower = None
                        for tower_var in tower_name_variations:
                            if activity_counts and tower_var in activity_counts:
                                found_tower = tower_var
                                break
                        
                        if found_tower:
                            ai_data = activity_counts[found_tower]
                            # Extract the Concreting total from AI response data
                            if isinstance(ai_data, list):
                                for category_data in ai_data:
                                    if isinstance(category_data, dict) and 'Category' in category_data:
                                        if category_data['Category'] == 'Civil Works':
                                            for activity_data in category_data['Activities']:
                                                if isinstance(activity_data, dict) and activity_data.get('Activity Name') == 'Concreting':
                                                    concreting_count = activity_data.get('Total', 0)
                                                    logger.info(f"Found raw Concreting count for {found_tower}: {concreting_count}")
                                                    break
                        else:
                            logger.info(f"No COS data found for {tower}, concreting_count = 0")
                        
                        # Apply Civil Works formula to concreting count first: (count*2)-1
                        final_concreting = (concreting_count * 2) - 1
                        logger.info(f"Applied Civil Works formula to Concreting: ({concreting_count}*2)-1 = {final_concreting}")
                        
                        # Calculate Slab Conduting as Final Concreting - 5
                        completed_flats = max(0, final_concreting - 5)  # Ensure non-negative
                        logger.info(f"Calculated Slab Conduting for {tower}: {final_concreting} - 5 = {completed_flats}")
                        
                        # Get closed checklist from Asite (normal processing)
                        asite_activity = cos_to_asite_mapping.get(activity, activity)
                        if isinstance(asite_activity, list):
                            asite_activities = asite_activity
                        else:
                            asite_activities = [asite_activity]

                        closed_checklist = 0
                        if structure_analysis is not None and not structure_analysis.empty:
                            for asite_act in asite_activities:
                                # Try case-insensitive match
                                matching_rows = structure_analysis[
                                    (structure_analysis['tower_key'] == tower_key) &
                                    (structure_analysis['activityName'].str.lower() == asite_act.lower())
                                ]
                                
                                if not matching_rows.empty:
                                    closed_checklist += matching_rows['CompletedCount'].sum()
                                    logger.info(f"Found {tower_key} - {asite_act}: {matching_rows['CompletedCount'].sum()} completed")
                                else:
                                    logger.info(f"No match found for {tower_key} - {asite_act}")

                        # Calculate Open/Missing
                        in_progress = 0
                        if completed_flats == 0 or closed_checklist > completed_flats:
                            open_missing = 0
                        else:
                            open_missing = abs(completed_flats - closed_checklist)

                        display_activity = asite_activities[0] if isinstance(asite_activity, list) else asite_activity

                        consolidated_rows.append({
                            "Tower": tower,
                            "Category": category,
                            "Activity Name": display_activity,
                            "Completed Work*(Count of Flat)": completed_flats,
                            "In progress": in_progress,
                            "Closed checklist": closed_checklist,
                            "Open/Missing check list": open_missing
                        })
                        
                        continue  # Skip to next activity
                    
                    # SPECIAL HANDLING FOR ACTIVITIES THAT MATCH CONCRETING
                    # Shuttering, Reinforcement, De-Shuttering have same completed count as Concreting
                    if activity in activities_same_as_concreting:
                        # Get the concreting count from activity_counts (COS data)
                        completed_flats = 0
                        tower_name_variations = [tower_key, tower, f"Tower {tower_num}"]
                        
                        found_tower = None
                        for tower_var in tower_name_variations:
                            if activity_counts and tower_var in activity_counts:
                                found_tower = tower_var
                                break
                        
                        if found_tower:
                            ai_data = activity_counts[found_tower]
                            # Extract the Concreting total from AI response data
                            if isinstance(ai_data, list):
                                for category_data in ai_data:
                                    if isinstance(category_data, dict) and 'Category' in category_data:
                                        if category_data['Category'] == 'Civil Works':
                                            for activity_data in category_data['Activities']:
                                                if isinstance(activity_data, dict) and activity_data.get('Activity Name') == 'Concreting':
                                                    completed_flats = activity_data.get('Total', 0)
                                                    logger.info(f"Found Concreting count for {found_tower}: {completed_flats}, using for {activity}")
                                                    break
                        else:
                            logger.info(f"No COS data found for {tower}, {activity} completed_flats = 0")
                        
                        # *** APPLY FORMULA (count*2)-1 FOR CIVIL WORKS ***
                        if category == "Civil Works":
                            completed_flats = (completed_flats * 2) - 1
                            logger.info(f"Applied Civil Works formula ((count*2)-1) for {activity}: {completed_flats}")
                        
                        # Get closed checklist from Asite (normal processing)
                        asite_activity = cos_to_asite_mapping.get(activity, activity)
                        if isinstance(asite_activity, list):
                            asite_activities = asite_activity
                        else:
                            asite_activities = [asite_activity]

                        closed_checklist = 0
                        if structure_analysis is not None and not structure_analysis.empty:
                            for asite_act in asite_activities:
                                # Try case-insensitive match
                                matching_rows = structure_analysis[
                                    (structure_analysis['tower_key'] == tower_key) &
                                    (structure_analysis['activityName'].str.lower() == asite_act.lower())
                                ]
                                
                                if not matching_rows.empty:
                                    closed_checklist += matching_rows['CompletedCount'].sum()
                                    logger.info(f"Found {tower_key} - {asite_act}: {matching_rows['CompletedCount'].sum()} completed")
                                else:
                                    logger.info(f"No match found for {tower_key} - {asite_act}")

                        # Calculate Open/Missing
                        in_progress = 0
                        if completed_flats == 0 or closed_checklist > completed_flats:
                            open_missing = 0
                        else:
                            open_missing = abs(completed_flats - closed_checklist)

                        display_activity = asite_activities[0] if isinstance(asite_activity, list) else asite_activity

                        consolidated_rows.append({
                            "Tower": tower,
                            "Category": category,
                            "Activity Name": display_activity,
                            "Completed Work*(Count of Flat)": completed_flats,
                            "In progress": in_progress,
                            "Closed checklist": closed_checklist,
                            "Open/Missing check list": open_missing
                        })
                        
                        continue  # Skip to next activity
                    
                    # REGULAR PROCESSING FOR OTHER ACTIVITIES (Concreting, Plumbing Works, etc.)
                    # Map COS activity name to Asite name(s)
                    asite_activity = cos_to_asite_mapping.get(activity, activity)
                    if isinstance(asite_activity, list):
                        asite_activities = asite_activity
                    else:
                        asite_activities = [asite_activity]

                    # Get completed count from structure_analysis (Asite data with /Pour filter)
                    closed_checklist = 0
                    if structure_analysis is not None and not structure_analysis.empty:
                        for asite_act in asite_activities:
                            # Try case-insensitive match
                            matching_rows = structure_analysis[
                                (structure_analysis['tower_key'] == tower_key) &
                                (structure_analysis['activityName'].str.lower() == asite_act.lower())
                            ]
                            
                            if not matching_rows.empty:
                                closed_checklist += matching_rows['CompletedCount'].sum()
                                logger.info(f"Found {tower_key} - {asite_act}: {matching_rows['CompletedCount'].sum()} completed")
                            else:
                                logger.info(f"No match found for {tower_key} - {asite_act}")
                    else:
                        logger.info(f"No Asite data available for {tower_key} - {activity}")

                    # Get completed flats count from activity_counts (COS data)
                    completed_flats = 0
                    tower_name_variations = [tower_key, tower, f"Tower {tower_num}"]
                    
                    found_tower = None
                    for tower_var in tower_name_variations:
                        if activity_counts and tower_var in activity_counts:
                            found_tower = tower_var
                            break
                    
                    if found_tower:
                        ai_data = activity_counts[found_tower]
                        # Extract the total from AI response data
                        if isinstance(ai_data, list):
                            for category_data in ai_data:
                                if isinstance(category_data, dict) and 'Activities' in category_data:
                                    for activity_data in category_data['Activities']:
                                        if isinstance(activity_data, dict) and activity_data.get('Activity Name') == activity:
                                            completed_flats = activity_data.get('Total', 0)
                                            logger.info(f"Found COS data for {found_tower} - {activity}: {completed_flats}")
                                            break
                    else:
                        logger.info(f"No COS data found for {tower} - {activity}, completed_flats = 0")

                    # *** APPLY FORMULA (count*2)-1 FOR CIVIL WORKS ***
                    if category == "Civil Works":
                        completed_flats = (completed_flats * 2) - 1
                        logger.info(f"Applied Civil Works formula ((count*2)-1) for {activity}: {completed_flats}")

                    # Calculate Open/Missing check list per clarified requirements
                    in_progress = 0  # Not calculated in the current code
                    if completed_flats == 0 or closed_checklist > completed_flats:
                        open_missing = 0
                    else:
                        open_missing = abs(completed_flats - closed_checklist)

                    # Use the first Asite activity name for display
                    display_activity = asite_activities[0] if isinstance(asite_activity, list) else asite_activity

                    consolidated_rows.append({
                        "Tower": tower,
                        "Category": category,
                        "Activity Name": display_activity,
                        "Completed Work*(Count of Flat)": completed_flats,
                        "In progress": in_progress,
                        "Closed checklist": closed_checklist,
                        "Open/Missing check list": open_missing
                    })

        # Create DataFrame
        df = pd.DataFrame(consolidated_rows)
        if df.empty:
            st.warning("No data available to generate consolidated checklist.")
            return None

        # Sort by Tower and Category for consistency
        df.sort_values(by=["Tower", "Category"], inplace=True)
        
        logger.info(f"\n{'='*60}")
        logger.info(f"Final consolidated data summary:")
        logger.info(f"Total rows: {len(df)}")
        logger.info(f"Towers included: {df['Tower'].unique().tolist()}")
        logger.info(f"{'='*60}\n")

        # Create a BytesIO buffer for the Excel file
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)
        worksheet = workbook.add_worksheet("Consolidated Checklist")

        # Define styles
        header_format = workbook.add_format({'bold': True, 'bg_color': '#D3D3D3'})
        total_format = workbook.add_format({'bold': True, 'bg_color': '#FFDAB9'})
        cell_format = workbook.add_format({'border': 1})

        # Column headers
        headers = ["Activity Name", "Completed", "In progress", "Closed checklist", "Open/Missing check list"]

        # Starting positions for each section
        col_start = 1  # Start from column B
        row_start = 0

        # Group by Tower
        grouped_by_tower = df.groupby('Tower', sort=False)

        for tower, tower_group in grouped_by_tower:
            # Reset column position for each tower
            col_pos = col_start

            # Group Categories within this Tower
            grouped_by_category = tower_group.groupby('Category', sort=False)

            # Process each Category side by side
            for category, cat_group in grouped_by_category:
                # Write category header
                worksheet.merge_range(row_start, col_pos, row_start, col_pos + 4, f"{tower} {category} Checklist Status", header_format)
                row_pos = row_start + 1

                # Write column headers
                for i, header in enumerate(headers, start=0):
                    worksheet.write(row_pos, col_pos + i, header, header_format)
                row_pos += 1

                # Write activity data
                for _, row in cat_group.iterrows():
                    worksheet.write(row_pos, col_pos, row["Activity Name"], cell_format)
                    worksheet.write(row_pos, col_pos + 1, row["Completed Work*(Count of Flat)"], cell_format)
                    worksheet.write(row_pos, col_pos + 2, row["In progress"], cell_format)
                    worksheet.write(row_pos, col_pos + 3, row["Closed checklist"], cell_format)
                    worksheet.write(row_pos, col_pos + 4, row["Open/Missing check list"], cell_format)
                    row_pos += 1

                # Write total pending checklist
                total_pending = cat_group["Open/Missing check list"].sum()
                worksheet.merge_range(row_pos, col_pos, row_pos, col_pos + 3, "Total pending check list", total_format)
                worksheet.write(row_pos, col_pos + 4, total_pending, total_format)
                row_pos += 2

                # Move to the next column position (side-by-side sections)
                col_pos += 6

            # Move to the next tower (below the current sections)
            row_start = row_pos

        # Auto-adjust column widths for Sheet 1
        for col in range(col_start, col_pos):
            worksheet.set_column(col, col, 20)

        # Create Sheet 2: Checklist with current month
        current_month = datetime.now().strftime("%B")  # Gets full month name (e.g., "January", "February")
        sheet2_name = f"Checklist {current_month}"
        worksheet2 = workbook.add_worksheet(sheet2_name)
        current_row = 0

        # Write title
        worksheet2.write(current_row, 0, f"Checklist: {current_month}", header_format)
        current_row += 1

        # Write headers
        headers = [
            "Site",
            "Total of Missing & Open Checklist-Civil",
            "Total of Missing & Open Checklist-MEP",
            "Total of Missing & Open Checklist-Interior Finishing",
            "TOTAL"
        ]
        for col, header in enumerate(headers, start=0):
            worksheet2.write(current_row, col, header, header_format)
        current_row += 1

        # Categorize categories into Civil, MEP, and Interior Finishing
        def map_category_to_type(category):
            if category in ["Civil Works"]:
                return "Civil"
            elif category in ["MEP Works"]:
                return "MEP"
            elif category in ["Interior Finishing Works"]:
                return "Interior"
            else:
                return "Civil"  # Default to Civil

        # Aggregate open/missing counts by tower and type (Civil/MEP) for ALL towers
        summary_data = {}
        for _, row in df.iterrows():
            tower = row["Tower"]
            category = row["Category"]
            open_missing = row["Open/Missing check list"]
            
            # Convert tower name to display format
            tower_num = tower.replace("Tower ", "")
            if len(tower_num) == 1:
                tower_num = f"0{tower_num}"
            site_name = f"Eden-Tower {tower_num}"

            type_ = map_category_to_type(category)
            
            if site_name not in summary_data:
                summary_data[site_name] = {"Civil": 0, "MEP": 0, "Interior": 0}
            
            summary_data[site_name][type_] += open_missing

        logger.info(f"Summary data for Sheet 2: {summary_data}")

        # Write summary data to Sheet 2 (all towers, even if zero)
        for site_name, counts in sorted(summary_data.items()):
            civil_count = counts["Civil"]
            mep_count = counts["MEP"]
            interior_count = counts["Interior"]
            total_count = civil_count + mep_count + interior_count
            
            worksheet2.write(current_row, 0, site_name, cell_format)
            worksheet2.write(current_row, 1, civil_count, cell_format)
            worksheet2.write(current_row, 2, mep_count, cell_format)
            worksheet2.write(current_row, 3, interior_count, cell_format)
            worksheet2.write(current_row, 4, total_count, cell_format)
            
            current_row += 1

        # Create Sheet 3: External Development Activities (Always create this sheet)
        worksheet3 = workbook.add_worksheet("External Development Activities")
        current_row = 0

        # Write title
        worksheet3.write(current_row, 0, "External Development Activities - Consolidated Report", header_format)
        current_row += 2

        # External Development Activities - Always include all activities
        external_dev_activities = [
            "Granular Sub-Base", 
            "Kerb Stone", 
            "Rain Water / Storm Line", 
            "Saucer Drain / Paver Block",
            "Sewer Line", 
            "Stamp Concrete", 
            "Storm Line", 
            "WMM"
        ]

        # Write headers for External Development
        ext_headers = ["Activity Name", "Total Count Across All Towers"]
        for col, header in enumerate(ext_headers, start=0):
            worksheet3.write(current_row, col, header, header_format)
        current_row += 1

        # Always process and display all external development activities
        logger.info("Creating External Development Activities table with all activities...")
        grand_total = 0
        
        for activity in external_dev_activities:
            total_count = 0
            
            # Check all towers for this activity
            if activity_counts:
                tower_keys = list(activity_counts.keys())
                logger.info(f"Processing activity: {activity} across towers: {tower_keys}")
                
                for tower_key in tower_keys:
                    ai_data = activity_counts[tower_key]
                    
                    # Extract the total from AI response data
                    if isinstance(ai_data, list):
                        for category_data in ai_data:
                            if isinstance(category_data, dict) and 'Category' in category_data:
                                if category_data['Category'] == 'External Development Activities':
                                    for activity_data in category_data['Activities']:
                                        if isinstance(activity_data, dict) and activity_data.get('Activity Name') == activity:
                                            activity_total = activity_data.get('Total', 0)
                                            total_count += activity_total
                                            logger.info(f"Found {activity} in {tower_key}: {activity_total}")
                                            break
            
            # Always write the activity to the sheet, even if count is 0
            worksheet3.write(current_row, 0, activity, cell_format)
            worksheet3.write(current_row, 1, total_count, cell_format)
            grand_total += total_count
            current_row += 1
            logger.info(f"Added {activity} to External Development sheet with total: {total_count}")

        # Add a summary row with grand total
        current_row += 1
        worksheet3.write(current_row, 0, "GRAND TOTAL", total_format)
        worksheet3.write(current_row, 1, grand_total, total_format)
        
        # Add note about future updates
        current_row += 2
        worksheet3.write(current_row, 0, "Note: This table will reflect updated counts when data source is refreshed", 
                        workbook.add_format({'italic': True, 'font_size': 10}))
        
        logger.info(f"External Development Activities sheet created with {len(external_dev_activities)} activities and grand total: {grand_total}")

        # Auto-adjust column widths for all sheets
        for col in range(5):
            worksheet2.set_column(col, col, 25)
        for col in range(2):
            worksheet3.set_column(col, col, 25)

        # Close the workbook
        workbook.close()
        output.seek(0)
        return output

    except Exception as e:
        logger.error(f"Error generating consolidated Excel: {str(e)}")
        st.error(f"❌ Error generating Excel file: {str(e)}")
        import traceback
        logger.error(f"Full traceback: {traceback.format_exc()}")
        return None
    

# Combined function to handle analysis and display
def run_analysis_and_display():
    try:
        st.write("Running status analysis...")
        AnalyzeStatusManually()
        st.success("Status analysis completed successfully!")

        # Ensure ai_response is initialized
        if 'ai_response' not in st.session_state or not isinstance(st.session_state.ai_response, dict):
            st.session_state.ai_response = {}
            logger.info("Initialized st.session_state.ai_response in run_analysis_and_display")

        # Step 1: Display activity counts and generate AI data
        st.write("Displaying activity counts and generating AI data...")
        display_activity_count()
        st.success("Activity counts displayed successfully!")

        # Step 2: Check AI data totals
        st.write("Checking AI data totals...")
        logger.info(f"st.session_state.ai_response contents: {st.session_state.ai_response}")
        if not st.session_state.ai_response:
            st.error("❌ No AI data available in st.session_state.ai_response. Attempting to regenerate.")
            logger.error("No AI data in st.session_state.ai_response after display_activity_count")
            # Re-run display_activity_count to attempt recovery
            display_activity_count()
            if not st.session_state.ai_response:
                st.error("❌ Failed to regenerate AI data. Please check data fetching and try again.")
                logger.error("Failed to regenerate AI data")
                return

        # Step 3: Generate consolidated checklist Excel file
        st.write("Generating consolidated checklist Excel file...")
        structure_analysis = st.session_state.get('structure_analysis', None)
        if structure_analysis is None:
            st.error("❌ No structure analysis data available. Please ensure analysis ran successfully.")
            logger.error("No structure_analysis in st.session_state")
            return

        with st.spinner("Generating Excel file... This may take a moment."):
            excel_file = generate_consolidated_Checklist_excel(structure_analysis, st.session_state.ai_response)
        
        if excel_file:
            timestamp = pd.Timestamp.now(tz='Asia/Kolkata').strftime('%Y%m%d_%H%M')
            file_name = f"Consolidated_Checklist_Eden_{timestamp}.xlsx"
            
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                st.sidebar.download_button(
                    label="📥 Download Checklist Excel",
                    data=excel_file,
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_excel_button",
                    help="Click to download the consolidated checklist in Excel format."
                )
            st.success("Excel file generated successfully! Click the button above to download.")
        else:
            st.error("Failed to generate Excel file. Please check the logs for details.")
            logger.error("Failed to generate Excel file")

    except Exception as e:
        st.error(f"Error during analysis, display, or Excel generation: {str(e)}")
        logger.error(f"Error during analysis, display, or Excel generation: {str(e)}")

# Streamlit UI
st.markdown(
    """
    <h1 style='font-family: "Arial Black", Gadget, sans-serif; 
               color: red; 
               font-size: 48px; 
               text-align: center;'>
        Eden CheckList - Report
    </h1>
    """,
    unsafe_allow_html=True
)

# Initialize and Fetch Data
st.sidebar.title("🔐 Asite Initialization")
email = st.sidebar.text_input("Email", "impwatson@gadieltechnologies.com", key="email_input")
password = st.sidebar.text_input("Password", "Srihari@790$", type="password", key="password_input")

if st.sidebar.button("Initialize and Fetch Data"):
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        success = loop.run_until_complete(initialize_and_fetch_data(email, password))
        if success:
            st.sidebar.success("Initialization and data fetching completed successfully!")
        else:
            st.sidebar.error("Initialization and data fetching failed!")
    except Exception as e:
        st.sidebar.error(f"Initialization and data fetching failed: {str(e)}")
    finally:
        loop.close()

# Analyze and Display
st.sidebar.title("📊 Status Analysis")
if st.sidebar.button("Analyze and Display Activity Counts"):
    with st.spinner("Running analysis and displaying activity counts..."):
        run_analysis_and_display()
