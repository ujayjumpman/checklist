import streamlit as st
import pandas as pd
import requests
import json
import openpyxl
import time
import math
from io import BytesIO
from datetime import datetime
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import ibm_boto3
from ibm_botocore.client import Config
import io

# Initialize lists to store counts of green (1) and non-green (0)
ews1 = []
ews2 = []
ews3 = []
lig1 = []
lig2 = []
lig3 = []

def count_green_cells(sheet, rows, cols, tower_name):
    """Generic function to count green cells with detailed logging"""
    green_count = 0
    green_cells = []
    
    st.write(f"Processing {tower_name} - Rows: {rows}, Columns: {cols}")
    
    for row in rows:
        for col in cols:
            cell = sheet[f"{col}{row}"]
            
            # Get background color of the cell
            fill = cell.fill
            if fill.start_color.type == 'rgb' and fill.start_color.rgb:
                bg_color = f"#{fill.start_color.rgb[-6:]}"
            else:
                bg_color = "#FFFFFF"

            # Count green cells (completed slabs)
            if bg_color == "#92D050":
                green_count += 1
                green_cells.append(f"{col}{row}")
                st.write(f"  {tower_name}: Found green at {col}{row}")
    
    st.write(f"  {tower_name}: Total green cells = {green_count}")
    st.write(f"  {tower_name}: Green cell locations = {green_cells}")
    return green_count


def EWS1(sheet, selected_year, selected_month):
    """EWS Tower 1: row 8 to 22 columns D, H, L, P (Pour 1)"""
    global ews1
    rows = list(range(8, 23))  # 8 to 22 inclusive
    cols = ['D', 'H', 'L', 'P']
    
    green_count = count_green_cells(sheet, rows, cols, "EWS Tower 1")
    ews1 = [1] * green_count  # Append 1 for each green cell


def EWS2(sheet, selected_year, selected_month):
    """EWS Tower 2: row 8 to 22 columns U, Y, AC, AG (Pour 2)"""
    global ews2
    rows = list(range(8, 23))  # 8 to 22 inclusive
    cols = ['U', 'Y', 'AC', 'AG']
    
    green_count = count_green_cells(sheet, rows, cols, "EWS Tower 2")
    ews2 = [1] * green_count


def EWS3(sheet, selected_year, selected_month):
    """EWS Tower 3: row 8 to 22 columns AL, AP, AT, AX (Pour 3)"""
    global ews3
    rows = list(range(8, 23))  # 8 to 22 inclusive
    cols = ['AL', 'AP', 'AT', 'AX']
    
    green_count = count_green_cells(sheet, rows, cols, "EWS Tower 3")
    ews3 = [1] * green_count


def LIG1(sheet, selected_year, selected_month):
    """LIG Tower 1: row 30 to 44 columns AL, AP, AT, AX (Pour 3)"""
    global lig1
    rows = list(range(30, 45))  # 30 to 44 inclusive
    cols = ['AL', 'AP', 'AT', 'AX']
    
    green_count = count_green_cells(sheet, rows, cols, "LIG Tower 1")
    lig1 = [1] * green_count


def LIG2(sheet, selected_year, selected_month):
    """LIG Tower 2: row 30 to 44 columns U, Y, AC, AG (Pour 2)"""
    global lig2
    rows = list(range(30, 45))  # 30 to 44 inclusive
    cols = ['U', 'Y', 'AC', 'AG']
    
    green_count = count_green_cells(sheet, rows, cols, "LIG Tower 2")
    lig2 = [1] * green_count


def LIG3(sheet, selected_year, selected_month):
    """LIG Tower 3: row 30 to 44 columns D, H, L, P (Pour 1)"""
    global lig3
    rows = list(range(30, 45))  # 30 to 44 inclusive
    cols = ['D', 'H', 'L', 'P']
    
    green_count = count_green_cells(sheet, rows, cols, "LIG Tower 3")
    lig3 = [1] * green_count


def Processjson(data):
    json_data = []
    for project, tower, green, non_green, finishing in zip(
        data["Project Name"],
        data["Tower"],
        data["Green (1)"],
        data["Non-Green (0)"],
        data["Finishing"]
    ):
        total = green + non_green
        structure = f"{math.ceil(green / total * 100)}%" if total > 0 else "0%"
        
        entry = {
            "Project": project,
            "Tower": tower,
            "Slab Count": green,
            "Structure": structure,
            "Finishing": finishing
        }
        json_data.append(entry)
    
    return json_data


def ProcessEWS_LIG(exceldatas, selected_year, selected_month):
    wb = load_workbook(exceldatas, data_only=True)
    sheet_names = wb.sheetnames
    sheet_name = "Revised Baseline 45daysNGT+Rai"

    if sheet_name not in sheet_names:
        st.error(f"Sheet '{sheet_name}' not found in the Excel file!")
        st.write(f"Available sheets: {sheet_names}")
        return json.dumps([])

    sheet = wb[sheet_name]

    # Clear all lists
    global ews1, ews2, ews3, lig1, lig2, lig3
    ews1.clear()
    ews2.clear()
    ews3.clear()
    lig1.clear()
    lig2.clear()
    lig3.clear()
   
    st.write("\n=== Starting Cell Processing ===\n")
    
    # Process each tower
    EWS1(sheet, selected_year, selected_month)
    EWS2(sheet, selected_year, selected_month)
    EWS3(sheet, selected_year, selected_month)
    LIG1(sheet, selected_year, selected_month)
    LIG2(sheet, selected_year, selected_month)
    LIG3(sheet, selected_year, selected_month)

    # Count all green cells for each tower
    ews1_count = len(ews1)
    ews2_count = len(ews2)
    ews3_count = len(ews3)
    lig1_count = len(lig1)
    lig2_count = len(lig2)
    lig3_count = len(lig3)

    st.write(f"\n### Final Counts:")
    st.write(f"EWS Tower 1 (D,H,L,P rows 8-22): {ews1_count} green cells")
    st.write(f"EWS Tower 2 (U,Y,AC,AG rows 8-22): {ews2_count} green cells")
    st.write(f"EWS Tower 3 (AL,AP,AT,AX rows 8-22): {ews3_count} green cells")
    st.write(f"LIG Tower 1 (AL,AP,AT,AX rows 30-44): {lig1_count} green cells")
    st.write(f"LIG Tower 2 (U,Y,AC,AG rows 30-44): {lig2_count} green cells")
    st.write(f"LIG Tower 3 (D,H,L,P rows 30-44): {lig3_count} green cells")

    data = {
        "Project Name": ["EWS", "EWS", "EWS", "LIG", "LIG", "LIG"],
        "Tower": ["EWST1", "EWST2", "EWST3", "LIGT1", "LIGT2", "LIGT3"],
        "Green (1)": [ews1_count, ews2_count, ews3_count, lig1_count, lig2_count, lig3_count],
        "Non-Green (0)": [0, 0, 0, 0, 0, 0],
        "Finishing": ["0%","0%","0%","0%","0%","0%"]
    }

    green_counts = data["Green (1)"]
    
    # Structure percentage based on 60 total possible (15 floors × 4 pours)
    total_possible = 60
    
    data["Structure"] = [f"{(count / total_possible * 100):.2f}%" for count in green_counts]
    
    json_data = Processjson(data)
    
    # Display the data table
    display_df = pd.DataFrame(data)
    st.write("\n### Slab Completion Summary:")
    st.dataframe(display_df)

    return json.dumps(json_data)
