#correct code as of now
# import streamlit as st
# import requests
# import json
# import urllib.parse
# import urllib3
# import certifi
# import pandas as pd
# from datetime import datetime
# import re
# import logging
# import os
# from dotenv import load_dotenv
# import aiohttp
# import asyncio
# from concurrent.futures import ThreadPoolExecutor, as_completed
# import time
# import openpyxl
# import io
# from dotenv import load_dotenv
# from uuid import uuid4
# import ibm_boto3
# from ibm_botocore.client import Config
# from tenacity import retry, stop_after_attempt, wait_exponential
# import xlsxwriter


# STRUCTURAL_STAGES = {
#     "Footing": ["footing"],
#     "Plinth Beam": ["plinth beam", "plinth"],
#     "Shear Wall and Column": ["ground floor shear wall", "ground floor column", "shear wall", "column"],
#     "1st Floor Slab": ["1st floor slab", "first floor slab"],
#     "1st Floor Shear Wall and Column": ["1st floor shear wall", "1st floor column"],
#     "2nd Floor Roof Slab": ["2nd floor", "roof slab", "second floor"],
#     "Terrace Work": ["terrace", "roof"]
# }

# STAGE_EXCLUSIONS = {
#     "Shear Wall and Column": ["1st floor", "first floor", "2nd floor", "second floor", "terrace"],
#     "1st Floor Slab": ["2nd floor", "second floor", "terrace", "roof slab"],
# }

# # Set up logging
# logging.basicConfig(level=logging.INFO)
# logger = logging.getLogger(__name__)

# # Configure logging
# logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
# logger = logging.getLogger(__name__)

# # Disable SSL warnings
# urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# # Load environment variables
# load_dotenv()

# # IBM COS Configuration
# COS_API_KEY = os.getenv("COS_API_KEY")
# COS_SERVICE_INSTANCE_ID = os.getenv("COS_SERVICE_INSTANCE_ID")
# COS_ENDPOINT = os.getenv("COS_ENDPOINT")
# COS_BUCKET = os.getenv("COS_BUCKET")

# # WatsonX configuration
# WATSONX_API_URL = os.getenv("WATSONX_API_URL_1")
# MODEL_ID = os.getenv("MODEL_ID_1")
# PROJECT_ID = os.getenv("PROJECT_ID_1")
# API_KEY = os.getenv("API_KEY_1")

# # API Endpoints
# LOGIN_URL = "https://dms.asite.com/apilogin/"
# IAM_TOKEN_URL = "https://iam.cloud.ibm.com/identity/token"

# # Login Function
# async def login_to_asite(email, password):
#     headers = {"Accept": "application/json", "Content-Type": "application/x-www-form-urlencoded"}
#     payload = {"emailId": email, "password": password}
#     response = requests.post(LOGIN_URL, headers=headers, data=payload, verify=certifi.where(), timeout=50)
#     if response.status_code == 200:
#         try:
#             session_id = response.json().get("UserProfile", {}).get("Sessionid")
#             logger.info(f"Login successful, Session ID: {session_id}")
#             st.session_state.sessionid = session_id
#             st.sidebar.success(f"✅ Login successful, Session ID: {session_id}")
#             return session_id
#         except json.JSONDecodeError:
#             logger.error("JSONDecodeError during login")
#             st.sidebar.error("❌ Failed to parse login response")
#             return None
#     logger.error(f"Login failed: {response.status_code} - {response.text}")
#     st.sidebar.error(f"❌ Login failed: {response.status_code} - {response.text}")
#     return None

# # Function to generate access token
# @retry(stop=stop_after_attempt(5), wait=wait_exponential(multiplier=2, min=10, max=60))
# def get_access_token(API_KEY):
#     headers = {"Content-Type": "application/x-www-form-urlencoded", "Accept": "application/json"}
#     data = {"grant_type": "urn:ibm:params:oauth:grant-type:apikey", "apikey": API_KEY}
#     response = requests.post(IAM_TOKEN_URL, headers=headers, data=data, verify=certifi.where(), timeout=50)
#     try:
#         if response.status_code == 200:
#             token_info = response.json()
#             logger.info("Access token generated successfully")
#             return token_info['access_token']
#         else:
#             logger.error(f"Failed to get access token: {response.status_code} - {response.text}")
#             st.error(f"❌ Failed to get access token: {response.status_code} - {response.text}")
#             raise Exception("Failed to get access token")
#     except Exception as e:
#         logger.error(f"Exception getting access token: {str(e)}")
#         st.error(f"❌ Error getting access token: {str(e)}")
#         return None

# # Initialize COS client
# @retry(stop=stop_after_attempt(5), wait=wait_exponential(multiplier=1, min=4, max=10))
# def initialize_cos_client():
#     try:
#         logger.info("Attempting to initialize COS client...")
#         cos_client = ibm_boto3.client(
#             's3',
#             ibm_api_key_id=COS_API_KEY,
#             ibm_service_instance_id=COS_SERVICE_INSTANCE_ID,
#             config=Config(
#                 signature_version='oauth',
#                 connect_timeout=180,
#                 read_timeout=180,
#                 retries={'max_attempts': 15}
#             ),
#             endpoint_url=COS_ENDPOINT
#         )
#         logger.info("COS client initialized successfully")
#         return cos_client
#     except Exception as e:
#         logger.error(f"Error initializing COS client: {str(e)}")
#         st.error(f"❌ Error initializing COS client: {str(e)}")
#         raise

# # Fetch Workspace ID
# async def GetWorkspaceID():
#     url = "https://dmsak.asite.com/api/workspace/workspacelist"
#     headers = {
#         'Cookie': f'ASessionID={st.session_state.sessionid}',
#         "Accept": "application/json",
#         "Content-Type": "application/x-www-form-urlencoded",
#     }
#     response = requests.get(url, headers=headers)
#     if response.status_code != 200:
#         st.error(f"Failed to fetch workspace list: {response.status_code} - {response.text}")
#         raise Exception(f"Failed to fetch workspace list: {response.status_code}")
#     try:
#         data = response.json()
#         st.session_state.workspaceid = data['asiteDataList']['workspaceVO'][3]['Workspace_Id']
#         st.write(f"Workspace ID: {st.session_state.workspaceid}")
#     except (KeyError, IndexError) as e:
#         st.error(f"Error parsing workspace ID: {str(e)}")
#         raise

# # Fetch Project IDs
# async def GetProjectId():
#     url = f"https://adoddleak.asite.com/commonapi/qaplan/getQualityPlanList;searchCriteria={{'criteria': [{{'field': 'planCreationDate','operator': 6,'values': ['11-Mar-2025']}}], 'projectId': {str(st.session_state.workspaceid)}, 'recordLimit': 1000, 'recordStart': 1}}"
#     headers = {
#         'Cookie': f'ASessionID={st.session_state.sessionid}',
#         "Accept": "application/json",
#         "Content-Type": "application/x-www-form-urlencoded",
#     }
#     response = requests.get(url, headers=headers)
#     if response.status_code != 200:
#         st.error(f"Failed to fetch project IDs: {response.status_code} - {response.text}")
#         raise Exception(f"Failed to fetch project IDs: {response.status_code}")
#     data = response.json()
#     if not data.get('data'):
#         st.error("No quality plans found for the specified date.")
#         raise Exception("No quality plans found")
#     st.session_state.Wave_City_Club_structure = data['data'][0]['planId']
#     st.write(f"Wave City Club Structure Project ID: {st.session_state.Wave_City_Club_structure}")

# # Asynchronous Fetch Function
# async def fetch_data(session, url, headers):
#     async with session.get(url, headers=headers) as response:
#         if response.status == 200:
#             return await response.json()
#         elif response.status == 204:
#             return None
#         else:
#             raise Exception(f"Error fetching data: {response.status} - {await response.text()}")

# # Fetch All Structure Data
# async def GetAllDatas():
#     record_limit = 1000
#     headers = {'Cookie': f'ASessionID={st.session_state.sessionid}'}
#     all_structure_data = []

#     async with aiohttp.ClientSession() as session:
#         start_record = 1
#         st.write("Fetching Wave_City_Club Structure data...")
#         while True:
#             url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanAssociation/?projectId={st.session_state.workspaceid}&planId={st.session_state.Wave_City_Club_structure}&recordStart={start_record}&recordLimit={record_limit}"
#             try:
#                 async with session.get(url, headers=headers) as response:
#                     if response.status == 204:
#                         st.write("No more Wave_City_Club Structure data available (204)")
#                         break
#                     data = await response.json()
#                     if 'associationList' in data and data['associationList']:
#                         all_structure_data.extend(data['associationList'])
#                     else:
#                         all_structure_data.extend(data if isinstance(data, list) else [])
#                     st.write(f"Fetched {len(all_structure_data[-record_limit:])} Wave_City_Club Structure records (Total: {len(all_structure_data)})")
#                     if len(all_structure_data[-record_limit:]) < record_limit:
#                         break
#                     start_record += record_limit
#             except Exception as e:
#                 st.error(f"❌ Error fetching Structure data: {str(e)}")
#                 break

#     df_structure = pd.DataFrame(all_structure_data)
    
#     desired_columns = ['activitySeq', 'qiLocationId']
#     if 'statusName' in df_structure.columns:
#         desired_columns.append('statusName')
#     elif 'statusColor' in df_structure.columns:
#         desired_columns.append('statusColor')
#         status_mapping = {'#4CAF50': 'Completed', '#4CB0F0': 'Not Started', '#4C0F0': 'Not Started'}
#         df_structure['statusName'] = df_structure['statusColor'].map(status_mapping).fillna('Unknown')
#         desired_columns.append('statusName')
#     else:
#         st.error("❌ Neither statusName nor statusColor found in data!")
#         return pd.DataFrame()

#     Wave_City_Club_structure = df_structure[desired_columns]

#     st.write(f"Wave_City_Club STRUCTURE ({', '.join(desired_columns)})")
#     st.write(f"Total records: {len(Wave_City_Club_structure)}")
#     st.write(Wave_City_Club_structure)
    
#     return Wave_City_Club_structure

# # Fetch Activity Data
# async def Get_Activity():
#     record_limit = 1000
#     headers = {
#         'Cookie': f'ASessionID={st.session_state.sessionid}',
#         "Accept": "application/json",
#         "Content-Type": "application/x-www-form-urlencoded",
#     }
    
#     all_structure_activity_data = []
    
#     async with aiohttp.ClientSession() as session:
#         start_record = 1
#         st.write("Fetching Activity data for Wave_City_Club Structure...")
#         while True:
#             url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanActivities/?projectId={st.session_state.workspaceid}&planId={st.session_state.Wave_City_Club_structure}&recordStart={start_record}&recordLimit={record_limit}"
#             try:
#                 data = await fetch_data(session, url, headers)
#                 if data is None:
#                     st.write("No more Structure Activity data available (204)")
#                     break
#                 if 'activityList' in data and data['activityList']:
#                     all_structure_activity_data.extend(data['activityList'])
#                 else:
#                     all_structure_activity_data.extend(data if isinstance(data, list) else [])
#                 st.write(f"Fetched {len(all_structure_activity_data[-record_limit:])} Structure Activity records (Total: {len(all_structure_activity_data)})")
#                 if len(all_structure_activity_data[-record_limit:]) < record_limit:
#                     break
#                 start_record += record_limit
#             except Exception as e:
#                 st.error(f"❌ Error fetching Structure Activity data: {str(e)}")
#                 break
 
#     structure_activity_data = pd.DataFrame(all_structure_activity_data)[['activityName', 'activitySeq', 'formTypeId']]

#     st.write("Wave_City_Club STRUCTURE ACTIVITY DATA (activityName and activitySeq)")
#     st.write(f"Total records: {len(structure_activity_data)}")
#     st.write(structure_activity_data)
      
#     return structure_activity_data

# # Fetch Location/Module Data
# async def Get_Location():
#     record_limit = 1000
#     headers = {
#         'Cookie': f'ASessionID={st.session_state.sessionid}',
#         "Accept": "application/json",
#         "Content-Type": "application/x-www-form-urlencoded",
#     }
    
#     all_structure_location_data = []
    
#     async with aiohttp.ClientSession() as session:
#         start_record = 1
#         total_records_fetched = 0
#         st.write("Fetching Wave_City_Club Structure Location/Module data...")
#         while True:
#             url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanLocation/?projectId={st.session_state.workspaceid}&planId={st.session_state.Wave_City_Club_structure}&recordStart={start_record}&recordLimit={record_limit}"
#             try:
#                 data = await fetch_data(session, url, headers)
#                 if data is None:
#                     st.write("No more Structure Location data available (204)")
#                     break
#                 if isinstance(data, list):
#                     location_data = [{'qiLocationId': item.get('qiLocationId', ''), 'qiParentId': item.get('qiParentId', ''), 'name': item.get('name', '')} 
#                                    for item in data if isinstance(item, dict)]
#                     all_structure_location_data.extend(location_data)
#                     total_records_fetched = len(all_structure_location_data)
#                     st.write(f"Fetched {len(location_data)} Structure Location records (Total: {total_records_fetched})")
#                 elif isinstance(data, dict) and 'locationList' in data and data['locationList']:
#                     location_data = [{'qiLocationId': loc.get('qiLocationId', ''), 'qiParentId': loc.get('qiParentId', ''), 'name': loc.get('name', '')} 
#                                    for loc in data['locationList']]
#                     all_structure_location_data.extend(location_data)
#                     total_records_fetched = len(all_structure_location_data)
#                     st.write(f"Fetched {len(location_data)} Structure Location records (Total: {total_records_fetched})")
#                 else:
#                     st.warning(f"No 'locationList' in Structure Location data or empty list.")
#                     break
#                 if len(location_data) < record_limit:
#                     break
#                 start_record += record_limit
#             except Exception as e:
#                 st.error(f"❌ Error fetching Structure Location data: {str(e)}")
#                 break
        
#     structure_df = pd.DataFrame(all_structure_location_data)
    
#     if 'name' in structure_df.columns and structure_df['name'].isna().all():
#         st.error("❌ All 'name' values in Structure Location data are missing or empty!")

#     st.write("Wave_City_Club STRUCTURE LOCATION/MODULE DATA")
#     st.write(f"Total records: {len(structure_df)}")
#     st.write(structure_df)
    
#     st.session_state.structure_location_data = structure_df
    
#     return structure_df

# # Process individual chunk
# def process_chunk(chunk, chunk_idx, dataset_name, location_df):
#     logger.info(f"Starting thread for {dataset_name} Chunk {chunk_idx + 1}")
#     generated_text = format_chunk_locally(chunk, chunk_idx, len(chunk), dataset_name, location_df)
#     logger.info(f"Completed thread for {dataset_name} Chunk {chunk_idx + 1}")
#     return generated_text, chunk_idx

# # Process data with manual counting
# def process_manually(analysis_df, total, dataset_name, chunk_size=1000, max_workers=4):
#     if analysis_df.empty:
#         st.warning(f"No completed activities found for {dataset_name}.")
#         return "No completed activities found."

#     unique_activities = analysis_df['activityName'].unique()
#     logger.info(f"Unique activities in {dataset_name} dataset: {list(unique_activities)}")
#     logger.info(f"Total records in {dataset_name} dataset: {len(analysis_df)}")

#     st.write(f"Saved Wave_City_Club {dataset_name} data to Wave_City_Club_{dataset_name.lower()}_data.json")
#     chunks = [analysis_df[i:i + chunk_size] for i in range(0, len(analysis_df), chunk_size)]

#     location_df = st.session_state.structure_location_data

#     chunk_results = {}
#     progress_bar = st.progress(0)
#     status_text = st.empty()

#     with ThreadPoolExecutor(max_workers=max_workers) as executor:
#         future_to_chunk = {
#             executor.submit(process_chunk, chunk, idx, dataset_name, location_df): idx 
#             for idx, chunk in enumerate(chunks)
#         }

#         completed_chunks = 0
#         for future in as_completed(future_to_chunk):
#             chunk_idx = future_to_chunk[future]
#             try:
#                 generated_text, idx = future.result()
#                 chunk_results[idx] = generated_text
#                 completed_chunks += 1
#                 progress_percent = completed_chunks / len(chunks)
#                 progress_bar.progress(progress_percent)
#                 status_text.text(f"Processed chunk {completed_chunks} of {len(chunks)} ({progress_percent:.1%} complete)")
#             except Exception as e:
#                 logger.error(f"Error processing chunk {chunk_idx + 1} for {dataset_name}: {str(e)}")
#                 st.error(f"❌ Error processing chunk {chunk_idx + 1}: {str(e)}")

#     parsed_data = {}
#     for chunk_idx in sorted(chunk_results.keys()):
#         generated_text = chunk_results[chunk_idx]
#         if not generated_text:
#             continue

#         current_tower = None
#         tower_activities = []
#         lines = generated_text.split("\n")
        
#         for line in lines:
#             line = line.strip()
#             if not line:
#                 continue
            
#             if line.startswith("Tower:"):
#                 try:
#                     tower_parts = line.split("Tower:", 1)
#                     if len(tower_parts) > 1:
#                         if current_tower and tower_activities:
#                             if current_tower not in parsed_data:
#                                 parsed_data[current_tower] = []
#                             parsed_data[current_tower].extend(tower_activities)
#                             tower_activities = []
#                         current_tower = tower_parts[1].strip()
#                 except Exception as e:
#                     logger.warning(f"Error parsing Tower line: {line}, error: {str(e)}")
#                     if not current_tower:
#                         current_tower = f"Unknown Tower {chunk_idx}"
                    
#             elif line.startswith("Total Completed Activities:"):
#                 continue
#             elif not line.strip().startswith("activityName"):
#                 try:
#                     parts = re.split(r'\s{2,}', line.strip())
#                     if len(parts) >= 2:
#                         activity_name = ' '.join(parts[:-1]).strip()
#                         count_str = parts[-1].strip()
#                         count_match = re.search(r'\d+', count_str)
#                         if count_match:
#                             count = int(count_match.group())
#                             if current_tower:
#                                 tower_activities.append({
#                                     "activityName": activity_name,
#                                     "completedCount": count
#                                 })
#                     else:
#                         match = re.match(r'^\s*(.+?)\s+(\d+)$', line)
#                         if match and current_tower:
#                             activity_name = match.group(1).strip()
#                             count = int(match.group(2).strip())
#                             tower_activities.append({
#                                 "activityName": activity_name,
#                                 "completedCount": count
#                             })
#                 except (ValueError, IndexError) as e:
#                     logger.warning(f"Skipping malformed activity line: {line}, error: {str(e)}")

#         if current_tower and tower_activities:
#             if current_tower not in parsed_data:
#                 parsed_data[current_tower] = []
#             parsed_data[current_tower].extend(tower_activities)

#     aggregated_data = {}
#     for tower_name, activities in parsed_data.items():
#         tower_short_name = tower_name.split('/')[1] if '/' in tower_name else tower_name
#         if tower_short_name not in aggregated_data:
#             aggregated_data[tower_short_name] = {}
        
#         for activity in activities:
#             name = activity.get("activityName", "Unknown")
#             count = activity.get("completedCount", 0)
#             if name in aggregated_data[tower_short_name]:
#                 aggregated_data[tower_short_name][name] += count
#             else:
#                 aggregated_data[tower_short_name][name] = count

#     combined_output_lines = []
#     sorted_towers = sorted(aggregated_data.keys())
    
#     for i, tower_short_name in enumerate(sorted_towers):
#         combined_output_lines.append(f"{tower_short_name:<11} activityName            CompletedCount")
#         activity_dict = aggregated_data[tower_short_name]
#         tower_total = 0
#         for name, count in sorted(activity_dict.items()):
#             combined_output_lines.append(f"{'':<11} {name:<23} {count:>14}")
#             tower_total += count
#         combined_output_lines.append(f"{'':<11} Total for {tower_short_name:<11}: {tower_total:>14}")
#         if i < len(sorted_towers) - 1:
#             combined_output_lines.append("")
    
#     combined_output = "\n".join(combined_output_lines)
#     return combined_output

# # Local formatting function for manual counting
# def format_chunk_locally(chunk, chunk_idx, chunk_size, dataset_name, location_df):
#     towers_data = {}
    
#     for _, row in chunk.iterrows():
#         tower_name = row['tower_name']
#         activity_name = row['activityName']
#         count = int(row['CompletedCount'])
        
#         if tower_name not in towers_data:
#             towers_data[tower_name] = []
            
#         towers_data[tower_name].append({
#             "activityName": activity_name,
#             "completedCount": count
#         })
    
#     output = ""
#     total_activities = 0
    
#     for tower_name, activities in sorted(towers_data.items()):
#         output += f"Tower: {tower_name}\n"
#         output += "activityName            CompletedCount\n"
#         activity_dict = {}
#         for activity in activities:
#             name = activity['activityName']
#             count = activity['completedCount']
#             activity_dict[name] = activity_dict.get(name, 0) + count
#         for name, count in sorted(activity_dict.items()):
#             output += f"{name:<30} {count}\n"
#             total_activities += count
    
#     output += f"Total Completed Activities: {total_activities}"
#     return output

# def process_data(df, activity_df, location_df, dataset_name, stage_name=None):
#     """
#     Modified process_data function that optionally filters by structural stage
    
#     Args:
#         df: The structure data (with statusName, qiLocationId, activitySeq)
#         activity_df: Activity data (with activityName, activitySeq)
#         location_df: Location data (with qiLocationId, name, qiParentId)
#         dataset_name: Name of the dataset (e.g., "Structure")
#         stage_name: Optional - Name of the structural stage (e.g., "Footing", "Plinth Beam")
    
#     Returns:
#         Tuple of (analysis DataFrame, total count)
#     """
#     completed = df[df['statusName'] == 'Completed']
#     if completed.empty:
#         logger.warning(f"No completed activities found in {dataset_name} data" + 
#                       (f" for stage {stage_name}." if stage_name else "."))
#         return pd.DataFrame(), 0

#     completed = completed.merge(location_df[['qiLocationId', 'name']], on='qiLocationId', how='left')
#     completed = completed.merge(activity_df[['activitySeq', 'activityName']], on='activitySeq', how='left')

#     if 'qiActivityId' not in completed.columns:
#         completed['qiActivityId'] = completed['qiLocationId'].astype(str) + '$$' + completed['activitySeq'].astype(str)

#     if completed['name'].isna().all():
#         logger.error(f"All 'name' values are missing in {dataset_name} data after merge!")
#         st.error(f"❌ All 'name' values are missing in {dataset_name} data after merge! Check location data.")
#         completed['name'] = 'Unknown'
#     else:
#         completed['name'] = completed['name'].fillna('Unknown')

#     completed['activityName'] = completed['activityName'].fillna('Unknown')

#     parent_child_dict = dict(zip(location_df['qiLocationId'], location_df['qiParentId']))
#     name_dict = dict(zip(location_df['qiLocationId'], location_df['name']))

#     def get_full_path(location_id):
#         path = []
#         current_id = location_id
#         max_depth = 10
#         depth = 0
        
#         while current_id and depth < max_depth:
#             if current_id not in parent_child_dict or current_id not in name_dict:
#                 logger.warning(f"Location ID {current_id} not found in parent_child_dict or name_dict. Path so far: {path}")
#                 break
            
#             parent_id = parent_child_dict.get(current_id)
#             name = name_dict.get(current_id, "Unknown")
            
#             if not parent_id:
#                 if name != "Quality":
#                     path.append(name)
#                     path.append("Quality")
#                 else:
#                     path.append(name)
#                 break
            
#             path.append(name)
#             current_id = parent_id
#             depth += 1
        
#         if depth >= max_depth:
#             logger.warning(f"Max depth reached while computing path for location_id {location_id}. Possible circular reference. Path: {path}")
        
#         if not path:
#             logger.warning(f"No path constructed for location_id {location_id}. Using 'Unknown'.")
#             return "Unknown"
        
#         full_path = '/'.join(reversed(path))
#         logger.debug(f"Full path for location_id {location_id}: {full_path}")
#         return full_path

#     completed['full_path'] = completed['qiLocationId'].apply(get_full_path)

#     # ============================================================================
#     # CRITICAL: STAGE FILTERING - Apply stage filter based on full_path if stage_name is provided
#     # ============================================================================
#     if stage_name:
#         def matches_stage(full_path, stage):
#             """Check if path contains stage keywords and doesn't contain exclusion keywords"""
#             if pd.isna(full_path):
#                 return False
#             path_lower = str(full_path).lower()
            
#             # Get inclusion keywords
#             keywords = STRUCTURAL_STAGES.get(stage, [])
#             if not keywords:
#                 return False
            
#             # Check if path contains any inclusion keyword
#             has_inclusion = any(keyword in path_lower for keyword in keywords)
#             if not has_inclusion:
#                 return False
            
#             # Check for exclusions
#             exclusions = STAGE_EXCLUSIONS.get(stage, [])
#             if exclusions:
#                 # If path contains any exclusion keyword, reject it
#                 has_exclusion = any(exclusion in path_lower for exclusion in exclusions)
#                 if has_exclusion:
#                     return False
            
#             return True
        
#         # Filter by stage BEFORE further processing
#         logger.info(f"Before stage filter ({stage_name}): {len(completed)} records")
#         completed = completed[completed['full_path'].apply(lambda x: matches_stage(x, stage_name))]
#         logger.info(f"After stage filter ({stage_name}): {len(completed)} records")
        
#         if completed.empty:
#             logger.warning(f"No completed activities found for stage {stage_name} in {dataset_name} data after stage filtering.")
#             st.warning(f"No completed activities found for stage {stage_name} in {dataset_name} data.")
#             return pd.DataFrame(), 0
        
#         # Log sample paths after filtering
#         logger.info(f"Sample paths after stage filtering for {stage_name}: {completed['full_path'].head(10).tolist()}")
#     # ============================================================================

#     # Filter by structural elements
#     def has_structural_element(full_path):
#         """Check if path contains structural work elements"""
#         structural_keywords = [
#             'footing', 'plinth beam', 'slab', 'shear wall', 'column', 
#             'beam', 'roof', 'floor', 'staircase', 'lift', 'water tank', 'terrace'
#         ]
#         path_lower = full_path.lower()
#         return any(keyword in path_lower for keyword in structural_keywords)
    
#     logger.info(f"Sample paths before structural filtering: {completed['full_path'].head(10).tolist()}")
#     completed = completed[completed['full_path'].apply(has_structural_element)]
    
#     if completed.empty:
#         logger.warning(f"No completed activities with structural elements found in {dataset_name} data" + 
#                       (f" for stage {stage_name}" if stage_name else "") + " after filtering.")
#         st.warning(f"No completed activities with structural elements found" + 
#                   (f" for stage {stage_name}" if stage_name else "") + f" in {dataset_name} data.")
#         return pd.DataFrame(), 0
    
#     logger.info(f"After structural element filtering" + 
#                (f" for {stage_name}" if stage_name else "") + f": {len(completed)} records remain")
#     logger.info(f"Sample paths after filtering: {completed['full_path'].head(10).tolist()}")

#     def get_tower_name(full_path):
#         """Extract block name from the path"""
#         parts = full_path.split('/')
        
#         if len(parts) < 2:
#             logger.warning(f"Unexpected path format: {full_path}")
#             return "Unknown"
        
#         block_part = parts[1].strip() if len(parts) > 1 else "Unknown"
#         logger.info(f"Extracting tower from path: {full_path} -> block_part: {block_part}")
        
#         return block_part

#     completed['tower_name'] = completed['full_path'].apply(get_tower_name)
    
#     unique_towers = completed['tower_name'].unique()
#     logger.info(f"Unique tower names found in {dataset_name}" + 
#                (f" for stage {stage_name}" if stage_name else "") + f": {list(unique_towers)}")
#     st.write(f"**Unique tower names found in {dataset_name}" + 
#             (f" for stage {stage_name}" if stage_name else "") + ":**")
#     st.write(list(unique_towers))

#     # Count by tower_name and activityName
#     analysis = completed.groupby(['tower_name', 'activityName']).size().reset_index(name='CompletedCount')
#     analysis = analysis.sort_values(by=['tower_name', 'activityName'], ascending=True)
#     total_completed = analysis['CompletedCount'].sum()

#     logger.info(f"Total completed activities for {dataset_name}" + 
#                (f" stage {stage_name}" if stage_name else "") + f" after processing: {total_completed}")
#     st.write(f"**Activity counts for {dataset_name}" + 
#             (f" - {stage_name}" if stage_name else "") + ":**")
#     st.write(analysis)
    
#     return analysis, total_completed


# # Main analysis function for Wave City Club Structure
# def AnalyzeStatusManually(email=None, password=None):
#     """
#     Modified analysis function that processes data for each structural stage separately
#     """
#     start_time = time.time()

#     if 'sessionid' not in st.session_state:
#         st.error("❌ Please log in first!")
#         return

#     required_data = [
#         'eden_structure',
#         'structure_activity_data',
#         'structure_location_data'
#     ]
    
#     for data_key in required_data:
#         if data_key not in st.session_state:
#             st.error(f"❌ Please fetch required data first! Missing: {data_key}")
#             return
#         if not isinstance(st.session_state[data_key], pd.DataFrame):
#             st.error(f"❌ {data_key} is not a DataFrame! Found type: {type(st.session_state[data_key])}")
#             return

#     structure_data = st.session_state.eden_structure
#     structure_activity = st.session_state.structure_activity_data
#     structure_locations = st.session_state.structure_location_data
    
#     # Validate required columns
#     for df, name in [(structure_data, "Structure")]:
#         if 'statusName' not in df.columns:
#             st.error(f"❌ statusName column not found in {name} data!")
#             return
#         if 'qiLocationId' not in df.columns:
#             st.error(f"❌ qiLocationId column not found in {name} data!")
#             return
#         if 'activitySeq' not in df.columns:
#             st.error(f"❌ activitySeq column not found in {name} data!")
#             return

#     for df, name in [(structure_locations, "Structure Location")]:
#         if 'qiLocationId' not in df.columns or 'name' not in df.columns:
#             st.error(f"❌ qiLocationId or name column not found in {name} data!")
#             return

#     for df, name in [(structure_activity, "Structure Activity")]:
#         if 'activitySeq' not in df.columns or 'activityName' not in df.columns:
#             st.error(f"❌ activitySeq or activityName column not found in {name} data!")
#             return

#     # Initialize storage for stage-wise analysis
#     st.session_state.stage_analysis = {}
#     st.session_state.stage_totals = {}

#     # Process each stage separately
#     st.write("### Processing Data by Structural Stages:")
    
#     for stage_name in STRUCTURAL_STAGES.keys():
#         st.write(f"\n#### Processing Stage: {stage_name}")
#         st.write("="*80)
        
#         # Process the structure data for this specific stage
#         stage_analysis, stage_total = process_data(
#             structure_data, 
#             structure_activity, 
#             structure_locations, 
#             "Structure", 
#             stage_name  # Pass stage_name to enable stage filtering
#         )
        
#         # Store the results
#         st.session_state.stage_analysis[stage_name] = stage_analysis
#         st.session_state.stage_totals[stage_name] = stage_total
        
#         st.write(f"**Stage {stage_name} - Full Output:**")
#         if not stage_analysis.empty:
#             stage_output = process_manually(stage_analysis, stage_total, f"Structure-{stage_name}")
#             if stage_output:
#                 st.text(stage_output)
#         else:
#             st.warning(f"No data found for stage: {stage_name}")
        
#         st.write("="*80)

#     # Store the original analysis for backward compatibility (use Footing as default)
#     st.session_state.structure_analysis = st.session_state.stage_analysis.get('Footing', pd.DataFrame())
#     st.session_state.structure_total = st.session_state.stage_totals.get('Footing', 0)

#     end_time = time.time()
#     st.write(f"\n### Total execution time: {end_time - start_time:.2f} seconds")
#     st.success("✅ Stage-wise analysis completed successfully!")

# def get_cos_files():
#     try:
#         # Initialize COS client (assuming initialize_cos_client is defined elsewhere)
#         cos_client = initialize_cos_client()
#         if not cos_client:
#             st.error("❌ Failed to initialize COS client.")
#             return None

#         # Update prefix to look for files in the Wave City Club folder
#         st.write(f"Attempting to list objects in bucket '{COS_BUCKET}' with prefix 'Wave City Club/'")
#         response = cos_client.list_objects_v2(Bucket=COS_BUCKET, Prefix="Wave City Club/")
#         if 'Contents' not in response:
#             st.error(f"❌ No files found in the 'Wave City Club' folder of bucket '{COS_BUCKET}'.")
#             logger.error("No objects found in Wave City Club folder")
#             return None

#         all_files = [obj['Key'] for obj in response.get('Contents', [])]
#         st.write("**All files in Wave City Club folder:**")
#         if all_files:
#             st.write("\n".join(all_files))
#         else:
#             st.write("No files found.")
#             logger.warning("Wave City Club folder is empty")
#             return None

#         # Update the regex pattern to match the new file name format
#         pattern = re.compile(
#             r"Wave City Club/Structure\s*Work\s*Tracker\s*Wave\s*City\s*Club\s*all\s*Block[\(\s]*(.*?)(?:[\)\s]*\.xlsx)$",
#             re.IGNORECASE
#         )
        
#         # Supported date formats for parsing
#         date_formats = ["%d-%m-%Y", "%d-%m-%y", "%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y"]

#         file_info = []
#         for obj in response.get('Contents', []):
#             key = obj['Key']
#             match = pattern.match(key)
#             if match:
#                 date_str = match.group(1).strip('()').strip()
#                 parsed_date = None
#                 for fmt in date_formats:
#                     try:
#                         parsed_date = datetime.strptime(date_str, fmt)
#                         break
#                     except ValueError:
#                         continue
#                 if parsed_date:
#                     file_info.append({'key': key, 'date': parsed_date})
#                 else:
#                     logger.warning(f"Could not parse date in filename: {key}")
#                     st.warning(f"Skipping file with unparseable date: {key}")
#             else:
#                 st.write(f"File '{key}' does not match the expected pattern 'Wave City Club/Structure Work Tracker Wave City Club all Block (DD-MM-YYYY).xlsx'")

#         if not file_info:
#             st.error("❌ No Excel files matched the expected pattern in the 'Wave City Club' folder.")
#             logger.error("No files matched the expected pattern")
#             return None

#         # Find the latest file based on the parsed date
#         latest_file = max(file_info, key=lambda x: x['date']) if file_info else None
#         if not latest_file:
#             st.error("❌ No valid Excel files found for Structure Work Tracker.")
#             logger.error("No valid files after date parsing")
#             return None

#         file_key = latest_file['key']
#         st.success(f"Found matching file: {file_key}")
#         return file_key
#     except Exception as e:
#         st.error(f"❌ Error fetching COS files: {str(e)}")
#         logger.error(f"Error fetching COS files: {str(e)}")
#         return None

# if 'cos_df_B1' not in st.session_state:
#     st.session_state.cos_df_B1 = None  # For B1 Banket Hall & Finedine
# if 'cos_df_B5' not in st.session_state:
#     st.session_state.cos_df_B5 = None
# if 'cos_df_B6' not in st.session_state:
#     st.session_state.cos_df_B6 = None
# if 'cos_df_B7' not in st.session_state:
#     st.session_state.cos_df_B7 = None
# if 'cos_df_B9' not in st.session_state:
#     st.session_state.cos_df_B9 = None
# if 'cos_df_B8' not in st.session_state:
#     st.session_state.cos_df_B8 = None
# if 'cos_df_B2_B3' not in st.session_state:
#     st.session_state.cos_df_B2_B3 = None  # For B2 & B3
# if 'cos_df_B4' not in st.session_state:
#     st.session_state.cos_df_B4 = None
# if 'cos_df_B11' not in st.session_state:
#     st.session_state.cos_df_B11 = None
# if 'cos_df_B10' not in st.session_state:
#     st.session_state.cos_df_B10 = None

# if 'cos_tname_B1' not in st.session_state:
#     st.session_state.cos_tname_B1 = None  # For B1 Banket Hall & Finedine
# if 'cos_tname_B5' not in st.session_state:
#     st.session_state.cos_tname_B5 = None
# if 'cos_tname_B6' not in st.session_state:
#     st.session_state.cos_tname_B6 = None
# if 'cos_tname_B7' not in st.session_state:
#     st.session_state.cos_tname_B7 = None
# if 'cos_tname_B9' not in st.session_state:
#     st.session_state.cos_tname_B9 = None
# if 'cos_tname_B8' not in st.session_state:
#     st.session_state.cos_tname_B8 = None
# if 'cos_tname_B2_B3' not in st.session_state:
#     st.session_state.cos_tname_B2_B3 = None  # For B2 & B3
# if 'cos_tname_B4' not in st.session_state:
#     st.session_state.cos_tname_B4 = None
# if 'cos_tname_B11' not in st.session_state:
#     st.session_state.cos_tname_B11 = None
# if 'cos_tname_B10' not in st.session_state:
#     st.session_state.cos_tname_B10 = None

# if 'ai_response' not in st.session_state:
#     st.session_state.ai_response = {}  # Initialize as empty dictionary

# # Process Excel files for Wave City Club blocks with updated sheet names and expected_columns
# def process_file(file_stream, filename):
#     """
#     Process COS Excel file and extract activity counts based on Actual Finish dates.
#     Column G = Activity Name (index 6)
#     Column L = Actual Finish Date (index 11)
#     """
#     try:
#         workbook = openpyxl.load_workbook(file_stream)
#         available_sheets = workbook.sheetnames
#         st.write(f"Available sheets in {filename}: {', '.join(available_sheets)}")

#         target_sheets = [
#             "B1 Banket Hall & Finedine ",
#             "B5", "B6", "B7", "B9", "B8", 
#             "B2 & B3",
#             "B4", "B11", "B10"
#         ]
        
#         results = []

#         for sheet_name in target_sheets:
#             if sheet_name not in available_sheets:
#                 st.warning(f"Sheet '{sheet_name}' not found in file: {filename}")
#                 continue

#             file_stream.seek(0)

#             try:
#                 # Read the sheet starting from row 2 (header at row 1)
#                 df = pd.read_excel(file_stream, sheet_name=sheet_name, header=1)
#                 st.write(f"\n📋 Processing sheet: {sheet_name}")
                
#                 # Trim column names
#                 df.columns = [col.strip() if isinstance(col, str) else col for col in df.columns]
                
#                 # Get Column G (Activity Name - index 6) and Column L (Actual Finish - index 11)
#                 if len(df.columns) >= 12:
#                     # Create a clean dataframe with just what we need
#                     clean_df = pd.DataFrame({
#                         'Activity Name': df.iloc[:, 6],  # Column G
#                         'Actual Finish': df.iloc[:, 11]  # Column L
#                     })
                    
#                     # Remove rows where Activity Name is empty
#                     clean_df = clean_df.dropna(subset=['Activity Name'])
#                     clean_df = clean_df[clean_df['Activity Name'].astype(str).str.strip() != '']
                    
#                     # Convert Actual Finish to datetime
#                     clean_df['Actual Finish'] = pd.to_datetime(clean_df['Actual Finish'], errors='coerce')
                    
#                     # Filter only rows with valid Actual Finish dates
#                     clean_df = clean_df[clean_df['Actual Finish'].notna()]
                    
#                     st.write(f"✅ Filtered data for {sheet_name}: **{len(clean_df)} rows** with Actual Finish dates")
#                     if len(clean_df) > 0:
#                         st.write("Sample activities:")
#                         st.write(clean_df.head(5))
                    
#                     results.append((clean_df, sheet_name))
#                 else:
#                     st.error(f"❌ Sheet {sheet_name} has insufficient columns: {len(df.columns)}")
#                     continue

#             except Exception as e:
#                 st.error(f"❌ Error processing sheet {sheet_name}: {str(e)}")
#                 continue

#         if not results:
#             st.error(f"❌ No valid sheets processed from file: {filename}")
#             return [(None, None)]

#         return results

#     except Exception as e:
#         st.error(f"❌ Error loading Excel file: {str(e)}")
#         return [(None, None)]

# def count_activities_by_foundation_concreting(df, sheet_name, stage_name=None):
#     """
#     Count activities based on stage-specific key activity logic:
#     - For each stage, look for a specific key activity in the tracker
#     - Check if that activity has a date in the Actual Finish column (column L, index 11)
#     - If found, count it and apply the same count to all related Civil Works activities
    
#     Stage-to-Activity Mapping:
#     - Footing: Foundation Concreting
#     - Plinth Beam: Plinth Beam Concreting
#     - Shear Wall and Column: GF Column Casting
#     - 1st Floor Slab: GF Roof Slab Casting
#     - 1st Floor Shear Wall and Column: FF Column Casting
#     - 2nd Floor Roof Slab: FF Roof Slab Casting
#     - Terrace Work: Terrace Work
#     """
#     if df is None or df.empty:
#         logger.warning(f"No data for {sheet_name}")
#         return {}
    
#     st.write(f"\n{'='*60}")
#     st.write(f"**🔍 Processing: {sheet_name}**")
#     if stage_name:
#         st.write(f"**📍 Stage: {stage_name}**")
#     st.write(f"{'='*60}")
    
#     # Map stages to their key activities to look for in tracker
#     stage_to_key_activity = {
#         "Footing": ["foundation concreting"],
#         "Plinth Beam": ["plinth beam concreting", "plinth concreting"],
#         "Shear Wall and Column": ["gf column casting", "ground floor column casting", "column casting"],
#         "1st Floor Slab": ["gf roof slab casting", "ground floor roof slab casting", "gf slab casting"],
#         "1st Floor Shear Wall and Column": ["ff column casting", "first floor column casting", "1st floor column casting"],
#         "2nd Floor Roof Slab": ["ff roof slab casting", "first floor roof slab casting", "2nd floor roof slab casting", "sf roof slab casting"],
#         "Terrace Work": ["terrace work", "terrace"]
#     }
    
#     # Keywords to identify foundation concreting and related activities (default fallback)
#     foundation_keywords = ['foundation' ,'Foundation']
    
#     # Target activities to count
#     target_activities = {
#         'Concreting': ['concreting', 'concrete'],
#         'Shuttering': ['shuttering', 'formwork', 'shutter'],
#         'Reinforcement': ['reinforcement', 'rebar', 'steel'],
#         'De-Shuttering': ['de-shuttering', 'deshuttering', 'de shuttering', 'removal']
#     }
    
#     activity_counts = {
#         'Concreting': 0,
#         'Shuttering': 0,
#         'Reinforcement': 0,
#         'De-Shuttering': 0,
#         'Slab conduting': 0  # Added Slab conduting
#     }
    
#     # Determine which key activities to look for based on stage
#     key_activity_keywords = []
#     if stage_name and stage_name in stage_to_key_activity:
#         key_activity_keywords = stage_to_key_activity[stage_name]
#         st.write(f"🎯 **Looking for key activities:** {', '.join(key_activity_keywords)}")
#     else:
#         # Default to foundation concreting for backward compatibility
#         key_activity_keywords = ['foundation concreting', 'foundation concrete']
#         st.write(f"🎯 **Using default:** Foundation Concreting")
    
#     # Step 1: Find key activity instances
#     key_activity_count = 0
#     key_activities_found = []
    
#     for idx, row in df.iterrows():
#         activity_name = str(row['Activity Name']).lower().strip()
        
#         # Check if this matches any of the key activity keywords
#         for keyword in key_activity_keywords:
#             if keyword.lower() in activity_name:
#                 key_activity_count += 1
#                 key_activities_found.append(row['Activity Name'])
#                 break
    
#     if key_activities_found:
#         st.write(f"✅ **Found {key_activity_count} key activity instances:**")
#         for act in key_activities_found[:5]:  # Show first 5
#             st.write(f"   • {act}")
#         if len(key_activities_found) > 5:
#             st.write(f"   ... and {len(key_activities_found) - 5} more")
    
#     # Step 2: If key activity found, apply count to all related activities
#     if key_activity_count > 0:
#         for activity in target_activities.keys():
#             activity_counts[activity] = key_activity_count
        
#         # IMPORTANT: Set Slab conduting equal to Concreting
#         activity_counts['Slab conduting'] = key_activity_count
        
#         st.write(f"\n📊 **Applied count {key_activity_count} to all Civil Works activities**")
#         st.write(f"   • Slab conduting set to match Concreting: {key_activity_count}")
#     else:
#         # If no key activity found, count each activity individually
#         st.write(f"\n⚠️ **No key activity found. Counting activities individually...**")
        
#         for activity_key, keywords in target_activities.items():
#             count = 0
#             found_activities = []
#             for idx, row in df.iterrows():
#                 activity_name = str(row['Activity Name']).lower().strip()
#                 if any(keyword in activity_name for keyword in keywords):
#                     count += 1
#                     found_activities.append(row['Activity Name'])
            
#             activity_counts[activity_key] = count
#             if found_activities:
#                 st.write(f"   {activity_key}: {count} activities found")
        
#         # IMPORTANT: Set Slab conduting equal to Concreting count
#         activity_counts['Slab conduting'] = activity_counts['Concreting']
#         st.write(f"   • Slab conduting set to match Concreting: {activity_counts['Concreting']}")
    
#     # Step 3: Display final counts
#     st.write(f"\n**📈 Final Activity Counts for {sheet_name}:**")
#     for activity, count in activity_counts.items():
#         st.write(f"   • {activity}: **{count}**")
    
#     st.write(f"{'='*60}\n")
#     return activity_counts



# # Function to handle activity count display
# def display_activity_count():
#     """
#     Updated version that uses COS tracker data (Column G and L) to count activities.
#     Uses Foundation Concreting as the base count for related Civil Works activities.
#     """
#     if 'file_key' not in st.session_state or not st.session_state.file_key:
#         st.error("❌ No COS file found. Please fetch COS data first.")
#         return
    
#     try:
#         # Initialize COS client
#         cos_client = initialize_cos_client()
#         if not cos_client:
#             st.error("❌ Failed to initialize COS client")
#             return
        
#         # Fetch the file
#         file_key = st.session_state.file_key
#         st.write(f"📂 Fetching file: **{file_key}**")
        
#         response = cos_client.get_object(Bucket=COS_BUCKET, Key=file_key)
#         file_bytes = io.BytesIO(response['Body'].read())
        
#         # Process the file with new logic
#         st.write("📄 Processing COS file with new activity counting logic...")
#         results = process_file(file_bytes, file_key)
        
#         # Activity categories
#         categories = {
#             "Civil Works": ["Concreting", "Shuttering", "Reinforcement", "De-Shuttering"],
#             "MEP Works": ["Plumbing Works", "Slab conduting", "Wall Conduiting", "Wiring & Switch Socket"],
#             "Interior Finishing Works": ["Floor Tiling", "POP & Gypsum Plaster", "Wall Tiling", "Waterproofing – Sunken"]
#         }
        
#         # Initialize ai_response dictionary
#         if 'ai_response' not in st.session_state or not isinstance(st.session_state.ai_response, dict):
#             st.session_state.ai_response = {}
#             logger.info("Initialized ai_response in display_activity_count")
        
#         # Process each block
#         all_activity_counts = {}
        
#         for df, sheet_name in results:
#             if df is not None and not df.empty:
#                 # Get activity counts for this sheet using foundation concreting logic
#                 activity_counts = count_activities_by_foundation_concreting(df, sheet_name)
                
#                 # Store with clean block name
#                 clean_name = sheet_name.strip()
#                 if clean_name == "B1 Banket Hall & Finedine ":
#                     clean_name = "B1 Banket Hall & Finedine"
                
#                 all_activity_counts[clean_name] = activity_counts
        
#         if not all_activity_counts:
#             st.error("❌ No activity counts found from COS tracker")
#             return
        
#         # Display results
#         st.write("## 📊 Activity Counts by Block (Based on Actual Finish Dates)")
        
#         # Process each block and create AI response structure
#         for block_name, activity_counts in sorted(all_activity_counts.items()):
#             st.write(f"### 🏢 {block_name}")
            
#             # Create AI response structure for this block
#             ai_data = []
            
#             for category, activities in categories.items():
#                 category_data = {
#                     "Category": category,
#                     "Activities": []
#                 }
                
#                 for activity in activities:
#                     # Get count from COS data if available
#                     count = activity_counts.get(activity, 0)
                    
#                     # IMPORTANT: If this is "Slab conduting", use Concreting's count
#                     if activity == "Slab conduting":
#                         count = activity_counts.get("Concreting", 0)
                    
#                     category_data["Activities"].append({
#                         "Activity Name": activity,
#                         "Total": int(count)
#                     })
                
#                 ai_data.append(category_data)
            
#             # Store in session state
#             st.session_state.ai_response[block_name] = ai_data
#             logger.info(f"Stored ai_response for {block_name}: {ai_data}")
            
#             # Display as table
#             display_data = []
#             for category_data in ai_data:
#                 category = category_data["Category"]
#                 for activity in category_data["Activities"]:
#                     display_data.append({
#                         "Category": category,
#                         "Activity Name": activity["Activity Name"],
#                         "Count": activity["Total"]
#                     })
            
#             display_df = pd.DataFrame(display_data)
            
#             # Show by category
#             for category in ["Civil Works", "MEP Works", "Interior Finishing Works"]:
#                 category_df = display_df[display_df["Category"] == category]
#                 if not category_df.empty:
#                     st.write(f"**{category}**")
#                     st.table(category_df[["Activity Name", "Count"]])
        
#         # Create consolidated summary
#         st.write("### 📈 Consolidated Activity Summary Across All Blocks")
        
#         category_mapping = {
#             "Concreting": "Civil Works",
#             "Shuttering": "Civil Works", 
#             "Reinforcement": "Civil Works",
#             "De-Shuttering": "Civil Works",
#             "Plumbing Works": "MEP Works",
#             "Slab conduting": "MEP Works",
#             "Wall Conduiting": "MEP Works", 
#             "Wiring & Switch Socket": "MEP Works",
#             "Floor Tiling": "Interior Finishing Works",
#             "POP & Gypsum Plaster": "Interior Finishing Works",
#             "Wall Tiling": "Interior Finishing Works",
#             "Waterproofing – Sunken": "Interior Finishing Works"
#         }
        
#         consolidated_summary = {}
#         for block_name, ai_data in st.session_state.ai_response.items():
#             for category_data in ai_data:
#                 for activity in category_data["Activities"]:
#                     activity_name = activity["Activity Name"]
#                     count = activity["Total"]
                    
#                     if activity_name not in consolidated_summary:
#                         consolidated_summary[activity_name] = 0
#                     consolidated_summary[activity_name] += count
        
#         # Display consolidated
#         consolidated_data = []
#         for activity_name, total_count in sorted(consolidated_summary.items()):
#             category = category_mapping.get(activity_name, "Other")
#             consolidated_data.append({
#                 "Category": category,
#                 "Activity Name": activity_name,
#                 "Total Count": total_count
#             })
        
#         consolidated_df = pd.DataFrame(consolidated_data)
        
#         for category in ["Civil Works", "MEP Works", "Interior Finishing Works"]:
#             category_df = consolidated_df[consolidated_df["Category"] == category]
#             if not category_df.empty:
#                 st.write(f"**{category}**")
#                 st.table(category_df[["Activity Name", "Total Count"]])
        
#         st.success("✅ Activity counts updated successfully from COS tracker!")
        
#     except Exception as e:
#         st.error(f"❌ Error fetching COS data: {str(e)}")
#         logger.error(f"Error fetching COS data: {str(e)}")
#         import traceback
#         st.code(traceback.format_exc())



# # Function to get access token for WatsonX API
# def get_access_token(api_key):
#     try:
#         headers = {"Content-Type": "application/x-www-form-urlencoded"}
#         data = {
#             "grant_type": "urn:ibm:params:oauth:grant-type:apikey",
#             "apikey": api_key
#         }
#         response = requests.post("https://iam.cloud.ibm.com/identity/token", headers=headers, data=data)
#         if response.status_code == 200:
#             return response.json().get("access_token")
#         else:
#             logger.error(f"Failed to get access token: {response.status_code} - {response.text}")
#             return None
#     except Exception as e:
#         logger.error(f"Error getting access token: {str(e)}")
#         return None

# # WatsonX Prompt Generation
# # WatsonX Prompt Generation (Updated with new categories)
# def generatePrompt(json_datas):
#     try:
#         if isinstance(json_datas, pd.DataFrame):
#             json_str = json_datas.reset_index().to_json(orient='records', indent=2)
#         else:
#             json_str = str(json_datas)

#         body = {
#             "input": f"""
#             Read the table data provided below and categorize the activities into the following categories: Civil Works, MEP Works, Interior Finishing Works, and External Development Activities. Compute the total count of each activity within its respective category and return the results as a JSON array, following the example format provided. For activities like "UP-First Fix" and "CP-First Fix", combine them as "Plumbing Works". If an activity is not found in the data, include it with a count of 0. Ensure the counts are accurate, the output is grouped by category, and the JSON structure is valid with no nested or repeated keys.

#             Table Data:
#             {json_str}

#             Categories and Activities:
#             - Civil Works: Concreting, Shuttering, Reinforcement, De-Shuttering
#             - MEP Works: Plumbing Works, Slab conduting, Wall Conduiting, Wiring & Switch Socket
#             - Interior Finishing Works: Floor Tiling, POP & Gypsum Plaster, Wall Tiling, Waterproofing – Sunken
#             - External Development Activities: Granular Sub-Base, Kerb Stone, Rain Water / Storm Line, Saucer Drain / Paver Block, Sewer Line, Stamp Concrete, Storm Line, WMM

#             Example JSON format needed:
#             [
#               {{
#                 "Category": "Civil Works",
#                 "Activities": [
#                   {{"Activity Name": "Concreting", "Total": 0}},
#                   {{"Activity Name": "Shuttering", "Total": 0}},
#                   {{"Activity Name": "Reinforcement", "Total": 0}},
#                   {{"Activity Name": "De-Shuttering", "Total": 0}}
#                 ]
#               }},
#               {{
#                 "Category": "MEP Works",
#                 "Activities": [
#                   {{"Activity Name": "Plumbing Works", "Total": 0}},
#                   {{"Activity Name": "Slab conduting", "Total": 0}},
#                   {{"Activity Name": "Wall Conduiting", "Total": 0}},
#                   {{"Activity Name": "Wiring & Switch Socket", "Total": 0}}
#                 ]
#               }},
#               {{
#                 "Category": "Interior Finishing Works",
#                 "Activities": [
#                   {{"Activity Name": "Floor Tiling", "Total": 0}},
#                   {{"Activity Name": "POP & Gypsum Plaster", "Total": 0}},
#                   {{"Activity Name": "Wall Tiling", "Total": 0}},
#                   {{"Activity Name": "Waterproofing – Sunken", "Total": 0}}
#                 ]
#               }},
#               {{
#                 "Category": "External Development Activities",
#                 "Activities": [
#                   {{"Activity Name": "Granular Sub-Base", "Total": 0}},
#                   {{"Activity Name": "Kerb Stone", "Total": 0}},
#                   {{"Activity Name": "Rain Water / Storm Line", "Total": 0}},
#                   {{"Activity Name": "Saucer Drain / Paver Block", "Total": 0}},
#                   {{"Activity Name": "Sewer Line", "Total": 0}},
#                   {{"Activity Name": "Stamp Concrete", "Total": 0}},
#                   {{"Activity Name": "Storm Line", "Total": 0}},
#                   {{"Activity Name": "WMM", "Total": 0}}
#                 ]
#               }}
#             ]

#             Return only the JSON array, no additional text, explanations, or code. Ensure the counts are accurate, activities are correctly categorized, and the JSON structure is valid.
#             """,
#             "parameters": {
#                 "decoding_method": "greedy",
#                 "max_new_tokens": 8100,
#                 "min_new_tokens": 0,
#                 "stop_sequences": [";"],
#                 "repetition_penalty": 1.05,
#                 "temperature": 0.5
#             },
#             "model_id": os.getenv("MODEL_ID_1"),
#             "project_id": os.getenv("PROJECT_ID_1")
#         }
        
#         access_token = get_access_token(os.getenv("API_KEY_1"))
#         if not access_token:
#             logger.error("Failed to obtain access token for WatsonX API")
#             return generate_fallback_totals(json_datas)
            
#         headers = {
#             "Accept": "application/json",
#             "Content-Type": "application/json",
#             "Authorization": f"Bearer {access_token}"
#         }
        
#         logger.info("Sending request to WatsonX API")
#         response = requests.post(os.getenv("WATSONX_API_URL_1"), headers=headers, json=body, timeout=60)
        
#         logger.info(f"WatsonX API response status: {response.status_code}")
#         logger.debug(f"WatsonX API response text: {response.text[:1000]}...")  # Log first 1000 chars
        
#         if response.status_code != 200:
#             logger.error(f"WatsonX API call failed: {response.status_code} - {response.text}")
#             st.warning(f"WatsonX API failed with status {response.status_code}: {response.text}. Using fallback method to calculate totals.")
#             return generate_fallback_totals(json_datas)
            
#         response_data = response.json()
#         logger.debug(f"WatsonX API response data: {response_data}")
        
#         if 'results' not in response_data or not response_data['results']:
#             logger.error("WatsonX API response does not contain 'results' key")
#             st.warning("WatsonX API response invalid. Using fallback method to calculate totals.")
#             return generate_fallback_totals(json_datas)

#         generated_text = response_data['results'][0].get('generated_text', '').strip()
#         logger.debug(f"Generated text from WatsonX: {generated_text[:1000]}...")  # Log first 1000 chars
        
#         if not generated_text:
#             logger.error("WatsonX API returned empty generated text")
#             st.warning("WatsonX API returned empty response. Using fallback method to calculate totals.")
#             return generate_fallback_totals(json_datas)

#         if not (generated_text.startswith('[') and generated_text.endswith(']')):
#             start_idx = generated_text.find('[')
#             end_idx = generated_text.rfind(']')
#             if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
#                 generated_text = generated_text[start_idx:end_idx+1]
#                 logger.info("Extracted JSON array from response")
#             else:
#                 logger.error(f"Could not extract valid JSON array from response: {generated_text[:1000]}...")
#                 return generate_fallback_totals(json_datas)
        
#         try:
#             parsed_json = json.loads(generated_text)
#             if not all(isinstance(item, dict) and 'Category' in item and 'Activities' in item for item in parsed_json):
#                 logger.warning("JSON structure doesn't match expected format")
#                 return generate_fallback_totals(json_datas)
#             logger.info("Successfully parsed WatsonX API response")
#             return generated_text
#         except json.JSONDecodeError as e:
#             logger.error(f"WatsonX API returned invalid JSON: {e}")
#             st.warning(f"WatsonX API returned invalid JSON. Error: {str(e)}. Using fallback method to calculate totals.")
#             error_position = int(str(e).split('(char ')[1].split(')')[0]) if '(char ' in str(e) else 0
#             context_start = max(0, error_position - 50)
#             context_end = min(len(generated_text), error_position + 50)
#             logger.error(f"JSON error context: ...{generated_text[context_start:error_position]}[ERROR HERE]{generated_text[error_position:context_end]}...")
#             return generate_fallback_totals(json_datas)
    
#     except Exception as e:
#         logger.error(f"Error in WatsonX API call: {str(e)}")
#         st.warning(f"Error in WatsonX API call: {str(e)}. Using fallback method to calculate totals.")
#         return generate_fallback_totals(json_datas)
    
# # Fallback Total Calculation
# def generate_fallback_totals(count_table):
#     try:
#         if not isinstance(count_table, pd.DataFrame):
#             logger.error("Fallback method received invalid input: not a DataFrame")
#             return json.dumps([
#                 {"Category": "Civil Works", "Activities": [
#                     {"Activity Name": "Concreting", "Total": 0},
#                     {"Activity Name": "Shuttering", "Total": 0},
#                     {"Activity Name": "Reinforcement", "Total": 0},
#                     {"Activity Name": "De-Shuttering", "Total": 0}
#                 ]},
#                 {"Category": "MEP Works", "Activities": [
#                     {"Activity Name": "Plumbing Works", "Total": 0},
#                     {"Activity Name": "Slab conduting", "Total": 0},
#                     {"Activity Name": "Wall Conduiting", "Total": 0},
#                     {"Activity Name": "Wiring & Switch Socket", "Total": 0}
#                 ]},
#                 {"Category": "Interior Finishing Works", "Activities": [
#                     {"Activity Name": "Floor Tiling", "Total": 0},
#                     {"Activity Name": "POP & Gypsum Plaster", "Total": 0},
#                     {"Activity Name": "Wall Tiling", "Total": 0},
#                     {"Activity Name": "Waterproofing – Sunken", "Total": 0}
#                 ]},
#                 {"Category": "External Development Activities", "Activities": [
#                     {"Activity Name": "Granular Sub-Base", "Total": 0},
#                     {"Activity Name": "Kerb Stone", "Total": 0},
#                     {"Activity Name": "Rain Water / Storm Line", "Total": 0},
#                     {"Activity Name": "Saucer Drain / Paver Block", "Total": 0},
#                     {"Activity Name": "Sewer Line", "Total": 0},
#                     {"Activity Name": "Stamp Concrete", "Total": 0},
#                     {"Activity Name": "Storm Line", "Total": 0},
#                     {"Activity Name": "WMM", "Total": 0}
#                 ]}
#             ], indent=2)

#         categories = {
#             "Civil Works": [
#                 "Concreting", "Shuttering", "Reinforcement", "De-Shuttering"
#             ],
#             "MEP Works": [
#                 "Plumbing Works", "Slab conduting", "Wall Conduiting", "Wiring & Switch Socket"
#             ],
#             "Interior Finishing Works": [
#                 "Floor Tiling", "POP & Gypsum Plaster", "Wall Tiling", "Waterproofing – Sunken"
#             ],
#             "External Development Activities": [
#                 "Granular Sub-Base", "Kerb Stone", "Rain Water / Storm Line", "Saucer Drain / Paver Block",
#                 "Sewer Line", "Stamp Concrete", "Storm Line", "WMM"
#             ]
#         }

#         result = []
#         for category, activities in categories.items():
#             category_data = {"Category": category, "Activities": []}
            
#             for activity in activities:
#                 if activity == "Plumbing Works":
#                     combined_count = count_table.loc["UP-First Fix and CP-First Fix", "Count"] if "UP-First Fix and CP-First Fix" in count_table.index else 0
#                     total = combined_count
#                 else:
#                     total = count_table.loc[activity, "Count"] if activity in count_table.index else 0
#                 category_data["Activities"].append({
#                     "Activity Name": activity,
#                     "Total": int(total) if pd.notna(total) else 0
#                 })
            
#             result.append(category_data)

#         return json.dumps(result, indent=2)
#     except Exception as e:
#         logger.error(f"Error in fallback total calculation: {str(e)}")
#         st.error(f"Error in fallback total calculation: {str(e)}")
#         return json.dumps([
#             {"Category": "Civil Works", "Activities": [
#                 {"Activity Name": "Concreting", "Total": 0},
#                 {"Activity Name": "Shuttering", "Total": 0},
#                 {"Activity Name": "Reinforcement", "Total": 0},
#                 {"Activity Name": "De-Shuttering", "Total": 0}
#             ]},
#             {"Category": "MEP Works", "Activities": [
#                 {"Activity Name": "Plumbing Works", "Total": 0},
#                 {"Activity Name": "Slab conduting", "Total": 0},
#                 {"Activity Name": "Wall Conduiting", "Total": 0},
#                 {"Activity Name": "Wiring & Switch Socket", "Total": 0}
#             ]},
#             {"Category": "Interior Finishing Works", "Activities": [
#                 {"Activity Name": "Floor Tiling", "Total": 0},
#                 {"Activity Name": "POP & Gypsum Plaster", "Total": 0},
#                 {"Activity Name": "Wall Tiling", "Total": 0},
#                 {"Activity Name": "Waterproofing – Sunken", "Total": 0}
#             ]},
#             {"Category": "External Development Activities", "Activities": [
#                 {"Activity Name": "Granular Sub-Base", "Total": 0},
#                 {"Activity Name": "Kerb Stone", "Total": 0},
#                 {"Activity Name": "Rain Water / Storm Line", "Total": 0},
#                 {"Activity Name": "Saucer Drain / Paver Block", "Total": 0},
#                 {"Activity Name": "Sewer Line", "Total": 0},
#                 {"Activity Name": "Stamp Concrete", "Total": 0},
#                 {"Activity Name": "Storm Line", "Total": 0},
#                 {"Activity Name": "WMM", "Total": 0}
#             ]}
#         ], indent=2)


# # Extract Totals from AI Data
# def getTotal(ai_data):
#     try:
#         if isinstance(ai_data, str):
#             ai_data = json.loads(ai_data)
            
#         if not isinstance(ai_data, list):
#             logger.error(f"AI data is not a list: {ai_data}")
#             return [0] * len(st.session_state.get('sheduledf', pd.DataFrame()).index)

#         share = []
#         for category_data in ai_data:
#             if isinstance(category_data, dict) and 'Activities' in category_data:
#                 for activity in category_data['Activities']:
#                     if isinstance(activity, dict) and 'Total' in activity:
#                         total = activity['Total']
#                         share.append(int(total) if isinstance(total, (int, float)) and pd.notna(total) else 0)
#                     else:
#                         share.append(0)
#             else:
#                 share.append(0)
#         return share
#     except Exception as e:
#         logger.error(f"Error parsing AI data: {str(e)}")
#         st.error(f"Error parsing AI data: {str(e)}")
#         return [0] * len(st.session_state.get('sheduledf', pd.DataFrame()).index)

# # Function to handle activity count display
# def run_analysis_and_display():
#     try:
#         st.write("Running stage-wise status analysis...")
#         AnalyzeStatusManually()
#         st.success("Stage-wise status analysis completed successfully!")

#         if 'ai_response' not in st.session_state or not isinstance(st.session_state.ai_response, dict):
#             st.session_state.ai_response = {}
#             logger.info("Initialized st.session_state.ai_response in run_analysis_and_display")

#         st.write("Displaying activity counts and generating AI data...")
#         display_activity_count()
#         st.success("Activity counts displayed successfully!")

#         st.write("Checking AI data totals...")
#         logger.info(f"st.session_state.ai_response contents: {st.session_state.ai_response}")
#         if not st.session_state.ai_response:
#             st.error("❌ No AI data available in st.session_state.ai_response. Attempting to regenerate.")
#             logger.error("No AI data in st.session_state.ai_response after display_activity_count")
#             display_activity_count()
#             if not st.session_state.ai_response:
#                 st.error("❌ Failed to regenerate AI data. Please check data fetching and try again.")
#                 logger.error("Failed to regenerate AI data")
#                 return

#         st.write("Generating consolidated checklist Excel file with stage-based sheets...")
        
#         # Check if stage analysis exists
#         if 'stage_analysis' not in st.session_state:
#             st.error("❌ No stage analysis data available. Please ensure stage analysis ran successfully.")
#             logger.error("No stage_analysis in st.session_state")
#             return

#         structure_analysis = st.session_state.get('structure_analysis', None)

#         with st.spinner("Generating Excel file with 7 stage sheets... This may take a moment."):
#             excel_file = generate_consolidated_Checklist_excel(structure_analysis, st.session_state.ai_response)
        
#         if excel_file:
#             timestamp = pd.Timestamp.now(tz='Asia/Kolkata').strftime('%Y%m%d_%H%M')
#             file_name = f"Consolidated_Checklist_WaveCityClub_Stages_{timestamp}.xlsx"
            
#             col1, col2, col3 = st.columns([1, 2, 1])
#             with col2:
#                 st.sidebar.download_button(
#                     label="📥 Download Stage-Based Checklist Excel",
#                     data=excel_file,
#                     file_name=file_name,
#                     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#                     key="download_excel_button_stages",
#                     help="Click to download the consolidated checklist with 7 stage-based sheets."
#                 )
#             st.success("Excel file with stage-based sheets generated successfully! Click the button above to download.")
#         else:
#             st.error("Failed to generate Excel file. Please check the logs for details.")
#             logger.error("Failed to generate Excel file")

#     except Exception as e:
#         st.error(f"Error during analysis, display, or Excel generation: {str(e)}")
#         logger.error(f"Error during analysis, display, or Excel generation: {str(e)}")
#         import traceback
#         st.error(traceback.format_exc())



# # Combined function for Initialize and Fetch Data
# async def initialize_and_fetch_data(email, password):
#     with st.spinner("Starting initialization and data fetching process..."):
#         # Step 1: Login
#         if not email or not password:
#             st.sidebar.error("Please provide both email and password!")
#             return False
#         try:
#             st.sidebar.write("Logging in...")
#             session_id = await login_to_asite(email, password)
#             if not session_id:
#                 st.sidebar.error("Login failed!")
#                 return False
#             st.sidebar.success("Login successful!")
#         except Exception as e:
#             st.sidebar.error(f"Login failed: {str(e)}")
#             return False

#         # Step 2: Get Workspace ID
#         try:
#             st.sidebar.write("Fetching Workspace ID...")
#             await GetWorkspaceID()
#             st.sidebar.success("Workspace ID fetched successfully!")
#         except Exception as e:
#             st.sidebar.error(f"Failed to fetch Workspace ID: {str(e)}")
#             return False

#         # Step 3: Get Project IDs
#         try:
#             st.sidebar.write("Fetching Project IDs...")
#             await GetProjectId()
#             st.sidebar.success("Project IDs fetched successfully!")
#         except Exception as e:
#             st.sidebar.error(f"Failed to fetch Project IDs: {str(e)}")
#             return False

#         # Step 4: Get All Data (Structure only)
#         try:
#             st.sidebar.write("Fetching All Data...")
#             Edenstructure = await GetAllDatas()
#             st.session_state.eden_structure = Edenstructure
#             st.sidebar.success("All Data fetched successfully!")
#         except Exception as e:
#             st.sidebar.error(f"Failed to fetch All Data: {str(e)}")
#             return False

#         # Step 5: Get Activity Data
#         try:
#             st.sidebar.write("Fetching Activity Data...")
#             structure_activity_data = await Get_Activity()
#             st.session_state.structure_activity_data = structure_activity_data
#             st.sidebar.success("Activity Data fetched successfully!")
#         except Exception as e:
#             st.sidebar.error(f"Failed to fetch Activity Data: {str(e)}")
#             return False

#         # Step 6: Get Location/Module Data
#         try:
#             st.sidebar.write("Fetching Location/Module Data...")
#             structure_location_data = await Get_Location()
#             st.session_state.structure_location_data = structure_location_data 
#             st.sidebar.success("Location/Module Data fetched successfully!")
#         except Exception as e:
#             st.sidebar.error(f"Failed to fetch Location/Module Data: {str(e)}")
#             return False

#         # Step 7: Fetch COS Files
#         try:
#             st.sidebar.write("Fetching COS files from Wave City Club folder...")
#             file_key = get_cos_files()
#             st.session_state.file_key = file_key
#             if file_key:
#                 st.success(f"Found 1 file in COS storage: {file_key}")
#                 try:
#                     st.write(f"Processing file: {file_key}")
#                     cos_client = initialize_cos_client()
#                     if not cos_client:
#                         st.error("Failed to initialize COS client during file fetch")
#                         logger.error("COS client initialization failed during file fetch")
#                         return False
#                     st.write("Fetching file from COS...")
#                     response = cos_client.get_object(Bucket=COS_BUCKET, Key=file_key)
#                     file_bytes = io.BytesIO(response['Body'].read())
#                     st.write("File fetched successfully. Processing sheets...")
#                     results = process_file(file_bytes, file_key)
#                     st.write(f"Processing results: {len(results)} sheets processed")
#                     for df, sheet_name in results:
#                         if df is not None:
#                             if sheet_name == "B1 Banket Hall & Finedine ":
#                                 st.session_state.cos_df_B1 = df
#                                 st.session_state.cos_tname_B1 = "B1 Banket Hall & Finedine"
#                                 st.write(f"Processed Data for {sheet_name} - {len(df)} rows:")
#                                 st.write(df.head())
#                             elif sheet_name == "B5":
#                                 st.session_state.cos_df_B5 = df
#                                 st.session_state.cos_tname_B5 = "B5"
#                                 st.write(f"Processed Data for {sheet_name} - {len(df)} rows:")
#                                 st.write(df.head())
#                             elif sheet_name == "B6":
#                                 st.session_state.cos_df_B6 = df
#                                 st.session_state.cos_tname_B6 = "B6"
#                                 st.write(f"Processed Data for {sheet_name} - {len(df)} rows:")
#                                 st.write(df.head())
#                             elif sheet_name == "B7":
#                                 st.session_state.cos_df_B7 = df
#                                 st.session_state.cos_tname_B7 = "B7"
#                                 st.write(f"Processed Data for {sheet_name} - {len(df)} rows:")
#                                 st.write(df.head())
#                             elif sheet_name == "B9":
#                                 st.session_state.cos_df_B9 = df
#                                 st.session_state.cos_tname_B9 = "B9"
#                                 st.write(f"Processed Data for {sheet_name} - {len(df)} rows:")
#                                 st.write(df.head())
#                             elif sheet_name == "B8":
#                                 st.session_state.cos_df_B8 = df
#                                 st.session_state.cos_tname_B8 = "B8"
#                                 st.write(f"Processed Data for {sheet_name} - {len(df)} rows:")
#                                 st.write(df.head())
#                             elif sheet_name == "B2 & B3":
#                                 st.session_state.cos_df_B2_B3 = df
#                                 st.session_state.cos_tname_B2_B3 = "B2 & B3"
#                                 st.write(f"Processed Data for {sheet_name} - {len(df)} rows:")
#                                 st.write(df.head())
#                             elif sheet_name == "B4":
#                                 st.session_state.cos_df_B4 = df
#                                 st.session_state.cos_tname_B4 = "B4"
#                                 st.write(f"Processed Data for {sheet_name} - {len(df)} rows:")
#                                 st.write(df.head())
#                             elif sheet_name == "B11":
#                                 st.session_state.cos_df_B11 = df
#                                 st.session_state.cos_tname_B11 = "B11"
#                                 st.write(f"Processed Data for {sheet_name} - {len(df)} rows:")
#                                 st.write(df.head())
#                             elif sheet_name == "B10":
#                                 st.session_state.cos_df_B10 = df
#                                 st.session_state.cos_tname_B10 = "B10"
#                                 st.write(f"Processed Data for {sheet_name} - {len(df)} rows:")
#                                 st.write(df.head())
#                         else:
#                             st.warning(f"No data processed for {sheet_name} in {file_key}.")
#                 except Exception as e:
#                     st.error(f"Error loading {file_key} from cloud storage: {str(e)}")
#                     logger.error(f"Error loading {file_key}: {str(e)}")
#                     return False
#             else:
#                 st.warning("No expected Excel files available in the 'Wave City Club' folder of the COS bucket.")
#                 return False
#         except Exception as e:
#             st.sidebar.error(f"Failed to fetch COS files: {str(e)}")
#             logger.error(f"Failed to fetch COS files: {str(e)}")
#             return False

#     st.sidebar.success("All steps completed successfully!")
#     return True


# def generate_consolidated_Checklist_excel(structure_analysis=None, activity_counts=None):
#     """
#     Updated Excel generation using pre-processed stage data from session_state
#     """
#     try:
#         if activity_counts is None:
#             activity_counts = st.session_state.get('ai_response', {})
#             if not activity_counts:
#                 st.error("❌ No activity counts data available.")
#                 logger.error("activity_counts is empty in generate_consolidated_Checklist_excel")
#                 return None
        
#         # Check if stage analysis exists
#         if 'stage_analysis' not in st.session_state:
#             st.error("❌ Stage analysis not found. Please run AnalyzeStatusManually first!")
#             logger.error("No stage_analysis in st.session_state")
#             return None

#         # Define categories and activities
#         categories = {
#             "Civil Works": ["Concreting", "Shuttering", "Reinforcement", "De-Shuttering"],
#             "MEP Works": ["Plumbing Works", "Slab conduting", "Wall Conduiting", "Wiring & Switch Socket"],
#             "Interior Finishing Works": ["Floor Tiling", "POP & Gypsum Plaster", "Wall Tiling", "Waterproofing – Sunken"]
#         }

#         cos_to_asite_mapping = {
#             "Concreting": "Concreting",
#             "Shuttering": "Shuttering", 
#             "Reinforcement": "Reinforcement",
#             "De-Shuttering": "De-Shuttering",
#             "Plumbing Works": "Plumbing Works",
#             "Slab conduting": "Slab conduting",
#             "Wall Conduiting": "Wall Conducting",
#             "Wiring & Switch Socket": "Wiring & Switch Socket",
#             "Floor Tiling": "Floor Tiling",
#             "POP & Gypsum Plaster": "POP & Gypsum Plaster",
#             "Wall Tiling": "Wall Tile",
#             "Waterproofing – Sunken": "Waterproofing - Sunken"
#         }

#         block_to_asite_filter = {
#             "B1 Banket Hall & Finedine": [
#                 "01. Block (B1) Banquet Hall ",
#                 "02. Block (B1) Fine Dine"
#             ],
#             "B2 & B3": [
#                 "03. Block 02 (B2) Changing room ",
#                 "04. Block 03 (B3) GYM "
#             ],
#             "B4": "04. Block 4 (B4) ",
#             "B5": "05. Block 05 (B5) Admin +Member Lounge +Creche+AV Room+Surveillance Room +Toilets ",
#             "B6": "06. Block 06 (B6) Toilets ",
#             "B7": "07. Block 07 (B7) Indoor Sports ",
#             "B8": "08. Block 08 (B8) Squash Court ",
#             "B9": "09. Block 09 (B9) Spa and Saloon",
#             "B10": "10. Block 09 (B10) Spa and Saloon",
#             "B11": "11. Block 11 (B11) "
#         }

#         blocks = [
#             "B1 Banket Hall & Finedine", "B5", "B6", "B7", "B9", "B8",
#             "B2 & B3", "B4", "B11", "B10"
#         ]

#         # Create Excel workbook
#         output = io.BytesIO()
#         workbook = xlsxwriter.Workbook(output)

#         header_format = workbook.add_format({'bold': True, 'bg_color': '#D3D3D3'})
#         total_format = workbook.add_format({'bold': True, 'bg_color': '#FFDAB9'})
#         cell_format = workbook.add_format({'border': 1})

#         # Generate each stage sheet using pre-processed data
#         for stage_name in STRUCTURAL_STAGES.keys():
#             worksheet = workbook.add_worksheet(stage_name)
#             logger.info(f"Creating sheet: {stage_name}")
            
#             # Get pre-processed stage analysis
#             stage_analysis = st.session_state.stage_analysis.get(stage_name, pd.DataFrame())
            
#             if stage_analysis.empty:
#                 worksheet.write(0, 0, f"No data available for {stage_name}", header_format)
#                 logger.warning(f"No data for stage {stage_name}")
#                 continue

#             # **NEW: Generate BLOCK-SPECIFIC activity counts for this stage**
#             stage_activity_counts = {}
            
#             # Map stages to their key activities
#             stage_to_key_activity = {
#                 "Footing": ["foundation concreting"],
#                 "Plinth Beam": ["plinth beam concreting", "plinth concreting"],
#                 "Shear Wall and Column": ["gf column casting", "ground floor column casting"],
#                 "1st Floor Slab": ["gf roof slab casting", "ground floor roof slab casting"],
#                 "1st Floor Shear Wall and Column": ["ff column casting", "first floor column casting"],
#                 "2nd Floor Roof Slab": ["ff roof slab casting", "first floor roof slab casting", "2nd floor roof slab casting"],
#                 "Terrace Work": ["terrace work", "terrace"]
#             }
            
#             key_activities = stage_to_key_activity.get(stage_name, ["foundation concreting"])
#             logger.info(f"Stage {stage_name}: Looking for activities: {key_activities}")
            
#             # Re-process tracker data for each block individually
#             try:
#                 cos_client = initialize_cos_client()
#                 if cos_client and 'file_key' in st.session_state and st.session_state.file_key:
#                     file_key = st.session_state.file_key
#                     response = cos_client.get_object(Bucket=COS_BUCKET, Key=file_key)
#                     file_bytes = io.BytesIO(response['Body'].read())
                    
#                     # Process file to get tracker data for each block
#                     results = process_file(file_bytes, file_key)
                    
#                     for df, sheet_name in results:
#                         if df is not None and not df.empty:
#                             # Clean block name
#                             clean_name = sheet_name.strip()
#                             if clean_name == "B1 Banket Hall & Finedine ":
#                                 clean_name = "B1 Banket Hall & Finedine"
                            
#                             # Count the key activity for THIS SPECIFIC BLOCK ONLY
#                             key_activity_count = 0
#                             found_activities = []
                            
#                             for idx, row in df.iterrows():
#                                 activity_name = str(row['Activity Name']).lower().strip()
                                
#                                 # Check if this row matches any of the key activities for this stage
#                                 for key_activity in key_activities:
#                                     if key_activity.lower() in activity_name:
#                                         key_activity_count += 1
#                                         found_activities.append(row['Activity Name'])
#                                         logger.info(f"Stage {stage_name}, Block {clean_name}: Found '{row['Activity Name']}'")
#                                         break
                            
#                             # Store count for this block - all Civil Works activities get this count
#                             block_activity_counts = {
#                                 'Concreting': key_activity_count,
#                                 'Shuttering': key_activity_count,
#                                 'Reinforcement': key_activity_count,
#                                 'De-Shuttering': key_activity_count,
#                                 'Slab conduting': key_activity_count
#                             }
                            
#                             stage_activity_counts[clean_name] = block_activity_counts
#                             logger.info(f"Stage {stage_name}, Block {clean_name}: Key activity count = {key_activity_count} (found {len(found_activities)} activities)")
                
#                 logger.info(f"Generated block-specific activity counts for {stage_name}: {stage_activity_counts}")
#             except Exception as e:
#                 logger.error(f"Error generating block-specific counts for {stage_name}: {str(e)}")
#                 # Fallback to using global activity_counts
#                 stage_activity_counts = activity_counts

#             consolidated_rows = []

#             # Process data for each block and category
#             for block in blocks:
#                 for category, activities in categories.items():
#                     for activity in activities:
#                         # **NEW: Skip Slab conduting for stages other than 1st Floor Slab and 2nd Floor Roof Slab**
#                         if activity == "Slab conduting" and stage_name not in ["1st Floor Slab", "2nd Floor Roof Slab"]:
#                             continue
                        
#                         asite_activity = cos_to_asite_mapping.get(activity, activity)
#                         asite_activities = asite_activity if isinstance(asite_activity, list) else [asite_activity]

#                         # Get closed_checklist from stage_analysis (already filtered by stage!)
#                         closed_checklist = 0
#                         asite_filters = block_to_asite_filter.get(block, block)
                        
#                         if isinstance(asite_filters, list):
#                             for asite_filter in asite_filters:
#                                 for asite_act in asite_activities:
#                                     matching_rows = stage_analysis[
#                                         (stage_analysis['tower_name'].str.strip() == asite_filter.strip()) &
#                                         (stage_analysis['activityName'] == asite_act)
#                                     ]
#                                     if not matching_rows.empty:
#                                         count = matching_rows['CompletedCount'].sum()
#                                         closed_checklist += count
#                                         logger.info(f"Sheet {stage_name}, Block {block} ('{asite_filter}'), Activity: {asite_act}, Count: {count}")
#                         else:
#                             for asite_act in asite_activities:
#                                 matching_rows = stage_analysis[
#                                     (stage_analysis['tower_name'].str.strip() == asite_filters.strip()) &
#                                     (stage_analysis['activityName'] == asite_act)
#                                 ]
#                                 if not matching_rows.empty:
#                                     count = matching_rows['CompletedCount'].sum()
#                                     closed_checklist += count
#                                     logger.info(f"Sheet {stage_name}, Block {block} ('{asite_filters}'), Activity: {asite_act}, Count: {count}")

#                         # Get COS data (Completed Work count) - use stage-specific counts
#                         completed_flats = 0
                        
#                         # **NEW: Only get COS data for Slab conduting in specific sheets**
#                         if activity == "Slab conduting" and stage_name in ["1st Floor Slab", "2nd Floor Roof Slab"]:
#                             if block in stage_activity_counts:
#                                 activity_data_dict = stage_activity_counts[block]
#                                 if isinstance(activity_data_dict, dict):
#                                     completed_flats = activity_data_dict.get(activity, 0)
#                         elif activity != "Slab conduting":
#                             # For all other activities, get COS data from stage-specific counts
#                             if block in stage_activity_counts:
#                                 activity_data_dict = stage_activity_counts[block]
#                                 if isinstance(activity_data_dict, dict):
#                                     completed_flats = activity_data_dict.get(activity, 0)

#                         # Calculate open/missing
#                         in_progress = 0
#                         if completed_flats == 0 or closed_checklist > completed_flats:
#                             open_missing = 0
#                         else:
#                             open_missing = abs(completed_flats - closed_checklist)

#                         display_activity = asite_activities[0]

#                         consolidated_rows.append({
#                             "Block": block,
#                             "Category": category,
#                             "Activity Name": display_activity,
#                             "Completed Work*(Count of Flat)": completed_flats,
#                             "In progress": in_progress,
#                             "Closed checklist": closed_checklist,
#                             "Open/Missing check list": open_missing
#                         })
            
#             # Store consolidated_rows in session state for summary calculation
#             if 'all_consolidated_rows' not in st.session_state:
#                 st.session_state.all_consolidated_rows = {}
#             st.session_state.all_consolidated_rows[stage_name] = consolidated_rows.copy()

#             # Write to worksheet
#             df = pd.DataFrame(consolidated_rows)
#             if df.empty:
#                 worksheet.write(0, 0, f"No activities found for {stage_name}", header_format)
#                 continue

#             df.sort_values(by=["Block", "Category"], inplace=True)

#             headers = ["Activity Name", "Completed", "In progress", "Closed checklist", "Open/Missing check list"]
#             col_start = 1
#             row_start = 0

#             grouped_by_block = df.groupby('Block')

#             for block, block_group in grouped_by_block:
#                 col_pos = col_start
#                 grouped_by_category = block_group.groupby('Category')

#                 for category, cat_group in grouped_by_category:
#                     worksheet.merge_range(row_start, col_pos, row_start, col_pos + 4, 
#                                         f"{block} {category} - {stage_name}", header_format)
#                     row_pos = row_start + 1

#                     for i, header in enumerate(headers):
#                         worksheet.write(row_pos, col_pos + i, header, header_format)
#                     row_pos += 1

#                     for _, row in cat_group.iterrows():
#                         worksheet.write(row_pos, col_pos, row["Activity Name"], cell_format)
#                         worksheet.write(row_pos, col_pos + 1, row["Completed Work*(Count of Flat)"], cell_format)
#                         worksheet.write(row_pos, col_pos + 2, row["In progress"], cell_format)
#                         worksheet.write(row_pos, col_pos + 3, row["Closed checklist"], cell_format)
#                         worksheet.write(row_pos, col_pos + 4, row["Open/Missing check list"], cell_format)
#                         row_pos += 1

#                     total_pending = cat_group["Open/Missing check list"].sum()
#                     worksheet.merge_range(row_pos, col_pos, row_pos, col_pos + 3, "Total pending check list", total_format)
#                     worksheet.write(row_pos, col_pos + 4, total_pending, total_format)
#                     row_pos += 2

#                     col_pos += 6

#                 row_start = row_pos

#             for col in range(col_start, col_pos):
#                 worksheet.set_column(col, col, 20)

#         # **NEW: Create Summary Sheet with dynamic month name**
#         current_month = pd.Timestamp.now(tz='Asia/Kolkata').strftime('%B')  # Full month name (e.g., "January")
#         summary_sheet_name = f"Checklist {current_month}"
#         worksheet_summary = workbook.add_worksheet(summary_sheet_name)
#         current_row = 0

#         worksheet_summary.write(current_row, 0, f"Checklist: {current_month}", header_format)
#         current_row += 1

#         summary_headers = [
#             "Site",
#             "Total of Missing & Open Checklist-Civil",
#             "Total of Missing & Open Checklist-MEP",
#             "Total of Missing & Open Checklist-Interior Finishing",
#             "TOTAL"
#         ]
#         for col, header in enumerate(summary_headers, start=0):
#             worksheet_summary.write(current_row, col, header, header_format)
#         current_row += 1

#         def map_category_to_type(category):
#             if category in ["Civil Works"]:
#                 return "Civil"
#             elif category in ["MEP Works"]:
#                 return "MEP"
#             elif category in ["Interior Finishing Works"]:
#                 return "Interior"
#             else:
#                 return "Civil"

#         summary_data = {}
        
#         # Store all consolidated rows from sheet generation to use in summary
#         if 'all_consolidated_rows' not in st.session_state:
#             st.session_state.all_consolidated_rows = {}
        
#         # Aggregate data from the already-calculated consolidated rows
#         for stage_name in STRUCTURAL_STAGES.keys():
#             if stage_name in st.session_state.all_consolidated_rows:
#                 stage_rows = st.session_state.all_consolidated_rows[stage_name]
                
#                 for row in stage_rows:
#                     block = row['Block']
#                     category = row['Category']
#                     open_missing = row['Open/Missing check list']
                    
#                     # Convert block name to display format
#                     if block == "B1 Banket Hall & Finedine":
#                         site_name = "WaveCityClub-Block 01 Banket Hall & Finedine"
#                     elif "&" in block:
#                         block_num = block.replace(" & ", "&")
#                         site_name = f"WaveCityClub-Block {block_num}"
#                     else:
#                         block_num = block[1:]
#                         if len(block_num) == 1:
#                             block_num = f"0{block_num}"
#                         site_name = f"WaveCityClub-Block {block_num}"
                    
#                     type_ = map_category_to_type(category)
                    
#                     if site_name not in summary_data:
#                         summary_data[site_name] = {"Civil": 0, "MEP": 0, "Interior": 0}
                    
#                     summary_data[site_name][type_] += open_missing

#         # Write summary data
#         for site_name, counts in sorted(summary_data.items()):
#             civil_count = counts["Civil"]
#             mep_count = counts["MEP"]
#             interior_count = counts["Interior"]
#             total_count = civil_count + mep_count + interior_count
            
#             worksheet_summary.write(current_row, 0, site_name, cell_format)
#             worksheet_summary.write(current_row, 1, civil_count, cell_format)
#             worksheet_summary.write(current_row, 2, mep_count, cell_format)
#             worksheet_summary.write(current_row, 3, interior_count, cell_format)
#             worksheet_summary.write(current_row, 4, total_count, cell_format)
#             current_row += 1

#         # Auto-adjust column widths
#         for col in range(5):
#             worksheet_summary.set_column(col, col, 25)

#         workbook.close()
#         output.seek(0)
        
#         logger.info("Successfully generated Excel file with 7 stage-based sheets")
#         return output

#     except Exception as e:
#         logger.error(f"Error generating consolidated Excel: {str(e)}")
#         st.error(f"❌ Error generating Excel file: {str(e)}")
#         import traceback
#         logger.error(f"Full traceback: {traceback.format_exc()}")
#         return None

# # Combined function to handle analysis and display
# def run_analysis_and_display():
#     try:
#         st.write("Running stage-wise status analysis...")
#         AnalyzeStatusManually()
#         st.success("Stage-wise status analysis completed successfully!")

#         if 'ai_response' not in st.session_state or not isinstance(st.session_state.ai_response, dict):
#             st.session_state.ai_response = {}
#             logger.info("Initialized st.session_state.ai_response in run_analysis_and_display")

#         st.write("Displaying activity counts and generating AI data...")
#         display_activity_count()
#         st.success("Activity counts displayed successfully!")

#         st.write("Checking AI data totals...")
#         logger.info(f"st.session_state.ai_response contents: {st.session_state.ai_response}")
#         if not st.session_state.ai_response:
#             st.error("❌ No AI data available in st.session_state.ai_response. Attempting to regenerate.")
#             logger.error("No AI data in st.session_state.ai_response after display_activity_count")
#             display_activity_count()
#             if not st.session_state.ai_response:
#                 st.error("❌ Failed to regenerate AI data. Please check data fetching and try again.")
#                 logger.error("Failed to regenerate AI data")
#                 return

#         st.write("Generating consolidated checklist Excel file with stage-based sheets...")
        
#         # Check if stage analysis exists
#         if 'stage_analysis' not in st.session_state:
#             st.error("❌ No stage analysis data available. Please ensure stage analysis ran successfully.")
#             logger.error("No stage_analysis in st.session_state")
#             return

#         structure_analysis = st.session_state.get('structure_analysis', None)

#         with st.spinner("Generating Excel file with 7 stage sheets... This may take a moment."):
#             excel_file = generate_consolidated_Checklist_excel(structure_analysis, st.session_state.ai_response)
        
#         if excel_file:
#             timestamp = pd.Timestamp.now(tz='Asia/Kolkata').strftime('%Y%m%d_%H%M')
#             file_name = f"Consolidated_Checklist_WaveCityClub_Stages_{timestamp}.xlsx"
            
#             col1, col2, col3 = st.columns([1, 2, 1])
#             with col2:
#                 st.sidebar.download_button(
#                     label="📥 Download Stage-Based Checklist Excel",
#                     data=excel_file,
#                     file_name=file_name,
#                     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#                     key="download_excel_button_stages",
#                     help="Click to download the consolidated checklist with 7 stage-based sheets."
#                 )
#             st.success("Excel file with stage-based sheets generated successfully! Click the button above to download.")
#         else:
#             st.error("Failed to generate Excel file. Please check the logs for details.")
#             logger.error("Failed to generate Excel file")

#     except Exception as e:
#         st.error(f"Error during analysis, display, or Excel generation: {str(e)}")
#         logger.error(f"Error during analysis, display, or Excel generation: {str(e)}")
#         import traceback
#         st.error(traceback.format_exc())

# # Streamlit UI
# st.markdown(
#     """
#     <h1 style='font-family: "Arial Black", Gadget, sans-serif; 
#                color: red; 
#                font-size: 48px; 
#                text-align: center;'>
#         CheckList - Report
#     </h1>
#     """,
#     unsafe_allow_html=True
# )

# # Initialize and Fetch Data
# st.sidebar.title("🔒 Asite Initialization")
# email = st.sidebar.text_input("Email", "impwatson@gadieltechnologies.com", key="email_input")
# password = st.sidebar.text_input("Password", "Srihari@790$", type="password", key="password_input")

# if st.sidebar.button("Initialize and Fetch Data"):
#     loop = asyncio.new_event_loop()
#     asyncio.set_event_loop(loop)
#     try:
#         success = loop.run_until_complete(initialize_and_fetch_data(email, password))
#         if success:
#             st.sidebar.success("Initialization and data fetching completed successfully!")
#         else:
#             st.sidebar.error("Initialization and data fetching failed!")
#     except Exception as e:
#         st.sidebar.error(f"Initialization and data fetching failed: {str(e)}")
#     finally:
#         loop.close()

# # Analyze and Display
# st.sidebar.title("📊 Status Analysis")
# if st.sidebar.button("Analyze and Display Activity Counts"):
#     with st.spinner("Running analysis and displaying activity counts..."):
#         run_analysis_and_display()






























import streamlit as st
import requests
import json
import urllib.parse
import urllib3
import certifi
import pandas as pd
from datetime import datetime
import re
import logging
import os
from dotenv import load_dotenv
import aiohttp
import asyncio
from concurrent.futures import ThreadPoolExecutor, as_completed
import time
import openpyxl
import io
from dotenv import load_dotenv
from uuid import uuid4
import ibm_boto3
from ibm_botocore.client import Config
from tenacity import retry, stop_after_attempt, wait_exponential
import xlsxwriter


STRUCTURAL_STAGES = {
    "Footing": ["footing"],
    "Plinth Beam": ["plinth beam", "plinth"],
    "Shear Wall and Column": ["ground floor shear wall", "ground floor column", "shear wall", "column"],
    "1st Floor Slab": ["1st floor slab", "first floor slab"],
    "1st Floor Shear Wall and Column": ["1st floor shear wall", "1st floor column"],
    "2nd Floor Roof Slab": ["2nd floor", "roof slab", "second floor"],
    "Terrace Work": ["terrace", "roof"]
}

STAGE_EXCLUSIONS = {
    "Shear Wall and Column": ["1st floor", "first floor", "2nd floor", "second floor", "terrace"],
    "1st Floor Slab": ["2nd floor", "second floor", "terrace", "roof slab"],
}

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Disable SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Load environment variables
load_dotenv()

# IBM COS Configuration
COS_API_KEY = os.getenv("COS_API_KEY")
COS_SERVICE_INSTANCE_ID = os.getenv("COS_SERVICE_INSTANCE_ID")
COS_ENDPOINT = os.getenv("COS_ENDPOINT")
COS_BUCKET = os.getenv("COS_BUCKET")

# WatsonX configuration
WATSONX_API_URL = os.getenv("WATSONX_API_URL_1")
MODEL_ID = os.getenv("MODEL_ID_1")
PROJECT_ID = os.getenv("PROJECT_ID_1")
API_KEY = os.getenv("API_KEY_1")

# API Endpoints
LOGIN_URL = "https://dms.asite.com/apilogin/"
IAM_TOKEN_URL = "https://iam.cloud.ibm.com/identity/token"

# Login Function
async def login_to_asite(email, password):
    headers = {"Accept": "application/json", "Content-Type": "application/x-www-form-urlencoded"}
    payload = {"emailId": email, "password": password}
    response = requests.post(LOGIN_URL, headers=headers, data=payload, verify=certifi.where(), timeout=50)
    if response.status_code == 200:
        try:
            session_id = response.json().get("UserProfile", {}).get("Sessionid")
            logger.info(f"Login successful, Session ID: {session_id}")
            st.session_state.sessionid = session_id
            st.sidebar.success(f"✅ Login successful, Session ID: {session_id}")
            return session_id
        except json.JSONDecodeError:
            logger.error("JSONDecodeError during login")
            st.sidebar.error("❌ Failed to parse login response")
            return None
    logger.error(f"Login failed: {response.status_code} - {response.text}")
    st.sidebar.error(f"❌ Login failed: {response.status_code} - {response.text}")
    return None

# Function to generate access token
@retry(stop=stop_after_attempt(5), wait=wait_exponential(multiplier=2, min=10, max=60))
def get_access_token(API_KEY):
    headers = {"Content-Type": "application/x-www-form-urlencoded", "Accept": "application/json"}
    data = {"grant_type": "urn:ibm:params:oauth:grant-type:apikey", "apikey": API_KEY}
    response = requests.post(IAM_TOKEN_URL, headers=headers, data=data, verify=certifi.where(), timeout=50)
    try:
        if response.status_code == 200:
            token_info = response.json()
            logger.info("Access token generated successfully")
            return token_info['access_token']
        else:
            logger.error(f"Failed to get access token: {response.status_code} - {response.text}")
            st.error(f"❌ Failed to get access token: {response.status_code} - {response.text}")
            raise Exception("Failed to get access token")
    except Exception as e:
        logger.error(f"Exception getting access token: {str(e)}")
        st.error(f"❌ Error getting access token: {str(e)}")
        return None

# Initialize COS client
@retry(stop=stop_after_attempt(5), wait=wait_exponential(multiplier=1, min=4, max=10))
def initialize_cos_client():
    try:
        logger.info("Attempting to initialize COS client...")
        cos_client = ibm_boto3.client(
            's3',
            ibm_api_key_id=COS_API_KEY,
            ibm_service_instance_id=COS_SERVICE_INSTANCE_ID,
            config=Config(
                signature_version='oauth',
                connect_timeout=180,
                read_timeout=180,
                retries={'max_attempts': 15}
            ),
            endpoint_url=COS_ENDPOINT
        )
        logger.info("COS client initialized successfully")
        return cos_client
    except Exception as e:
        logger.error(f"Error initializing COS client: {str(e)}")
        st.error(f"❌ Error initializing COS client: {str(e)}")
        raise

# Fetch Workspace ID
async def GetWorkspaceID():
    url = "https://dmsak.asite.com/api/workspace/workspacelist"
    headers = {
        'Cookie': f'ASessionID={st.session_state.sessionid}',
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    response = requests.get(url, headers=headers)
    if response.status_code != 200:
        st.error(f"Failed to fetch workspace list: {response.status_code} - {response.text}")
        raise Exception(f"Failed to fetch workspace list: {response.status_code}")
    try:
        data = response.json()
        st.session_state.workspaceid = data['asiteDataList']['workspaceVO'][3]['Workspace_Id']
        st.write(f"Workspace ID: {st.session_state.workspaceid}")
    except (KeyError, IndexError) as e:
        st.error(f"Error parsing workspace ID: {str(e)}")
        raise

# Fetch Project IDs
async def GetProjectId():
    url = f"https://adoddleak.asite.com/commonapi/qaplan/getQualityPlanList;searchCriteria={{'criteria': [{{'field': 'planCreationDate','operator': 6,'values': ['11-Mar-2025']}}], 'projectId': {str(st.session_state.workspaceid)}, 'recordLimit': 1000, 'recordStart': 1}}"
    headers = {
        'Cookie': f'ASessionID={st.session_state.sessionid}',
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    response = requests.get(url, headers=headers)
    if response.status_code != 200:
        st.error(f"Failed to fetch project IDs: {response.status_code} - {response.text}")
        raise Exception(f"Failed to fetch project IDs: {response.status_code}")
    data = response.json()
    if not data.get('data'):
        st.error("No quality plans found for the specified date.")
        raise Exception("No quality plans found")
    st.session_state.Wave_City_Club_structure = data['data'][0]['planId']
    st.write(f"Wave City Club Structure Project ID: {st.session_state.Wave_City_Club_structure}")

# Asynchronous Fetch Function
async def fetch_data(session, url, headers):
    async with session.get(url, headers=headers) as response:
        if response.status == 200:
            return await response.json()
        elif response.status == 204:
            return None
        else:
            raise Exception(f"Error fetching data: {response.status} - {await response.text()}")

# Fetch All Structure Data
async def GetAllDatas():
    record_limit = 1000
    headers = {'Cookie': f'ASessionID={st.session_state.sessionid}'}
    all_structure_data = []

    async with aiohttp.ClientSession() as session:
        start_record = 1
        st.write("Fetching Wave_City_Club Structure data...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanAssociation/?projectId={st.session_state.workspaceid}&planId={st.session_state.Wave_City_Club_structure}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                async with session.get(url, headers=headers) as response:
                    if response.status == 204:
                        st.write("No more Wave_City_Club Structure data available (204)")
                        break
                    data = await response.json()
                    if 'associationList' in data and data['associationList']:
                        all_structure_data.extend(data['associationList'])
                    else:
                        all_structure_data.extend(data if isinstance(data, list) else [])
                    st.write(f"Fetched {len(all_structure_data[-record_limit:])} Wave_City_Club Structure records (Total: {len(all_structure_data)})")
                    if len(all_structure_data[-record_limit:]) < record_limit:
                        break
                    start_record += record_limit
            except Exception as e:
                st.error(f"❌ Error fetching Structure data: {str(e)}")
                break

    df_structure = pd.DataFrame(all_structure_data)
    
    desired_columns = ['activitySeq', 'qiLocationId']
    if 'statusName' in df_structure.columns:
        desired_columns.append('statusName')
    elif 'statusColor' in df_structure.columns:
        desired_columns.append('statusColor')
        status_mapping = {'#4CAF50': 'Completed', '#4CB0F0': 'Not Started', '#4C0F0': 'Not Started'}
        df_structure['statusName'] = df_structure['statusColor'].map(status_mapping).fillna('Unknown')
        desired_columns.append('statusName')
    else:
        st.error("❌ Neither statusName nor statusColor found in data!")
        return pd.DataFrame()

    Wave_City_Club_structure = df_structure[desired_columns]

    st.write(f"Wave_City_Club STRUCTURE ({', '.join(desired_columns)})")
    st.write(f"Total records: {len(Wave_City_Club_structure)}")
    st.write(Wave_City_Club_structure)
    
    return Wave_City_Club_structure

# Fetch Activity Data
async def Get_Activity():
    record_limit = 1000
    headers = {
        'Cookie': f'ASessionID={st.session_state.sessionid}',
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    
    all_structure_activity_data = []
    
    async with aiohttp.ClientSession() as session:
        start_record = 1
        st.write("Fetching Activity data for Wave_City_Club Structure...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanActivities/?projectId={st.session_state.workspaceid}&planId={st.session_state.Wave_City_Club_structure}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more Structure Activity data available (204)")
                    break
                if 'activityList' in data and data['activityList']:
                    all_structure_activity_data.extend(data['activityList'])
                else:
                    all_structure_activity_data.extend(data if isinstance(data, list) else [])
                st.write(f"Fetched {len(all_structure_activity_data[-record_limit:])} Structure Activity records (Total: {len(all_structure_activity_data)})")
                if len(all_structure_activity_data[-record_limit:]) < record_limit:
                    break
                start_record += record_limit
            except Exception as e:
                st.error(f"❌ Error fetching Structure Activity data: {str(e)}")
                break
 
    structure_activity_data = pd.DataFrame(all_structure_activity_data)[['activityName', 'activitySeq', 'formTypeId']]

    st.write("Wave_City_Club STRUCTURE ACTIVITY DATA (activityName and activitySeq)")
    st.write(f"Total records: {len(structure_activity_data)}")
    st.write(structure_activity_data)
      
    return structure_activity_data

# Fetch Location/Module Data
async def Get_Location():
    record_limit = 1000
    headers = {
        'Cookie': f'ASessionID={st.session_state.sessionid}',
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    
    all_structure_location_data = []
    
    async with aiohttp.ClientSession() as session:
        start_record = 1
        total_records_fetched = 0
        st.write("Fetching Wave_City_Club Structure Location/Module data...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanLocation/?projectId={st.session_state.workspaceid}&planId={st.session_state.Wave_City_Club_structure}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more Structure Location data available (204)")
                    break
                if isinstance(data, list):
                    location_data = [{'qiLocationId': item.get('qiLocationId', ''), 'qiParentId': item.get('qiParentId', ''), 'name': item.get('name', '')} 
                                   for item in data if isinstance(item, dict)]
                    all_structure_location_data.extend(location_data)
                    total_records_fetched = len(all_structure_location_data)
                    st.write(f"Fetched {len(location_data)} Structure Location records (Total: {total_records_fetched})")
                elif isinstance(data, dict) and 'locationList' in data and data['locationList']:
                    location_data = [{'qiLocationId': loc.get('qiLocationId', ''), 'qiParentId': loc.get('qiParentId', ''), 'name': loc.get('name', '')} 
                                   for loc in data['locationList']]
                    all_structure_location_data.extend(location_data)
                    total_records_fetched = len(all_structure_location_data)
                    st.write(f"Fetched {len(location_data)} Structure Location records (Total: {total_records_fetched})")
                else:
                    st.warning(f"No 'locationList' in Structure Location data or empty list.")
                    break
                if len(location_data) < record_limit:
                    break
                start_record += record_limit
            except Exception as e:
                st.error(f"❌ Error fetching Structure Location data: {str(e)}")
                break
        
    structure_df = pd.DataFrame(all_structure_location_data)
    
    if 'name' in structure_df.columns and structure_df['name'].isna().all():
        st.error("❌ All 'name' values in Structure Location data are missing or empty!")

    st.write("Wave_City_Club STRUCTURE LOCATION/MODULE DATA")
    st.write(f"Total records: {len(structure_df)}")
    st.write(structure_df)
    
    st.session_state.structure_location_data = structure_df
    
    return structure_df

# Process individual chunk
def process_chunk(chunk, chunk_idx, dataset_name, location_df):
    logger.info(f"Starting thread for {dataset_name} Chunk {chunk_idx + 1}")
    generated_text = format_chunk_locally(chunk, chunk_idx, len(chunk), dataset_name, location_df)
    logger.info(f"Completed thread for {dataset_name} Chunk {chunk_idx + 1}")
    return generated_text, chunk_idx

# Process data with manual counting
def process_manually(analysis_df, total, dataset_name, chunk_size=1000, max_workers=4):
    if analysis_df.empty:
        st.warning(f"No completed activities found for {dataset_name}.")
        return "No completed activities found."

    unique_activities = analysis_df['activityName'].unique()
    logger.info(f"Unique activities in {dataset_name} dataset: {list(unique_activities)}")
    logger.info(f"Total records in {dataset_name} dataset: {len(analysis_df)}")

    st.write(f"Saved Wave_City_Club {dataset_name} data to Wave_City_Club_{dataset_name.lower()}_data.json")
    chunks = [analysis_df[i:i + chunk_size] for i in range(0, len(analysis_df), chunk_size)]

    location_df = st.session_state.structure_location_data

    chunk_results = {}
    progress_bar = st.progress(0)
    status_text = st.empty()

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_chunk = {
            executor.submit(process_chunk, chunk, idx, dataset_name, location_df): idx 
            for idx, chunk in enumerate(chunks)
        }

        completed_chunks = 0
        for future in as_completed(future_to_chunk):
            chunk_idx = future_to_chunk[future]
            try:
                generated_text, idx = future.result()
                chunk_results[idx] = generated_text
                completed_chunks += 1
                progress_percent = completed_chunks / len(chunks)
                progress_bar.progress(progress_percent)
                status_text.text(f"Processed chunk {completed_chunks} of {len(chunks)} ({progress_percent:.1%} complete)")
            except Exception as e:
                logger.error(f"Error processing chunk {chunk_idx + 1} for {dataset_name}: {str(e)}")
                st.error(f"❌ Error processing chunk {chunk_idx + 1}: {str(e)}")

    parsed_data = {}
    for chunk_idx in sorted(chunk_results.keys()):
        generated_text = chunk_results[chunk_idx]
        if not generated_text:
            continue

        current_tower = None
        tower_activities = []
        lines = generated_text.split("\n")
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            if line.startswith("Tower:"):
                try:
                    tower_parts = line.split("Tower:", 1)
                    if len(tower_parts) > 1:
                        if current_tower and tower_activities:
                            if current_tower not in parsed_data:
                                parsed_data[current_tower] = []
                            parsed_data[current_tower].extend(tower_activities)
                            tower_activities = []
                        current_tower = tower_parts[1].strip()
                except Exception as e:
                    logger.warning(f"Error parsing Tower line: {line}, error: {str(e)}")
                    if not current_tower:
                        current_tower = f"Unknown Tower {chunk_idx}"
                    
            elif line.startswith("Total Completed Activities:"):
                continue
            elif not line.strip().startswith("activityName"):
                try:
                    parts = re.split(r'\s{2,}', line.strip())
                    if len(parts) >= 2:
                        activity_name = ' '.join(parts[:-1]).strip()
                        count_str = parts[-1].strip()
                        count_match = re.search(r'\d+', count_str)
                        if count_match:
                            count = int(count_match.group())
                            if current_tower:
                                tower_activities.append({
                                    "activityName": activity_name,
                                    "completedCount": count
                                })
                    else:
                        match = re.match(r'^\s*(.+?)\s+(\d+)$', line)
                        if match and current_tower:
                            activity_name = match.group(1).strip()
                            count = int(match.group(2).strip())
                            tower_activities.append({
                                "activityName": activity_name,
                                "completedCount": count
                            })
                except (ValueError, IndexError) as e:
                    logger.warning(f"Skipping malformed activity line: {line}, error: {str(e)}")

        if current_tower and tower_activities:
            if current_tower not in parsed_data:
                parsed_data[current_tower] = []
            parsed_data[current_tower].extend(tower_activities)

    aggregated_data = {}
    for tower_name, activities in parsed_data.items():
        tower_short_name = tower_name.split('/')[1] if '/' in tower_name else tower_name
        if tower_short_name not in aggregated_data:
            aggregated_data[tower_short_name] = {}
        
        for activity in activities:
            name = activity.get("activityName", "Unknown")
            count = activity.get("completedCount", 0)
            if name in aggregated_data[tower_short_name]:
                aggregated_data[tower_short_name][name] += count
            else:
                aggregated_data[tower_short_name][name] = count

    combined_output_lines = []
    sorted_towers = sorted(aggregated_data.keys())
    
    for i, tower_short_name in enumerate(sorted_towers):
        combined_output_lines.append(f"{tower_short_name:<11} activityName            CompletedCount")
        activity_dict = aggregated_data[tower_short_name]
        tower_total = 0
        for name, count in sorted(activity_dict.items()):
            combined_output_lines.append(f"{'':<11} {name:<23} {count:>14}")
            tower_total += count
        combined_output_lines.append(f"{'':<11} Total for {tower_short_name:<11}: {tower_total:>14}")
        if i < len(sorted_towers) - 1:
            combined_output_lines.append("")
    
    combined_output = "\n".join(combined_output_lines)
    return combined_output

# Local formatting function for manual counting
def format_chunk_locally(chunk, chunk_idx, chunk_size, dataset_name, location_df):
    towers_data = {}
    
    for _, row in chunk.iterrows():
        tower_name = row['tower_name']
        activity_name = row['activityName']
        count = int(row['CompletedCount'])
        
        if tower_name not in towers_data:
            towers_data[tower_name] = []
            
        towers_data[tower_name].append({
            "activityName": activity_name,
            "completedCount": count
        })
    
    output = ""
    total_activities = 0
    
    for tower_name, activities in sorted(towers_data.items()):
        output += f"Tower: {tower_name}\n"
        output += "activityName            CompletedCount\n"
        activity_dict = {}
        for activity in activities:
            name = activity['activityName']
            count = activity['completedCount']
            activity_dict[name] = activity_dict.get(name, 0) + count
        for name, count in sorted(activity_dict.items()):
            output += f"{name:<30} {count}\n"
            total_activities += count
    
    output += f"Total Completed Activities: {total_activities}"
    return output

def process_data(df, activity_df, location_df, dataset_name, stage_name=None):
    """
    Modified process_data function that optionally filters by structural stage
    
    Args:
        df: The structure data (with statusName, qiLocationId, activitySeq)
        activity_df: Activity data (with activityName, activitySeq)
        location_df: Location data (with qiLocationId, name, qiParentId)
        dataset_name: Name of the dataset (e.g., "Structure")
        stage_name: Optional - Name of the structural stage (e.g., "Footing", "Plinth Beam")
    
    Returns:
        Tuple of (analysis DataFrame, total count)
    """
    completed = df[df['statusName'] == 'Completed']
    if completed.empty:
        logger.warning(f"No completed activities found in {dataset_name} data" + 
                      (f" for stage {stage_name}." if stage_name else "."))
        return pd.DataFrame(), 0

    completed = completed.merge(location_df[['qiLocationId', 'name']], on='qiLocationId', how='left')
    completed = completed.merge(activity_df[['activitySeq', 'activityName']], on='activitySeq', how='left')

    if 'qiActivityId' not in completed.columns:
        completed['qiActivityId'] = completed['qiLocationId'].astype(str) + '$$' + completed['activitySeq'].astype(str)

    if completed['name'].isna().all():
        logger.error(f"All 'name' values are missing in {dataset_name} data after merge!")
        st.error(f"❌ All 'name' values are missing in {dataset_name} data after merge! Check location data.")
        completed['name'] = 'Unknown'
    else:
        completed['name'] = completed['name'].fillna('Unknown')

    completed['activityName'] = completed['activityName'].fillna('Unknown')

    parent_child_dict = dict(zip(location_df['qiLocationId'], location_df['qiParentId']))
    name_dict = dict(zip(location_df['qiLocationId'], location_df['name']))

    def get_full_path(location_id):
        path = []
        current_id = location_id
        max_depth = 10
        depth = 0
        
        while current_id and depth < max_depth:
            if current_id not in parent_child_dict or current_id not in name_dict:
                logger.warning(f"Location ID {current_id} not found in parent_child_dict or name_dict. Path so far: {path}")
                break
            
            parent_id = parent_child_dict.get(current_id)
            name = name_dict.get(current_id, "Unknown")
            
            if not parent_id:
                if name != "Quality":
                    path.append(name)
                    path.append("Quality")
                else:
                    path.append(name)
                break
            
            path.append(name)
            current_id = parent_id
            depth += 1
        
        if depth >= max_depth:
            logger.warning(f"Max depth reached while computing path for location_id {location_id}. Possible circular reference. Path: {path}")
        
        if not path:
            logger.warning(f"No path constructed for location_id {location_id}. Using 'Unknown'.")
            return "Unknown"
        
        full_path = '/'.join(reversed(path))
        logger.debug(f"Full path for location_id {location_id}: {full_path}")
        return full_path

    completed['full_path'] = completed['qiLocationId'].apply(get_full_path)

    # ============================================================================
    # CRITICAL: STAGE FILTERING - Apply stage filter based on full_path if stage_name is provided
    # ============================================================================
    if stage_name:
        def matches_stage(full_path, stage):
            """Check if path contains stage keywords and doesn't contain exclusion keywords"""
            if pd.isna(full_path):
                return False
            path_lower = str(full_path).lower()
            
            # Get inclusion keywords
            keywords = STRUCTURAL_STAGES.get(stage, [])
            if not keywords:
                return False
            
            # Check if path contains any inclusion keyword
            has_inclusion = any(keyword in path_lower for keyword in keywords)
            if not has_inclusion:
                return False
            
            # Check for exclusions
            exclusions = STAGE_EXCLUSIONS.get(stage, [])
            if exclusions:
                # If path contains any exclusion keyword, reject it
                has_exclusion = any(exclusion in path_lower for exclusion in exclusions)
                if has_exclusion:
                    return False
            
            return True
        
        # Filter by stage BEFORE further processing
        logger.info(f"Before stage filter ({stage_name}): {len(completed)} records")
        completed = completed[completed['full_path'].apply(lambda x: matches_stage(x, stage_name))]
        logger.info(f"After stage filter ({stage_name}): {len(completed)} records")
        
        if completed.empty:
            logger.warning(f"No completed activities found for stage {stage_name} in {dataset_name} data after stage filtering.")
            st.warning(f"No completed activities found for stage {stage_name} in {dataset_name} data.")
            return pd.DataFrame(), 0
        
        # Log sample paths after filtering
        logger.info(f"Sample paths after stage filtering for {stage_name}: {completed['full_path'].head(10).tolist()}")
    # ============================================================================

    # Filter by structural elements
    def has_structural_element(full_path):
        """Check if path contains structural work elements"""
        structural_keywords = [
            'footing', 'plinth beam', 'slab', 'shear wall', 'column', 
            'beam', 'roof', 'floor', 'staircase', 'lift', 'water tank', 'terrace'
        ]
        path_lower = full_path.lower()
        return any(keyword in path_lower for keyword in structural_keywords)
    
    logger.info(f"Sample paths before structural filtering: {completed['full_path'].head(10).tolist()}")
    completed = completed[completed['full_path'].apply(has_structural_element)]
    
    if completed.empty:
        logger.warning(f"No completed activities with structural elements found in {dataset_name} data" + 
                      (f" for stage {stage_name}" if stage_name else "") + " after filtering.")
        st.warning(f"No completed activities with structural elements found" + 
                  (f" for stage {stage_name}" if stage_name else "") + f" in {dataset_name} data.")
        return pd.DataFrame(), 0
    
    logger.info(f"After structural element filtering" + 
               (f" for {stage_name}" if stage_name else "") + f": {len(completed)} records remain")
    logger.info(f"Sample paths after filtering: {completed['full_path'].head(10).tolist()}")

    def get_tower_name(full_path):
        """Extract block name from the path"""
        parts = full_path.split('/')
        
        if len(parts) < 3:
            logger.warning(f"Unexpected path format (less than 3 parts): {full_path}")
            # Fallback: try to find a part that looks like a block
            for part in parts:
                if 'block' in part.lower():
                    return part.strip()
            return parts[1].strip() if len(parts) > 1 else "Unknown"
        
        # Typically: Quality/Wave city club structure/08. Block 08 (B8) Squash Court/...
        # We want parts[2] which is the block name
        block_part = parts[2].strip()
        logger.info(f"Extracting tower from path: {full_path} -> block_part: {block_part}")
        
        return block_part

    completed['tower_name'] = completed['full_path'].apply(get_tower_name)
    
    unique_towers = completed['tower_name'].unique()
    logger.info(f"Unique tower names found in {dataset_name}" + 
               (f" for stage {stage_name}" if stage_name else "") + f": {list(unique_towers)}")
    st.write(f"**Unique tower names found in {dataset_name}" + 
            (f" for stage {stage_name}" if stage_name else "") + ":**")
    st.write(list(unique_towers))

    # Count by tower_name and activityName
    analysis = completed.groupby(['tower_name', 'activityName']).size().reset_index(name='CompletedCount')
    analysis = analysis.sort_values(by=['tower_name', 'activityName'], ascending=True)
    total_completed = analysis['CompletedCount'].sum()

    logger.info(f"Total completed activities for {dataset_name}" + 
               (f" stage {stage_name}" if stage_name else "") + f" after processing: {total_completed}")
    st.write(f"**Activity counts for {dataset_name}" + 
            (f" - {stage_name}" if stage_name else "") + ":**")
    st.write(analysis)
    
    return analysis, total_completed


# Main analysis function for Wave City Club Structure
def AnalyzeStatusManually(email=None, password=None):
    """
    Modified analysis function that processes data for each structural stage separately
    """
    start_time = time.time()

    if 'sessionid' not in st.session_state:
        st.error("❌ Please log in first!")
        return

    required_data = [
        'eden_structure',
        'structure_activity_data',
        'structure_location_data'
    ]
    
    for data_key in required_data:
        if data_key not in st.session_state:
            st.error(f"❌ Please fetch required data first! Missing: {data_key}")
            return
        if not isinstance(st.session_state[data_key], pd.DataFrame):
            st.error(f"❌ {data_key} is not a DataFrame! Found type: {type(st.session_state[data_key])}")
            return

    structure_data = st.session_state.eden_structure
    structure_activity = st.session_state.structure_activity_data
    structure_locations = st.session_state.structure_location_data
    
    # Validate required columns
    for df, name in [(structure_data, "Structure")]:
        if 'statusName' not in df.columns:
            st.error(f"❌ statusName column not found in {name} data!")
            return
        if 'qiLocationId' not in df.columns:
            st.error(f"❌ qiLocationId column not found in {name} data!")
            return
        if 'activitySeq' not in df.columns:
            st.error(f"❌ activitySeq column not found in {name} data!")
            return

    for df, name in [(structure_locations, "Structure Location")]:
        if 'qiLocationId' not in df.columns or 'name' not in df.columns:
            st.error(f"❌ qiLocationId or name column not found in {name} data!")
            return

    for df, name in [(structure_activity, "Structure Activity")]:
        if 'activitySeq' not in df.columns or 'activityName' not in df.columns:
            st.error(f"❌ activitySeq or activityName column not found in {name} data!")
            return

    # Initialize storage for stage-wise analysis
    st.session_state.stage_analysis = {}
    st.session_state.stage_totals = {}

    # Process each stage separately
    st.write("### Processing Data by Structural Stages:")
    
    for stage_name in STRUCTURAL_STAGES.keys():
        st.write(f"\n#### Processing Stage: {stage_name}")
        st.write("="*80)
        
        # Process the structure data for this specific stage
        stage_analysis, stage_total = process_data(
            structure_data, 
            structure_activity, 
            structure_locations, 
            "Structure", 
            stage_name  # Pass stage_name to enable stage filtering
        )
        
        # Store the results
        st.session_state.stage_analysis[stage_name] = stage_analysis
        st.session_state.stage_totals[stage_name] = stage_total
        
        st.write(f"**Stage {stage_name} - Full Output:**")
        if not stage_analysis.empty:
            stage_output = process_manually(stage_analysis, stage_total, f"Structure-{stage_name}")
            if stage_output:
                st.text(stage_output)
        else:
            st.warning(f"No data found for stage: {stage_name}")
        
        st.write("="*80)

    # Store the original analysis for backward compatibility (use Footing as default)
    st.session_state.structure_analysis = st.session_state.stage_analysis.get('Footing', pd.DataFrame())
    st.session_state.structure_total = st.session_state.stage_totals.get('Footing', 0)

    end_time = time.time()
    st.write(f"\n### Total execution time: {end_time - start_time:.2f} seconds")
    st.success("✅ Stage-wise analysis completed successfully!")

def get_cos_files():
    try:
        # Initialize COS client (assuming initialize_cos_client is defined elsewhere)
        cos_client = initialize_cos_client()
        if not cos_client:
            st.error("❌ Failed to initialize COS client.")
            return None

        # Update prefix to look for files in the Wave City Club folder
        st.write(f"Attempting to list objects in bucket '{COS_BUCKET}' with prefix 'Wave City Club/'")
        response = cos_client.list_objects_v2(Bucket=COS_BUCKET, Prefix="Wave City Club/")
        if 'Contents' not in response:
            st.error(f"❌ No files found in the 'Wave City Club' folder of bucket '{COS_BUCKET}'.")
            logger.error("No objects found in Wave City Club folder")
            return None

        all_files = [obj['Key'] for obj in response.get('Contents', [])]
        st.write("**All files in Wave City Club folder:**")
        if all_files:
            st.write("\n".join(all_files))
        else:
            st.write("No files found.")
            logger.warning("Wave City Club folder is empty")
            return None

        # Update the regex pattern to match the new file name format
        pattern = re.compile(
            r"Wave City Club/Structure\s*Work\s*Tracker\s*Wave\s*City\s*Club\s*all\s*Block[\(\s]*(.*?)(?:[\)\s]*\.xlsx)$",
            re.IGNORECASE
        )
        
        # Supported date formats for parsing
        date_formats = ["%d-%m-%Y", "%d-%m-%y", "%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y"]

        file_info = []
        for obj in response.get('Contents', []):
            key = obj['Key']
            match = pattern.match(key)
            if match:
                date_str = match.group(1).strip('()').strip()
                parsed_date = None
                for fmt in date_formats:
                    try:
                        parsed_date = datetime.strptime(date_str, fmt)
                        break
                    except ValueError:
                        continue
                if parsed_date:
                    file_info.append({'key': key, 'date': parsed_date})
                else:
                    logger.warning(f"Could not parse date in filename: {key}")
                    st.warning(f"Skipping file with unparseable date: {key}")
            else:
                st.write(f"File '{key}' does not match the expected pattern 'Wave City Club/Structure Work Tracker Wave City Club all Block (DD-MM-YYYY).xlsx'")

        if not file_info:
            st.error("❌ No Excel files matched the expected pattern in the 'Wave City Club' folder.")
            logger.error("No files matched the expected pattern")
            return None

        # Find the latest file based on the parsed date
        latest_file = max(file_info, key=lambda x: x['date']) if file_info else None
        if not latest_file:
            st.error("❌ No valid Excel files found for Structure Work Tracker.")
            logger.error("No valid files after date parsing")
            return None

        file_key = latest_file['key']
        st.success(f"Found matching file: {file_key}")
        return file_key
    except Exception as e:
        st.error(f"❌ Error fetching COS files: {str(e)}")
        logger.error(f"Error fetching COS files: {str(e)}")
        return None

if 'cos_df_B1' not in st.session_state:
    st.session_state.cos_df_B1 = None  # For B1 Banket Hall & Finedine
if 'cos_df_B5' not in st.session_state:
    st.session_state.cos_df_B5 = None
if 'cos_df_B6' not in st.session_state:
    st.session_state.cos_df_B6 = None
if 'cos_df_B7' not in st.session_state:
    st.session_state.cos_df_B7 = None
if 'cos_df_B9' not in st.session_state:
    st.session_state.cos_df_B9 = None
if 'cos_df_B8' not in st.session_state:
    st.session_state.cos_df_B8 = None
if 'cos_df_B2_B3' not in st.session_state:
    st.session_state.cos_df_B2_B3 = None  # For B2 & B3
if 'cos_df_B4' not in st.session_state:
    st.session_state.cos_df_B4 = None
if 'cos_df_B11' not in st.session_state:
    st.session_state.cos_df_B11 = None
if 'cos_df_B10' not in st.session_state:
    st.session_state.cos_df_B10 = None

if 'cos_tname_B1' not in st.session_state:
    st.session_state.cos_tname_B1 = None  # For B1 Banket Hall & Finedine
if 'cos_tname_B5' not in st.session_state:
    st.session_state.cos_tname_B5 = None
if 'cos_tname_B6' not in st.session_state:
    st.session_state.cos_tname_B6 = None
if 'cos_tname_B7' not in st.session_state:
    st.session_state.cos_tname_B7 = None
if 'cos_tname_B9' not in st.session_state:
    st.session_state.cos_tname_B9 = None
if 'cos_tname_B8' not in st.session_state:
    st.session_state.cos_tname_B8 = None
if 'cos_tname_B2_B3' not in st.session_state:
    st.session_state.cos_tname_B2_B3 = None  # For B2 & B3
if 'cos_tname_B4' not in st.session_state:
    st.session_state.cos_tname_B4 = None
if 'cos_tname_B11' not in st.session_state:
    st.session_state.cos_tname_B11 = None
if 'cos_tname_B10' not in st.session_state:
    st.session_state.cos_tname_B10 = None

if 'ai_response' not in st.session_state:
    st.session_state.ai_response = {}  # Initialize as empty dictionary

# Process Excel files for Wave City Club blocks with updated sheet names and expected_columns
def process_file(file_stream, filename):
    """
    Process COS Excel file and extract activity counts based on Actual Finish dates.
    Column G = Activity Name (index 6)
    Column L = Actual Finish Date (index 11)
    """
    try:
        workbook = openpyxl.load_workbook(file_stream)
        available_sheets = workbook.sheetnames
        st.write(f"Available sheets in {filename}: {', '.join(available_sheets)}")

        target_sheets = [
            "B1 Banket Hall & Finedine ",
            "B5", "B6", "B7", "B9", "B8", 
            "B2 & B3",
            "B4", "B11", "B10"
        ]
        
        results = []

        for sheet_name in target_sheets:
            if sheet_name not in available_sheets:
                st.warning(f"Sheet '{sheet_name}' not found in file: {filename}")
                continue

            file_stream.seek(0)

            try:
                # Read the sheet starting from row 2 (header at row 1)
                df = pd.read_excel(file_stream, sheet_name=sheet_name, header=1)
                st.write(f"\n📋 Processing sheet: {sheet_name}")
                
                # Trim column names
                df.columns = [col.strip() if isinstance(col, str) else col for col in df.columns]
                
                # Get Column G (Activity Name - index 6) and Column L (Actual Finish - index 11)
                if len(df.columns) >= 12:
                    # Create a clean dataframe with just what we need
                    clean_df = pd.DataFrame({
                        'Activity Name': df.iloc[:, 6],  # Column G
                        'Actual Finish': df.iloc[:, 11]  # Column L
                    })
                    
                    # Remove rows where Activity Name is empty
                    clean_df = clean_df.dropna(subset=['Activity Name'])
                    clean_df = clean_df[clean_df['Activity Name'].astype(str).str.strip() != '']
                    
                    # Convert Actual Finish to datetime
                    clean_df['Actual Finish'] = pd.to_datetime(clean_df['Actual Finish'], errors='coerce')
                    
                    # Filter only rows with valid Actual Finish dates
                    clean_df = clean_df[clean_df['Actual Finish'].notna()]
                    
                    st.write(f"✅ Filtered data for {sheet_name}: **{len(clean_df)} rows** with Actual Finish dates")
                    if len(clean_df) > 0:
                        st.write("Sample activities:")
                        st.write(clean_df.head(5))
                    
                    results.append((clean_df, sheet_name))
                else:
                    st.error(f"❌ Sheet {sheet_name} has insufficient columns: {len(df.columns)}")
                    continue

            except Exception as e:
                st.error(f"❌ Error processing sheet {sheet_name}: {str(e)}")
                continue

        if not results:
            st.error(f"❌ No valid sheets processed from file: {filename}")
            return [(None, None)]

        return results

    except Exception as e:
        st.error(f"❌ Error loading Excel file: {str(e)}")
        return [(None, None)]

def count_activities_by_foundation_concreting(df, sheet_name, stage_name=None):
    """
    Count activities based on stage-specific key activity logic:
    - For each stage, look for a specific key activity in the tracker
    - Check if that activity has a date in the Actual Finish column (column L, index 11)
    - If found, count it and apply the same count to all related Civil Works activities
    
    Stage-to-Activity Mapping:
    - Footing: Foundation Concreting
    - Plinth Beam: Plinth Beam Concreting
    - Shear Wall and Column: GF Column Casting
    - 1st Floor Slab: GF Roof Slab Casting
    - 1st Floor Shear Wall and Column: FF Column Casting
    - 2nd Floor Roof Slab: FF Roof Slab Casting
    - Terrace Work: Terrace Work
    """
    if df is None or df.empty:
        logger.warning(f"No data for {sheet_name}")
        return {}
    
    st.write(f"\n{'='*60}")
    st.write(f"**🔍 Processing: {sheet_name}**")
    if stage_name:
        st.write(f"**📍 Stage: {stage_name}**")
    st.write(f"{'='*60}")
    
    # Map stages to their key activities to look for in tracker
    stage_to_key_activity = {
        "Footing": ["foundation concreting"],
        "Plinth Beam": ["plinth beam concreting", "plinth concreting"],
        "Shear Wall and Column": ["gf column casting", "ground floor column casting", "column casting"],
        "1st Floor Slab": ["gf roof slab casting", "ground floor roof slab casting", "gf slab casting"],
        "1st Floor Shear Wall and Column": ["ff column casting", "first floor column casting", "1st floor column casting"],
        "2nd Floor Roof Slab": ["ff roof slab casting", "first floor roof slab casting", "2nd floor roof slab casting", "sf roof slab casting"],
        "Terrace Work": ["terrace work", "terrace"]
    }
    
    # Keywords to identify foundation concreting and related activities (default fallback)
    foundation_keywords = ['foundation' ,'Foundation']
    
    # Target activities to count
    target_activities = {
        'Concreting': ['concreting', 'concrete'],
        'Shuttering': ['shuttering', 'formwork', 'shutter'],
        'Reinforcement': ['reinforcement', 'rebar', 'steel'],
        'De-Shuttering': ['de-shuttering', 'deshuttering', 'de shuttering', 'removal']
    }
    
    activity_counts = {
        'Concreting': 0,
        'Shuttering': 0,
        'Reinforcement': 0,
        'De-Shuttering': 0,
        'Slab conduting': 0  # Added Slab conduting
    }
    
    # Determine which key activities to look for based on stage
    key_activity_keywords = []
    if stage_name and stage_name in stage_to_key_activity:
        key_activity_keywords = stage_to_key_activity[stage_name]
        st.write(f"🎯 **Looking for key activities:** {', '.join(key_activity_keywords)}")
    else:
        # Default to foundation concreting for backward compatibility
        key_activity_keywords = ['foundation concreting', 'foundation concrete']
        st.write(f"🎯 **Using default:** Foundation Concreting")
    
    # Step 1: Find key activity instances
    key_activity_count = 0
    key_activities_found = []
    
    for idx, row in df.iterrows():
        activity_name = str(row['Activity Name']).lower().strip()
        
        # Check if this matches any of the key activity keywords
        for keyword in key_activity_keywords:
            if keyword.lower() in activity_name:
                key_activity_count += 1
                key_activities_found.append(row['Activity Name'])
                break
    
    if key_activities_found:
        st.write(f"✅ **Found {key_activity_count} key activity instances:**")
        for act in key_activities_found[:5]:  # Show first 5
            st.write(f"   • {act}")
        if len(key_activities_found) > 5:
            st.write(f"   ... and {len(key_activities_found) - 5} more")
    
    # Step 2: If key activity found, apply count to all related activities
    if key_activity_count > 0:
        for activity in target_activities.keys():
            activity_counts[activity] = key_activity_count
        
        # IMPORTANT: Set Slab conduting equal to Concreting
        activity_counts['Slab conduting'] = key_activity_count
        
        st.write(f"\n📊 **Applied count {key_activity_count} to all Civil Works activities**")
        st.write(f"   • Slab conduting set to match Concreting: {key_activity_count}")
    else:
        # If no key activity found, count each activity individually
        st.write(f"\n⚠️ **No key activity found. Counting activities individually...**")
        
        for activity_key, keywords in target_activities.items():
            count = 0
            found_activities = []
            for idx, row in df.iterrows():
                activity_name = str(row['Activity Name']).lower().strip()
                if any(keyword in activity_name for keyword in keywords):
                    count += 1
                    found_activities.append(row['Activity Name'])
            
            activity_counts[activity_key] = count
            if found_activities:
                st.write(f"   {activity_key}: {count} activities found")
        
        # IMPORTANT: Set Slab conduting equal to Concreting count
        activity_counts['Slab conduting'] = activity_counts['Concreting']
        st.write(f"   • Slab conduting set to match Concreting: {activity_counts['Concreting']}")
    
    # Step 3: Display final counts
    st.write(f"\n**📈 Final Activity Counts for {sheet_name}:**")
    for activity, count in activity_counts.items():
        st.write(f"   • {activity}: **{count}**")
    
    st.write(f"{'='*60}\n")
    return activity_counts



# Function to handle activity count display
def display_activity_count():
    """
    Updated version that uses COS tracker data (Column G and L) to count activities.
    Uses Foundation Concreting as the base count for related Civil Works activities.
    """
    if 'file_key' not in st.session_state or not st.session_state.file_key:
        st.error("❌ No COS file found. Please fetch COS data first.")
        return
    
    try:
        # Initialize COS client
        cos_client = initialize_cos_client()
        if not cos_client:
            st.error("❌ Failed to initialize COS client")
            return
        
        # Fetch the file
        file_key = st.session_state.file_key
        st.write(f"📂 Fetching file: **{file_key}**")
        
        response = cos_client.get_object(Bucket=COS_BUCKET, Key=file_key)
        file_bytes = io.BytesIO(response['Body'].read())
        
        # Process the file with new logic
        st.write("📄 Processing COS file with new activity counting logic...")
        results = process_file(file_bytes, file_key)
        
        # Activity categories
        categories = {
            "Civil Works": ["Concreting", "Shuttering", "Reinforcement", "De-Shuttering"],
            "MEP Works": ["Plumbing Works", "Slab conduting", "Wall Conduiting", "Wiring & Switch Socket"],
            "Interior Finishing Works": ["Floor Tiling", "POP & Gypsum Plaster", "Wall Tiling", "Waterproofing – Sunken"]
        }
        
        # Initialize ai_response dictionary
        if 'ai_response' not in st.session_state or not isinstance(st.session_state.ai_response, dict):
            st.session_state.ai_response = {}
            logger.info("Initialized ai_response in display_activity_count")
        
        # Process each block
        all_activity_counts = {}
        
        for df, sheet_name in results:
            if df is not None and not df.empty:
                # Get activity counts for this sheet using foundation concreting logic
                activity_counts = count_activities_by_foundation_concreting(df, sheet_name)
                
                # Store with clean block name
                clean_name = sheet_name.strip()
                if clean_name == "B1 Banket Hall & Finedine ":
                    clean_name = "B1 Banket Hall & Finedine"
                
                all_activity_counts[clean_name] = activity_counts
        
        if not all_activity_counts:
            st.error("❌ No activity counts found from COS tracker")
            return
        
        # Display results
        st.write("## 📊 Activity Counts by Block (Based on Actual Finish Dates)")
        
        # Process each block and create AI response structure
        for block_name, activity_counts in sorted(all_activity_counts.items()):
            st.write(f"### 🏢 {block_name}")
            
            # Create AI response structure for this block
            ai_data = []
            
            for category, activities in categories.items():
                category_data = {
                    "Category": category,
                    "Activities": []
                }
                
                for activity in activities:
                    # Get count from COS data if available
                    count = activity_counts.get(activity, 0)
                    
                    # IMPORTANT: If this is "Slab conduting", use Concreting's count
                    if activity == "Slab conduting":
                        count = activity_counts.get("Concreting", 0)
                    
                    category_data["Activities"].append({
                        "Activity Name": activity,
                        "Total": int(count)
                    })
                
                ai_data.append(category_data)
            
            # Store in session state
            st.session_state.ai_response[block_name] = ai_data
            logger.info(f"Stored ai_response for {block_name}: {ai_data}")
            
            # Display as table
            display_data = []
            for category_data in ai_data:
                category = category_data["Category"]
                for activity in category_data["Activities"]:
                    display_data.append({
                        "Category": category,
                        "Activity Name": activity["Activity Name"],
                        "Count": activity["Total"]
                    })
            
            display_df = pd.DataFrame(display_data)
            
            # Show by category
            for category in ["Civil Works", "MEP Works", "Interior Finishing Works"]:
                category_df = display_df[display_df["Category"] == category]
                if not category_df.empty:
                    st.write(f"**{category}**")
                    st.table(category_df[["Activity Name", "Count"]])
        
        # Create consolidated summary
        st.write("### 📈 Consolidated Activity Summary Across All Blocks")
        
        category_mapping = {
            "Concreting": "Civil Works",
            "Shuttering": "Civil Works", 
            "Reinforcement": "Civil Works",
            "De-Shuttering": "Civil Works",
            "Plumbing Works": "MEP Works",
            "Slab conduting": "MEP Works",
            "Wall Conduiting": "MEP Works", 
            "Wiring & Switch Socket": "MEP Works",
            "Floor Tiling": "Interior Finishing Works",
            "POP & Gypsum Plaster": "Interior Finishing Works",
            "Wall Tiling": "Interior Finishing Works",
            "Waterproofing – Sunken": "Interior Finishing Works"
        }
        
        consolidated_summary = {}
        for block_name, ai_data in st.session_state.ai_response.items():
            for category_data in ai_data:
                for activity in category_data["Activities"]:
                    activity_name = activity["Activity Name"]
                    count = activity["Total"]
                    
                    if activity_name not in consolidated_summary:
                        consolidated_summary[activity_name] = 0
                    consolidated_summary[activity_name] += count
        
        # Display consolidated
        consolidated_data = []
        for activity_name, total_count in sorted(consolidated_summary.items()):
            category = category_mapping.get(activity_name, "Other")
            consolidated_data.append({
                "Category": category,
                "Activity Name": activity_name,
                "Total Count": total_count
            })
        
        consolidated_df = pd.DataFrame(consolidated_data)
        
        for category in ["Civil Works", "MEP Works", "Interior Finishing Works"]:
            category_df = consolidated_df[consolidated_df["Category"] == category]
            if not category_df.empty:
                st.write(f"**{category}**")
                st.table(category_df[["Activity Name", "Total Count"]])
        
        st.success("✅ Activity counts updated successfully from COS tracker!")
        
    except Exception as e:
        st.error(f"❌ Error fetching COS data: {str(e)}")
        logger.error(f"Error fetching COS data: {str(e)}")
        import traceback
        st.code(traceback.format_exc())



# Function to get access token for WatsonX API
def get_access_token(api_key):
    try:
        headers = {"Content-Type": "application/x-www-form-urlencoded"}
        data = {
            "grant_type": "urn:ibm:params:oauth:grant-type:apikey",
            "apikey": api_key
        }
        response = requests.post("https://iam.cloud.ibm.com/identity/token", headers=headers, data=data)
        if response.status_code == 200:
            return response.json().get("access_token")
        else:
            logger.error(f"Failed to get access token: {response.status_code} - {response.text}")
            return None
    except Exception as e:
        logger.error(f"Error getting access token: {str(e)}")
        return None

# WatsonX Prompt Generation
# WatsonX Prompt Generation (Updated with new categories)
def generatePrompt(json_datas):
    try:
        if isinstance(json_datas, pd.DataFrame):
            json_str = json_datas.reset_index().to_json(orient='records', indent=2)
        else:
            json_str = str(json_datas)

        body = {
            "input": f"""
            Read the table data provided below and categorize the activities into the following categories: Civil Works, MEP Works, Interior Finishing Works, and External Development Activities. Compute the total count of each activity within its respective category and return the results as a JSON array, following the example format provided. For activities like "UP-First Fix" and "CP-First Fix", combine them as "Plumbing Works". If an activity is not found in the data, include it with a count of 0. Ensure the counts are accurate, the output is grouped by category, and the JSON structure is valid with no nested or repeated keys.

            Table Data:
            {json_str}

            Categories and Activities:
            - Civil Works: Concreting, Shuttering, Reinforcement, De-Shuttering
            - MEP Works: Plumbing Works, Slab conduting, Wall Conduiting, Wiring & Switch Socket
            - Interior Finishing Works: Floor Tiling, POP & Gypsum Plaster, Wall Tiling, Waterproofing – Sunken
            - External Development Activities: Granular Sub-Base, Kerb Stone, Rain Water / Storm Line, Saucer Drain / Paver Block, Sewer Line, Stamp Concrete, Storm Line, WMM

            Example JSON format needed:
            [
              {{
                "Category": "Civil Works",
                "Activities": [
                  {{"Activity Name": "Concreting", "Total": 0}},
                  {{"Activity Name": "Shuttering", "Total": 0}},
                  {{"Activity Name": "Reinforcement", "Total": 0}},
                  {{"Activity Name": "De-Shuttering", "Total": 0}}
                ]
              }},
              {{
                "Category": "MEP Works",
                "Activities": [
                  {{"Activity Name": "Plumbing Works", "Total": 0}},
                  {{"Activity Name": "Slab conduting", "Total": 0}},
                  {{"Activity Name": "Wall Conduiting", "Total": 0}},
                  {{"Activity Name": "Wiring & Switch Socket", "Total": 0}}
                ]
              }},
              {{
                "Category": "Interior Finishing Works",
                "Activities": [
                  {{"Activity Name": "Floor Tiling", "Total": 0}},
                  {{"Activity Name": "POP & Gypsum Plaster", "Total": 0}},
                  {{"Activity Name": "Wall Tiling", "Total": 0}},
                  {{"Activity Name": "Waterproofing – Sunken", "Total": 0}}
                ]
              }},
              {{
                "Category": "External Development Activities",
                "Activities": [
                  {{"Activity Name": "Granular Sub-Base", "Total": 0}},
                  {{"Activity Name": "Kerb Stone", "Total": 0}},
                  {{"Activity Name": "Rain Water / Storm Line", "Total": 0}},
                  {{"Activity Name": "Saucer Drain / Paver Block", "Total": 0}},
                  {{"Activity Name": "Sewer Line", "Total": 0}},
                  {{"Activity Name": "Stamp Concrete", "Total": 0}},
                  {{"Activity Name": "Storm Line", "Total": 0}},
                  {{"Activity Name": "WMM", "Total": 0}}
                ]
              }}
            ]

            Return only the JSON array, no additional text, explanations, or code. Ensure the counts are accurate, activities are correctly categorized, and the JSON structure is valid.
            """,
            "parameters": {
                "decoding_method": "greedy",
                "max_new_tokens": 8100,
                "min_new_tokens": 0,
                "stop_sequences": [";"],
                "repetition_penalty": 1.05,
                "temperature": 0.5
            },
            "model_id": os.getenv("MODEL_ID_1"),
            "project_id": os.getenv("PROJECT_ID_1")
        }
        
        access_token = get_access_token(os.getenv("API_KEY_1"))
        if not access_token:
            logger.error("Failed to obtain access token for WatsonX API")
            return generate_fallback_totals(json_datas)
            
        headers = {
            "Accept": "application/json",
            "Content-Type": "application/json",
            "Authorization": f"Bearer {access_token}"
        }
        
        logger.info("Sending request to WatsonX API")
        response = requests.post(os.getenv("WATSONX_API_URL_1"), headers=headers, json=body, timeout=60)
        
        logger.info(f"WatsonX API response status: {response.status_code}")
        logger.debug(f"WatsonX API response text: {response.text[:1000]}...")  # Log first 1000 chars
        
        if response.status_code != 200:
            logger.error(f"WatsonX API call failed: {response.status_code} - {response.text}")
            st.warning(f"WatsonX API failed with status {response.status_code}: {response.text}. Using fallback method to calculate totals.")
            return generate_fallback_totals(json_datas)
            
        response_data = response.json()
        logger.debug(f"WatsonX API response data: {response_data}")
        
        if 'results' not in response_data or not response_data['results']:
            logger.error("WatsonX API response does not contain 'results' key")
            st.warning("WatsonX API response invalid. Using fallback method to calculate totals.")
            return generate_fallback_totals(json_datas)

        generated_text = response_data['results'][0].get('generated_text', '').strip()
        logger.debug(f"Generated text from WatsonX: {generated_text[:1000]}...")  # Log first 1000 chars
        
        if not generated_text:
            logger.error("WatsonX API returned empty generated text")
            st.warning("WatsonX API returned empty response. Using fallback method to calculate totals.")
            return generate_fallback_totals(json_datas)

        if not (generated_text.startswith('[') and generated_text.endswith(']')):
            start_idx = generated_text.find('[')
            end_idx = generated_text.rfind(']')
            if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
                generated_text = generated_text[start_idx:end_idx+1]
                logger.info("Extracted JSON array from response")
            else:
                logger.error(f"Could not extract valid JSON array from response: {generated_text[:1000]}...")
                return generate_fallback_totals(json_datas)
        
        try:
            parsed_json = json.loads(generated_text)
            if not all(isinstance(item, dict) and 'Category' in item and 'Activities' in item for item in parsed_json):
                logger.warning("JSON structure doesn't match expected format")
                return generate_fallback_totals(json_datas)
            logger.info("Successfully parsed WatsonX API response")
            return generated_text
        except json.JSONDecodeError as e:
            logger.error(f"WatsonX API returned invalid JSON: {e}")
            st.warning(f"WatsonX API returned invalid JSON. Error: {str(e)}. Using fallback method to calculate totals.")
            error_position = int(str(e).split('(char ')[1].split(')')[0]) if '(char ' in str(e) else 0
            context_start = max(0, error_position - 50)
            context_end = min(len(generated_text), error_position + 50)
            logger.error(f"JSON error context: ...{generated_text[context_start:error_position]}[ERROR HERE]{generated_text[error_position:context_end]}...")
            return generate_fallback_totals(json_datas)
    
    except Exception as e:
        logger.error(f"Error in WatsonX API call: {str(e)}")
        st.warning(f"Error in WatsonX API call: {str(e)}. Using fallback method to calculate totals.")
        return generate_fallback_totals(json_datas)
    
# Fallback Total Calculation
def generate_fallback_totals(count_table):
    try:
        if not isinstance(count_table, pd.DataFrame):
            logger.error("Fallback method received invalid input: not a DataFrame")
            return json.dumps([
                {"Category": "Civil Works", "Activities": [
                    {"Activity Name": "Concreting", "Total": 0},
                    {"Activity Name": "Shuttering", "Total": 0},
                    {"Activity Name": "Reinforcement", "Total": 0},
                    {"Activity Name": "De-Shuttering", "Total": 0}
                ]},
                {"Category": "MEP Works", "Activities": [
                    {"Activity Name": "Plumbing Works", "Total": 0},
                    {"Activity Name": "Slab conduting", "Total": 0},
                    {"Activity Name": "Wall Conduiting", "Total": 0},
                    {"Activity Name": "Wiring & Switch Socket", "Total": 0}
                ]},
                {"Category": "Interior Finishing Works", "Activities": [
                    {"Activity Name": "Floor Tiling", "Total": 0},
                    {"Activity Name": "POP & Gypsum Plaster", "Total": 0},
                    {"Activity Name": "Wall Tiling", "Total": 0},
                    {"Activity Name": "Waterproofing – Sunken", "Total": 0}
                ]},
                {"Category": "External Development Activities", "Activities": [
                    {"Activity Name": "Granular Sub-Base", "Total": 0},
                    {"Activity Name": "Kerb Stone", "Total": 0},
                    {"Activity Name": "Rain Water / Storm Line", "Total": 0},
                    {"Activity Name": "Saucer Drain / Paver Block", "Total": 0},
                    {"Activity Name": "Sewer Line", "Total": 0},
                    {"Activity Name": "Stamp Concrete", "Total": 0},
                    {"Activity Name": "Storm Line", "Total": 0},
                    {"Activity Name": "WMM", "Total": 0}
                ]}
            ], indent=2)

        categories = {
            "Civil Works": [
                "Concreting", "Shuttering", "Reinforcement", "De-Shuttering"
            ],
            "MEP Works": [
                "Plumbing Works", "Slab conduting", "Wall Conduiting", "Wiring & Switch Socket"
            ],
            "Interior Finishing Works": [
                "Floor Tiling", "POP & Gypsum Plaster", "Wall Tiling", "Waterproofing – Sunken"
            ],
            "External Development Activities": [
                "Granular Sub-Base", "Kerb Stone", "Rain Water / Storm Line", "Saucer Drain / Paver Block",
                "Sewer Line", "Stamp Concrete", "Storm Line", "WMM"
            ]
        }

        result = []
        for category, activities in categories.items():
            category_data = {"Category": category, "Activities": []}
            
            for activity in activities:
                if activity == "Plumbing Works":
                    combined_count = count_table.loc["UP-First Fix and CP-First Fix", "Count"] if "UP-First Fix and CP-First Fix" in count_table.index else 0
                    total = combined_count
                else:
                    total = count_table.loc[activity, "Count"] if activity in count_table.index else 0
                category_data["Activities"].append({
                    "Activity Name": activity,
                    "Total": int(total) if pd.notna(total) else 0
                })
            
            result.append(category_data)

        return json.dumps(result, indent=2)
    except Exception as e:
        logger.error(f"Error in fallback total calculation: {str(e)}")
        st.error(f"Error in fallback total calculation: {str(e)}")
        return json.dumps([
            {"Category": "Civil Works", "Activities": [
                {"Activity Name": "Concreting", "Total": 0},
                {"Activity Name": "Shuttering", "Total": 0},
                {"Activity Name": "Reinforcement", "Total": 0},
                {"Activity Name": "De-Shuttering", "Total": 0}
            ]},
            {"Category": "MEP Works", "Activities": [
                {"Activity Name": "Plumbing Works", "Total": 0},
                {"Activity Name": "Slab conduting", "Total": 0},
                {"Activity Name": "Wall Conduiting", "Total": 0},
                {"Activity Name": "Wiring & Switch Socket", "Total": 0}
            ]},
            {"Category": "Interior Finishing Works", "Activities": [
                {"Activity Name": "Floor Tiling", "Total": 0},
                {"Activity Name": "POP & Gypsum Plaster", "Total": 0},
                {"Activity Name": "Wall Tiling", "Total": 0},
                {"Activity Name": "Waterproofing – Sunken", "Total": 0}
            ]},
            {"Category": "External Development Activities", "Activities": [
                {"Activity Name": "Granular Sub-Base", "Total": 0},
                {"Activity Name": "Kerb Stone", "Total": 0},
                {"Activity Name": "Rain Water / Storm Line", "Total": 0},
                {"Activity Name": "Saucer Drain / Paver Block", "Total": 0},
                {"Activity Name": "Sewer Line", "Total": 0},
                {"Activity Name": "Stamp Concrete", "Total": 0},
                {"Activity Name": "Storm Line", "Total": 0},
                {"Activity Name": "WMM", "Total": 0}
            ]}
        ], indent=2)


# Extract Totals from AI Data
def getTotal(ai_data):
    try:
        if isinstance(ai_data, str):
            ai_data = json.loads(ai_data)
            
        if not isinstance(ai_data, list):
            logger.error(f"AI data is not a list: {ai_data}")
            return [0] * len(st.session_state.get('sheduledf', pd.DataFrame()).index)

        share = []
        for category_data in ai_data:
            if isinstance(category_data, dict) and 'Activities' in category_data:
                for activity in category_data['Activities']:
                    if isinstance(activity, dict) and 'Total' in activity:
                        total = activity['Total']
                        share.append(int(total) if isinstance(total, (int, float)) and pd.notna(total) else 0)
                    else:
                        share.append(0)
            else:
                share.append(0)
        return share
    except Exception as e:
        logger.error(f"Error parsing AI data: {str(e)}")
        st.error(f"Error parsing AI data: {str(e)}")
        return [0] * len(st.session_state.get('sheduledf', pd.DataFrame()).index)

# Function to handle activity count display
def run_analysis_and_display():
    try:
        st.write("Running stage-wise status analysis...")
        AnalyzeStatusManually()
        st.success("Stage-wise status analysis completed successfully!")

        if 'ai_response' not in st.session_state or not isinstance(st.session_state.ai_response, dict):
            st.session_state.ai_response = {}
            logger.info("Initialized st.session_state.ai_response in run_analysis_and_display")

        st.write("Displaying activity counts and generating AI data...")
        display_activity_count()
        st.success("Activity counts displayed successfully!")

        st.write("Checking AI data totals...")
        logger.info(f"st.session_state.ai_response contents: {st.session_state.ai_response}")
        if not st.session_state.ai_response:
            st.error("❌ No AI data available in st.session_state.ai_response. Attempting to regenerate.")
            logger.error("No AI data in st.session_state.ai_response after display_activity_count")
            display_activity_count()
            if not st.session_state.ai_response:
                st.error("❌ Failed to regenerate AI data. Please check data fetching and try again.")
                logger.error("Failed to regenerate AI data")
                return

        st.write("Generating consolidated checklist Excel file with stage-based sheets...")
        
        # Check if stage analysis exists
        if 'stage_analysis' not in st.session_state:
            st.error("❌ No stage analysis data available. Please ensure stage analysis ran successfully.")
            logger.error("No stage_analysis in st.session_state")
            return

        structure_analysis = st.session_state.get('structure_analysis', None)

        with st.spinner("Generating Excel file with 7 stage sheets... This may take a moment."):
            excel_file = generate_consolidated_Checklist_excel(structure_analysis, st.session_state.ai_response)
        
        if excel_file:
            timestamp = pd.Timestamp.now(tz='Asia/Kolkata').strftime('%Y%m%d_%H%M')
            file_name = f"Consolidated_Checklist_WaveCityClub_Stages_{timestamp}.xlsx"
            
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                st.sidebar.download_button(
                    label="📥 Download Stage-Based Checklist Excel",
                    data=excel_file,
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_excel_button_stages",
                    help="Click to download the consolidated checklist with 7 stage-based sheets."
                )
            st.success("Excel file with stage-based sheets generated successfully! Click the button above to download.")
        else:
            st.error("Failed to generate Excel file. Please check the logs for details.")
            logger.error("Failed to generate Excel file")

    except Exception as e:
        st.error(f"Error during analysis, display, or Excel generation: {str(e)}")
        logger.error(f"Error during analysis, display, or Excel generation: {str(e)}")
        import traceback
        st.error(traceback.format_exc())



# Combined function for Initialize and Fetch Data
async def initialize_and_fetch_data(email, password):
    with st.spinner("Starting initialization and data fetching process..."):
        # Step 1: Login
        if not email or not password:
            st.sidebar.error("Please provide both email and password!")
            return False
        try:
            st.sidebar.write("Logging in...")
            session_id = await login_to_asite(email, password)
            if not session_id:
                st.sidebar.error("Login failed!")
                return False
            st.sidebar.success("Login successful!")
        except Exception as e:
            st.sidebar.error(f"Login failed: {str(e)}")
            return False

        # Step 2: Get Workspace ID
        try:
            st.sidebar.write("Fetching Workspace ID...")
            await GetWorkspaceID()
            st.sidebar.success("Workspace ID fetched successfully!")
        except Exception as e:
            st.sidebar.error(f"Failed to fetch Workspace ID: {str(e)}")
            return False

        # Step 3: Get Project IDs
        try:
            st.sidebar.write("Fetching Project IDs...")
            await GetProjectId()
            st.sidebar.success("Project IDs fetched successfully!")
        except Exception as e:
            st.sidebar.error(f"Failed to fetch Project IDs: {str(e)}")
            return False

        # Step 4: Get All Data (Structure only)
        try:
            st.sidebar.write("Fetching All Data...")
            Edenstructure = await GetAllDatas()
            st.session_state.eden_structure = Edenstructure
            st.sidebar.success("All Data fetched successfully!")
        except Exception as e:
            st.sidebar.error(f"Failed to fetch All Data: {str(e)}")
            return False

        # Step 5: Get Activity Data
        try:
            st.sidebar.write("Fetching Activity Data...")
            structure_activity_data = await Get_Activity()
            st.session_state.structure_activity_data = structure_activity_data
            st.sidebar.success("Activity Data fetched successfully!")
        except Exception as e:
            st.sidebar.error(f"Failed to fetch Activity Data: {str(e)}")
            return False

        # Step 6: Get Location/Module Data
        try:
            st.sidebar.write("Fetching Location/Module Data...")
            structure_location_data = await Get_Location()
            st.session_state.structure_location_data = structure_location_data 
            st.sidebar.success("Location/Module Data fetched successfully!")
        except Exception as e:
            st.sidebar.error(f"Failed to fetch Location/Module Data: {str(e)}")
            return False

        # Step 7: Fetch COS Files
        try:
            st.sidebar.write("Fetching COS files from Wave City Club folder...")
            file_key = get_cos_files()
            st.session_state.file_key = file_key
            if file_key:
                st.success(f"Found 1 file in COS storage: {file_key}")
                try:
                    st.write(f"Processing file: {file_key}")
                    cos_client = initialize_cos_client()
                    if not cos_client:
                        st.error("Failed to initialize COS client during file fetch")
                        logger.error("COS client initialization failed during file fetch")
                        return False
                    st.write("Fetching file from COS...")
                    response = cos_client.get_object(Bucket=COS_BUCKET, Key=file_key)
                    file_bytes = io.BytesIO(response['Body'].read())
                    st.write("File fetched successfully. Processing sheets...")
                    results = process_file(file_bytes, file_key)
                    st.write(f"Processing results: {len(results)} sheets processed")
                    for df, sheet_name in results:
                        if df is not None:
                            if sheet_name == "B1 Banket Hall & Finedine ":
                                st.session_state.cos_df_B1 = df
                                st.session_state.cos_tname_B1 = "B1 Banket Hall & Finedine"
                                st.write(f"Processed Data for {sheet_name} - {len(df)} rows:")
                                st.write(df.head())
                            elif sheet_name == "B5":
                                st.session_state.cos_df_B5 = df
                                st.session_state.cos_tname_B5 = "B5"
                                st.write(f"Processed Data for {sheet_name} - {len(df)} rows:")
                                st.write(df.head())
                            elif sheet_name == "B6":
                                st.session_state.cos_df_B6 = df
                                st.session_state.cos_tname_B6 = "B6"
                                st.write(f"Processed Data for {sheet_name} - {len(df)} rows:")
                                st.write(df.head())
                            elif sheet_name == "B7":
                                st.session_state.cos_df_B7 = df
                                st.session_state.cos_tname_B7 = "B7"
                                st.write(f"Processed Data for {sheet_name} - {len(df)} rows:")
                                st.write(df.head())
                            elif sheet_name == "B9":
                                st.session_state.cos_df_B9 = df
                                st.session_state.cos_tname_B9 = "B9"
                                st.write(f"Processed Data for {sheet_name} - {len(df)} rows:")
                                st.write(df.head())
                            elif sheet_name == "B8":
                                st.session_state.cos_df_B8 = df
                                st.session_state.cos_tname_B8 = "B8"
                                st.write(f"Processed Data for {sheet_name} - {len(df)} rows:")
                                st.write(df.head())
                            elif sheet_name == "B2 & B3":
                                st.session_state.cos_df_B2_B3 = df
                                st.session_state.cos_tname_B2_B3 = "B2 & B3"
                                st.write(f"Processed Data for {sheet_name} - {len(df)} rows:")
                                st.write(df.head())
                            elif sheet_name == "B4":
                                st.session_state.cos_df_B4 = df
                                st.session_state.cos_tname_B4 = "B4"
                                st.write(f"Processed Data for {sheet_name} - {len(df)} rows:")
                                st.write(df.head())
                            elif sheet_name == "B11":
                                st.session_state.cos_df_B11 = df
                                st.session_state.cos_tname_B11 = "B11"
                                st.write(f"Processed Data for {sheet_name} - {len(df)} rows:")
                                st.write(df.head())
                            elif sheet_name == "B10":
                                st.session_state.cos_df_B10 = df
                                st.session_state.cos_tname_B10 = "B10"
                                st.write(f"Processed Data for {sheet_name} - {len(df)} rows:")
                                st.write(df.head())
                        else:
                            st.warning(f"No data processed for {sheet_name} in {file_key}.")
                except Exception as e:
                    st.error(f"Error loading {file_key} from cloud storage: {str(e)}")
                    logger.error(f"Error loading {file_key}: {str(e)}")
                    return False
            else:
                st.warning("No expected Excel files available in the 'Wave City Club' folder of the COS bucket.")
                return False
        except Exception as e:
            st.sidebar.error(f"Failed to fetch COS files: {str(e)}")
            logger.error(f"Failed to fetch COS files: {str(e)}")
            return False

    st.sidebar.success("All steps completed successfully!")
    return True


def generate_consolidated_Checklist_excel(structure_analysis=None, activity_counts=None):
    """
    Updated Excel generation using pre-processed stage data from session_state
    """
    try:
        if activity_counts is None:
            activity_counts = st.session_state.get('ai_response', {})
            if not activity_counts:
                st.error("❌ No activity counts data available.")
                logger.error("activity_counts is empty in generate_consolidated_Checklist_excel")
                return None
        
        # Check if stage analysis exists
        if 'stage_analysis' not in st.session_state:
            st.error("❌ Stage analysis not found. Please run AnalyzeStatusManually first!")
            logger.error("No stage_analysis in st.session_state")
            return None

        # Define categories and activities
        categories = {
            "Civil Works": ["Concreting", "Shuttering", "Reinforcement", "De-Shuttering"],
            "MEP Works": ["Plumbing Works", "Slab conduting", "Wall Conduiting", "Wiring & Switch Socket"],
            "Interior Finishing Works": ["Floor Tiling", "POP & Gypsum Plaster", "Wall Tiling", "Waterproofing – Sunken"]
        }

        cos_to_asite_mapping = {
            "Concreting": "Concreting",
            "Shuttering": "Shuttering", 
            "Reinforcement": "Reinforcement",
            "De-Shuttering": "De-Shuttering",
            "Plumbing Works": "Plumbing Works",
            "Slab conduting": "Slab conduting",
            "Wall Conduiting": "Wall Conducting",
            "Wiring & Switch Socket": "Wiring & Switch Socket",
            "Floor Tiling": "Floor Tiling",
            "POP & Gypsum Plaster": "POP & Gypsum Plaster",
            "Wall Tiling": "Wall Tile",
            "Waterproofing – Sunken": "Waterproofing - Sunken"
        }

        block_to_asite_filter = {
            "B1 Banket Hall & Finedine": [
                "01. Block (B1) Banquet Hall ",
                "02. Block (B1) Fine Dine"
            ],
            "B2 & B3": [
                "03. Block 02 (B2) Changing room ",
                "04. Block 03 (B3) GYM "
            ],
            "B4": "04. Block 4 (B4) ",
            "B5": "05. Block 05 (B5) Admin +Member Lounge +Creche+AV Room+Surveillance Room +Toilets ",
            "B6": "06. Block 06 (B6) Toilets ",
            "B7": "07. Block 07 (B7) Indoor Sports ",
            "B8": "08. Block 08 (B8) Squash Court ",
            "B9": "09. Block 09 (B9) Spa and Saloon",
            "B10": "10. Block 09 (B10) Spa and Saloon",
            "B11": "11. Block 11 (B11) "
        }

        blocks = [
            "B1 Banket Hall & Finedine", "B5", "B6", "B7", "B9", "B8",
            "B2 & B3", "B4", "B11", "B10"
        ]

        # Create Excel workbook
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output)

        header_format = workbook.add_format({'bold': True, 'bg_color': '#D3D3D3'})
        total_format = workbook.add_format({'bold': True, 'bg_color': '#FFDAB9'})
        cell_format = workbook.add_format({'border': 1})

        # Generate each stage sheet using pre-processed data
        for stage_name in STRUCTURAL_STAGES.keys():
            worksheet = workbook.add_worksheet(stage_name)
            logger.info(f"Creating sheet: {stage_name}")
            
            # Get pre-processed stage analysis
            stage_analysis = st.session_state.stage_analysis.get(stage_name, pd.DataFrame())
            
            if stage_analysis.empty:
                worksheet.write(0, 0, f"No data available for {stage_name}", header_format)
                logger.warning(f"No data for stage {stage_name}")
                continue

            # **NEW: Generate BLOCK-SPECIFIC activity counts for this stage**
            stage_activity_counts = {}
            
            # Map stages to their key activities
            stage_to_key_activity = {
                "Footing": ["foundation concreting"],
                "Plinth Beam": ["plinth beam concreting", "plinth concreting"],
                "Shear Wall and Column": ["gf column casting", "ground floor column casting"],
                "1st Floor Slab": ["gf roof slab casting", "ground floor roof slab casting"],
                "1st Floor Shear Wall and Column": ["ff column casting", "first floor column casting"],
                "2nd Floor Roof Slab": ["ff roof slab casting", "first floor roof slab casting", "2nd floor roof slab casting"],
                "Terrace Work": ["terrace", "roof", "parapet", "waterproofing", "mumty", "railing"]
            }
            
            key_activities = stage_to_key_activity.get(stage_name, ["foundation concreting"])
            logger.info(f"Stage {stage_name}: Looking for activities: {key_activities}")
            
            # Re-process tracker data for each block individually
            try:
                cos_client = initialize_cos_client()
                if cos_client and 'file_key' in st.session_state and st.session_state.file_key:
                    file_key = st.session_state.file_key
                    response = cos_client.get_object(Bucket=COS_BUCKET, Key=file_key)
                    file_bytes = io.BytesIO(response['Body'].read())
                    
                    # Process file to get tracker data for each block
                    results = process_file(file_bytes, file_key)
                    
                    for df, sheet_name in results:
                        if df is not None and not df.empty:
                            # Clean block name
                            clean_name = sheet_name.strip()
                            if clean_name == "B1 Banket Hall & Finedine ":
                                clean_name = "B1 Banket Hall & Finedine"
                            
                            # Count the key activity for THIS SPECIFIC BLOCK ONLY
                            key_activity_count = 0
                            found_activities = []
                            
                            for idx, row in df.iterrows():
                                activity_name = str(row['Activity Name']).lower().strip()
                                
                                # Check if this row matches any of the key activities for this stage
                                for key_activity in key_activities:
                                    if key_activity.lower() in activity_name:
                                        key_activity_count += 1
                                        found_activities.append(row['Activity Name'])
                                        logger.info(f"Stage {stage_name}, Block {clean_name}: Found '{row['Activity Name']}'")
                                        break
                            
                            # Store count for this block - all Civil Works activities get this count
                            block_activity_counts = {
                                'Concreting': key_activity_count,
                                'Shuttering': key_activity_count,
                                'Reinforcement': key_activity_count,
                                'De-Shuttering': key_activity_count,
                                'Slab conduting': key_activity_count
                            }
                            
                            stage_activity_counts[clean_name] = block_activity_counts
                            logger.info(f"Stage {stage_name}, Block {clean_name}: Key activity count = {key_activity_count} (found {len(found_activities)} activities)")
                
                logger.info(f"Generated block-specific activity counts for {stage_name}: {stage_activity_counts}")
            except Exception as e:
                logger.error(f"Error generating block-specific counts for {stage_name}: {str(e)}")
                # Fallback to using global activity_counts
                stage_activity_counts = activity_counts

            consolidated_rows = []

            # Process data for each block and category
            for block in blocks:
                for category, activities in categories.items():
                    for activity in activities:
                        # **NEW: Skip Slab conduting for stages other than 1st Floor Slab and 2nd Floor Roof Slab**
                        if activity == "Slab conduting" and stage_name not in ["1st Floor Slab", "2nd Floor Roof Slab"]:
                            continue
                        
                        asite_activity = cos_to_asite_mapping.get(activity, activity)
                        asite_activities = asite_activity if isinstance(asite_activity, list) else [asite_activity]

                        # Get closed_checklist from stage_analysis (already filtered by stage!)
                        closed_checklist = 0
                        asite_filters = block_to_asite_filter.get(block, block)
                        
                        if isinstance(asite_filters, list):
                            for asite_filter in asite_filters:
                                for asite_act in asite_activities:
                                    matching_rows = stage_analysis[
                                        (stage_analysis['tower_name'].str.strip() == asite_filter.strip()) &
                                        (stage_analysis['activityName'] == asite_act)
                                    ]
                                    if not matching_rows.empty:
                                        count = matching_rows['CompletedCount'].sum()
                                        closed_checklist += count
                                        logger.info(f"Sheet {stage_name}, Block {block} ('{asite_filter}'), Activity: {asite_act}, Count: {count}")
                        else:
                            for asite_act in asite_activities:
                                matching_rows = stage_analysis[
                                    (stage_analysis['tower_name'].str.strip() == asite_filters.strip()) &
                                    (stage_analysis['activityName'] == asite_act)
                                ]
                                if not matching_rows.empty:
                                    count = matching_rows['CompletedCount'].sum()
                                    closed_checklist += count
                                    logger.info(f"Sheet {stage_name}, Block {block} ('{asite_filters}'), Activity: {asite_act}, Count: {count}")

                        # Get COS data (Completed Work count) - use stage-specific counts
                        completed_flats = 0
                        
                        # **NEW: Only get COS data for Slab conduting in specific sheets**
                        if activity == "Slab conduting" and stage_name in ["1st Floor Slab", "2nd Floor Roof Slab"]:
                            if block in stage_activity_counts:
                                activity_data_dict = stage_activity_counts[block]
                                if isinstance(activity_data_dict, dict):
                                    completed_flats = activity_data_dict.get(activity, 0)
                        elif activity != "Slab conduting":
                            # For all other activities, get COS data from stage-specific counts
                            if block in stage_activity_counts:
                                activity_data_dict = stage_activity_counts[block]
                                if isinstance(activity_data_dict, dict):
                                    completed_flats = activity_data_dict.get(activity, 0)

                        # Calculate open/missing
                        in_progress = 0
                        if completed_flats == 0 or closed_checklist > completed_flats:
                            open_missing = 0
                        else:
                            open_missing = abs(completed_flats - closed_checklist)

                        display_activity = asite_activities[0]

                        consolidated_rows.append({
                            "Block": block,
                            "Category": category,
                            "Activity Name": display_activity,
                            "Completed Work*(Count of Flat)": completed_flats,
                            "In progress": in_progress,
                            "Closed checklist": closed_checklist,
                            "Open/Missing check list": open_missing
                        })
            
            # Store consolidated_rows in session state for summary calculation
            if 'all_consolidated_rows' not in st.session_state:
                st.session_state.all_consolidated_rows = {}
            st.session_state.all_consolidated_rows[stage_name] = consolidated_rows.copy()

            # Write to worksheet
            df = pd.DataFrame(consolidated_rows)
            if df.empty:
                worksheet.write(0, 0, f"No activities found for {stage_name}", header_format)
                continue

            df.sort_values(by=["Block", "Category"], inplace=True)

            headers = ["Activity Name", "Completed", "In progress", "Closed checklist", "Open/Missing check list"]
            col_start = 1
            row_start = 0

            grouped_by_block = df.groupby('Block')

            for block, block_group in grouped_by_block:
                col_pos = col_start
                grouped_by_category = block_group.groupby('Category')

                for category, cat_group in grouped_by_category:
                    worksheet.merge_range(row_start, col_pos, row_start, col_pos + 4, 
                                        f"{block} {category} - {stage_name}", header_format)
                    row_pos = row_start + 1

                    for i, header in enumerate(headers):
                        worksheet.write(row_pos, col_pos + i, header, header_format)
                    row_pos += 1

                    for _, row in cat_group.iterrows():
                        worksheet.write(row_pos, col_pos, row["Activity Name"], cell_format)
                        worksheet.write(row_pos, col_pos + 1, row["Completed Work*(Count of Flat)"], cell_format)
                        worksheet.write(row_pos, col_pos + 2, row["In progress"], cell_format)
                        worksheet.write(row_pos, col_pos + 3, row["Closed checklist"], cell_format)
                        worksheet.write(row_pos, col_pos + 4, row["Open/Missing check list"], cell_format)
                        row_pos += 1

                    total_pending = cat_group["Open/Missing check list"].sum()
                    worksheet.merge_range(row_pos, col_pos, row_pos, col_pos + 3, "Total pending check list", total_format)
                    worksheet.write(row_pos, col_pos + 4, total_pending, total_format)
                    row_pos += 2

                    col_pos += 6

                row_start = row_pos

            for col in range(col_start, col_pos):
                worksheet.set_column(col, col, 20)

        # **NEW: Create Summary Sheet with dynamic month name**
        current_month = pd.Timestamp.now(tz='Asia/Kolkata').strftime('%B')  # Full month name (e.g., "January")
        summary_sheet_name = f"Checklist {current_month}"
        worksheet_summary = workbook.add_worksheet(summary_sheet_name)
        current_row = 0

        worksheet_summary.write(current_row, 0, f"Checklist: {current_month}", header_format)
        current_row += 1

        summary_headers = [
            "Site",
            "Total of Missing & Open Checklist-Civil",
            "Total of Missing & Open Checklist-MEP",
            "Total of Missing & Open Checklist-Interior Finishing",
            "TOTAL"
        ]
        for col, header in enumerate(summary_headers, start=0):
            worksheet_summary.write(current_row, col, header, header_format)
        current_row += 1

        def map_category_to_type(category):
            if category in ["Civil Works"]:
                return "Civil"
            elif category in ["MEP Works"]:
                return "MEP"
            elif category in ["Interior Finishing Works"]:
                return "Interior"
            else:
                return "Civil"

        summary_data = {}
        
        # Store all consolidated rows from sheet generation to use in summary
        if 'all_consolidated_rows' not in st.session_state:
            st.session_state.all_consolidated_rows = {}
        
        # Aggregate data from the already-calculated consolidated rows
        for stage_name in STRUCTURAL_STAGES.keys():
            if stage_name in st.session_state.all_consolidated_rows:
                stage_rows = st.session_state.all_consolidated_rows[stage_name]
                
                for row in stage_rows:
                    block = row['Block']
                    category = row['Category']
                    open_missing = row['Open/Missing check list']
                    
                    # Convert block name to display format
                    if block == "B1 Banket Hall & Finedine":
                        site_name = "WaveCityClub-Block 01 Banket Hall & Finedine"
                    elif "&" in block:
                        block_num = block.replace(" & ", "&")
                        site_name = f"WaveCityClub-Block {block_num}"
                    else:
                        block_num = block[1:]
                        if len(block_num) == 1:
                            block_num = f"0{block_num}"
                        site_name = f"WaveCityClub-Block {block_num}"
                    
                    type_ = map_category_to_type(category)
                    
                    if site_name not in summary_data:
                        summary_data[site_name] = {"Civil": 0, "MEP": 0, "Interior": 0}
                    
                    summary_data[site_name][type_] += open_missing

        # Write summary data
        for site_name, counts in sorted(summary_data.items()):
            civil_count = counts["Civil"]
            mep_count = counts["MEP"]
            interior_count = counts["Interior"]
            total_count = civil_count + mep_count + interior_count
            
            worksheet_summary.write(current_row, 0, site_name, cell_format)
            worksheet_summary.write(current_row, 1, civil_count, cell_format)
            worksheet_summary.write(current_row, 2, mep_count, cell_format)
            worksheet_summary.write(current_row, 3, interior_count, cell_format)
            worksheet_summary.write(current_row, 4, total_count, cell_format)
            current_row += 1

        # Auto-adjust column widths
        for col in range(5):
            worksheet_summary.set_column(col, col, 25)

        workbook.close()
        output.seek(0)
        
        logger.info("Successfully generated Excel file with 7 stage-based sheets")
        return output

    except Exception as e:
        logger.error(f"Error generating consolidated Excel: {str(e)}")
        st.error(f"❌ Error generating Excel file: {str(e)}")
        import traceback
        logger.error(f"Full traceback: {traceback.format_exc()}")
        return None

# Combined function to handle analysis and display
def run_analysis_and_display():
    try:
        st.write("Running stage-wise status analysis...")
        AnalyzeStatusManually()
        st.success("Stage-wise status analysis completed successfully!")

        if 'ai_response' not in st.session_state or not isinstance(st.session_state.ai_response, dict):
            st.session_state.ai_response = {}
            logger.info("Initialized st.session_state.ai_response in run_analysis_and_display")

        st.write("Displaying activity counts and generating AI data...")
        display_activity_count()
        st.success("Activity counts displayed successfully!")

        st.write("Checking AI data totals...")
        logger.info(f"st.session_state.ai_response contents: {st.session_state.ai_response}")
        if not st.session_state.ai_response:
            st.error("❌ No AI data available in st.session_state.ai_response. Attempting to regenerate.")
            logger.error("No AI data in st.session_state.ai_response after display_activity_count")
            display_activity_count()
            if not st.session_state.ai_response:
                st.error("❌ Failed to regenerate AI data. Please check data fetching and try again.")
                logger.error("Failed to regenerate AI data")
                return

        st.write("Generating consolidated checklist Excel file with stage-based sheets...")
        
        # Check if stage analysis exists
        if 'stage_analysis' not in st.session_state:
            st.error("❌ No stage analysis data available. Please ensure stage analysis ran successfully.")
            logger.error("No stage_analysis in st.session_state")
            return

        structure_analysis = st.session_state.get('structure_analysis', None)

        with st.spinner("Generating Excel file with 7 stage sheets... This may take a moment."):
            excel_file = generate_consolidated_Checklist_excel(structure_analysis, st.session_state.ai_response)
        
        if excel_file:
            timestamp = pd.Timestamp.now(tz='Asia/Kolkata').strftime('%Y%m%d_%H%M')
            file_name = f"Consolidated_Checklist_WaveCityClub_Stages_{timestamp}.xlsx"
            
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                st.sidebar.download_button(
                    label="📥 Download Stage-Based Checklist Excel",
                    data=excel_file,
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_excel_button_stages",
                    help="Click to download the consolidated checklist with 7 stage-based sheets."
                )
            st.success("Excel file with stage-based sheets generated successfully! Click the button above to download.")
        else:
            st.error("Failed to generate Excel file. Please check the logs for details.")
            logger.error("Failed to generate Excel file")

    except Exception as e:
        st.error(f"Error during analysis, display, or Excel generation: {str(e)}")
        logger.error(f"Error during analysis, display, or Excel generation: {str(e)}")
        import traceback
        st.error(traceback.format_exc())

# Streamlit UI
st.markdown(
    """
    <h1 style='font-family: "Arial Black", Gadget, sans-serif; 
               color: red; 
               font-size: 48px; 
               text-align: center;'>
        CheckList - Report
    </h1>
    """,
    unsafe_allow_html=True
)

# Initialize and Fetch Data
st.sidebar.title("🔒 Asite Initialization")
email = st.sidebar.text_input("Email", "impwatson@gadieltechnologies.com", key="email_input")
password = st.sidebar.text_input("Password", "Srihari@790$", type="password", key="password_input")

if st.sidebar.button("Initialize and Fetch Data"):
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        success = loop.run_until_complete(initialize_and_fetch_data(email, password))
        if success:
            st.sidebar.success("Initialization and data fetching completed successfully!")
        else:
            st.sidebar.error("Initialization and data fetching failed!")
    except Exception as e:
        st.sidebar.error(f"Initialization and data fetching failed: {str(e)}")
    finally:
        loop.close()

# Analyze and Display
st.sidebar.title("📊 Status Analysis")
if st.sidebar.button("Analyze and Display Activity Counts"):
    with st.spinner("Running analysis and displaying activity counts..."):
        run_analysis_and_display()


