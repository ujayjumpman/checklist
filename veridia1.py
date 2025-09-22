import streamlit as st
import pandas as pd
import requests
import json
import openpyxl
import time
import math
from io import BytesIO
from datetime import datetime
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import ibm_boto3
from ibm_botocore.client import Config
import io
from openpyxl.utils import column_index_from_string

tower2 = []
tower3 = []
tower4 = []
tower5 = []
tower6 = []
tower7 = []


def Tower2(sheet, ignore_year, ignore_month):
    rows = [4, 5, 6, 7, 9, 10, 14, 15, 16, 17, 19, 20]
    cols = ['B', 'D', 'F', 'H', 'J', 'L', 'N', 'P']
    for row in rows:
        for col in cols:
            cell = sheet[f"{col}{row}"]
            cell_value = cell.value
            is_date = False
            cell_date = None
            if isinstance(cell_value, datetime):
                is_date = True
                cell_date = cell_value
            elif isinstance(cell_value, str):
                try:
                    cell_date = datetime.strptime(cell_value, "%Y-%m-%d")
                    is_date = True
                except ValueError:
                    is_date = False
            if is_date and cell_date.year == ignore_year and cell_date.month == ignore_month:
                st.write(f"Cell {col}{row} skipped: Date {cell_value} matches ignored year {ignore_year} and month {ignore_month}")
                continue
            fill = cell.fill
            if fill.fill_type == "solid" and fill.start_color:
                color = fill.start_color.rgb
                if color == "FF92D050":
                    tower2.append(1)
                else:
                    tower2.append(0)
            else:
                tower2.append(0)

def Tower3(sheet, ignore_year, ignore_month):
    rows = [4, 5, 6, 7, 9, 10, 14, 15, 16, 17, 19, 20]
    cols = ['T', 'X', 'AB', 'AF']
    for row in rows:
        for col in cols:
            cell = sheet[f"{col}{row}"]
            cell_value = cell.value
            is_date = False
            cell_date = None
            if isinstance(cell_value, datetime):
                is_date = True
                cell_date = cell_value
            elif isinstance(cell_value, str):
                try:
                    cell_date = datetime.strptime(cell_value, "%Y-%m-%d")
                    is_date = True
                except ValueError:
                    is_date = False
            if is_date and cell_date.year == ignore_year and cell_date.month == ignore_month:
                st.write(f"Cell {col}{row} skipped: Date {cell_value} matches ignored year {ignore_year} and month {ignore_month}")
                continue
            fill = cell.fill
            if fill.fill_type == "solid" and fill.start_color:
                color = fill.start_color.rgb
                if color == "FF92D050":
                    tower3.append(1)
                else:
                    tower3.append(0)
            else:
                tower3.append(0)

def Tower4(sheet, ignore_year, ignore_month):
    rows = [4, 5, 6, 7, 9, 10, 14, 15, 16, 17, 19, 20]
    cols = ['AL', 'AP', 'AT', 'AX', 'BB', 'BF', 'BJ', 'BN']
    for row in rows:
        for col in cols:
            cell = sheet[f"{col}{row}"]
            cell_value = cell.value
            is_date = False
            cell_date = None
            if isinstance(cell_value, datetime):
                is_date = True
                cell_date = cell_value
            elif isinstance(cell_value, str):
                try:
                    cell_date = datetime.strptime(cell_value, "%Y-%m-%d")
                    is_date = True
                except ValueError:
                    is_date = False
            if is_date and cell_date.year == ignore_year and cell_date.month == ignore_month:
                st.write(f"Cell {col}{row} skipped: Date {cell_value} matches ignored year {ignore_year} and month {ignore_month}")
                continue
            fill = cell.fill
            if fill.fill_type == "solid" and fill.start_color:
                color = fill.start_color.rgb
                if color == "FF92D050":
                    tower4.append(1)
                else:
                    tower4.append(0)
            else:
                tower4.append(0)

def Tower5(sheet, ignore_year, ignore_month):
    rows = [4, 5, 6, 7, 9, 10, 14, 15, 16, 17, 19, 20]
    # FIXED: Only check these 7 columns (one side) instead of all 14
    cols = ['DE', 'DI', 'DM', 'DQ', 'DU', 'DY', 'EC']
    for row in rows:
        for col in cols:
            cell = sheet[f"{col}{row}"]
            cell_value = cell.value
            is_date = False
            cell_date = None
            if isinstance(cell_value, datetime):
                is_date = True
                cell_date = cell_value
            elif isinstance(cell_value, str):
                try:
                    cell_date = datetime.strptime(cell_value, "%Y-%m-%d")
                    is_date = True
                except ValueError:
                    is_date = False
            if is_date and cell_date.year == ignore_year and cell_date.month == ignore_month:
                st.write(f"Cell {col}{row} skipped: Date {cell_value} matches ignored year {ignore_year} and month {ignore_month}")
                continue
            fill = cell.fill
            if fill.fill_type == "solid" and fill.start_color:
                color = fill.start_color.rgb
                if color == "FF92D050":
                    tower5.append(1)
                else:
                    tower5.append(0)
            else:
                tower5.append(0)


def Tower6(sheet, ignore_year, ignore_month):
    rows = [4, 5, 6, 7, 9, 10, 14, 15, 16, 17, 19, 20]
    cols = ['FM', 'FQ', 'FU', 'FY', 'GC', 'GG', 'GK']
    for row in rows:
        for col in cols:
            cell = sheet[f"{col}{row}"]
            cell_value = cell.value
            is_date = False
            cell_date = None
            if isinstance(cell_value, datetime):
                is_date = True
                cell_date = cell_value
            elif isinstance(cell_value, str):
                try:
                    cell_date = datetime.strptime(cell_value, "%Y-%m-%d")
                    is_date = True
                except ValueError:
                    is_date = False
            if is_date and cell_date.year == ignore_year and cell_date.month == ignore_month:
                st.write(f"Cell {col}{row} skipped: Date {cell_value} matches ignored year {ignore_year} and month {ignore_month}")
                continue
            fill = cell.fill
            if fill.fill_type == "solid" and fill.start_color:
                color = fill.start_color.rgb
                if color == "FF92D050":
                    tower6.append(1)
                else:
                    tower6.append(0)
            else:
                tower6.append(0)

def Tower7(sheet, ignore_year, ignore_month):
    rows = [4, 5, 6, 7, 9, 10, 14, 15, 16, 17, 19, 20]
    cols = ['EI', 'EM', 'EQ', 'EU', 'EY', 'FC', 'FG']
    for row in rows:
        for col in cols:
            cell = sheet[f"{col}{row}"]
            cell_value = cell.value
            is_date = False
            cell_date = None
            if isinstance(cell_value, datetime):
                is_date = True
                cell_date = cell_value
            elif isinstance(cell_value, str):
                try:
                    cell_date = datetime.strptime(cell_value, "%Y-%m-%d")
                    is_date = True
                except ValueError:
                    is_date = False
            if is_date and cell_date.year == ignore_year and cell_date.month == ignore_month:
                st.write(f"Cell {col}{row} skipped: Date {cell_value} matches ignored year {ignore_year} and month {ignore_month}")
                continue
            fill = cell.fill
            if fill.fill_type == "solid" and fill.start_color:
                color = fill.start_color.rgb
                if color == "FF92D050":
                    tower7.append(1)
                else:
                    tower7.append(0)
            else:
                tower7.append(0)


def ProcessVeridia(exceldatas, ignore_year, ignore_month):
    wb = load_workbook(exceldatas)
    sheet_name = "Revised baseline with 60d NGT"
    sheet = wb[sheet_name]

    tower2.clear()
    Tower2(sheet, ignore_year, ignore_month)
    tower3.clear()
    Tower3(sheet, ignore_year, ignore_month)
    tower4.clear()
    Tower4(sheet, ignore_year, ignore_month)
    tower5.clear()
    Tower5(sheet, ignore_year, ignore_month)
    tower6.clear()
    Tower6(sheet, ignore_year, ignore_month)
    tower7.clear()
    Tower7(sheet, ignore_year, ignore_month)

    data = {
        "Project Name": ["VERIDIA"] * 6,
        "Tower": ["TOWER 2", "TOWER 3", "TOWER 4", "TOWER 5", "TOWER 6", "TOWER 7"],
        "Green (1)": [tower2.count(1), tower3.count(1), tower4.count(1), tower5.count(1), tower6.count(1), tower7.count(1)],
        "Non-Green (0)": [tower2.count(0), tower3.count(0), tower4.count(0), tower5.count(0), tower6.count(0), tower7.count(0)],
       
    }

    project_and_green = [{"Tower": project, "Green (1)": green} for project, green in zip(data["Tower"], data["Green (1)"])]
    json_data = json.dumps(project_and_green, indent=4)


    st.write(json_data)
    return json_data


    