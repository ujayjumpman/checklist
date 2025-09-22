# import streamlit as st
# import pandas as pd
# import requests
# import json
# import openpyxl
# import time
# import math
# from io import BytesIO
# from datetime import datetime
# import io
# from openpyxl import load_workbook
# from openpyxl.styles import PatternFill, Font, Alignment
# from openpyxl.utils import get_column_letter
# import ibm_boto3
# from ibm_botocore.client import Config
# import io
# from openpyxl.utils import column_index_from_string

# towerf = []
# towerg = []
# towerh = []





# def TowerF(sheet, ignore_year, ignore_month):
#     # st.write("Analyzing Eligo Tower F")
#     rows = [5, 6, 7, 8, 9, 10, 11, 12]
#     cols = ['B', 'D', 'F', 'H']
    
#     for row in rows:
#         for col in cols:
#             cell = sheet[f"{col}{row}"]
#             cell_value = cell.value
#             is_date = False
#             cell_date = None
#             if isinstance(cell_value, datetime):
#                 is_date = True
#                 cell_date = cell_value
#             elif isinstance(cell_value, str):
#                 try:
#                     cell_date = datetime.strptime(cell_value, "%Y-%m-%d")
#                     is_date = True
#                 except ValueError:
#                     is_date = False
#             if is_date and cell_date.year == ignore_year and cell_date.month == ignore_month:
#                 st.write(f"Cell {col}{row} skipped: Date {cell_value} matches ignored year {ignore_year} and month {ignore_month}")
#                 continue
#             fill = cell.fill
#             if fill.fill_type == "solid" and fill.start_color:
#                 color = fill.start_color.rgb
#                 if color == "FF92D050":
#                     towerf.append(1)
#                 else:
#                     towerf.append(0)
#             else:
#                 towerf.append(0)

# def TowerG(sheet, ignore_year, ignore_month):
#     # st.write("Analyzing Eligo Tower G")
#     rows = [5, 6, 7, 8, 9, 10, 11, 12]
#     cols = ['L', 'N', 'P', 'R', 'T', 'V']

#     for row in rows:
#         for col in cols:
#             cell = sheet[f"{col}{row}"]
#             cell_value = cell.value
#             is_date = False
#             cell_date = None
#             if isinstance(cell_value, datetime):
#                 is_date = True
#                 cell_date = cell_value
#             elif isinstance(cell_value, str):
#                 try:
#                     cell_date = datetime.strptime(cell_value, "%Y-%m-%d")
#                     is_date = True
#                 except ValueError:
#                     is_date = False
#             if is_date and cell_date.year == ignore_year and cell_date.month == ignore_month:
#                 st.write(f"Cell {col}{row} skipped: Date {cell_value} matches ignored year {ignore_year} and month {ignore_month}")
#                 continue
#             fill = cell.fill
#             if fill.fill_type == "solid" and fill.start_color:
#                 color = fill.start_color.rgb
#                 if color == "FF92D050":
#                     towerg.append(1)
#                 else:
#                     towerg.append(0)
#             else:
#                 towerg.append(0)

   
# def TowerH(sheet, ignore_year, ignore_month):
#     # st.write("Analyzing Eligo Tower H")
#     rows = [5, 6, 7, 8, 9, 10, 11, 12]
#     cols = ['Z', 'AB', 'AD', 'AF', 'AH', 'AJ', 'AL', 'AN', 'AP', 'AR', 'AT', 'AV', 'AX', 'AZ']

#     for row in rows:
#         for col in cols:
#             cell = sheet[f"{col}{row}"]
#             cell_value = cell.value
#             is_date = False
#             cell_date = None
#             if isinstance(cell_value, datetime):
#                 is_date = True
#                 cell_date = cell_value
#             elif isinstance(cell_value, str):
#                 try:
#                     cell_date = datetime.strptime(cell_value, "%Y-%m-%d")
#                     is_date = True
#                 except ValueError:
#                     is_date = False
#             if is_date and cell_date.year == ignore_year and cell_date.month == ignore_month:
#                 st.write(f"Cell {col}{row} skipped: Date {cell_value} matches ignored year {ignore_year} and month {ignore_month}")
#                 continue
#             fill = cell.fill
#             if fill.fill_type == "solid" and fill.start_color:
#                 color = fill.start_color.rgb
#                 if color == "FF92D050":
#                     towerh.append(1)
#                 else:
#                     towerh.append(0)
#             else:
#                 towerh.append(0)


# def Processjson(data):
#     json_data = []
#     for project, tower, green, non_green, finishing in zip(
#     data["Project Name"],
#     data["Tower"],
#     data["Green (1)"],
#     data["Non-Green (0)"],
#     data["Finishing"]
# ):
#         total = green + non_green
#         structure = f"{(green / total * 100):.2f}%" if total > 0 else "0.00%"
        
#         entry = {
#             "Project": project,
#             "Tower Name": tower,
#             "Structure": structure,
#             "Finishing": finishing
#         }
#         json_data.append(entry)
    
#     return json_data



# # def ProcessGandH(exceldatas, ignore_year, ignore_month):

# #     wb = load_workbook(exceldatas, data_only=True)
# #     sheet_names = wb.sheetnames
# #     sheet_name = "Revised Baselines- 25 days SC"

# #     sheet = wb[sheet_name]
# #     #Revised Baselines- 25 days SC
# #     TowerF(sheet, ignore_year, ignore_month)
# #     TowerG(sheet, ignore_year, ignore_month)
# #     TowerH(sheet, ignore_year, ignore_month)

# #     # st.write(towerf.count(1))
# #     # st.write(towerg.count(1))
# #     # st.write(towerh.count(1))
# #     data = {
# #     "Project Name":["ELIGO", "ELIGO", "ELIGO"],
# #     "Tower": ["TOWER F", "TOWER G", "TOWER H"],
# #     "Green (1)": [towerf.count(1), towerg.count(1), towerh.count(1)],
# #     "Non-Green (0)": [towerf.count(0), towerg.count(0), towerh.count(0)],
# #     "Finishing":[st.session_state.towerf_finishing,st.session_state.towerg_finishing,st.session_state.towerh_finishing]
# # }
# #      # Calculate average percentage of green
# #     green_counts = data["Green (1)"]
# #     non_green_counts = data["Non-Green (0)"]
# #     averages = []

# #     for green, non_green in zip(green_counts, non_green_counts):
# #         total = green + non_green
# #         avg = (green / total) * 100 if total > 0 else 0  # avoids division by zero
# #         averages.append(avg)

# #     data["Structure"] = data["Average (%)"] = [f"{(green / (green + non_green) * 100):.2f}%" if (green + non_green) > 0 else "0.00%"
# #                        for green, non_green in zip(green_counts, non_green_counts)]
    
# #     # st.table(data)
# #     # df = pd.DataFrame(data)
# #     json_data = Processjson(data)
# #     return json_data



# def ProcessGandH(exceldatas, ignore_year, ignore_month):
#     wb = load_workbook(exceldatas)  
#     sheet_name = "Revised Baselines- 25 days SC"
#     sheet = wb[sheet_name]

#     towerf.clear()
#     TowerF(sheet, ignore_year, ignore_month)
#     towerg.clear()
#     TowerG(sheet, ignore_year, ignore_month)
#     towerh.clear()
#     TowerH(sheet, ignore_year, ignore_month)


#     data = {
#         "Project Name": ["ELIGO", "ELIGO", "ELIGO"],
#         "Tower": ["TOWER F", "TOWER G", "TOWER H"],
#         "Green (1)": [towerf.count(1), towerg.count(1), towerh.count(1)],
#         "Non-Green (0)": [towerf.count(0), towerg.count(0), towerh.count(0)],
       
#     }

#     project_and_green = [{"Tower": project, "Green (1)": green} for project, green in zip(data["Tower"], data["Green (1)"])]
#     json_data = json.dumps(project_and_green, indent=4)


#     # st.write(json_data)
#     return json_data











import streamlit as st
import pandas as pd
import requests
import json
import openpyxl
import time
import math
from io import BytesIO
from datetime import datetime
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import ibm_boto3
from ibm_botocore.client import Config
import io
from openpyxl.utils import column_index_from_string

towerf = []
towerg = []
towerh = []

def TowerF(sheet, ignore_year, ignore_month):
    # st.write("Analyzing Eligo Tower F")
    rows = [5, 6, 7, 8, 9, 10, 11, 12]
    cols = ['D', 'H']  # Updated columns
    
    for row in rows:
        for col in cols:
            cell = sheet[f"{col}{row}"]
            cell_value = cell.value
            is_date = False
            cell_date = None
            if isinstance(cell_value, datetime):
                is_date = True
                cell_date = cell_value
            elif isinstance(cell_value, str):
                try:
                    cell_date = datetime.strptime(cell_value, "%Y-%m-%d")
                    is_date = True
                except ValueError:
                    is_date = False
            if is_date and cell_date.year == ignore_year and cell_date.month == ignore_month:
                st.write(f"Cell {col}{row} skipped: Date {cell_value} matches ignored year {ignore_year} and month {ignore_month}")
                continue
            fill = cell.fill
            if fill.fill_type == "solid" and fill.start_color:
                color = fill.start_color.rgb
                if color == "FF92D050":
                    towerf.append(1)
                else:
                    towerf.append(0)
            else:
                towerf.append(0)

def TowerG(sheet, ignore_year, ignore_month):
    # st.write("Analyzing Eligo Tower G")
    rows = [5, 6, 7, 8, 9, 10, 11, 12]
    cols = ['N', 'R', 'V']  # Updated columns

    for row in rows:
        for col in cols:
            cell = sheet[f"{col}{row}"]
            cell_value = cell.value
            is_date = False
            cell_date = None
            if isinstance(cell_value, datetime):
                is_date = True
                cell_date = cell_value
            elif isinstance(cell_value, str):
                try:
                    cell_date = datetime.strptime(cell_value, "%Y-%m-%d")
                    is_date = True
                except ValueError:
                    is_date = False
            if is_date and cell_date.year == ignore_year and cell_date.month == ignore_month:
                st.write(f"Cell {col}{row} skipped: Date {cell_value} matches ignored year {ignore_year} and month {ignore_month}")
                continue
            fill = cell.fill
            if fill.fill_type == "solid" and fill.start_color:
                color = fill.start_color.rgb
                if color == "FF92D050":
                    towerg.append(1)
                else:
                    towerg.append(0)
            else:
                towerg.append(0)

def TowerH(sheet, ignore_year, ignore_month):
    # st.write("Analyzing Eligo Tower H")
    rows = [5, 6, 7, 8, 9, 10, 11, 12]
    cols = ['AB', 'AF', 'AJ', 'AN', 'AR', 'AV', 'AZ']  # Updated columns

    for row in rows:
        for col in cols:
            cell = sheet[f"{col}{row}"]
            cell_value = cell.value
            is_date = False
            cell_date = None
            if isinstance(cell_value, datetime):
                is_date = True
                cell_date = cell_value
            elif isinstance(cell_value, str):
                try:
                    cell_date = datetime.strptime(cell_value, "%Y-%m-%d")
                    is_date = True
                except ValueError:
                    is_date = False
            if is_date and cell_date.year == ignore_year and cell_date.month == ignore_month:
                st.write(f"Cell {col}{row} skipped: Date {cell_value} matches ignored year {ignore_year} and month {ignore_month}")
                continue
            fill = cell.fill
            if fill.fill_type == "solid" and fill.start_color:
                color = fill.start_color.rgb
                if color == "FF92D050":
                    towerh.append(1)
                else:
                    towerh.append(0)
            else:
                towerh.append(0)

def Processjson(data):
    json_data = []
    for project, tower, green, non_green, finishing in zip(
    data["Project Name"],
    data["Tower"],
    data["Green (1)"],
    data["Non-Green (0)"],
    data["Finishing"]
):
        total = green + non_green
        structure = f"{(green / total * 100):.2f}%" if total > 0 else "0.00%"
        
        entry = {
            "Project": project,
            "Tower Name": tower,
            "Structure": structure,
            "Finishing": finishing
        }
        json_data.append(entry)
    
    return json_data

def ProcessGandH(exceldatas, ignore_year, ignore_month):
    wb = load_workbook(exceldatas)  
    sheet_name = "Revised Baselines- 25 days SC"
    sheet = wb[sheet_name]

    towerf.clear()
    TowerF(sheet, ignore_year, ignore_month)
    towerg.clear()
    TowerG(sheet, ignore_year, ignore_month)
    towerh.clear()
    TowerH(sheet, ignore_year, ignore_month)

    data = {
        "Project Name": ["ELIGO", "ELIGO", "ELIGO"],
        "Tower": ["TOWER F", "TOWER G", "TOWER H"],
        "Green (1)": [towerf.count(1), towerg.count(1), towerh.count(1)],
        "Non-Green (0)": [towerf.count(0), towerg.count(0), towerh.count(0)],
    }

    project_and_green = [{"Tower": project, "Green (1)": green} for project, green in zip(data["Tower"], data["Green (1)"])]
    json_data = json.dumps(project_and_green, indent=4)

    # st.write(json_data)
    return json_data

   
        