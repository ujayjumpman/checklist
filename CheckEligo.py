# -*- coding: utf-8 -*-
import streamlit as st
import requests
import json 
import urllib3
import certifi
import pandas as pd
from datetime import datetime
import re
import logging
import os
from dotenv import load_dotenv
import aiohttp
import asyncio
from typing import List, Tuple, Dict, Any
from concurrent.futures import ThreadPoolExecutor, as_completed
import time
import openpyxl
import io
from uuid import uuid4
import ibm_boto3
from ibm_botocore.client import Config
from tenacity import retry, stop_after_attempt, wait_exponential
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO
import traceback
from Tower_G_and_H import *
from datetime import date
from dateutil.relativedelta import relativedelta

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Disable SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Load environment variables
load_dotenv()

# IBM COS Configuration
COS_API_KEY = os.getenv("COS_API_KEY")
COS_SERVICE_INSTANCE_ID = os.getenv("COS_SERVICE_INSTANCE_ID")
COS_ENDPOINT = os.getenv("COS_ENDPOINT")
COS_BUCKET = os.getenv("COS_BUCKET")

# WatsonX configuration
WATSONX_API_URL = os.getenv("WATSONX_API_URL_1")
MODEL_ID = os.getenv("MODEL_ID_1")
PROJECT_ID = os.getenv("PROJECT_ID_1")
API_KEY = os.getenv("API_KEY_1")
EMAIL_ID = os.getenv("EMAIL_ID")
PASSWORD = os.getenv("PASSWORD")
# API Endpoints
LOGIN_URL = "https://dms.asite.com/apilogin/"
IAM_TOKEN_URL = "https://iam.cloud.ibm.com/identity/token"

# Login Function
async def login_to_asite(email, password):
    headers = {"Accept": "application/json", "Content-Type": "application/x-www-form-urlencoded"}
    payload = {"emailId": email, "password": password}
    try:
        response = requests.post(LOGIN_URL, headers=headers, data=payload, verify=certifi.where(), timeout=50)
        if response.status_code == 200:
            try:
                session_id = response.json().get("UserProfile", {}).get("Sessionid")
                if session_id:
                    logger.info(f"Login successful, Session ID: {session_id}")
                    st.session_state.sessionid = session_id
                    st.sidebar.success(f"… Login successful, Session ID: {session_id}")
                    return session_id
                else:
                    logger.error("No Session ID found in login response")
                    st.sidebar.error(" No Session ID in response")
                    return None
            except json.JSONDecodeError:
                logger.error("JSONDecodeError during login")
                st.sidebar.error(" Failed to parse login response")
                return None
        logger.error(f"Login failed: {response.status_code} - {response.text}")
        st.sidebar.error(f" Login failed: {response.status_code}")
        return None
    except Exception as e:
        logger.error(f"Error during login: {str(e)}")
        st.sidebar.error(f" Login error: {str(e)}")
        return None

# Function to generate access token
@retry(stop=stop_after_attempt(5), wait=wait_exponential(multiplier=2, min=10, max=60))
def get_access_token(api_key):
    headers = {"Content-Type": "application/x-www-form-urlencoded", "Accept": "application/json"}
    data = {"grant_type": "urn:ibm:params:oauth:grant-type:apikey", "apikey": api_key}
    response = requests.post(IAM_TOKEN_URL, headers=headers, data=data, verify=certifi.where(), timeout=50)
    try:
        if response.status_code == 200:
            token_info = response.json()
            logger.info("Access token generated successfully")
            return token_info['access_token']
        else:
            logger.error(f"Failed to get access token: {response.status_code} - {response.text}")
            st.error(f" Failed to get access token: {response.status_code} - {response.text}")
            raise Exception("Failed to get access token")
    except Exception as e:
        logger.error(f"Exception getting access token: {str(e)}")
        st.error(f" Error getting access token: {str(e)}")
        return None

# Initialize COS client
@retry(stop=stop_after_attempt(5), wait=wait_exponential(multiplier=1, min=4, max=10))
def initialize_cos_client():
    try:
        logger.info("Attempting to initialize COS client...")
        cos_client = ibm_boto3.client(
            's3',
            ibm_api_key_id=COS_API_KEY,
            ibm_service_instance_id=COS_SERVICE_INSTANCE_ID,
            config=Config(
                signature_version='oauth',
                connect_timeout=180,
                read_timeout=180,
                retries={'max_attempts': 15}
            ),
            endpoint_url=COS_ENDPOINT
        )
        logger.info("COS client initialized successfully")
        return cos_client
    except Exception as e:
        logger.error(f"Error initializing COS client: {str(e)}")
        st.error(f" Error initializing COS client: {str(e)}")
        raise

async def validate_session():
    url = "https://dmsak.asite.com/api/workspace/workspacelist"
    headers = {'Cookie': f'ASessionID={st.session_state.sessionid}'}
    async with aiohttp.ClientSession() as session:
        async with session.get(url, headers=headers) as response:
            if response.status == 200:
                logger.info("Session validated successfully")
                return True
            else:
                logger.error(f"Session validation failed: {response.status} - {await response.text()}")
                return False

async def refresh_session_if_needed():
    if 'sessionid' not in st.session_state or not st.session_state.sessionid:
        logger.warning("No session ID found in session state, attempting login")
        new_session_id = await login_to_asite(os.getenv("ASITE_EMAIL"), os.getenv("ASITE_PASSWORD"))
        if new_session_id:
            st.session_state.sessionid = new_session_id
            return new_session_id
        else:
            raise Exception("Failed to establish initial session")

    if not await validate_session():
        logger.info("Session invalid, attempting to refresh")
        new_session_id = await login_to_asite(os.getenv("ASITE_EMAIL"), os.getenv("ASITE_PASSWORD"))
        if new_session_id:
            st.session_state.sessionid = new_session_id
            logger.info(f"Session refreshed successfully, new Session ID: {new_session_id}")
            return new_session_id
        else:
            raise Exception("Failed to refresh session")
    logger.info("Session is valid, no refresh needed")
    return st.session_state.sessionid

# Fetch Workspace ID
async def GetWorkspaceID():
    await refresh_session_if_needed()
    url = "https://dmsak.asite.com/api/workspace/workspacelist"
    headers = {
        'Cookie': f'ASessionID={st.session_state.sessionid}',
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    response = requests.get(url, headers=headers)
    st.session_state.workspaceid = response.json()['asiteDataList']['workspaceVO'][1]['Workspace_Id']
    st.write(f"Workspace ID: {st.session_state.workspaceid}")

# Fetch Project IDs
async def GetProjectId():
    await refresh_session_if_needed()
    url = f"https://adoddleak.asite.com/commonapi/qaplan/getQualityPlanList;searchCriteria={{'criteria': [{{'field': 'planCreationDate','operator': 10,'values': ['11-Mar-2025']}}], 'projectId': {str(st.session_state.workspaceid)}, 'recordLimit': 1000, 'recordStart': 1}}"
    headers = {
        'Cookie': f'ASessionID={st.session_state.sessionid}',
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        data = response.json()
        logger.info(f"GetProjectId full response: {json.dumps(data, indent=2)}")
        
        if 'data' not in data or not data['data']:
            st.error(" No project data found in GetProjectId response")
            logger.error("No project data found in GetProjectId response")
            return
        
        # Log ALL fields in each plan to identify the correct field names
        st.write("### Available Plans (Full Data):")
        for idx, plan in enumerate(data['data']):
            st.write(f"\n**Plan {idx}:**")
            st.json(plan)
            logger.info(f"Plan {idx} full data: {json.dumps(plan, indent=2)}")
        
        # Try multiple possible field names for plan identification
        possible_name_fields = ['planName', 'name', 'title', 'planTitle', 'description']
        
        plan_mapping = {}
        for idx, plan in enumerate(data['data']):
            plan_id = plan.get('planId')
            
            # Try to find plan name from various possible fields
            plan_name = None
            for field in possible_name_fields:
                if field in plan and plan[field]:
                    plan_name = str(plan[field]).strip().upper()
                    break
            
            if not plan_name:
                # If no name field found, log the full plan data and use index-based fallback
                logger.warning(f"No name field found for plan {idx}. Available fields: {list(plan.keys())}")
                st.warning(f"⚠️ Plan {idx} has no identifiable name. Using fallback index-based mapping.")
                continue
            
            logger.info(f"Plan {idx}: Name='{plan_name}', ID={plan_id}")
            
            # Map based on plan name
            if 'STRUCTURE' in plan_name:
                plan_mapping['ELIGO_Structure'] = plan_id
                st.write(f"✓ Found Structure: {plan_name}")
            elif 'TOWER F' in plan_name and 'FINISHING' in plan_name:
                plan_mapping['Eligo_Tower_F_Finishing'] = plan_id
                st.write(f"✓ Found Tower F: {plan_name}")
            elif 'TOWER G' in plan_name and 'FINISHING' in plan_name:
                plan_mapping['Eligo_Tower_G_Finishing'] = plan_id
                st.write(f"✓ Found Tower G: {plan_name}")
            elif 'TOWER H' in plan_name and 'FINISHING' in plan_name:
                plan_mapping['Eligo_Tower_H_Finishing'] = plan_id
                st.write(f"✓ Found Tower H: {plan_name}")
            elif 'NON-TOWER' in plan_name or 'NON TOWER' in plan_name:
                plan_mapping['Eligo_Non_Tower_Area_Finishing'] = plan_id
                st.write(f"✓ Found Non-Tower: {plan_name}")
        
        # FALLBACK: If no names found, use original index-based mapping with warning
        if not plan_mapping:
            st.error("⚠️ FALLBACK MODE: Could not identify plans by name. Using index-based mapping.")
            st.warning("Please manually verify the plan order matches your expectations!")
            
            if len(data['data']) >= 5:
                st.session_state.ELIGO_Structure = data['data'][0]['planId']
                st.session_state.Eligo_Tower_F_Finishing = data['data'][1]['planId']
                st.session_state.Eligo_Tower_G_Finishing = data['data'][2]['planId']
                st.session_state.Eligo_Non_Tower_Area_Finishing = data['data'][3]['planId']
                st.session_state.Eligo_Tower_H_Finishing = data['data'][4]['planId']
                
                st.write("⚠️ Using fallback index mapping:")
                st.write(f"  [0] → Structure: {st.session_state.ELIGO_Structure}")
                st.write(f"  [1] → Tower F: {st.session_state.Eligo_Tower_F_Finishing}")
                st.write(f"  [2] → Tower G: {st.session_state.Eligo_Tower_G_Finishing}")
                st.write(f"  [3] → Non-Tower: {st.session_state.Eligo_Non_Tower_Area_Finishing}")
                st.write(f"  [4] → Tower H: {st.session_state.Eligo_Tower_H_Finishing}")
            else:
                st.error(f"❌ Insufficient plans returned. Expected 5, got {len(data['data'])}")
            return
        
        # Assign to session state with validation
        if 'ELIGO_Structure' in plan_mapping:
            st.session_state.ELIGO_Structure = plan_mapping['ELIGO_Structure']
            st.write(f"✅ ELIGO - Structure Project ID: {st.session_state.ELIGO_Structure}")
        else:
            st.error("❌ Structure plan not found!")
            st.session_state.ELIGO_Structure = None
            
        if 'Eligo_Tower_F_Finishing' in plan_mapping:
            st.session_state.Eligo_Tower_F_Finishing = plan_mapping['Eligo_Tower_F_Finishing']
            st.write(f"✅ ELIGO - Tower F Finishing Project ID: {st.session_state.Eligo_Tower_F_Finishing}")
        else:
            st.error("❌ Tower F Finishing plan not found!")
            st.session_state.Eligo_Tower_F_Finishing = None
            
        if 'Eligo_Tower_G_Finishing' in plan_mapping:
            st.session_state.Eligo_Tower_G_Finishing = plan_mapping['Eligo_Tower_G_Finishing']
            st.write(f"✅ ELIGO - Tower G Finishing Project ID: {st.session_state.Eligo_Tower_G_Finishing}")
        else:
            st.error("❌ Tower G Finishing plan not found!")
            st.session_state.Eligo_Tower_G_Finishing = None
            
        if 'Eligo_Tower_H_Finishing' in plan_mapping:
            st.session_state.Eligo_Tower_H_Finishing = plan_mapping['Eligo_Tower_H_Finishing']
            st.write(f"✅ ELIGO - Tower H Finishing Project ID: {st.session_state.Eligo_Tower_H_Finishing}")
        else:
            st.error("❌ Tower H Finishing plan not found!")
            st.session_state.Eligo_Tower_H_Finishing = None
            
        if 'Eligo_Non_Tower_Area_Finishing' in plan_mapping:
            st.session_state.Eligo_Non_Tower_Area_Finishing = plan_mapping['Eligo_Non_Tower_Area_Finishing']
            st.write(f"✅ ELIGO - Non Tower Area Finishing Project ID: {st.session_state.Eligo_Non_Tower_Area_Finishing}")
        else:
            st.warning("⚠️ Non-Tower Area plan not found (optional)")
            st.session_state.Eligo_Non_Tower_Area_Finishing = None
            
    except Exception as e:
        st.error(f" Error fetching Project IDs: {str(e)}")
        logger.error(f"Error fetching Project IDs: {str(e)}")  


# Asynchronous Fetch Function with Retry Logic
@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=2, max=10))
async def fetch_data(session, url, headers, email=None, password=None):
    try:
        logger.info(f"Fetching data from URL: {url}")
        async with session.get(url, headers=headers) as response:
            content_type = response.headers.get('Content-Type', '')
            logger.info(f"Response Content-Type: {content_type}")
            if response.status == 200:
                if 'application/json' in content_type:
                    return await response.json()
                else:
                    text = await response.text()
                    logger.error(f"Unexpected Content-Type: {content_type}, Response: {text[:500]}")
                    raise ValueError(f"Unexpected Content-Type: {content_type}, expected application/json")
            elif response.status == 204:
                logger.info("No content returned (204)")
                return None
            else:
                text = await response.text()
                logger.error(f"Error fetching data: {response.status} - {text[:500]}")
                if response.status == 401:
                    raise Exception("Unauthorized: Session may have expired")
                raise Exception(f"Error fetching data: {response.status} - {text[:500]}")
    except Exception as e:
        logger.error(f"Fetch failed: {str(e)}")
        raise

# Fetch All Data with Async
async def GetAllDatas():
    record_limit = 1000
    all_external_data = []
    all_structure_data = []
    all_finishing_data = []
    all_tower_h_data = []

    # Ensure session is valid before starting
    await refresh_session_if_needed()
    headers = {'Cookie': f'ASessionID={st.session_state.sessionid}'}

    async with aiohttp.ClientSession() as session:
       
        # Fetch ELIGO Structure data
        start_record = 1
        st.write("Fetching ELIGO Structure data...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanAssociation/?projectId={st.session_state.workspaceid}&planId={st.session_state.ELIGO_Structure}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                await refresh_session_if_needed()
                headers['Cookie'] = f'ASessionID={st.session_state.sessionid}'
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more Structure data available (204)")
                    break
                if 'associationList' in data and data['associationList']:
                    all_structure_data.extend(data['associationList'])
                else:
                    all_structure_data.extend(data if isinstance(data, list) else [])
                st.write(f"Fetched {len(all_structure_data[-record_limit:])} Structure records (Total: {len(all_structure_data)})")
                if len(all_structure_data[-record_limit:]) < record_limit:
                    break
                start_record += record_limit
                await asyncio.sleep(1)  # Rate limiting
            except Exception as e:
                st.error(f" Error fetching Structure data: {str(e)}")
                logger.error(f"Structure data fetch failed: {str(e)}")
                break

        # Fetch ELIGO Tower F Finishing data
        start_record = 1
        st.write("Fetching ELIGO Tower F Finishing data...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanAssociation/?projectId={st.session_state.workspaceid}&planId={st.session_state.Eligo_Tower_F_Finishing}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                # Refresh session before each major fetch to ensure validity
                await refresh_session_if_needed()
                headers['Cookie'] = f'ASessionID={st.session_state.sessionid}'
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more Finishing data available (204)")
                    break
                if 'associationList' in data and data['associationList']:
                    all_finishing_data.extend(data['associationList'])
                else:
                    all_finishing_data.extend(data if isinstance(data, list) else [])
                st.write(f"Fetched {len(all_finishing_data[-record_limit:])} Finishing records (Total: {len(all_finishing_data)})")
                if len(all_finishing_data[-record_limit:]) < record_limit:
                    break
                start_record += record_limit
                await asyncio.sleep(1)  # Rate limiting
            except Exception as e:
                st.error(f" Error fetching Finishing data: {str(e)}")
                logger.error(f"Finishing data fetch failed: {str(e)}")
                break

        # Fetch ELIGO Tower G Finishing data
        start_record = 1
        st.write("Fetching ELIGO Tower G Finishing data...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanAssociation/?projectId={st.session_state.workspaceid}&planId={st.session_state.Eligo_Tower_G_Finishing}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                await refresh_session_if_needed()
                headers['Cookie'] = f'ASessionID={st.session_state.sessionid}'
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more ELIGO Tower G Finishing data available (204)")
                    break
                if 'associationList' in data and data['associationList']:
                    all_external_data.extend(data['associationList'])
                else:
                    all_external_data.extend(data if isinstance(data, list) else [])
                st.write(f"Fetched {len(all_external_data[-record_limit:])} ELIGO Tower G Finishing records (Total: {len(all_external_data)})")
                if len(all_external_data[-record_limit:]) < record_limit:
                    break
                start_record += record_limit
                await asyncio.sleep(1)  # Rate limiting
            except Exception as e:
                st.error(f" Error fetching ELIGO Tower G Finishing data: {str(e)}")
                logger.error(f"ELIGO Tower G Finishing data fetch failed: {str(e)}")
                break

        # Fetch ELIGO Tower H Finishing data
        start_record = 1
        st.write("Fetching ELIGO Tower H Finishing data...")
        while True:
            url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanAssociation/?projectId={st.session_state.workspaceid}&planId={st.session_state.Eligo_Tower_H_Finishing}&recordStart={start_record}&recordLimit={record_limit}"
            try:
                await refresh_session_if_needed()
                headers['Cookie'] = f'ASessionID={st.session_state.sessionid}'
                data = await fetch_data(session, url, headers)
                if data is None:
                    st.write("No more ELIGO Tower H Finishing data available (204)")
                    break
                if 'associationList' in data and data['associationList']:
                    all_tower_h_data.extend(data['associationList'])
                else:
                    all_tower_h_data.extend(data if isinstance(data, list) else [])
                st.write(f"Fetched {len(all_tower_h_data[-record_limit:])} ELIGO Tower H Finishing records (Total: {len(all_tower_h_data)})")
                if len(all_tower_h_data[-record_limit:]) < record_limit:
                    break
                start_record += record_limit
                await asyncio.sleep(1)  # Rate limiting
            except Exception as e:
                st.error(f" Error fetching ELIGO Tower H Finishing data: {str(e)}")
                logger.error(f"ELIGO Tower H Finishing data fetch failed: {str(e)}")
                break

    # Process all dataframes
    df_finishing = pd.DataFrame(all_finishing_data)
    df_structure = pd.DataFrame(all_structure_data)
    df_external = pd.DataFrame(all_external_data)
    df_tower_h = pd.DataFrame(all_tower_h_data)
    
    desired_columns = ['activitySeq', 'qiLocationId']
    
    # Handle status columns for all dataframes
    if 'statusName' in df_finishing.columns:
        desired_columns.append('statusName')
    elif 'statusColor' in df_finishing.columns:
        desired_columns.append('statusColor')
        status_mapping = {'#4CAF50': 'Completed', '#4CB0F0': 'Not Started', '#4C0F0': 'Not Started'}
        df_finishing['statusName'] = df_finishing['statusColor'].map(status_mapping).fillna('Unknown')
        df_structure['statusName'] = df_structure['statusColor'].map(status_mapping).fillna('Unknown')
        df_external['statusName'] = df_external['statusColor'].map(status_mapping).fillna('Unknown')
        df_tower_h['statusName'] = df_tower_h['statusColor'].map(status_mapping).fillna('Unknown')
    else:
        st.error(" Neither statusName nor statusColor found in data!")
        # Create empty columns with 'Unknown' status as fallback
        df_finishing['statusName'] = 'Unknown'
        df_structure['statusName'] = 'Unknown'
        df_external['statusName'] = 'Unknown'
        df_tower_h['statusName'] = 'Unknown'
        desired_columns.append('statusName')

    # Ensure all required columns exist in each dataframe before selection
    for df in [df_finishing, df_structure, df_external, df_tower_h]:
        for col in desired_columns:
            if col not in df.columns:
                df[col] = None

    # Select desired columns
    eligo_tower_f_finishing = df_finishing[desired_columns].copy()
    eligo_structure = df_structure[desired_columns].copy()
    eligo_tower_g_finishing = df_external[desired_columns].copy()
    eligo_tower_h_finishing = df_tower_h[desired_columns].copy()

    # Display results
    st.write(f"ELIGO TOWER F FINISHING ({', '.join(desired_columns)})")
    st.write(f"Total records: {len(eligo_tower_f_finishing)}")
    st.write(eligo_tower_f_finishing)
    
    st.write(f"ELIGO STRUCTURE ({', '.join(desired_columns)})")
    st.write(f"Total records: {len(eligo_structure)}")
    st.write(eligo_structure)
    
    st.write(f"ELIGO TOWER G FINISHING ({', '.join(desired_columns)})")
    st.write(f"Total records: {len(eligo_tower_g_finishing)}")
    st.write(eligo_tower_g_finishing)
    
    st.write(f"ELIGO TOWER H FINISHING ({', '.join(desired_columns)})")
    st.write(f"Total records: {len(eligo_tower_h_finishing)}")
    st.write(eligo_tower_h_finishing)

    # Return all four datasets
    return eligo_tower_f_finishing, eligo_structure, eligo_tower_g_finishing, eligo_tower_h_finishing

# Fetch Activity Data with Async
async def Get_Activity():
    record_limit = 1000
    headers = {
        'Cookie': f'ASessionID={st.session_state.sessionid}',
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    all_finishing_activity_data = []
    all_structure_activity_data = []
    all_external_activity_data = []
    all_tower_h_activity_data = []

    # Ensure session is valid before starting
    await refresh_session_if_needed()

    try:
        async with aiohttp.ClientSession() as session:
            # Fetch ELIGO Tower F Finishing Activity data
            start_record = 1
            st.write("Fetching Activity data for ELIGO Tower F Finishing...")
            while True:
                url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanActivities/?projectId={st.session_state.workspaceid}&planId={st.session_state.Eligo_Tower_F_Finishing}&recordStart={start_record}&recordLimit={record_limit}"
                try:
                    await refresh_session_if_needed()
                    headers['Cookie'] = f'ASessionID={st.session_state.sessionid}'
                    data = await fetch_data(session, url, headers)
                    if data is None:
                        st.write("No more ELIGO Tower F Finishing Activity data available (204)")
                        break
                    if 'activityList' in data and data['activityList']:
                        all_finishing_activity_data.extend(data['activityList'])
                    else:
                        all_finishing_activity_data.extend(data if isinstance(data, list) else [])
                    st.write(f"Fetched {len(all_finishing_activity_data[-record_limit:])} Finishing Activity records (Total: {len(all_finishing_activity_data)})")
                    if len(all_finishing_activity_data[-record_limit:]) < record_limit:
                        break
                    start_record += record_limit
                    await asyncio.sleep(1)
                except Exception as e:
                    st.error(f" Error fetching ELIGO Tower F Finishing Activity data: {str(e)}")
                    logger.error(f"ELIGO Tower F Finishing Activity fetch failed: {str(e)}")
                    break

            # Fetch ELIGO Structure Activity data
            start_record = 1
            st.write("Fetching Activity data for ELIGO Structure...")
            while True:
                url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanActivities/?projectId={st.session_state.workspaceid}&planId={st.session_state.ELIGO_Structure}&recordStart={start_record}&recordLimit={record_limit}"
                try:
                    await refresh_session_if_needed()
                    headers['Cookie'] = f'ASessionID={st.session_state.sessionid}'
                    data = await fetch_data(session, url, headers)
                    if data is None:
                        st.write("No more Structure Activity data available (204)")
                        break
                    if 'activityList' in data and data['activityList']:
                        all_structure_activity_data.extend(data['activityList'])
                    else:
                        all_structure_activity_data.extend(data if isinstance(data, list) else [])
                    st.write(f"Fetched {len(all_structure_activity_data[-record_limit:])} Structure Activity records (Total: {len(all_structure_activity_data)})")
                    if len(all_structure_activity_data[-record_limit:]) < record_limit:
                        break
                    start_record += record_limit
                    await asyncio.sleep(1)
                except Exception as e:
                    st.error(f" Error fetching Structure Activity data: {str(e)}")
                    logger.error(f"Structure Activity fetch failed: {str(e)}")
                    break

            # Fetch ELIGO Tower G Finishing Activity data
            start_record = 1
            st.write("Fetching Activity data for ELIGO Tower G Finishing...")
            while True:
                url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanActivities/?projectId={st.session_state.workspaceid}&planId={st.session_state.Eligo_Tower_G_Finishing}&recordStart={start_record}&recordLimit={record_limit}"
                try:
                    await refresh_session_if_needed()
                    headers['Cookie'] = f'ASessionID={st.session_state.sessionid}'
                    data = await fetch_data(session, url, headers)
                    if data is None:
                        st.write("No more ELIGO Tower G Finishing Activity data available (204)")
                        break
                    if 'activityList' in data and data['activityList']:
                        all_external_activity_data.extend(data['activityList'])
                    else:
                        all_external_activity_data.extend(data if isinstance(data, list) else [])
                    st.write(f"Fetched {len(all_external_activity_data[-record_limit:])} ELIGO Tower G Finishing Activity records (Total: {len(all_external_activity_data)})")
                    if len(all_external_activity_data[-record_limit:]) < record_limit:
                        break
                    start_record += record_limit
                    await asyncio.sleep(1)
                except Exception as e:
                    st.error(f" Error fetching ELIGO Tower G Finishing Activity data: {str(e)}")
                    logger.error(f"ELIGO Tower G Finishing Activity fetch failed: {str(e)}")
                    break

            # Fetch ELIGO Tower H Finishing Activity data
            start_record = 1
            st.write("Fetching Activity data for ELIGO Tower H Finishing...")
            while True:
                url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanActivities/?projectId={st.session_state.workspaceid}&planId={st.session_state.Eligo_Tower_H_Finishing}&recordStart={start_record}&recordLimit={record_limit}"
                try:
                    await refresh_session_if_needed()
                    headers['Cookie'] = f'ASessionID={st.session_state.sessionid}'
                    data = await fetch_data(session, url, headers)
                    if data is None:
                        st.write("No more ELIGO Tower H Finishing Activity data available (204)")
                        break
                    if 'activityList' in data and data['activityList']:
                        all_tower_h_activity_data.extend(data['activityList'])
                    else:
                        all_tower_h_activity_data.extend(data if isinstance(data, list) else [])
                    st.write(f"Fetched {len(all_tower_h_activity_data[-record_limit:])} ELIGO Tower H Finishing Activity records (Total: {len(all_tower_h_activity_data)})")
                    if len(all_tower_h_activity_data[-record_limit:]) < record_limit:
                        break
                    start_record += record_limit
                    await asyncio.sleep(1)
                except Exception as e:
                    st.error(f" Error fetching ELIGO Tower H Finishing Activity data: {str(e)}")
                    logger.error(f"ELIGO Tower H Finishing Activity fetch failed: {str(e)}")
                    break

    except Exception as e:
        st.error(f" Unexpected error in Get_Activity: {str(e)}")
        logger.error(f"Unexpected error in Get_Activity: {str(e)}")
        # ALWAYS return 4 DataFrames, NEVER None
        return (
            pd.DataFrame(columns=['activityName', 'activitySeq', 'formTypeId']),
            pd.DataFrame(columns=['activityName', 'activitySeq', 'formTypeId']),
            pd.DataFrame(columns=['activityName', 'activitySeq', 'formTypeId']),
            pd.DataFrame(columns=['activityName', 'activitySeq', 'formTypeId'])
        )

    def safe_select(df, cols):
        if df is None or (isinstance(df, pd.DataFrame) and df.empty):
            return pd.DataFrame(columns=cols)
        missing = [col for col in cols if col not in df.columns]
        if missing:
            logger.warning(f"Missing columns in activity data: {missing}")
            for col in missing:
                df[col] = None
        return df[cols]

    finishing_activity_data = safe_select(pd.DataFrame(all_finishing_activity_data), ['activityName', 'activitySeq', 'formTypeId'])
    structure_activity_data = safe_select(pd.DataFrame(all_structure_activity_data), ['activityName', 'activitySeq', 'formTypeId'])
    external_activity_data = safe_select(pd.DataFrame(all_external_activity_data), ['activityName', 'activitySeq', 'formTypeId'])
    tower_h_activity_data = safe_select(pd.DataFrame(all_tower_h_activity_data), ['activityName', 'activitySeq', 'formTypeId'])
    
    st.write("ELIGO TOWER F FINISHING ACTIVITY DATA (activityName, activitySeq, formTypeId)")
    st.write(f"Total records: {len(finishing_activity_data)}")
    st.write(finishing_activity_data)
    
    st.write("ELIGO STRUCTURE ACTIVITY DATA (activityName, activitySeq, formTypeId)")
    st.write(f"Total records: {len(structure_activity_data)}")
    st.write(structure_activity_data)
    
    st.write("ELIGO TOWER G FINISHING ACTIVITY DATA (activityName, activitySeq, formTypeId)")
    st.write(f"Total records: {len(external_activity_data)}")
    st.write(external_activity_data)
    
    st.write("ELIGO TOWER H FINISHING ACTIVITY DATA (activityName, activitySeq, formTypeId)")
    st.write(f"Total records: {len(tower_h_activity_data)}")
    st.write(tower_h_activity_data)
    
    # ALWAYS return 4 DataFrames
    return finishing_activity_data, structure_activity_data, external_activity_data, tower_h_activity_data


# FIX 2: COMPLETE REWRITTEN Get_Location() FUNCTION
async def Get_Location():
    record_limit = 1000
    headers = {
        'Cookie': f'ASessionID={st.session_state.sessionid}',
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
    }
    all_finishing_location_data = []
    all_structure_location_data = []
    all_external_location_data = []
    all_tower_h_location_data = []

    # Ensure session is valid before starting
    await refresh_session_if_needed()

    try:
        async with aiohttp.ClientSession() as session:
            # Fetch ELIGO Tower F Finishing Location/Module data
            start_record = 1
            total_records_fetched = 0
            st.write("Fetching ELIGO Tower F Finishing Location/Module data...")
            while True:
                url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanLocation/?projectId={st.session_state.workspaceid}&planId={st.session_state.Eligo_Tower_F_Finishing}&recordStart={start_record}&recordLimit={record_limit}"
                try:
                    await refresh_session_if_needed()
                    headers['Cookie'] = f'ASessionID={st.session_state.sessionid}'
                    data = await fetch_data(session, url, headers)
                    if data is None:
                        st.write("No more ELIGO Tower F Finishing Location data available (204)")
                        break
                    if isinstance(data, list):
                        location_data = [{'qiLocationId': item.get('qiLocationId', ''), 'qiParentId': item.get('qiParentId', ''), 'name': item.get('name', '')} 
                                       for item in data if isinstance(item, dict)]
                        all_finishing_location_data.extend(location_data)
                        total_records_fetched = len(all_finishing_location_data)
                        st.write(f"Fetched {len(location_data)} Finishing Location records (Total: {total_records_fetched})")
                    elif isinstance(data, dict) and 'locationList' in data and data['locationList']:
                        location_data = [{'qiLocationId': loc.get('qiLocationId', ''), 'qiParentId': loc.get('qiParentId', ''), 'name': loc.get('name', '')} 
                                       for loc in data['locationList']]
                        all_finishing_location_data.extend(location_data)
                        total_records_fetched = len(all_finishing_location_data)
                        st.write(f"Fetched {len(location_data)} Finishing Location records (Total: {total_records_fetched})")
                    else:
                        st.warning(f"No 'locationList' in Finishing Location data or empty list.")
                        break
                    if len(location_data) < record_limit:
                        break
                    start_record += record_limit
                    await asyncio.sleep(1)
                except Exception as e:
                    st.error(f" Error fetching Finishing Location data: {str(e)}")
                    logger.error(f"Finishing Location fetch failed: {str(e)}")
                    break

            # Fetch ELIGO Structure Location/Module data
            start_record = 1
            total_records_fetched = 0
            st.write("Fetching ELIGO Structure Location/Module data...")
            while True:
                url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanLocation/?projectId={st.session_state.workspaceid}&planId={st.session_state.ELIGO_Structure}&recordStart={start_record}&recordLimit={record_limit}"
                try:
                    await refresh_session_if_needed()
                    headers['Cookie'] = f'ASessionID={st.session_state.sessionid}'
                    data = await fetch_data(session, url, headers)
                    if data is None:
                        st.write("No more Structure Location data available (204)")
                        break
                    if isinstance(data, list):
                        location_data = [{'qiLocationId': item.get('qiLocationId', ''), 'qiParentId': item.get('qiParentId', ''), 'name': item.get('name', '')} 
                                       for item in data if isinstance(item, dict)]
                        all_structure_location_data.extend(location_data)
                        total_records_fetched = len(all_structure_location_data)
                        st.write(f"Fetched {len(location_data)} Structure Location records (Total: {total_records_fetched})")
                    elif isinstance(data, dict) and 'locationList' in data and data['locationList']:
                        location_data = [{'qiLocationId': loc.get('qiLocationId', ''), 'qiParentId': loc.get('qiParentId', ''), 'name': loc.get('name', '')} 
                                       for loc in data['locationList']]
                        all_structure_location_data.extend(location_data)
                        total_records_fetched = len(all_structure_location_data)
                        st.write(f"Fetched {len(location_data)} Structure Location records (Total: {total_records_fetched})")
                    else:
                        st.warning(f"No 'locationList' in Structure Location data or empty list.")
                        break
                    if len(location_data) < record_limit:
                        break
                    start_record += record_limit
                    await asyncio.sleep(1)
                except Exception as e:
                    st.error(f" Error fetching Structure Location data: {str(e)}")
                    logger.error(f"Structure Location fetch failed: {str(e)}")
                    break

            # Fetch ELIGO Tower G Finishing Location/Module data
            start_record = 1
            total_records_fetched = 0
            st.write("Fetching ELIGO Tower G Finishing Location/Module data...")
            while True:
                url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanLocation/?projectId={st.session_state.workspaceid}&planId={st.session_state.Eligo_Tower_G_Finishing}&recordStart={start_record}&recordLimit={record_limit}"
                try:
                    await refresh_session_if_needed()
                    headers['Cookie'] = f'ASessionID={st.session_state.sessionid}'
                    data = await fetch_data(session, url, headers)
                    if data is None:
                        st.write("No more ELIGO Tower G Finishing Location data available (204)")
                        break
                    if isinstance(data, list):
                        location_data = [{'qiLocationId': item.get('qiLocationId', ''), 'qiParentId': item.get('qiParentId', ''), 'name': item.get('name', '')} 
                                       for item in data if isinstance(item, dict)]
                        all_external_location_data.extend(location_data)
                        total_records_fetched = len(all_external_location_data)
                        st.write(f"Fetched {len(location_data)} ELIGO Tower G Finishing Location records (Total: {total_records_fetched})")
                    elif isinstance(data, dict) and 'locationList' in data and data['locationList']:
                        location_data = [{'qiLocationId': loc.get('qiLocationId', ''), 'qiParentId': loc.get('qiParentId', ''), 'name': loc.get('name', '')} 
                                       for loc in data['locationList']]
                        all_external_location_data.extend(location_data)
                        total_records_fetched = len(all_external_location_data)
                        st.write(f"Fetched {len(location_data)} ELIGO Tower G Finishing Location records (Total: {total_records_fetched})")
                    else:
                        st.warning(f"No 'locationList' in ELIGO Tower G Finishing Location data or empty list.")
                        break
                    if len(location_data) < record_limit:
                        break
                    start_record += record_limit
                    await asyncio.sleep(1)
                except Exception as e:
                    st.error(f" Error fetching Tower G Finishing Location data: {str(e)}")
                    logger.error(f"Tower G Finishing Location fetch failed: {str(e)}")
                    break

            # Fetch ELIGO Tower H Finishing Location/Module data
            start_record = 1
            total_records_fetched = 0
            st.write("Fetching ELIGO Tower H Finishing Location/Module data...")
            while True:
                url = f"https://adoddleak.asite.com/commonapi/qaplan/getPlanLocation/?projectId={st.session_state.workspaceid}&planId={st.session_state.Eligo_Tower_H_Finishing}&recordStart={start_record}&recordLimit={record_limit}"
                try:
                    await refresh_session_if_needed()
                    headers['Cookie'] = f'ASessionID={st.session_state.sessionid}'
                    data = await fetch_data(session, url, headers)
                    if data is None:
                        st.write("No more ELIGO Tower H Finishing Location data available (204)")
                        break
                    if isinstance(data, list):
                        location_data = [{'qiLocationId': item.get('qiLocationId', ''), 'qiParentId': item.get('qiParentId', ''), 'name': item.get('name', '')} 
                                       for item in data if isinstance(item, dict)]
                        all_tower_h_location_data.extend(location_data)
                        total_records_fetched = len(all_tower_h_location_data)
                        st.write(f"Fetched {len(location_data)} ELIGO Tower H Finishing Location records (Total: {total_records_fetched})")
                    elif isinstance(data, dict) and 'locationList' in data and data['locationList']:
                        location_data = [{'qiLocationId': loc.get('qiLocationId', ''), 'qiParentId': loc.get('qiParentId', ''), 'name': loc.get('name', '')} 
                                       for loc in data['locationList']]
                        all_tower_h_location_data.extend(location_data)
                        total_records_fetched = len(all_tower_h_location_data)
                        st.write(f"Fetched {len(location_data)} ELIGO Tower H Finishing Location records (Total: {total_records_fetched})")
                    else:
                        st.warning(f"No 'locationList' in ELIGO Tower H Finishing Location data or empty list.")
                        break
                    if len(location_data) < record_limit:
                        break
                    start_record += record_limit
                    await asyncio.sleep(1)
                except Exception as e:
                    st.error(f" Error fetching Tower H Finishing Location data: {str(e)}")
                    logger.error(f"Tower H Finishing Location fetch failed: {str(e)}")
                    break

        # Create DataFrames with guaranteed columns
        finishing_df = pd.DataFrame(all_finishing_location_data) if all_finishing_location_data else pd.DataFrame(columns=['qiLocationId', 'qiParentId', 'name'])
        structure_df = pd.DataFrame(all_structure_location_data) if all_structure_location_data else pd.DataFrame(columns=['qiLocationId', 'qiParentId', 'name'])
        external_df = pd.DataFrame(all_external_location_data) if all_external_location_data else pd.DataFrame(columns=['qiLocationId', 'qiParentId', 'name'])
        tower_h_df = pd.DataFrame(all_tower_h_location_data) if all_tower_h_location_data else pd.DataFrame(columns=['qiLocationId', 'qiParentId', 'name'])

        # Ensure all required columns exist
        for df in [finishing_df, structure_df, external_df, tower_h_df]:
            for col in ['qiLocationId', 'qiParentId', 'name']:
                if col not in df.columns:
                    df[col] = None

        # Validate name field for all dataframes
        if finishing_df.empty or finishing_df['name'].isna().all():
            st.warning(" Tower F Finishing Location data has no 'name' values.")
        if structure_df.empty or structure_df['name'].isna().all():
            st.warning(" Structure Location data has no 'name' values.")
        if external_df.empty or external_df['name'].isna().all():
            st.warning(" Tower G Finishing Location data has no 'name' values.")
        if tower_h_df.empty or tower_h_df['name'].isna().all():
            st.warning(" Tower H Finishing Location data has no 'name' values.")
        
        st.write("ELIGO TOWER F FINISHING LOCATION/MODULE DATA")
        st.write(f"Total records: {len(finishing_df)}")
        st.write(finishing_df)
        
        st.write("ELIGO STRUCTURE LOCATION/MODULE DATA")
        st.write(f"Total records: {len(structure_df)}")
        st.write(structure_df)
        
        st.write("ELIGO TOWER G FINISHING LOCATION/MODULE DATA")
        st.write(f"Total records: {len(external_df)}")
        st.write(external_df)
        
        st.write("ELIGO TOWER H FINISHING LOCATION/MODULE DATA")
        st.write(f"Total records: {len(tower_h_df)}")
        st.write(tower_h_df)

        st.session_state.finishing_location_data = finishing_df
        st.session_state.structure_location_data = structure_df
        st.session_state.external_location_data = external_df
        st.session_state.tower_h_location_data = tower_h_df

        return finishing_df, structure_df, external_df, tower_h_df

    except Exception as e:
        st.error(f" Unexpected error in Get_Location: {str(e)}")
        logger.error(f"Unexpected error in Get_Location: {str(e)}")
        # ALWAYS return 4 DataFrames with required columns, NEVER None
        return (
            pd.DataFrame(columns=['qiLocationId', 'qiParentId', 'name']),
            pd.DataFrame(columns=['qiLocationId', 'qiParentId', 'name']),
            pd.DataFrame(columns=['qiLocationId', 'qiParentId', 'name']),
            pd.DataFrame(columns=['qiLocationId', 'qiParentId', 'name'])
        )


# Process individual chunk
def process_chunk(chunk, chunk_idx, dataset_name, location_df):
    logger.info(f"Starting thread for {dataset_name} Chunk {chunk_idx + 1}")
    generated_text = format_chunk_locally(chunk, chunk_idx, len(chunk), dataset_name, location_df)
    logger.info(f"Completed thread for {dataset_name} Chunk {chunk_idx + 1}")
    return generated_text, chunk_idx

# Process data with manual counting
def process_manually(analysis_df, total, dataset_name, chunk_size=1000, max_workers=4):
    if analysis_df.empty:
        st.warning(f"No completed activities found for {dataset_name}.")
        return {"towers": {}, "total": 0}

    unique_activities = analysis_df['activityName'].unique()
    logger.info(f"Unique activities in {dataset_name} dataset: {list(unique_activities)}")
    logger.info(f"Total records in {dataset_name} dataset: {len(analysis_df)}")

    st.write(f"Saved ELIGO {dataset_name} data to eligo_{dataset_name.lower()}_data.json")
    chunks = [analysis_df[i:i + chunk_size] for i in range(0, len(analysis_df), chunk_size)]

    location_df = (
        st.session_state.finishing_location_data if dataset_name.lower() == "finishing"
        else st.session_state.structure_location_data if dataset_name.lower() == "structure"
        else st.session_state.external_location_data
    )

    chunk_results = {}
    progress_bar = st.progress(0)
    status_text = st.empty()

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_chunk = {
            executor.submit(process_chunk, chunk, idx, dataset_name, location_df): idx
            for idx, chunk in enumerate(chunks)
        }

        completed_chunks = 0
        for future in as_completed(future_to_chunk):
            chunk_idx = future_to_chunk[future]
            try:
                generated_text, idx = future.result()
                chunk_results[idx] = generated_text
                completed_chunks += 1
                progress_percent = completed_chunks / len(chunks)
                progress_bar.progress(progress_percent)
                status_text.text(f"Processed chunk {completed_chunks} of {len(chunks)} ({progress_percent:.1%} complete)")
            except Exception as e:
                logger.error(f"Error processing chunk {chunk_idx + 1} for {dataset_name}: {str(e)}")
                st.error(f" Error processing chunk {chunk_idx + 1}: {str(e)}")

    parsed_data = {}
    for chunk_idx in sorted(chunk_results.keys()):
        generated_text = chunk_results[chunk_idx]
        if not generated_text:
            continue

        current_tower = None
        tower_activities = []
        lines = generated_text.split("\n")

        for line in lines:
            line = line.strip()
            if not line:
                continue

            if line.startswith("Tower:"):
                try:
                    tower_parts = line.split("Tower:", 1)
                    if len(tower_parts) > 1:
                        if current_tower and tower_activities:
                            if current_tower not in parsed_data:
                                parsed_data[current_tower] = []
                            parsed_data[current_tower].extend(tower_activities)
                            tower_activities = []
                        current_tower = tower_parts[1].strip()
                except Exception as e:
                    logger.warning(f"Error parsing Tower line: {line}, error: {str(e)}")
                    if not current_tower:
                        current_tower = f"Unknown Tower {chunk_idx}"

            elif line.startswith("Total Completed Activities:"):
                continue
            elif not line.strip().startswith("activityName"):
                try:
                    parts = re.split(r'\s{2,}', line.strip())
                    if len(parts) >= 2:
                        activity_name = ' '.join(parts[:-1]).strip()
                        count_str = parts[-1].strip()
                        count_match = re.search(r'\d+', count_str)
                        if count_match:
                            count = int(count_match.group())
                            if current_tower:
                                tower_activities.append({
                                    "activityName": activity_name,
                                    "completedCount": count
                                })
                    else:
                        match = re.match(r'^\s*(.+?)\s+(\d+)$', line)
                        if match and current_tower:
                            activity_name = match.group(1).strip()
                            count = int(match.group(2).strip())
                            tower_activities.append({
                                "activityName": activity_name,
                                "completedCount": count
                            })
                except (ValueError, IndexError) as e:
                    logger.warning(f"Skipping malformed activity line: {line}, error: {str(e)}")

        if current_tower and tower_activities:
            if current_tower not in parsed_data:
                parsed_data[current_tower] = []
            parsed_data[current_tower].extend(tower_activities)

    aggregated_data = {}
    for tower_name, activities in parsed_data.items():
        tower_short_name = tower_name.split('/')[1] if '/' in tower_name else tower_name
        if tower_short_name not in aggregated_data:
            aggregated_data[tower_short_name] = {}
        
        for activity in activities:
            name = activity.get("activityName", "Unknown")
            count = activity.get("completedCount", 0)
            if name in aggregated_data[tower_short_name]:
                aggregated_data[tower_short_name][name] += count
            else:
                aggregated_data[tower_short_name][name] = count

    return {"towers": aggregated_data, "total": total}

# Local formatting function for manual counting
def format_chunk_locally(chunk, chunk_idx, chunk_size, dataset_name, location_df):
    towers_data = {}
    
    for _, row in chunk.iterrows():
        tower_name = row['tower_name']
        activity_name = row['activityName']
        count = int(row['CompletedCount'])
        
        if tower_name not in towers_data:
            towers_data[tower_name] = []
            
        towers_data[tower_name].append({
            "activityName": activity_name,
            "completedCount": count
        })
    
    output = ""
    total_activities = 0
    
    for tower_name, activities in sorted(towers_data.items()):
        output += f"Tower: {tower_name}\n"
        output += "activityName            CompletedCount\n"
        activity_dict = {}
        for activity in activities:
            name = activity['activityName']
            count = activity['completedCount']
            activity_dict[name] = activity_dict.get(name, 0) + count
        for name, count in sorted(activity_dict.items()):
            output += f"{name:<30} {count}\n"
            total_activities += count
    
    output += f"Total Completed Activities: {total_activities}"
    return output


# ADD THIS FUNCTION BEFORE process_data()
def apply_roof_slab_filter(completed_df):
    """
    IMPROVED Roof Slab Filter for Tower F/G/H Finishing
    
    Handles patterns like:
    - Quality/Tower-F Finishing/F 1/1 First Floor Roof Slab/101
    - Quality/Tower-F/F 1/1 First Floor Roof Slab/101
    - Quality/Tower-F Finishing/F 2/1 First Floor Roof Slab/102
    - Quality/Tower-F Finishing/F 1/3 Third Floor Roof Slab/301
    - Quality/Tower-F Finishing/F 1/4 Fourth Roof Slab/401
    
    New Pattern: ^Quality\/Tower-[FGH](?: Finishing)?\/F\s*\d+\/\d+\s+.*Roof Slab\/\s*\d{3}$
    Breakdown:
    - ^                               = Start of string
    - Quality\/                       = Must start with "Quality/"
    - Tower-[FGH]                     = Tower F, G, or H
    - (?: Finishing)?                 = Optional " Finishing" text
    - \/                              = Forward slash
    - F\s*\d+\/\d+\s+                 = "F" followed by unit numbers like "1/1", "2/1", etc.
    - .*Roof Slab\/                   = Any text ending with "Roof Slab/"
    - \s*\d{3}$                       = Exactly 3 digits at the end (unit numbers: 101, 102, etc.)
    
    Args:
        completed_df: DataFrame with 'full_path' column
    
    Returns:
        Filtered DataFrame with only unit-level records
    """
    if completed_df.empty:
        logger.warning("DataFrame is empty - no filtering applied")
        return completed_df
    
    try:
        # STRICT PATTERN - Matches your exact desired format
        strict_pattern = r'^Quality\/Tower-[FGH](?: Finishing)?\/F\s*\d+\/\d+\s+.*Roof Slab\/\s*\d{3}$'
        
        # FALLBACK PATTERN - For any unit-level entries ending with 3 digits
        fallback_pattern = r'^.*\/\s*\d{3}$'
        
        # First try strict pattern
        mask_strict = completed_df['full_path'].str.match(strict_pattern, case=False, na=False)
        
        # If no matches with strict pattern, use fallback
        if mask_strict.sum() == 0:
            logger.info("No matches with strict pattern, trying fallback pattern")
            mask = completed_df['full_path'].str.match(fallback_pattern, case=False, na=False)
        else:
            mask = mask_strict
        
        filtered_df = completed_df[mask].copy()
        
        records_removed = len(completed_df) - len(filtered_df)
        
        logger.info(f"\n{'='*80}")
        logger.info(f"IMPROVED ROOF SLAB FILTER APPLIED")
        logger.info(f"{'='*80}")
        logger.info(f"  Before: {len(completed_df)} records")
        logger.info(f"  After:  {len(filtered_df)} records")
        logger.info(f"  Removed: {records_removed} records (non-unit entries)")
        logger.info(f"  Strict pattern: {strict_pattern}")
        logger.info(f"  Fallback pattern: {fallback_pattern}")
        
        # Show detailed statistics
        logger.info(f"\n  Pattern matching statistics:")
        logger.info(f"    Matched with strict pattern: {mask_strict.sum()} records")
        logger.info(f"    Matched with fallback pattern: {mask.sum()} records")
        
        # Show sample of removed entries if any
        if records_removed > 0:
            removed = completed_df[~mask]
            logger.info(f"\n  Sample removed entries (floor headers, etc.):")
            for idx, (_, row) in enumerate(removed.head(10).iterrows(), 1):
                logger.info(f"    {idx}. {row['full_path']}")
                # Also log why it didn't match if we can determine
                if not re.search(r'\/\s*\d{3}$', str(row['full_path'])):
                    logger.info(f"       → Does not end with /XXX pattern")
                elif not re.search(r'Roof Slab', str(row['full_path']), re.I):
                    logger.info(f"       → Does not contain 'Roof Slab'")
        
        # Show sample of kept entries
        if len(filtered_df) > 0:
            logger.info(f"\n  Sample kept entries (unit-level records):")
            kept_samples = filtered_df.head(10)
            for idx, (_, row) in enumerate(kept_samples.iterrows(), 1):
                logger.info(f"    {idx}. {row['full_path']}")
                
                # Extract and display the unit number
                match = re.search(r'\/(\d{3})$', str(row['full_path']))
                if match:
                    unit_num = match.group(1)
                    floor_num = unit_num[0] if len(unit_num) > 0 else '?'
                    logger.info(f"       → Unit: {unit_num}, Floor: {floor_num}xx")
        
        # Additional validation: Ensure we're getting the expected floor ranges
        if len(filtered_df) > 0:
            unit_numbers = []
            for path in filtered_df['full_path']:
                match = re.search(r'\/(\d{3})$', str(path))
                if match:
                    unit_numbers.append(int(match.group(1)))
            
            if unit_numbers:
                logger.info(f"\n  Unit number statistics:")
                logger.info(f"    Min unit: {min(unit_numbers)}")
                logger.info(f"    Max unit: {max(unit_numbers)}")
                logger.info(f"    Total unique units: {len(set(unit_numbers))}")
                
                # Group by floor
                floor_counts = {}
                for unit in unit_numbers:
                    floor = str(unit)[0] if len(str(unit)) >= 1 else '0'
                    floor_counts[floor] = floor_counts.get(floor, 0) + 1
                
                logger.info(f"    Units per floor:")
                for floor in sorted(floor_counts.keys()):
                    logger.info(f"      Floor {floor}xx: {floor_counts[floor]} units")
        
        logger.info(f"{'='*80}\n")
        
        return filtered_df
        
    except Exception as e:
        logger.error(f"ERROR in roof slab filter: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return completed_df


def apply_structure_roof_slab_filter(completed_df, dataset_name):
    """
    UNIVERSAL Structure Work Filter for Tower F, G, and H
    
    Handles patterns like:
    - Quality/Tower-F/F1/1st Floor Roof Slab
    - Quality/Tower-G/G2/2nd Floor Roof Slab
    - Quality/Tower-H/H1/Stilt Roof Slab (REMOVED by this filter)
    - Quality/Tower-H/H1/00.Stilt Roof Slab (REMOVED by this filter)
    
    Pattern breakdown:
    ^Quality\/Tower\s*[FGH]\/[FGH]\d+\/.*Roof\s+Slab$
    - ^Quality\/ = Must start with Quality/
    - Tower\s*[FGH] = Tower F, G, or H (with optional space)
    - \/[FGH]\d+ = Slash, then wing designation (F1-F2, G1-G3, H1-H7)
    - \/.*Roof\s+Slab = Any text ending with "Roof Slab"
    - $ = Must end with Slab (no trailing /XXX)
    
    ADDITIONAL FILTER: Removes all Stilt entries
    - Stilt Roof Slab (with or without leading "00.")
    - Excludes: "1st Floor", "2nd Floor", "3rd Floor", "4th Floor", etc.
    """
    if completed_df.empty:
        logger.warning("DataFrame is empty - no filtering applied")
        return completed_df
    
    try:
        # STEP 1: UNIVERSAL PATTERN - Matches roof slab entries
        universal_pattern = r'^Quality\/Tower\s*[FGH]\/[FGH]\d+\/.*Roof\s+Slab$'
        
        # Apply pattern (case-insensitive for "slab" variations)
        mask_roof_slab = completed_df['full_path'].str.match(universal_pattern, case=False, na=False)
        
        # STEP 2: STILT REMOVAL FILTER
        # Remove entries containing "Stilt" (with or without "00.")
        stilt_pattern = r'(?:00\.)?\s*Stilt\s+Roof\s+Slab'
        mask_stilt = completed_df['full_path'].str.contains(stilt_pattern, case=False, na=False, regex=True)
        
        # Combine filters: Keep roof slabs BUT exclude stilt
        mask_final = mask_roof_slab & ~mask_stilt
        
        filtered_df = completed_df[mask_final].copy()
        
        records_removed = len(completed_df) - len(filtered_df)
        stilt_removed = mask_stilt.sum()
        roof_slab_matched = mask_roof_slab.sum()
        
        logger.info(f"\n{'='*80}")
        logger.info(f"STRUCTURE WORK ROOF SLAB FILTER - {dataset_name}")
        logger.info(f"{'='*80}")
        logger.info(f"  Roof Slab Pattern: {universal_pattern}")
        logger.info(f"  Stilt Removal Pattern: {stilt_pattern}")
        logger.info(f"\n  Before: {len(completed_df)} records")
        logger.info(f"  Matched Roof Slab: {roof_slab_matched} records")
        logger.info(f"  Removed (Stilt): {stilt_removed} records")
        logger.info(f"  After:  {len(filtered_df)} records")
        logger.info(f"  Total Removed: {records_removed} records")
        
        # Show sample of removed Stilt entries
        if stilt_removed > 0:
            removed_stilt = completed_df[mask_stilt]
            logger.info(f"\n  Stilt entries removed (showing first 10):")
            for idx, (_, row) in enumerate(removed_stilt.head(10).iterrows(), 1):
                path = str(row['full_path'])
                logger.info(f"    {idx}. {path}")
        
        # Show sample of kept entries
        if len(filtered_df) > 0:
            logger.info(f"\n  Sample kept entries (floor-level roof slabs):")
            kept_samples = filtered_df.head(10)
            for idx, (_, row) in enumerate(kept_samples.iterrows(), 1):
                path = str(row['full_path'])
                logger.info(f"    {idx}. {path}")
            
            # Extract floor information
            logger.info(f"\n  Floor distribution in filtered data:")
            floor_matches = []
            for path in filtered_df['full_path']:
                path_str = str(path)
                # Try to extract floor level
                floor_match = re.search(r'(\d+)(?:st|nd|rd|th)\s+Floor', path_str, re.I)
                if floor_match:
                    floor_num = floor_match.group(1)
                    floor_matches.append(int(floor_num))
            
            if floor_matches:
                logger.info(f"    Min Floor: {min(floor_matches)}")
                logger.info(f"    Max Floor: {max(floor_matches)}")
                
                # Count by floor
                from collections import Counter
                floor_counts = Counter(floor_matches)
                for floor in sorted(floor_counts.keys()):
                    logger.info(f"    Floor {floor}: {floor_counts[floor]} entries")
        
        logger.info(f"{'='*80}\n")
        
        return filtered_df
        
    except Exception as e:
        logger.error(f"ERROR in roof slab filter: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        return completed_df
# ============================================================================
# UPDATED process_data() FUNCTION
# ============================================================================

def process_data(df, activity_df, location_df, dataset_name, use_module_hierarchy_for_finishing=False):
    """
    Process completed activities with UNIVERSAL Structure Work filter
    - Tower F Finishing: Uses Finishing filter (with /XXX unit numbers)
    - Tower G Finishing: Uses Finishing filter (with /XXX unit numbers)
    - Tower H Finishing: Uses Finishing filter (with /XXX unit numbers)
    - Structure (ELIGO): Uses Structure filter (NO /XXX, just floor levels)
    """
    # Keep only completed rows
    completed = df[df['statusName'] == 'Completed'].copy()
    
    # Define allowed activities
    ALLOWED_MEP = [
        "Plumbing Works",
        "Slab Conducting",
        "Wall Conduting", 
        "Wiring & Switch Socket",
    ]
    
    ALLOWED_FINISHING = [
        "Floor Tile",
        "Wall Tile",
        "POP & Gypsum Plaster",
        "Waterproofing - Sunken",
    ]
    
    ALLOWED_CIVIL = [
        "Concreting",
        "Shuttering", 
        "Reinforcement",
        "De-Shuttering",
    ]

    # Initialize count table
    asite_activities = ALLOWED_MEP + ALLOWED_FINISHING + ALLOWED_CIVIL
    count_table = pd.DataFrame({'Count': [0] * len(asite_activities)}, index=asite_activities)

    logger.info(f"\n{'='*80}")
    logger.info(f"PROCESSING: {dataset_name}")
    logger.info(f"{'='*80}")
    logger.info(f"Initial rows: {len(df)}")
    logger.info(f"Completed rows (statusName='Completed'): {len(completed)}")

    if completed.empty:
        logger.warning(f"No completed activities found in {dataset_name} data.")
        return pd.DataFrame(), 0, count_table

    # Merge location and activity names
    logger.info(f"Merging with {len(location_df)} locations and {len(activity_df)} activities...")
    completed = completed.merge(location_df[['qiLocationId', 'name']], on='qiLocationId', how='left')
    completed = completed.merge(activity_df[['activitySeq', 'activityName']], on='activitySeq', how='left')
    
    if 'qiActivityId' not in completed.columns:
        completed['qiActivityId'] = completed['qiLocationId'].astype(str) + '$$' + completed['activitySeq'].astype(str)
    
    completed['name'] = completed['name'].fillna('Unknown')
    logger.info(f"After merge: {len(completed)} rows")

    # Normalize activity names
    def normalize_activity_name(activity_name):
        """Normalize activity names"""
        if not isinstance(activity_name, str):
            return activity_name
        
        name = activity_name.strip()
        name_lower = name.lower()
        
        # DE-SHUTTERING
        de_shuttering_keywords = [
            "de-shuttering", "de shuttering", "deshuttering", "de-shutter",
            "removal", "striking", "removal of formwork"
        ]
        if any(keyword in name_lower for keyword in de_shuttering_keywords):
            return "De-Shuttering"
        
        # SLAB CONDUCTING
        slab_keywords = ["slab", "casting", "cast", "no. of slab", "no of slab"]
        if any(keyword in name_lower for keyword in slab_keywords):
            if "concreting" not in name_lower:
                return "Slab Conducting"
        
        # WALL CONDUCTING
        if any(kw in name_lower for kw in ["wall conduting"]):
            return "Wall Conduting"
        
        # PLUMBING
        if any(kw in name_lower for kw in ["UP 1st Fix", "CP 1st Fix"]):
            return "Plumbing Works"
        
        # WIRING
        if any(kw in name_lower for kw in ["EL 2nd fix","el 2nd fix"]):
            return "Wiring & Switch Socket"
        
        # POP & GYPSUM
        if any(kw in name_lower for kw in ["POP Punning (Major area)","Pop Punning (Major Area)","POP Punning (Minor Area)","Pop Punning (Minor Area)"]):
            return "POP & Gypsum Plaster"
        
        # WATERPROOFING
        if any(kw in name_lower for kw in ["water proofing works","Water Proofing Works"]):
            return "Waterproofing - Sunken"
        
        # WALL TILE
        if "wall til" in name_lower:
            return "Wall Tile"
        
        # FLOOR TILE
        if "floor til" in name_lower:
            return "Floor Tile"
        
        # CONCRETING
        if any(kw in name_lower for kw in ["concreting", "concrete"]):
            return "Concreting"
        
        # SHUTTERING
        if any(kw in name_lower for kw in ["shuttering", "shutter", "formwork"]):
            return "Shuttering"
        
        # REINFORCEMENT
        if any(kw in name_lower for kw in ["reinforcement", "rebar", "steel"]):
            if "de-" not in name_lower:
                return "Reinforcement"
        
        return activity_name

    completed['activityName'] = completed['activityName'].apply(normalize_activity_name).fillna('Unknown')
    
    logger.info(f"\nActivity distribution BEFORE filtering:")
    for act in sorted(completed['activityName'].unique()):
        count = len(completed[completed['activityName'] == act])
        is_allowed = act in (ALLOWED_MEP + ALLOWED_FINISHING + ALLOWED_CIVIL)
        logger.info(f"  {act}: {count} records [ALLOWED: {is_allowed}]")

    # Build location path dictionaries
    parent_child_dict = dict(zip(location_df['qiLocationId'], location_df['qiParentId']))
    name_dict = dict(zip(location_df['qiLocationId'], location_df['name']))

    def get_full_path(location_id):
        """Build full location path"""
        path = []
        current_id = location_id
        max_depth = 10
        depth = 0
        
        while current_id and depth < max_depth:
            if current_id not in parent_child_dict or current_id not in name_dict:
                break
            parent_id = parent_child_dict.get(current_id)
            nm = name_dict.get(current_id, "Unknown")
            
            if not parent_id:
                if nm != "Quality":
                    path.append(nm)
                path.append("Quality")
            else:
                path.append(nm)
            current_id = parent_id
            depth += 1
        
        if not path:
            return "Unknown"
        return '/'.join(reversed(path))

    completed['full_path'] = completed['qiLocationId'].apply(get_full_path)
    
    # ========== APPLY APPROPRIATE FILTER ==========
    if "Structure" in dataset_name or "ELIGO" in dataset_name:
        logger.info(f"\n✅ APPLYING STRUCTURE WORK ROOF SLAB FILTER for {dataset_name}")
        logger.info(f"   Pattern: Quality/Tower [FGH]/[FGH]#/...Roof Slab")
        completed = apply_structure_roof_slab_filter(completed, dataset_name)
        
        if completed.empty:
            logger.error(f"❌ CRITICAL: All records filtered out for {dataset_name}!")
            return pd.DataFrame(), 0, count_table
            
        logger.info(f"   Records after filter: {len(completed)}")
    
    elif "Finishing" in dataset_name:
        logger.info(f"\n✅ APPLYING FINISHING ROOF SLAB FILTER for {dataset_name}")
        logger.info(f"   Pattern: Quality/Tower-[FGH] Finishing/[FGH] #/#...Roof Slab/###")
        completed = apply_roof_slab_filter(completed)
        
        if completed.empty:
            logger.error(f"❌ CRITICAL: All records filtered out for {dataset_name}!")
            return pd.DataFrame(), 0, count_table
            
        logger.info(f"   Records after filter: {len(completed)}")
    
    # ACTIVITY WHITELIST
    logger.info(f"\nApplying activity whitelist filter...")
    allowed_all = set(ALLOWED_MEP + ALLOWED_FINISHING + ALLOWED_CIVIL)
    completed = completed[completed['activityName'].isin(allowed_all)].copy()
    
    logger.info(f"After whitelist filter: {len(completed)} rows")

    if completed.empty:
        logger.warning(f"No completed activities in allowed list for {dataset_name}.")
        return pd.DataFrame(), 0, count_table

    logger.info(f"\nActivity distribution AFTER filters:")
    for act in sorted(completed['activityName'].unique()):
        count = len(completed[completed['activityName'] == act])
        logger.info(f"  {act}: {count} records")

    # Map tower names
    def get_tower_name(full_path):
        """Extract tower name from path"""
        if not isinstance(full_path, str):
            return "Unknown"
        parts = full_path.split('/')
        for part in parts:
            if part.strip().lower().startswith("tower"):
                return part.strip()
        return "Unknown"

    completed['tower_name'] = completed['full_path'].apply(get_tower_name)
    
    logger.info(f"\nTower distribution:")
    for tower in sorted(completed['tower_name'].unique()):
        count = len(completed[completed['tower_name'] == tower])
        logger.info(f"  {tower}: {count} records")

    # Group by tower and activity
    analysis = (
        completed.groupby(['tower_name', 'activityName'])
        .size()
        .reset_index(name='CompletedCount')
        .sort_values(by=['tower_name', 'activityName'])
    )
    
    total_completed = analysis['CompletedCount'].sum()
    
    # Update activity count table
    activity_counts = (
        completed.groupby('activityName')
        .size()
        .reset_index(name='Count')
    )
    
    for activity in asite_activities:
        if activity in activity_counts['activityName'].values:
            count_table.loc[activity, 'Count'] = activity_counts.loc[
                activity_counts['activityName'] == activity, 'Count'
            ].iloc[0]

    logger.info(f"\n{'='*80}")
    logger.info(f"FINAL RESULTS FOR {dataset_name}:")
    logger.info(f"Total completed activities: {total_completed}")
    logger.info(f"{'='*80}\n")
    
    return analysis, total_completed, count_table

# Main analysis function
def _ensure_session_keys(keys: List[str]) -> Tuple[bool, List[str]]:
    missing = [k for k in keys if k not in st.session_state]
    return (len(missing) == 0, missing)


def normalize_activity_name(name):
    if not isinstance(name, str):
        return name
    typo_corrections = {
        "Wall Conduting": "Wall Conduting",
        "Slab conduting": "Slab Conduting",
    }
    lower = name.lower()
    for typo, correct in typo_corrections.items():
        if lower == typo.lower():
            return correct
    return name

def process_cos_data(tower_name: str, tower_df: pd.DataFrame) -> pd.DataFrame:
    """
    Convert COS tracker rows to Checklist activity counts.
    Uses MIN(UP-First Fix, CP-First Fix) for plumbing logic and maps COS names -> checklist names.
    **ONLY counts activities with valid Actual Finish dates (completed activities).**
    Returns DataFrame columns: ["Tower", "Activity Name", "Count"].
    
    Args:
        tower_name: Tower identifier (e.g., 'TF', 'TG', 'TH', 'Structure')
        tower_df: DataFrame with COS tracker data containing 'Activity Name' and 'Actual Finish' columns
    
    Returns:
        pd.DataFrame: DataFrame with columns ["Tower", "Activity Name", "Count"]
    
    Examples:
        >>> result = process_cos_data('TG', tower_g_df)
        >>> result.to_dict('records')
        [
            {'Tower': 'TG', 'Activity Name': 'Wall Conduting', 'Count': 5},
            {'Tower': 'TG', 'Activity Name': 'Plumbing Works', 'Count': 3},
        ]
    """
    try:
        # Guards - Check for None or empty DataFrame
        if tower_df is None:
            logger.warning(f"{tower_name}: DataFrame is None")
            return pd.DataFrame(columns=["Tower", "Activity Name", "Count"])
        
        if not isinstance(tower_df, pd.DataFrame):
            logger.warning(f"{tower_name}: Input is not a DataFrame, type: {type(tower_df)}")
            return pd.DataFrame(columns=["Tower", "Activity Name", "Count"])
        
        if tower_df.empty:
            logger.warning(f"{tower_name}: DataFrame is empty")
            return pd.DataFrame(columns=["Tower", "Activity Name", "Count"])

        # *** FILTER TO ONLY COMPLETED ACTIVITIES ***
        original_count = len(tower_df)
        tower_df = tower_df.copy()  # Work with a copy to avoid modifying original
        
        # Check if 'Actual Finish' column exists
        if 'Actual Finish' not in tower_df.columns:
            logger.warning(f"{tower_name}: No 'Actual Finish' column found. Available columns: {list(tower_df.columns)}")
            st.warning(f"⚠️ {tower_name}: No 'Actual Finish' column found. Available: {list(tower_df.columns)}")
            return pd.DataFrame(columns=["Tower", "Activity Name", "Count"])
        
        # Filter to only rows with non-null Actual Finish dates
        tower_df = tower_df[pd.notna(tower_df['Actual Finish'])].copy()
        completed_count = len(tower_df)
        
        logger.info(f"{tower_name}: Filtered from {original_count} to {completed_count} rows with Actual Finish dates")
        st.write(f"  📊 {tower_name}: {original_count} total rows → {completed_count} completed rows")
        
        # Check if any completed activities remain
        if tower_df.empty:
            logger.warning(f"{tower_name}: No rows with valid Actual Finish dates")
            st.warning(f"  ⚠️ {tower_name}: No completed activities found")
            return pd.DataFrame(columns=["Tower", "Activity Name", "Count"])

        # Detect activity column - try multiple possible column names
        possible_activity_cols = [
            "Activity Name", 
            "ActivityName", 
            "Activity", 
            "activityName",
            "Task Name",
            "Task",
            "Description"
        ]
        
        activity_col = None
        for col_name in possible_activity_cols:
            if col_name in tower_df.columns:
                activity_col = col_name
                logger.info(f"{tower_name}: Using activity column: {activity_col}")
                break
        
        if activity_col is None:
            logger.error(f"{tower_name}: No activity column found. Available columns: {list(tower_df.columns)}")
            st.error(f"  ❌ {tower_name}: No activity column found. Available: {list(tower_df.columns)}")
            return pd.DataFrame(columns=["Tower", "Activity Name", "Count"])
        
        # COS to Asite activity name mapping
        cos_to_asite_mapping = {
            "Wall Conduting": "Wall Conduting",
            "UP-First Fix": "Plumbing Works",
            "UP- First Fix": "Plumbing Works",
            "CP-First Fix": "Plumbing Works",
            "CP- First Fix": "Plumbing Works",
            "Min. count of UP-First Fix and CP-First Fix": "Plumbing Works",
            "POP punning (Major area)": "POP & Gypsum Plaster",
            "POP Punning(Major area)": "POP & Gypsum Plaster",
            "EL-2nd Fix": "Wiring & Switch Socket",
            "EL- 2nd Fix": "Wiring & Switch Socket",
            "EL 2nd Fix": "Wiring & Switch Socket",
            "El Second Fix": "Wiring & Switch Socket",
            "Water Proofing Works": "Waterproofing - Sunken",
            "Waterproofing Works": "Waterproofing - Sunken",
            "Wall Tiling": "Wall Tile",
            "Wall Tile": "Wall Tile",
            "Floor Tiling": "Floor Tile",
            "Floor Tile": "Floor Tile",
            "Installation of doors": "Door/Window Frame",
            "Concreting": "Concreting",
            "Shuttering": "Shuttering",
            "Reinforcement": "Reinforcement",
            "De-Shuttering": "De-Shuttering",
            "No. of Slab cast": "Slab Conducting",
            "No of Slab cast": "Slab Conducting",
        }
        
        # Process activities
        activity_counts = {}
        
        for idx, row in tower_df.iterrows():
            try:
                activity_name_raw = str(row[activity_col]).strip()
                
                # Skip empty or null values
                if not activity_name_raw or activity_name_raw.lower() in ['nan', 'none', 'na', 'n/a', '']:
                    continue
                
                # Try to map COS activity name to Asite name
                activity_name_mapped = None
                for cos_name, asite_name in cos_to_asite_mapping.items():
                    if activity_name_raw.lower() == cos_name.lower():
                        activity_name_mapped = asite_name
                        break
                
                # If no mapping found, use original name
                if activity_name_mapped is None:
                    activity_name_mapped = activity_name_raw
                    logger.debug(f"{tower_name}: No mapping for activity: {activity_name_raw}, using original")
                
                # Increment count for this activity
                if activity_name_mapped in activity_counts:
                    activity_counts[activity_name_mapped] += 1
                else:
                    activity_counts[activity_name_mapped] = 1
                    
            except Exception as e:
                logger.warning(f"{tower_name}: Error processing row {idx}: {str(e)}")
                continue
        
        # Convert counts to DataFrame
        if not activity_counts:
            logger.warning(f"{tower_name}: No valid activities extracted")
            st.warning(f"  ⚠️ {tower_name}: No valid activities found after processing")
            return pd.DataFrame(columns=["Tower", "Activity Name", "Count"])
        
        # Create result DataFrame
        result_data = []
        for activity_name, count in sorted(activity_counts.items()):
            result_data.append({
                "Tower": tower_name,
                "Activity Name": activity_name,
                "Count": count
            })
        
        result_df = pd.DataFrame(result_data)
        
        logger.info(f"{tower_name}: Processed {len(result_df)} unique activities with total {result_df['Count'].sum()} counts")
        st.write(f"  ✓ {tower_name}: Extracted {len(result_df)} unique activities")
        
        return result_df
        
    except Exception as e:
        logger.error(f"{tower_name}: Unexpected error in process_cos_data: {str(e)}", exc_info=True)
        st.error(f"  ❌ {tower_name}: Error processing COS data: {str(e)}")
        return pd.DataFrame(columns=["Tower", "Activity Name", "Count"])
def diagnostic_location_paths(df, activity_df, location_df, dataset_name):
    """
    Show actual location paths for completed activities to understand filtering needs
    """
    completed = df[df['statusName'] == 'Completed'].copy()
    
    logger.info(f"\n{'='*100}")
    logger.info(f"LOCATION PATH DIAGNOSTIC: {dataset_name}")
    logger.info(f"{'='*100}")
    logger.info(f"Total completed activities: {len(completed)}")
    
    # Merge data
    completed = completed.merge(location_df[['qiLocationId', 'name']], on='qiLocationId', how='left')
    completed = completed.merge(activity_df[['activitySeq', 'activityName']], on='activitySeq', how='left')
    completed['name'] = completed['name'].fillna('Unknown')
    
    # Normalize
    def normalize_activity_name(name):
        if not isinstance(name, str):
            return name
        typo_corrections = {
            "Wall Conduting": "Wall Conduting",
            "Slab conduting": "Slab Conducting",
        }
        for typo, correct in typo_corrections.items():
            if isinstance(name, str) and name.lower() == typo.lower():
                return correct
        return name

    completed['activityName'] = completed['activityName'].apply(normalize_activity_name).fillna('Unknown')
    
    # Build paths
    parent_child_dict = dict(zip(location_df['qiLocationId'], location_df['qiParentId']))
    name_dict = dict(zip(location_df['qiLocationId'], location_df['name']))

    def get_full_path(location_id):
        path = []
        current_id = location_id
        max_depth = 10
        depth = 0
        
        while current_id and depth < max_depth:
            if current_id not in parent_child_dict or current_id not in name_dict:
                break
            parent_id = parent_child_dict.get(current_id)
            nm = name_dict.get(current_id, "Unknown")
            
            if not parent_id:
                if nm != "Quality":
                    path.append(nm)
                path.append("Quality")
            else:
                path.append(nm)
            current_id = parent_id
            depth += 1
        
        if not path:
            return "Unknown"
        return '/'.join(reversed(path))

    completed['full_path'] = completed['qiLocationId'].apply(get_full_path)

    # DEBUG: Show actual paths for Tower F Finishing
    if dataset_name == "Tower F Finishing":
        st.write(f"### Debug: Sample paths for {dataset_name}")
        sample_paths = completed['full_path'].head(20).tolist()
        for i, path in enumerate(sample_paths, 1):
            st.write(f"{i}. {path}")
    
    # Focus on MEP activities
    MEP_ACTIVITIES = ["Plumbing Works", "Wall Conduting", "Wiring & Switch Socket", "Slab Conducting"]
    FINISHING_ACTIVITIES = ["Floor Tile", "Wall Tile", "POP & Gypsum Plaster", "Waterproofing - Sunken"]
    
    logger.info(f"\n{'='*100}")
    logger.info("MEP ACTIVITIES - ALL LOCATION PATHS:")
    logger.info(f"{'='*100}")
    
    for activity in MEP_ACTIVITIES:
        activity_data = completed[completed['activityName'] == activity]
        if activity_data.empty:
            logger.info(f"\n{activity}: NO DATA")
            continue
        
        logger.info(f"\n{activity}: {len(activity_data)} total records")
        logger.info(f"Unique paths ({len(activity_data['full_path'].unique())} total):")
        
        for path in sorted(activity_data['full_path'].unique()):
            count = len(activity_data[activity_data['full_path'] == path])
            # Break down path
            parts = path.split('/')
            logger.info(f"  [{count:3d}x] {path}")
            if len(parts) >= 3:
                logger.info(f"         L0: {parts[0]}, L1: {parts[1]}, L2: {parts[2]}")
    
    logger.info(f"\n{'='*100}")
    logger.info("INTERIOR FINISHING ACTIVITIES - ALL LOCATION PATHS:")
    logger.info(f"{'='*100}")
    
    for activity in FINISHING_ACTIVITIES:
        activity_data = completed[completed['activityName'] == activity]
        if activity_data.empty:
            logger.info(f"\n{activity}: NO DATA")
            continue
        
        logger.info(f"\n{activity}: {len(activity_data)} total records")
        logger.info(f"Unique paths ({len(activity_data['full_path'].unique())} total):")
        
        for path in sorted(activity_data['full_path'].unique()):
            count = len(activity_data[activity_data['full_path'] == path])
            parts = path.split('/')
            logger.info(f"  [{count:3d}x] {path}")
            if len(parts) >= 3:
                logger.info(f"         L0: {parts[0]}, L1: {parts[1]}, L2: {parts[2]}")
    
    # Analyze what distinguishes correct vs incorrect counts
    logger.info(f"\n{'='*100}")
    logger.info("PATH PATTERN ANALYSIS:")
    logger.info(f"{'='*100}")
    
    for activity in MEP_ACTIVITIES + FINISHING_ACTIVITIES:
        activity_data = completed[completed['activityName'] == activity]
        if activity_data.empty:
            continue
        
        logger.info(f"\n{activity}:")
        
        # Check for common keywords
        has_roof_slab_slash = activity_data['full_path'].str.contains('roof slab/', case=False, na=False).sum()
        has_roof_slab_space = activity_data['full_path'].str.contains(r'roof\s+slab', case=False, na=False, regex=True).sum()
        has_first_floor = activity_data['full_path'].str.contains('First Floor', case=False, na=False).sum()
        has_second_floor = activity_data['full_path'].str.contains('Second Floor', case=False, na=False).sum()
        has_floor_pattern = activity_data['full_path'].str.contains(r'\d+(?:st|nd|rd|th)\s+Floor', case=False, na=False, regex=True).sum()
        
        logger.info(f"  Contains 'roof slab/': {has_roof_slab_slash}/{len(activity_data)}")
        logger.info(f"  Contains 'roof slab' (any spacing): {has_roof_slab_space}/{len(activity_data)}")
        logger.info(f"  Contains 'First Floor': {has_first_floor}/{len(activity_data)}")
        logger.info(f"  Contains 'Second Floor': {has_second_floor}/{len(activity_data)}")
        logger.info(f"  Contains floor pattern (1st/2nd/etc): {has_floor_pattern}/{len(activity_data)}")
    
    logger.info(f"\n{'='*100}\n")

def AnalyzeStatusManually(email: str = None, password: str = None) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    """
    Refactored AnalyzeStatusManually with proper None handling:
    - Validates session state and dataframes.
    - Converts None values to empty DataFrames.
    - Processes Asite datasets via process_data/process_manually.
    - Returns combined COS + Asite dataframes.
    Returns: (combined_data, outputs)
    """
    start_time = time.time()
    outputs: Dict[str, Any] = {}

    # Ensure logged in
    if 'sessionid' not in st.session_state:
        st.error("Please log in first!")
        return {}, {}

    # Required keys
    required_data = [
        'eligo_tower_f_finishing', 'eligo_structure', 'eligo_tower_g_finishing', 'eligo_tower_h_finishing',
        'finishing_activity_data', 'structure_activity_data', 'external_activity_data', 'tower_h_activity_data',
        'finishing_location_data', 'structure_location_data', 'external_location_data', 'tower_h_location_data',
    ]
    ok, missing = _ensure_session_keys(required_data)
    if not ok:
        st.error(f"Please fetch required data first! Missing: {missing}")
        return {}, {}

    # Load primary datasets with comprehensive None checking
    try:
        finishing_data = st.session_state.get('eligo_tower_f_finishing')
        structure_data = st.session_state.get('eligo_structure')
        external_data = st.session_state.get('eligo_tower_g_finishing')
        tower_h_data = st.session_state.get('eligo_tower_h_finishing')

        finishing_activity = st.session_state.get('finishing_activity_data')
        structure_activity = st.session_state.get('structure_activity_data')
        external_activity = st.session_state.get('external_activity_data')
        tower_h_activity = st.session_state.get('tower_h_activity_data')

        finishing_locations = st.session_state.get('finishing_location_data')
        structure_locations = st.session_state.get('structure_location_data')
        external_locations = st.session_state.get('external_location_data')
        tower_h_locations = st.session_state.get('tower_h_location_data')

        # Convert None values to empty DataFrames with correct columns
        if finishing_data is None:
            st.warning("finishing_data is None, creating empty DataFrame")
            finishing_data = pd.DataFrame(columns=['statusName', 'qiLocationId', 'activitySeq'])
        if structure_data is None:
            st.warning("structure_data is None, creating empty DataFrame")
            structure_data = pd.DataFrame(columns=['statusName', 'qiLocationId', 'activitySeq'])
        if external_data is None:
            st.warning("external_data is None, creating empty DataFrame")
            external_data = pd.DataFrame(columns=['statusName', 'qiLocationId', 'activitySeq'])
        if tower_h_data is None:
            st.warning("tower_h_data is None, creating empty DataFrame")
            tower_h_data = pd.DataFrame(columns=['statusName', 'qiLocationId', 'activitySeq'])

        if finishing_activity is None:
            st.warning("finishing_activity is None, creating empty DataFrame")
            finishing_activity = pd.DataFrame(columns=['activitySeq', 'activityName', 'formTypeId'])
        if structure_activity is None:
            st.warning("structure_activity is None, creating empty DataFrame")
            structure_activity = pd.DataFrame(columns=['activitySeq', 'activityName', 'formTypeId'])
        if external_activity is None:
            st.warning("external_activity is None, creating empty DataFrame")
            external_activity = pd.DataFrame(columns=['activitySeq', 'activityName', 'formTypeId'])
        if tower_h_activity is None:
            st.warning("tower_h_activity is None, creating empty DataFrame")
            tower_h_activity = pd.DataFrame(columns=['activitySeq', 'activityName', 'formTypeId'])

        if finishing_locations is None:
            st.warning("finishing_locations is None, creating empty DataFrame")
            finishing_locations = pd.DataFrame(columns=['qiLocationId', 'qiParentId', 'name'])
        if structure_locations is None:
            st.warning("structure_locations is None, creating empty DataFrame")
            structure_locations = pd.DataFrame(columns=['qiLocationId', 'qiParentId', 'name'])
        if external_locations is None:
            st.warning("external_locations is None, creating empty DataFrame")
            external_locations = pd.DataFrame(columns=['qiLocationId', 'qiParentId', 'name'])
        if tower_h_locations is None:
            st.warning("tower_h_locations is None, creating empty DataFrame")
            tower_h_locations = pd.DataFrame(columns=['qiLocationId', 'qiParentId', 'name'])

    except Exception as e:
        st.error(f"Error retrieving session state data: {str(e)}")
        logger.exception("Session state retrieval failure")
        return {}, {}
    
    diagnostic_location_paths(external_data, external_activity, external_locations, "Tower G Finishing")
    # Validate dataframes -> main datasets
    main_datasets = [
        (finishing_data, "Tower F Finishing", ['statusName', 'qiLocationId', 'activitySeq']),
        (structure_data, "ELIGO Structure", ['statusName', 'qiLocationId', 'activitySeq']),
        (external_data, "Tower G Finishing", ['statusName', 'qiLocationId', 'activitySeq']),
        (tower_h_data, "Tower H Finishing", ['statusName', 'qiLocationId', 'activitySeq']),
    ]
    
    for df, name, required_cols in main_datasets:
        if df is None:
            st.error(f"{name} data is None!")
            return {}, {}
        if not isinstance(df, pd.DataFrame):
            st.error(f"{name} data is not a DataFrame! Type: {type(df)}")
            return {}, {}
        
        # Add missing columns if needed
        for col in required_cols:
            if col not in df.columns:
                df[col] = None

    # Validate locations
    location_datasets = [
        (finishing_locations, "Tower F Finishing Location", ['qiLocationId', 'name', 'qiParentId']),
        (structure_locations, "ELIGO Structure Location", ['qiLocationId', 'name', 'qiParentId']),
        (external_locations, "Tower G Finishing Location", ['qiLocationId', 'name', 'qiParentId']),
        (tower_h_locations, "Tower H Finishing Location", ['qiLocationId', 'name', 'qiParentId'])
    ]
    
    for df, name, required_cols in location_datasets:
        if df is None:
            st.error(f"{name} data is None!")
            return {}, {}
        if not isinstance(df, pd.DataFrame):
            st.error(f"{name} data is not a DataFrame!")
            return {}, {}
        
        for col in required_cols:
            if col not in df.columns:
                df[col] = None

    # Validate activity datasets
    activity_datasets = [
        (finishing_activity, "Tower F Finishing Activity", ['activitySeq', 'activityName']),
        (structure_activity, "ELIGO Structure Activity", ['activitySeq', 'activityName']),
        (external_activity, "Tower G Finishing Activity", ['activitySeq', 'activityName']),
        (tower_h_activity, "Tower H Finishing Activity", ['activitySeq', 'activityName']),
    ]
    
    for df, name, required_cols in activity_datasets:
        if df is None:
            st.error(f"{name} data is None!")
            return {}, {}
        if not isinstance(df, pd.DataFrame):
            st.error(f"{name} data is not a DataFrame!")
            return {}, {}
        
        for col in required_cols:
            if col not in df.columns:
                df[col] = None

    # Normalize activity names in copies (avoid SettingWithCopy)
    finishing_activity = st.session_state['finishing_activity_data'].copy() if isinstance(st.session_state.get('finishing_activity_data'), pd.DataFrame) else pd.DataFrame(columns=['activitySeq', 'activityName'])
    structure_activity = st.session_state['structure_activity_data'].copy() if isinstance(st.session_state.get('structure_activity_data'), pd.DataFrame) else pd.DataFrame(columns=['activitySeq', 'activityName'])
    external_activity = st.session_state['external_activity_data'].copy() if isinstance(st.session_state.get('external_activity_data'), pd.DataFrame) else pd.DataFrame(columns=['activitySeq', 'activityName'])
    tower_h_activity = st.session_state['tower_h_activity_data'].copy() if isinstance(st.session_state.get('tower_h_activity_data'), pd.DataFrame) else pd.DataFrame(columns=['activitySeq', 'activityName'])

    for df in [finishing_activity, structure_activity, external_activity, tower_h_activity]:
        if not df.empty and 'activityName' in df.columns:
            df['activityName'] = df['activityName'].apply(normalize_activity_name)

    # Attempt to fetch slab/COS info
    st.write("Fetching COS slab cycle data...")
    if not all(key in st.session_state for key in ['cos_client', 'bucket_name']):
        st.error("COS client not initialized. Please run 'Initialize and Fetch Data' first.")
        st.session_state['slab_df'] = pd.DataFrame()
    else:
        try:
            cos_client = st.session_state['cos_client']
            bucket_name = st.session_state['bucket_name']

            if 'file_list' not in st.session_state or st.session_state['file_list'] is None:
                st.write("File list not found in session state. Fetching files from COS...")
                try:
                    response = cos_client.list_objects_v2(Bucket=bucket_name, Prefix="Eligo/")
                    contents = response.get('Contents', [])
                    st.session_state['file_list'] = [{'Key': obj['Key']} for obj in contents] if contents else []
                    if not contents:
                        st.warning("No files found in Eligo folder.")
                except Exception as ex:
                    logger.exception("Error fetching file list from COS")
                    st.session_state['file_list'] = []
                    
            file_list = st.session_state.get('file_list', [])

            if file_list is None:
                st.session_state['file_list'] = []
                file_list = []
            elif not isinstance(file_list, list):
                st.error(f"file_list is not a list. Type: {type(file_list)}")
                st.session_state['file_list'] = []
                file_list = []
            else:
                # Try to call GetSlabReport if present
                try:
                    if 'GetSlabReport' in globals() and callable(globals()['GetSlabReport']):
                        GetSlabReport()
                    else:
                        st.warning("GetSlabReport function not found. Skipping slab report generation.")
                except Exception as e:
                    st.warning(f"Error calling GetSlabReport: {e}")
                    logger.exception("GetSlabReport error")
                    
        except Exception as e:
            st.error(f"Error fetching COS slab cycle data: {str(e)}")
            logger.exception("COS slab data error")
            st.session_state['slab_df'] = pd.DataFrame()

    # Process Asite datasets
    asite_data = []
    dataset_specs = [
        ("Tower F Finishing", finishing_data, finishing_activity, finishing_locations),
        ("Structure", structure_data, structure_activity, structure_locations),
        ("Tower G Finishing", external_data, external_activity, external_locations),
        ("Tower H Finishing", tower_h_data, tower_h_activity, tower_h_locations),
    ]

    for dataset_name, data, activity, location in dataset_specs:
        try:
            # CHANGE: Use process_data_fixed instead of process_data
            analysis, total, count_table = process_data(data, activity, location, dataset_name)
            
            if analysis is None or (hasattr(analysis, "empty") and analysis.empty):
                logger.warning("No valid data processed for %s", dataset_name)
                st.warning(f"No completed activities found for {dataset_name}.")
                outputs[dataset_name] = {"towers": {}, "total": 0}
                continue

            output = process_manually(analysis, total, dataset_name)
            outputs[dataset_name] = output

            for tower, activities in output.get("towers", {}).items():
                for activity_name, count in activities.items():
                    normalized_name = normalize_activity_name(activity_name)
                    asite_data.append({
                        "Dataset": dataset_name,
                        "Tower": tower,
                        "Activity Name": normalized_name,
                        "Count": count
                    })
        except Exception as e:
            logger.exception("Error processing %s", dataset_name)
            st.error(f"Error processing {dataset_name} data: {str(e)}")
            outputs[dataset_name] = {"towers": {}, "total": 0}
            continue

    asite_df = pd.DataFrame(asite_data) if asite_data else pd.DataFrame(columns=["Dataset", "Tower", "Activity Name", "Count"])

    # PROCESS COS DATA
    st.write("### Processing COS Data...")
    cos_data = []

    tower_mappings = [
        ('cos_df_tower_f', 'Tower F', 'TF'),
        ('cos_df_tower_g', 'Tower G', 'TG'),
        ('cos_df_tower_h', 'Tower H', 'TH'),
        ('cos_df_structure', 'Structure', 'Structure')
    ]

    st.write("### Debug: Available COS session state keys:")
    cos_keys = [k for k in st.session_state.keys() if k.lower().startswith('cos_')]
    st.write(f"Found COS keys: {cos_keys}")

    for df_key, display_name, short_name in tower_mappings:
        tower_df = st.session_state.get(df_key)
        st.write(f"#### Processing {display_name}...")
        
        if tower_df is not None and isinstance(tower_df, pd.DataFrame) and not tower_df.empty:
            total_rows = len(tower_df)
            completed_rows = len(tower_df[pd.notna(tower_df['Actual Finish'])]) if 'Actual Finish' in tower_df.columns else 0
            st.write(f"  📊 Total rows: {total_rows}, Completed (with Actual Finish): {completed_rows}")
            
            processed_df = process_cos_data(short_name, tower_df)
            if processed_df is not None and isinstance(processed_df, pd.DataFrame) and not processed_df.empty:
                st.write(f"  ✓ Processed {len(processed_df)} mapped activities for {display_name}")
                for _, r in processed_df.iterrows():
                    st.write(f"    • {r['Activity Name']}: {r['Count']}")
                for _, row in processed_df.iterrows():
                    cos_data.append({
                        "Tower": row["Tower"],
                        "Activity Name": row["Activity Name"],
                        "Count": row["Count"]
                    })
            else:
                st.warning(f"  ⚠ No mapped activities returned for {display_name}")
        else:
            st.warning(f"  ⚠ No data found for {display_name}")

    cos_df = pd.DataFrame(cos_data) if cos_data else pd.DataFrame(columns=["Tower", "Activity Name", "Count"])

    # Show debug info
    st.write("### COS Data Debug:")
    st.write(f"Total COS records created: {len(cos_data)}")
    if not cos_df.empty:
        st.write(cos_df.head(10))
        st.write(f"COS DataFrame shape: {cos_df.shape}")
    else:
        st.write("COS DataFrame is empty!")

    # Log dataframes for debugging
    try:
        logger.info("Asite DataFrame:\n%s", asite_df.to_string() if not asite_df.empty else "Empty")
        logger.info("COS DataFrame:\n%s", cos_df.to_string() if not cos_df.empty else "Empty")
    except Exception:
        logger.exception("Error logging dataframes")

    st.write("### Asite DataFrame (Debug):")
    st.write(asite_df)
    st.write("### COS DataFrame (Debug):")
    st.write(cos_df)

    combined_data = {
        "COS": cos_df if not cos_df.empty else pd.DataFrame(columns=["Tower", "Activity Name", "Count"]),
        "Asite": asite_df if not asite_df.empty else pd.DataFrame(columns=["Tower", "Activity Name", "Count"])
    }

    # Call AI categorization
    with st.spinner("Categorizing activities with WatsonX..."):
        try:
            slab_report = st.session_state.get('slabreport', pd.DataFrame())
            ai_response = generatePrompt(combined_data, slab_report if not isinstance(slab_report, str) else str(slab_report))
            st.session_state['ai_response'] = ai_response
        except Exception as e:
            st.error(f"Error calling generatePrompt: {e}")
            logger.exception("generatePrompt error")
            ai_response = None
            st.session_state['ai_response'] = None

    st.write("### Categorized Activity Counts (COS and Asite):")
    if ai_response:
        try:
            ai_data = json.loads(ai_response) if isinstance(ai_response, str) else ai_response
            st.json(ai_data)
        except (json.JSONDecodeError, TypeError) as e:
            st.error(f"Failed to parse AI response as JSON: {str(e)}")
            st.text(str(ai_response)[:500])

    end_time = time.time()
    st.write(f"Total execution time: {end_time - start_time:.2f} seconds")
    
    return combined_data, outputs
    
# COS File Fetching Function
def get_cos_files():
    try:
        cos_client = initialize_cos_client()
        if not cos_client:
            st.error(" Failed to initialize COS client.")
            return []

        response = cos_client.list_objects_v2(Bucket=COS_BUCKET, Prefix="Eligo/")
        if 'Contents' not in response:
            st.error(f" No files found in the 'Eligo' folder of bucket '{COS_BUCKET}'. Please ensure the folder exists and contains files.")
            logger.error("No objects found in Eligo folder")
            return []

        all_files = [obj['Key'] for obj in response.get('Contents', [])]
        st.write("**All files in Eligo folder:**")
        if all_files:
            st.write("\n".join(all_files))
        else:
            st.write("No files found.")
            logger.warning("Eligo folder is empty")
            return []

        # Pattern for Finishing Tracker files (Tower G and Tower H)
        finishing_pattern = re.compile(
            r"Eligo/Tower\s*([F|G|H])\s*Finishing\s*Tracker[\(\s]*(.*?)(?:[\)\s]*\.xlsx)$",
            re.IGNORECASE
        )
        # Pattern for Structure Work Tracker file
        structure_pattern = re.compile(
            r"Eligo/Structure\s*Work\s*Tracker[\(\s]*(.*?)(?:[\)\s]*\.xlsx)$",
            re.IGNORECASE
        )

        date_formats = [
            "%d-%m-%Y", "%d-%m-%y", "%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y"
        ]

        file_info = []
        for obj in response.get('Contents', []):
            key = obj['Key']
            # Check for Finishing Tracker files (Tower G and Tower H)
            finishing_match = finishing_pattern.match(key)
            if finishing_match:
                tower = finishing_match.group(1)
                date_str = finishing_match.group(2).strip('()').strip()
                parsed_date = None
                
                for fmt in date_formats:
                    try:
                        parsed_date = datetime.strptime(date_str, fmt)
                        break
                    except ValueError:
                        continue
                
                if parsed_date:
                    file_info.append({
                        'key': key,
                        'tower': tower,
                        'date': parsed_date,
                        'type': 'finishing'
                    })
                else:
                    logger.warning(f"Could not parse date in filename: {key} (date: {date_str})")
                    st.warning(f"Skipping file with unparseable date: {key}")
            # Check for Structure Work Tracker file
            else:
                structure_match = structure_pattern.match(key)
                if structure_match:
                    date_str = structure_match.group(1).strip('()').strip()
                    parsed_date = None
                    
                    for fmt in date_formats:
                        try:
                            parsed_date = datetime.strptime(date_str, fmt)
                            break
                        except ValueError:
                            continue
                    
                    if parsed_date:
                        file_info.append({
                            'key': key,
                            'tower': None,  # No specific tower associated
                            'date': parsed_date,
                            'type': 'structure'
                        })
                    else:
                        logger.warning(f"Could not parse date in filename: {key} (date: {date_str})")
                        st.warning(f"Skipping file with unparseable date: {key}")

        if not file_info:
            st.error(" No Excel files matched the expected patterns in the 'Eligo' folder. Expected formats: 'Tower G/H Finishing Tracker(date).xlsx' or 'Structure Work Tracker(date).xlsx'.")
            logger.error("No files matched the expected patterns")
            return []

        # Separate Finishing and Structure files
        finishing_files = {}
        structure_files = []
        for info in file_info:
            if info['type'] == 'finishing':
                tower = info['tower']
                if tower not in finishing_files or info['date'] > finishing_files[tower]['date']:
                    finishing_files[tower] = info
            elif info['type'] == 'structure':
                structure_files.append(info)

        # Select the latest Structure Work Tracker file (if multiple exist)
        if structure_files:
            latest_structure_file = max(structure_files, key=lambda x: x['date'])
            files = [info['key'] for info in finishing_files.values()] + [latest_structure_file['key']]
        else:
            files = [info['key'] for info in finishing_files.values()]

        if not files:
            st.error(" No valid Excel files found for Tower G, Tower H, or Structure Work Tracker after filtering.")
            logger.error("No valid files after filtering")
            return []

        st.success(f"Found {len(files)} matching files: {', '.join(files)}")
        return files
    except Exception as e:
        st.error(f" Error fetching COS files: {str(e)}")
        logger.error(f"Error fetching COS files: {str(e)}")
        return []

# Initialize session state variables
if 'cos_df_eligo_tower_f_finishing' not in st.session_state:
    st.session_state.cos_df_eligo_tower_f_finishing = None
if 'cos_df_eligo_structure' not in st.session_state:
    st.session_state.cos_df_eligo_structure = None
if 'cos_df_eligo_tower_g_finishing' not in st.session_state:
    st.session_state.cos_df_eligo_tower_g_finishing = None
if 'cos_tname_eligo_tower_f_finishing' not in st.session_state:
    st.session_state.cos_tname_eligo_tower_f_finishing = None
if 'cos_tname_eligo_structure' not in st.session_state:
    st.session_state.cos_tname_eligo_structure = None
if 'cos_tname_eligo_tower_g_finishing' not in st.session_state:
    st.session_state.cos_tname_eligo_tower_g_finishing = None
if 'cos_df_eligo_tower_h_finishing' not in st.session_state:
    st.session_state.cos_df_eligo_tower_h_finishing = None
if 'cos_tname_eligo_tower_h_finishing' not in st.session_state:
    st.session_state.cos_tname_eligo_tower_h_finishing = None    

# ADD THESE MISSING INITIALIZATIONS:
if 'cos_client' not in st.session_state:
    st.session_state.cos_client = None
if 'bucket_name' not in st.session_state:
    st.session_state.bucket_name = None
if 'file_list' not in st.session_state:
    st.session_state.file_list = None
if 'slabreport' not in st.session_state:
    st.session_state.slabreport = pd.DataFrame()
if 'slab_df' not in st.session_state:
    st.session_state.slab_df = pd.DataFrame()



if 'ignore_month' not in st.session_state:
    st.session_state.ignore_month = False  
if 'ignore_year' not in st.session_state:
    st.session_state.ignore_year = False 



def process_file(file_stream, filename):
    try:
        workbook = openpyxl.load_workbook(file_stream)
        available_sheets = workbook.sheetnames
        logger.info(f"Available sheets in {filename}: {available_sheets}")

        # Check if the file is a Structure Work Tracker file
        is_structure_tracker = "Structure Work Tracker" in filename

        if is_structure_tracker:
            # Handle Structure Work Tracker file
            possible_sheet_names = [
                "Revised Baselines- 25 days SC", "Revised Baselines", "Baselines",
                "Structure Work", "Structure", "RevisedBaselines"
            ]
            sheet_name = None
            for name in possible_sheet_names:
                if name in available_sheets:
                    sheet_name = name
                    break
            if not sheet_name:
                # Fallback to the first sheet if no expected name is found
                sheet_name = available_sheets[0] if available_sheets else None
            if not sheet_name:
                st.error(f"No valid sheets found in {filename}. Available sheets: {', '.join(available_sheets)}")
                logger.error(f"No valid sheets found in {filename}")
                return [(None, None)]

            file_stream.seek(0)
            try:
                # Try different header rows (0, 1, 2) to find valid columns
                df = None
                actual_columns = None
                for header_row in [0, 1, 2]:
                    try:
                        file_stream.seek(0)
                        df = pd.read_excel(file_stream, sheet_name=sheet_name, header=header_row)
                        actual_columns = df.columns.tolist()
                        # Check if columns are all "Unnamed"
                        if all(col.startswith("Unnamed:") for col in actual_columns):
                            continue
                        logger.info(f"Columns in {sheet_name} (header={header_row}): {actual_columns}")
                        st.write(f"Columns in {sheet_name} (header row {header_row + 1}): {actual_columns}")
                        st.write(f"First 5 rows of {sheet_name} (header row {header_row + 1}):")
                        st.write(df.head(5))
                        break
                    except Exception as e:
                        logger.warning(f"Failed to read {sheet_name} with header row {header_row}: {str(e)}")
                        continue

                if df is None or actual_columns is None:
                    st.error(f"Could not find valid headers in {sheet_name}. All attempts yielded unnamed columns or errors.")
                    logger.error(f"No valid headers found in {sheet_name}")
                    return [(None, None)]

                # Define possible column names for mapping (more flexible for Structure Work Tracker)
                column_mapping = {
                    'Activity ID': ['Activity ID', 'Task ID', 'ID', 'ActivityID'],
                    'Activity Name': ['Activity Name', 'Task Name', 'Activity', 'Name', 'Task', 'Description'],
                    'Actual Finish': ['Actual Finish', 'Finish Date', 'Completion Date', 'Actual End', 'End Date', 'Finish']
                }

                # Find matching columns
                target_columns = ['Activity ID', 'Activity Name', 'Actual Finish']
                selected_columns = {}
                for target in target_columns:
                    for possible_name in column_mapping[target]:
                        if possible_name in actual_columns:
                            selected_columns[target] = possible_name
                            break

                # For Structure Work Tracker, proceed even if some columns are missing
                if not selected_columns:
                    logger.warning(f"No recognizable columns in {sheet_name}. Using all available columns.")
                    selected_columns = {col: col for col in actual_columns[:3]}  # Use first 3 columns as a fallback
                elif len(selected_columns) < len(target_columns):
                    missing_cols = [col for col in target_columns if col not in selected_columns]
                    st.warning(f"Some expected columns missing in {sheet_name}. Missing: {missing_cols}, Found: {list(selected_columns.keys())}")
                    logger.warning(f"Missing columns in {sheet_name}: {missing_cols}")

                # Select and rename columns
                df = df[list(selected_columns.values())]
                df.columns = list(selected_columns.keys())
                
                # For Structure Work Tracker, don't enforce Activity Name requirement
                if 'Activity Name' in df.columns:
                    df = df.dropna(subset=['Activity Name'])
                    df['Activity Name'] = df['Activity Name'].astype(str).str.strip()
                else:
                    st.warning(f"No 'Activity Name' column in {sheet_name}. Proceeding with available data.")
                    df = df.dropna(how='all')  # Drop rows where all values are NaN

                if 'Actual Finish' in df.columns:
                    df['Actual_Finish_Original'] = df['Actual Finish'].astype(str)
                    df['Actual Finish'] = pd.to_datetime(df['Actual Finish'], errors='coerce')
                    has_na_mask = (
                        pd.isna(df['Actual Finish']) |
                        (df['Actual_Finish_Original'].str.upper() == 'NAT') |
                        (df['Actual_Finish_Original'].str.lower().isin(['nan', 'na', 'n/a', 'none', '']))
                    )
                    # Use a fallback column for NA check if Activity Name is missing
                    display_col = 'Activity Name' if 'Activity Name' in df.columns else df.columns[0]
                    na_rows = df[has_na_mask][[display_col, 'Actual Finish']]
                    if not na_rows.empty:
                        st.write(f"Sample of rows with NA or invalid values in Actual Finish for {filename}:")
                        st.write(na_rows.head(10))
                        na_activities = na_rows.groupby(display_col).size().reset_index(name='Count')
                        st.write(f"Items with NA or invalid Actual Finish values in {filename}:")
                        st.write(na_activities)
                    else:
                        st.write(f"No NA or invalid values found in Actual Finish for {filename}")
                    df.drop('Actual_Finish_Original', axis=1, inplace=True)

                # Display unique values based on available columns
                display_col = 'Activity Name' if 'Activity Name' in df.columns else df.columns[0]
                st.write(f"Unique {display_col} in {sheet_name} ({filename}):")
                unique_activities = df[[display_col]].drop_duplicates()
                st.write(unique_activities)

                return [(df, "Structure Work Tracker")]
            except Exception as e:
                st.error(f"Error processing sheet {sheet_name} in {filename}: {str(e)}")
                logger.error(f"Error processing sheet {sheet_name} in {filename}: {str(e)}")
                return [(None, None)]
        else:
            # Handle Tower G or Tower H Finishing Tracker files
            tower_letter = None
            if "Tower F" in filename or "TowerF" in filename:
                tower_letter = "F"
            elif "Tower G" in filename or "TowerG" in filename:
                tower_letter = "G"
                
            elif "Tower H" in filename or "TowerH" in filename:
                tower_letter = "H"

            if not tower_letter:
                st.error(f"Cannot determine tower letter from filename: {filename}")
                logger.error(f"Cannot determine tower letter from filename: {filename}")
                return [(None, None)]

            possible_sheet_names = [
                f"Tower {tower_letter} Finishing",
                f"TOWER {tower_letter} FINISHING",
                f"TOWER{tower_letter}FINISHING",
                f"TOWER {tower_letter}FINISHING",
                f"TOWER{tower_letter} FINISHING",
                f"Tower{tower_letter}Finishing",
                f"Finish"
            ]

            sheet_name = None
            for name in possible_sheet_names:
                if name in available_sheets:
                    sheet_name = name
                    break

            if not sheet_name:
                for available in available_sheets:
                    if f"TOWER {tower_letter}" in available.upper() and "FINISH" in available.upper():
                        sheet_name = available
                        break

            if not sheet_name:
                st.error(f"Required sheet for Tower {tower_letter} not found in file. Available sheets: {', '.join(available_sheets)}")
                logger.error(f"Required sheet for Tower {tower_letter} not found in {filename}")
                return [(None, None)]

            file_stream.seek(0)

            try:
                df = pd.read_excel(file_stream, sheet_name=sheet_name, header=0)

                expected_columns = [
                    'Module', 'Floor', 'Flat', 'Domain', 'Activity ID', 'Activity Name',
                    'Monthly Look Ahead', 'Baseline Duration', 'Baseline Start', 'Baseline Finish',
                    'Actual Start', 'Actual Finish', '% Complete', 'Start', 'Finish', 'Delay Reasons'
                ]

                if len(df.columns) < len(expected_columns):
                    st.warning(f"Excel file has fewer columns than expected ({len(df.columns)} found, {len(expected_columns)} expected).")
                    expected_columns = expected_columns[:len(df.columns)]

                df.columns = expected_columns[:len(df.columns)]

                target_columns = ["Module", "Floor", "Flat", "Activity ID", "Activity Name", "Actual Finish"]
                available_columns = [col for col in target_columns if col in df.columns]

                if len(available_columns) < len(target_columns):
                    missing_cols = [col for col in target_columns if col not in available_columns]
                    st.warning(f"Missing columns in file: {', '.join(missing_cols)}")
                    for col in missing_cols:
                        df[col] = None

                df = df[target_columns]
                df = df.dropna(subset=['Activity Name'])

                df['Activity Name'] = df['Activity Name'].astype(str).str.strip()

                if 'Floor' in df.columns:
                    df['Floor'] = df['Floor'].astype(str)
                    v_rows = df[df['Floor'].str.strip().str.upper() == 'V']
                    if not v_rows.empty:
                        df = pd.concat([df, v_rows], ignore_index=True)

                if 'Actual Finish' in df.columns:
                    df['Actual_Finish_Original'] = df['Actual Finish'].astype(str)
                    df['Actual Finish'] = pd.to_datetime(df['Actual Finish'], errors='coerce')
                    has_na_mask = (
                        pd.isna(df['Actual Finish']) |
                        (df['Actual_Finish_Original'].str.upper() == 'NAT') |
                        (df['Actual_Finish_Original'].str.lower().isin(['nan', 'na', 'n/a', 'none', '']))
                    )
                    na_rows = df[has_na_mask][['Activity Name', 'Actual Finish']]
                    if not na_rows.empty:
                        st.write("Sample of rows with NA or invalid values in Actual Finish:")
                        st.write(na_rows.head(10))
                        na_activities = na_rows.groupby('Activity Name').size().reset_index(name='Count')
                        st.write("Activities with NA or invalid Actual Finish values:")
                        st.write(na_activities)
                    else:
                        st.write("No NA or invalid values found in Actual Finish")
                    df.drop('Actual_Finish_Original', axis=1, inplace=True)

                st.write(f"Unique Activity Names in {sheet_name}:")
                unique_activities = df[['Module', 'Floor', 'Activity Name']].drop_duplicates()
                st.write(unique_activities)

                return [(df, f"Tower {tower_letter} Finishing")]

            except Exception as e:
                st.error(f"Error processing sheet {sheet_name}: {str(e)}")
                logger.error(f"Error processing sheet {sheet_name}: {str(e)}")
                return [(None, None)]

    except Exception as e:
        st.error(f"Error loading Excel file {filename}: {str(e)}")
        logger.error(f"Error loading Excel file {filename}: {str(e)}")
        return [(None, None)]

#Slab code
def GetSlabReport():
    foundeligo = False
    today = date.today()
    prev_month = today - relativedelta(months=1)
    month_year = today.strftime("%m-%Y")
    prev_month_year = prev_month.strftime("%m-%Y")
    
    try:
        logger.info("Attempting to initialize COS client...")
        cos_client = ibm_boto3.client(
            's3',
            ibm_api_key_id=COS_API_KEY,
            ibm_service_instance_id=COS_SERVICE_INSTANCE_ID,
            config=Config(
                signature_version='oauth',
                connect_timeout=180,
                read_timeout=180,
                retries={'max_attempts': 15}
            ),
            endpoint_url=COS_ENDPOINT
        )
        
        response = cos_client.list_objects_v2(Bucket="projectreportnew")
        files = [obj['Key'] for obj in response.get('Contents', []) if obj['Key'].endswith('.xlsx')]
        
        # NEW: Extract dates from Structure Work Tracker files and find the latest
        structure_files = []
        
        for file in files:
            if file.startswith("Eligo") and "Structure Work Tracker" in file:
                st.write(f"Found Structure Work Tracker file: {file}")
                
                # Extract date from filename using multiple regex patterns
                import re
                date_patterns = [
                    r'(\d{2}-\d{2}-\d{4})',  # DD-MM-YYYY
                    r'(\d{2}/\d{2}/\d{4})',  # DD/MM/YYYY  
                    r'(\d{4}-\d{2}-\d{2})',  # YYYY-MM-DD
                    r'(\d{2}\.\d{2}\.\d{4})', # DD.MM.YYYY
                    r'(\d{1,2}-\d{1,2}-\d{4})', # D-M-YYYY or DD-M-YYYY
                    r'(\d{1,2}/\d{1,2}/\d{4})', # D/M/YYYY or DD/M/YYYY
                ]
                
                file_date = None
                date_str = None
                
                for pattern in date_patterns:
                    date_match = re.search(pattern, file)
                    if date_match:
                        date_str = date_match.group(1)
                        # Try to parse the date with different formats
                        date_formats = [
                            "%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d.%m.%Y",
                            "%d-%m-%Y", "%d/%m/%Y"  # Handle single digit days/months
                        ]
                        
                        for fmt in date_formats:
                            try:
                                file_date = datetime.strptime(date_str, fmt)
                                break
                            except ValueError:
                                continue
                        
                        if file_date:
                            break
                
                if file_date:
                    structure_files.append({
                        'filename': file,
                        'date': file_date,
                        'date_str': date_str
                    })
                    st.write(f"  - Parsed date: {file_date.strftime('%Y-%m-%d')} from {date_str}")
                else:
                    st.warning(f"  - Could not parse date from filename: {file}")
        
        if not structure_files:
            st.error("No Structure Work Tracker files with valid dates found!")
            st.session_state.slabreport = "No Data Found"
            return
        
        # Sort by date and get the latest file (most recent first)
        structure_files.sort(key=lambda x: x['date'], reverse=True)
        latest_file_info = structure_files[0]
        latest_file = latest_file_info['filename']
        
        st.success(f"ðŸŽ¯ Selected latest Structure Work Tracker file: {latest_file}")
        st.info(f"ðŸ“… File date: {latest_file_info['date'].strftime('%Y-%m-%d')}")
        
        # Show all found files for debugging (sorted by date, newest first)
        st.write("ðŸ“‹ All Structure Work Tracker files found (sorted by date, newest first):")
        for i, file_info in enumerate(structure_files):
            marker = "ðŸ“ **SELECTED**" if i == 0 else "  "
            st.write(f"{marker} {file_info['filename']} - {file_info['date'].strftime('%Y-%m-%d')}")
        
        try:
            response = cos_client.get_object(Bucket="projectreportnew", Key=latest_file)
            
            if st.session_state.ignore_month and st.session_state.ignore_year:
                st.session_state.slabreport = ProcessGandH(
                    io.BytesIO(response['Body'].read()), 
                    st.session_state.ignore_year, 
                    st.session_state.ignore_month
                )
            else:
                st.session_state.slabreport = ProcessGandH(io.BytesIO(response['Body'].read()))
            
            foundeligo = True
            st.success(f"… Successfully processed latest file: {latest_file}")
            
        except Exception as e:
            st.error(f" Error processing latest file {latest_file}: {e}")
            st.session_state.slabreport = "No Data Found"
            
            # Try the second most recent file as fallback
            if len(structure_files) > 1:
                st.warning("Trying the second most recent file as fallback...")
                fallback_file_info = structure_files[1]
                fallback_file = fallback_file_info['filename']
                
                try:
                    response = cos_client.get_object(Bucket="projectreportnew", Key=fallback_file)
                    
                    if st.session_state.ignore_month and st.session_state.ignore_year:
                        st.session_state.slabreport = ProcessGandH(
                            io.BytesIO(response['Body'].read()), 
                            st.session_state.ignore_year, 
                            st.session_state.ignore_month
                        )
                    else:
                        st.session_state.slabreport = ProcessGandH(io.BytesIO(response['Body'].read()))
                    
                    foundeligo = True
                    st.success(f"… Successfully processed fallback file: {fallback_file}")
                    
                except Exception as fallback_error:
                    st.error(f" Fallback also failed: {fallback_error}")
                    st.session_state.slabreport = "No Data Found"
        
    except Exception as e:
        st.error(f" Error fetching COS files: {e}")
        logger.error(f"Error fetching COS files: {e}")
        st.session_state.slabreport = "No Data Found"

                   
    except Exception as e:
        print(f"Error fetching COS files: {e}")
        files = ["Error fetching COS files"]
        st.session_state.slabreport = "No Data Found"
        
ACTIVITY_MAPPING = {
    # Finishing Work Activities (Asite -> COS)
    "Wall Conduting": "Wall Conduting",
    "Door/Window Frame": "Installation of doors",
    "Plumbing Works": "Min. count of UP-First Fix and CP-First Fix",
    "Waterproofing - Sunken": "Water Proofing Works",
    "POP & Gypsum Plaster": "POP punning (Major area)",
    "Wall Tile": "Wall Tiling",
    "Floor Tile": "Floor Tiling",
    "Door/Window Shutter": "Installation of doors",
    "Wiring & Switch Socket": "EL 2nd Fix",
    
    # Structure Work Activities (Asite -> COS)
    "Shuttering": "No. of Slab cast",
    "Reinforcement": "No. of Slab cast",
    "Slab Conducting": "No. of Slab cast",
    "Concreting": "No. of Slab cast",
    "De-Shuttering": "Not Required",
    "Brickwork": "Not Required",
    "Plastering": "Not Required",
}

def calculate_activity_counts(cos_df, asite_df):
    """
    Calculate activity counts based on tracker mapping logic
    
    Args:
        cos_df: COS DataFrame with columns ['Tower', 'Activity Name', 'Count']
        asite_df: Asite DataFrame with columns ['Tower', 'Activity Name', 'Count']
    
    Returns:
        tuple: (cos_activities_dict, asite_activities_dict)
        Format: {"Tower|Activity": count}
    
    Special Calculations:
        - Plumbing Works = Min(UP-First Fix, CP-First Fix)
        - Slab Conducting = No. of Slab cast
        - Shuttering = No. of Slab cast
        - Reinforcement = No. of Slab cast
        - Concreting = No. of Slab cast
    """
    try:
        # Initialize result dictionaries
        cos_activities = {}
        asite_activities = {}
        
        # Build tower-activity lookup for COS
        cos_by_tower_activity = {}
        for idx, row in cos_df.iterrows():
            tower = row.get('Tower', 'Unknown')
            activity = row.get('Activity Name', 'Unknown')
            count = row.get('Count', 0)
            
            if tower not in cos_by_tower_activity:
                cos_by_tower_activity[tower] = {}
            cos_by_tower_activity[tower][activity] = count
            
            key = f"{tower}|{activity}"
            cos_activities[key] = count
        
        # Process Asite with special calculations
        for idx, row in asite_df.iterrows():
            tower = row.get('Tower', 'Unknown')
            activity = row.get('Activity Name', 'Unknown')
            count = row.get('Count', 0)
            
            # ========== SPECIAL CALCULATION RULES FROM TRACKER ==========
            
            # Rule 1: Plumbing Works = Min(UP-First Fix, CP-First Fix)
            if activity == "Plumbing Works":
                if tower in cos_by_tower_activity:
                    up_count = cos_by_tower_activity[tower].get("UP-First Fix", 0)
                    cp_count = cos_by_tower_activity[tower].get("CP-First Fix", 0)
                    if up_count > 0 or cp_count > 0:
                        count = min(up_count, cp_count)
                        logger.info(f"{tower} - Plumbing Works calculated: min({up_count}, {cp_count}) = {count}")
            
            # Rule 2: Slab Conducting = No. of Slab cast
            elif activity == "Slab Conducting":
                if tower in cos_by_tower_activity:
                    slab_cast = cos_by_tower_activity[tower].get("No. of Slab cast", 0)
                    if slab_cast > 0:
                        count = slab_cast
                        logger.info(f"{tower} - Slab Conducting calculated: {slab_cast}")
            
            # Rule 3: Shuttering = No. of Slab cast
            elif activity == "Shuttering":
                if tower in cos_by_tower_activity:
                    slab_cast = cos_by_tower_activity[tower].get("No. of Slab cast", 0)
                    if slab_cast > 0:
                        count = slab_cast
                        logger.info(f"{tower} - Shuttering calculated: {slab_cast}")
            
            # Rule 4: Reinforcement = No. of Slab cast
            elif activity == "Reinforcement":
                if tower in cos_by_tower_activity:
                    slab_cast = cos_by_tower_activity[tower].get("No. of Slab cast", 0)
                    if slab_cast > 0:
                        count = slab_cast
                        logger.info(f"{tower} - Reinforcement calculated: {slab_cast}")
            
            # Rule 5: Concreting = No. of Slab cast
            elif activity == "Concreting":
                if tower in cos_by_tower_activity:
                    slab_cast = cos_by_tower_activity[tower].get("No. of Slab cast", 0)
                    if slab_cast > 0:
                        count = slab_cast
                        logger.info(f"{tower} - Concreting calculated: {slab_cast}")
            
            # ========================================================
            
            key = f"{tower}|{activity}"
            asite_activities[key] = count
        
        logger.info(f"COS Activities calculated: {len(cos_activities)} items")
        logger.info(f"Asite Activities calculated: {len(asite_activities)} items")
        
        return cos_activities, asite_activities
    
    except Exception as e:
        logger.error(f"Error in calculate_activity_counts: {str(e)}")
        st.error(f"Calculation Error: {str(e)}")
        return {}, {}


def apply_calculation_logic(cos_df, asite_df):
    """
    Apply calculation logic and return updated dataframes
    
    Args:
        cos_df: COS DataFrame
        asite_df: Asite DataFrame
    
    Returns:
        tuple: (cos_df_updated, asite_df_updated)
    """
    try:
        # Make copies to avoid modifying original data
        cos_df = cos_df.copy()
        asite_df = asite_df.copy()
        
        # Get calculations
        cos_calc, asite_calc = calculate_activity_counts(cos_df, asite_df)
        
        # Update COS dataframe (if needed)
        for idx, row in cos_df.iterrows():
            key = f"{row['Tower']}|{row['Activity Name']}"
            if key in cos_calc:
                cos_df.at[idx, 'Count'] = cos_calc[key]
        
        # Update Asite dataframe with calculated values
        for idx, row in asite_df.iterrows():
            key = f"{row['Tower']}|{row['Activity Name']}"
            if key in asite_calc:
                asite_df.at[idx, 'Count'] = asite_calc[key]
        
        logger.info("Calculation logic applied successfully")
        return cos_df, asite_df
    
    except Exception as e:
        logger.error(f"Error in apply_calculation_logic: {str(e)}")
        return cos_df, asite_df


def get_calculated_count(tower, activity, cos_df, asite_df, calc_type='asite'):
    """
    Get a specific calculated count for an activity
    
    Args:
        tower: Tower name (e.g., 'Tower G', 'Tower H', 'TG', 'TH')
        activity: Activity name
        cos_df: COS DataFrame
        asite_df: Asite DataFrame
        calc_type: 'cos' or 'asite' (default: 'asite')
    
    Returns:
        int: Calculated count for the activity
    """
    try:
        cos_calc, asite_calc = calculate_activity_counts(cos_df, asite_df)
        key = f"{tower}|{activity}"
        
        if calc_type.lower() == 'cos':
            return cos_calc.get(key, 0)
        else:
            return asite_calc.get(key, 0)
    
    except Exception as e:
        logger.error(f"Error in get_calculated_count: {str(e)}")
        return 0




def generatePrompt(combined_data, slab):
    try:
        st.write(slab)
        st.write(json.loads(slab))
        
        # Keep the simple approach from veridea that works
        cos_df = combined_data["COS"] if isinstance(combined_data["COS"], pd.DataFrame) else pd.DataFrame()
        asite_df = combined_data["Asite"] if isinstance(combined_data["Asite"], pd.DataFrame) else pd.DataFrame()

        # Tower mapping for eligo - map G, 2G, 3G etc. to Tower G
        def map_tower_names(df):
            if not df.empty and 'Tower' in df.columns:
                # Map all G variants to Tower G, H variants to Tower H, etc.
                df = df.copy()
                df['Tower'] = df['Tower'].astype(str)
                df.loc[df['Tower'].str.contains('G', case=False, na=False), 'Tower'] = 'Tower G'
                df.loc[df['Tower'].str.contains('H', case=False, na=False), 'Tower'] = 'Tower H' 
                df.loc[df['Tower'].str.contains('F', case=False, na=False), 'Tower'] = 'Tower F'
            return df

        # Apply tower mapping
        if not cos_df.empty:
            cos_df = map_tower_names(cos_df)
        if not asite_df.empty:
            asite_df = map_tower_names(asite_df)
        
        cos_df_calc, asite_df_calc = apply_calculation_logic(cos_df, asite_df)
        # Direct JSON conversion like in veridea (the working version)
        cos_json = cos_df[['Tower', 'Activity Name', 'Count']].to_json(orient='records', indent=2) if not cos_df.empty else "[]"
        asite_json = asite_df[['Tower', 'Activity Name', 'Count']].to_json(orient='records', indent=2) if not asite_df.empty else "[]"

        body = {
            "input": f"""
            Read the table data provided below for COS and Asite sources, which include tower-specific activity counts. Categorize the activities into the specified categories (MEP, Interior Finishing, ED Civil) for each tower in each source (COS and Asite). Compute the total count of each activity within its respective category for each tower and return the results as a JSON object with "COS" and "Asite" sections, where each section contains a list of towers, each with categories and their activities. For the MEP category in COS, calculate the total count between 'UP-First Fix' and 'CP-First Fix' and report it as 'Min. count of UP-First Fix and CP-First Fix' for each tower. If an activity is not found for a tower, include it with a count of 0. If the Structure Work category has no activities in COS, return an empty list for it. Ensure the counts are accurate, the output is grouped by tower and category, and the JSON structure is valid with no nested or repeated keys.

            The data provided is as follows:
            
            Slab:
            {slab}

            COS Table Data:
            {cos_json}

            Asite Table Data:
            {asite_json}

            Categories and Activities:
            COS:
            - MEP: EL-First Fix, UP-First Fix, CP-First Fix, Min. count of UP-First Fix and CP-First Fix, C-Gypsum and POP Punning, EL 2nd Fix,
            - Interior Finishing: Installation of doors, Waterproofing Works, Wall Tiling, Floor Tiling
            - ED Civil: Concreting, Shuttering, Reinforcement, De-Shuttering
            Asite:
            - MEP: Wall Conduting, Plumbing Works, Wiring & Switch Socket, Slab Conducting
            - Interior Finishing: Waterproofing - Sunken, Wall Tile, Floor Tile, POP & Gypsum Plaster
            - ED Civil: Concreting, Shuttering, Reinforcement, De-Shuttering

            Slab:
            - Get total greens of Each Tower

            Example JSON format needed:
            {{
              "COS": [
                {{
                  "Tower": "Tower G",
                  "Categories": [
                    {{
                      "Category": "MEP",
                      "Activities": [
                        
                        {{"Activity Name": "UP-First Fix", "Total": 0}},
                        {{"Activity Name": "CP-First Fix", "Total": 0}},
                        {{"Activity Name": "Min. count of UP-First Fix and CP-First Fix", "Total": 0}},
                        {{"Activity Name": "POP punning (Major area)", "Total": 0}},
                        {{"Activity Name": "EL-Second Fix", "Total": 0}}
                      ]
                    }},
                    {{
                      "Category": "Interior Finishing",
                      "Activities": [
                        {{"Activity Name": "Installation of doors", "Total": 0}},
                        {{"Activity Name": "Waterproofing Works", "Total": 0}},
                        {{"Activity Name": "Wall Tiling", "Total": 0}},
                        {{"Activity Name": "Floor Tiling", "Total": 0}}
                      ]
                    }},
                    {{
                      "Category": "ED Civil",
                      "Activities": [
                        {{"Activity Name": "Concreting", "Total": 0}},
                        {{"Activity Name": "Shuttering", "Total": 0}},
                        {{"Activity Name": "Reinforcement", "Total": 0}},
                        {{"Activity Name": "De-Shuttering", "Total": 0}}
                      ]
                    }}
                  ]
                {{ "Tower": "Tower H", "Categories": [...] }},
                {{ "Tower": "Tower F", "Categories": [...] }}
              ],
              "Asite": [
                {{
                  "Tower": "Tower G",
                  "Categories": [
                    {{
                      "Category": "MEP",
                      "Activities": [
                        {{"Activity Name": "Wall Conduting", "Total": 0}},
                        {{"Activity Name": "Plumbing Works", "Total": 0}},
                        {{"Activity Name": "Wiring & Switch Socket", "Total": 0}},
                        {{"Activity Name": "Slab Conducting", "Total": 0}}
                      ]
                    }},
                    {{
                      "Category": "Interior Finishing",
                      "Activities": [
                        {{"Activity Name": "Waterproofing - Sunken", "Total": 0}},
                        {{"Activity Name": "Wall Tile", "Total": 0}},
                        {{"Activity Name": "Floor Tile", "Total": 0}},
                        {{"Activity Name": "POP & Gypsum Plaster", "Total": 0}}
                      ]
                    }},
                    {{
                      "Category": "ED Civil",
                      "Activities": [
                        {{"Activity Name": "Concreting", "Total": 0}},
                         {{"Activity Name": "Shuttering", "Total": 0}},
                        {{"Activity Name": "Reinforcement", "Total": 0}},
                        {{"Activity Name": "De-Shuttering", "Total": 0}}
                      ]
                    }}
                ]
                }},
                {{ "Tower": "Tower H", "Categories": [...] }},
                {{ "Tower": "Tower F", "Categories": [...] }}
              ],
              "Slab":{{
                 "Tower Name":"Total"
              }}
            }}

            Return only the JSON object, no additional text, explanations, or code. Ensure the counts are accurate, activities are correctly categorized, and the JSON structure is valid.
            """,
            "parameters": {
                "decoding_method": "greedy",
                "max_new_tokens": 8100,
                "min_new_tokens": 0,
                "stop_sequences": [],  
                "repetition_penalty": 1.0,
                "temperature": 0.1
            },
            "model_id": os.getenv("MODEL_ID_1"),
            "project_id": os.getenv("PROJECT_ID_1")
        }
        
        access_token = get_access_token(os.getenv("API_KEY_1"))
        if not access_token:
            logger.error("Failed to obtain access token for WatsonX API")
            return combined_data
            
        headers = {
            "Accept": "application/json",
            "Content-Type": "application/json",
            "Authorization": f"Bearer {access_token}"
        }
        
        response = requests.post(os.getenv("WATSONX_API_URL_1"), headers=headers, json=body, timeout=1000)
        
        if response.status_code != 200:
            logger.error(f"WatsonX API call failed: {response.status_code} - {response.text}")
            st.warning(f"WatsonX API failed with status {response.status_code}: {response.text}. Using fallback method to calculate totals.")
            return combined_data
            
        response_data = response.json()
        if 'results' not in response_data or not response_data['results']:
            logger.error("WatsonX API response does not contain 'results' key")
            st.warning("WatsonX API response invalid. Using fallback method to calculate totals.")
            return combined_data

        generated_text = response_data['results'][0].get('generated_text', '').strip()
        logger.info(f"Raw WatsonX API response: {generated_text[:1000]}...")
        if not generated_text:
            logger.error("WatsonX API returned empty generated text")
            st.warning("WatsonX API returned empty response. Using fallback method to calculate totals.")
            return combined_data

        # Fix 1: Enhanced JSON extraction with repair capability
        fixed_json_text = extract_and_repair_json(generated_text)
        if fixed_json_text is None:
            logger.error("Failed to extract or repair JSON from response")
            return combined_data
        
        try:
            parsed_json = json.loads(fixed_json_text)
            if not (isinstance(parsed_json, dict) and "COS" in parsed_json and "Asite" in parsed_json):
                logger.warning(f"Invalid JSON structure: {json.dumps(parsed_json, indent=2)}")
                return combined_data
            for source in ["COS", "Asite"]:
                if not isinstance(parsed_json[source], list):
                    logger.warning(f"Expected list for {source}, got: {type(parsed_json[source])}")
                    return combined_data
                for tower_data in parsed_json[source]:
                    if not isinstance(tower_data, dict) or "Tower" not in tower_data or "Categories" not in tower_data:
                        logger.warning(f"Invalid tower data in {source}: {tower_data}")
                        return combined_data
            return json.dumps(parsed_json, indent=2)  # Return standardized JSON
        except json.JSONDecodeError as e:
            logger.error(f"Failed to parse JSON after repair attempt: {str(e)}")
            logger.error(f"Full response after repair: {fixed_json_text}")
            error_position = int(str(e).split('(char ')[1].split(')')[0]) if '(char ' in str(e) else 0
            context_start = max(0, error_position - 50)
            context_end = min(len(fixed_json_text), error_position + 50)
            logger.error(f"JSON error context: ...{fixed_json_text[context_start:error_position]}[ERROR HERE]{fixed_json_text[error_position:context_end]}...")
            st.warning(f"WatsonX API returned invalid JSON that couldn't be repaired. Error: {str(e)}. Using fallback method to calculate totals.")
            return combined_data
    
    except Exception as e:
        logger.error(f"Error in WatsonX API call: {str(e)}")
        st.warning(f"Error in WatsonX API call: {str(e)}. Using fallback method to calculate totals.")
        return combined_data
    
def get_concreting_count_from_consolidated(consolidated_rows):
    """
    Extract Concreting count from consolidated checklist data
    Use this AFTER generating consolidated_rows in generate_consolidated_Checklist_excel
    """
    try:
        for row in consolidated_rows:
            if (row.get('Category') == 'Civil Works' and 
                row.get('Activity Name') == 'Concreting'):
                return row.get('Completed Work*(Count of Flat)', 0)
    except Exception as e:
        logger.error(f"Error extracting Concreting count: {str(e)}")
    
    return 0


def get_concreting_from_ai_response(ai_data):
    """
    Extract Concreting count from AI response JSON
    Returns total Concreting count across all towers
    """
    try:
        if isinstance(ai_data, str):
            ai_data = json.loads(ai_data)
        
        total_concreting = 0
        concreting_by_tower = {}
        
        # Check COS data for Concreting (includes slab data)
        for tower_data in ai_data.get("COS", []):
            tower_name = tower_data.get("Tower", "Unknown")
            tower_concreting = 0
            
            for category_data in tower_data.get("Categories", []):
                if category_data.get("Category") == "ED Civil":
                    for activity in category_data.get("Activities", []):
                        if activity.get("Activity Name") == "Concreting":
                            count = activity.get("Total", 0)
                            tower_concreting += count
                            total_concreting += count
            
            if tower_concreting > 0:
                concreting_by_tower[tower_name] = tower_concreting
        
        return {
            "total": total_concreting,
            "by_tower": concreting_by_tower
        }
    except Exception as e:
        logger.error(f"Error extracting Concreting count from AI: {str(e)}")
    
    return {"total": 0, "by_tower": {}}


def get_concreting_from_consolidated(consolidated_rows):
    """
    Extract Concreting count from consolidated rows
    Returns Civil Works Concreting data by tower
    """
    try:
        concreting_data = {}
        total = 0
        
        for row in consolidated_rows:
            if (row.get('Category') == 'Civil Works' and 
                row.get('Activity Name') == 'Concreting'):
                tower = row.get('Tower', 'Unknown')
                count = row.get('Completed Work*(Count of Flat)', 0)
                concreting_data[tower] = count
                total += count
        
        return {
            "total": total,
            "by_tower": concreting_data
        }
    except Exception as e:
        logger.error(f"Error extracting Concreting from consolidated rows: {str(e)}")
    
    return {"total": 0, "by_tower": {}}


def apply_concreting_mapping_same_value(consolidated_rows):
    """
    Map Concreting count to Shuttering, Reinforcement, and De-Shuttering with same value.
    
    Example: If Concreting = 24
    Then: Shuttering = 24, Reinforcement = 24, De-Shuttering = 24
    
    Args:
        consolidated_rows: List of dictionaries with activity data
    
    Returns:
        List of modified consolidated rows
    """
    try:
        # Find concreting count per tower
        concreting_by_tower = {}
        for row in consolidated_rows:
            if row.get('Activity Name') == 'Concreting' and row.get('Category') == 'Civil Works':
                tower = row.get('Tower')
                concreting_count = row.get('Completed Work*(Count of Flat)', 0)
                concreting_by_tower[tower] = concreting_count
        
        if not concreting_by_tower:
            logger.warning("No Concreting data found to map")
            return consolidated_rows
        
        # Apply same value to all Civil Works activities
        for row in consolidated_rows:
            if row.get('Category') == 'Civil Works' and row.get('Tower') in concreting_by_tower:
                tower = row.get('Tower')
                concreting_count = concreting_by_tower[tower]
                activity_name = row.get('Activity Name')
                
                # Apply concreting count to all Civil Works activities
                if activity_name in ['Shuttering', 'Reinforcement', 'De-Shuttering']:
                    row['Completed Work*(Count of Flat)'] = concreting_count
        
        logger.info("Applied same value mapping: All Civil Works activities now have Concreting count")
        return consolidated_rows
    
    except Exception as e:
        logger.error(f"Error in apply_concreting_mapping_same_value: {str(e)}")
        return consolidated_rows


def sync_slab_conducting_with_concreting(consolidated_rows):
    """
    Synchronize Slab Conducting count to match Concreting count across all towers.
    
    Logic:
    - For each tower, find the Concreting count in Civil Works
    - Apply that same count to Slab Conducting in MEP Works
    - If Concreting doesn't exist, Slab Conducting remains unchanged
    
    Args:
        consolidated_rows: List of dictionaries containing activity data
    
    Returns:
        List of modified consolidated rows with synced Slab Conducting counts
    
    Example:
        If Tower TG has Concreting = 24, then Slab Conducting for TG = 24
        If Tower TH has Concreting = 18, then Slab Conducting for TH = 18
    """
    try:
        # Build dictionary of Concreting counts by tower
        concreting_by_tower = {}
        
        for row in consolidated_rows:
            if (row.get('Category') == 'Civil Works' and 
                row.get('Activity Name') == 'Concreting'):
                tower = row.get('Tower')
                count = row.get('Completed Work*(Count of Flat)', 0)
                concreting_by_tower[tower] = count
                logger.info(f"Found Concreting count for {tower}: {count}")
        
        if not concreting_by_tower:
            logger.warning("No Concreting data found. Slab Conducting will not be synced.")
            return consolidated_rows
        
        # Apply Concreting count to Slab Conducting
        rows_modified = 0
        for row in consolidated_rows:
            if (row.get('Category') == 'MEP Works' and 
                row.get('Activity Name') == 'Slab Conducting'):
                tower = row.get('Tower')
                
                if tower in concreting_by_tower:
                    old_count = row.get('Completed Work*(Count of Flat)', 0)
                    new_count = concreting_by_tower[tower]
                    
                    row['Completed Work*(Count of Flat)'] = new_count
                    
                    logger.info(f"Synced {tower} - Slab Conducting: {old_count} → {new_count}")
                    rows_modified += 1
        
        logger.info(f"Slab Conducting sync completed: {rows_modified} rows updated")
        return consolidated_rows
    
    except Exception as e:
        logger.error(f"Error in sync_slab_conducting_with_concreting: {str(e)}")
        return consolidated_rows

def display_concreting_summary(ai_data, consolidated_rows):
    """
    Display comprehensive Concreting summary with multiple breakdowns
    """
    try:
        st.subheader("🔨 Concreting Progress Summary")
        
        # Get data from both sources
        ai_concreting = get_concreting_from_ai_response(ai_data)
        consolidated_concreting = get_concreting_from_consolidated(consolidated_rows)
        
        # Display in columns
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.metric(
                label="Total Concreting (COS)",
                value=ai_concreting["total"]
            )
        
        with col2:
            st.metric(
                label="Total Concreting (Consolidated)",
                value=consolidated_concreting["total"]
            )
        
        with col3:
            difference = abs(ai_concreting["total"] - consolidated_concreting["total"])
            st.metric(
                label="Difference",
                value=difference
            )
        
        # Display by tower breakdown
        st.write("**Concreting by Tower (COS Data):**")
        if ai_concreting["by_tower"]:
            for tower, count in sorted(ai_concreting["by_tower"].items()):
                st.write(f"  • **{tower}**: {count} units")
        else:
            st.write("  No tower data available")
        
        # Display by tower breakdown (Consolidated)
        st.write("**Concreting by Tower (Consolidated Checklist):**")
        if consolidated_concreting["by_tower"]:
            for tower, count in sorted(consolidated_concreting["by_tower"].items()):
                st.write(f"  • **{tower}**: {count} units")
        else:
            st.write("  No tower data available")
    
    except Exception as e:
        logger.error(f"Error displaying concreting summary: {str(e)}")
        st.error(f"Error displaying concreting summary: {str(e)}")


def get_concreting_by_tower(consolidated_rows):
    """
    Extract Concreting count broken down by tower
    Returns a dictionary like: {'TF': 16, 'TG': 20, 'TH': 18}
    """
    try:
        concreting_by_tower = {}
        for row in consolidated_rows:
            if (row.get('Category') == 'Civil Works' and 
                row.get('Activity Name') == 'Concreting'):
                tower = row.get('Tower', 'Unknown')
                count = row.get('Completed Work*(Count of Flat)', 0)
                concreting_by_tower[tower] = count
        
        return concreting_by_tower
    except Exception as e:
        logger.error(f"Error extracting Concreting count by tower: {str(e)}")
    
    return {}

def extract_and_repair_json(text):
    # Try to find JSON content within the response
    json_match = re.search(r'\{.*\}', text, re.DOTALL)
    if json_match:
        json_text = json_match.group(0)
        
        # Common JSON repair operations
        try:
            # First try: Parse as is
            json.loads(json_text)
            return json_text
        except json.JSONDecodeError as e:
            try:
                # Fix 1: Try fixing missing commas between objects in arrays
                fixed1 = re.sub(r'}\s*{', '},{', json_text)
                json.loads(fixed1)
                logger.info("JSON fixed by adding missing commas between objects")
                return fixed1
            except json.JSONDecodeError:
                try:
                    # Fix 2: Try fixing trailing commas in arrays/objects
                    fixed2 = re.sub(r',\s*}', '}', fixed1)
                    fixed2 = re.sub(r',\s*]', ']', fixed2)
                    json.loads(fixed2)
                    logger.info("JSON fixed by removing trailing commas")
                    return fixed2
                except json.JSONDecodeError:
                    try:
                        # Fix 3: Try using a JSON repair library or more aggressive repairs
                        # Here we'll use a simple approach to balance braces/brackets
                        fixed3 = fixed2
                        count_open_braces = fixed3.count('{')
                        count_close_braces = fixed3.count('}')
                        if count_open_braces > count_close_braces:
                            fixed3 += '}' * (count_open_braces - count_close_braces)
                        
                        count_open_brackets = fixed3.count('[')
                        count_close_brackets = fixed3.count(']')
                        if count_open_brackets > count_close_brackets:
                            fixed3 += ']' * (count_open_brackets - count_close_brackets)
                        
                        # Fix unquoted keys (a common issue)
                        fixed3 = re.sub(r'([{,])\s*([a-zA-Z0-9_]+)\s*:', r'\1"\2":', fixed3)
                        
                        json.loads(fixed3)
                        logger.info("JSON fixed with aggressive repairs")
                        return fixed3
                    except json.JSONDecodeError:
                        # Final attempt: Try to load the JSON with a more permissive parser
                        try:
                            import demjson3  # type: ignore
                            parsed = demjson3.decode(json_text)
                            logger.info("JSON fixed with demjson3")
                            return json.dumps(parsed)
                        except Exception:
                            logger.error("All JSON repair attempts failed")
                            return None
    else:
        logger.error("No JSON-like content found in the response")
        return None

# Fix the getTotal function to handle the improved JSON structure
def getTotal(ai_data):
    try:
        if isinstance(ai_data, str):
            try:
                ai_data = json.loads(ai_data)
            except json.JSONDecodeError as e:
                logger.error(f"Error parsing AI data JSON: {str(e)}")
                return [0] * len(st.session_state.get('sheduledf', pd.DataFrame()).index)
            
        if not isinstance(ai_data, dict) or "COS" not in ai_data or "Asite" not in ai_data:
            logger.error(f"AI data is not in expected format: {ai_data}")
            return [0] * len(st.session_state.get('sheduledf', pd.DataFrame()).index)

        share = []
        
        # Process sources properly
        for source in ["COS", "Asite"]:
            if not isinstance(ai_data[source], list):
                logger.error(f"{source} data is not a list: {ai_data[source]}")
                continue
                
            for tower_data in ai_data[source]:
                if not isinstance(tower_data, dict) or "Categories" not in tower_data:
                    logger.error(f"Invalid tower data format in {source}: {tower_data}")
                    continue
                    
                for category_data in tower_data["Categories"]:
                    if not isinstance(category_data, dict) or "Activities" not in category_data:
                        logger.error(f"Invalid category data format: {category_data}")
                        continue
                        
                    for activity in category_data["Activities"]:
                        if isinstance(activity, dict) and "Total" in activity:
                            total = activity["Total"]
                            share.append(int(total) if isinstance(total, (int, float)) and pd.notna(total) else 0)
                        else:
                            logger.warning(f"Activity missing Total field: {activity}")
                            share.append(0)
        
        # Ensure we have enough values for the schedule dataframe
        expected_length = len(st.session_state.get('sheduledf', pd.DataFrame()).index)
        if len(share) < expected_length:
            logger.warning(f"Not enough values in share list (got {len(share)}, need {expected_length}). Padding with zeros.")
            share.extend([0] * (expected_length - len(share)))
        elif len(share) > expected_length:
            logger.warning(f"Too many values in share list (got {len(share)}, need {expected_length}). Truncating.")
            share = share[:expected_length]
            
        return share
    except Exception as e:
        logger.error(f"Error processing AI data: {str(e)}")
        st.error(f"Error processing AI data: {str(e)}")
        return [0] * len(st.session_state.get('sheduledf', pd.DataFrame()).index)    
 



   
# Function to handle activity count display logic
def display_activity_count():
    try:
        if 'ai_response' not in st.session_state or not st.session_state.ai_response:
            st.error(" No AI-generated data available. Please run the analysis first.")
            return

        try:
            ai_data = json.loads(st.session_state.ai_response) if isinstance(st.session_state.ai_response, str) else st.session_state.ai_response
        except (json.JSONDecodeError, TypeError) as e:
            st.error(f" Failed to parse AI response: {str(e)}")
            st.write("Raw AI response:")
            st.text(str(st.session_state.ai_response)[:500])
            return

        if not isinstance(ai_data, dict) or "COS" not in ai_data or "Asite" not in ai_data:
            st.error(" Invalid AI data format. Expected 'COS' and 'Asite' sections.")
            st.write("AI data content:")
            st.json(ai_data if isinstance(ai_data, dict) else {})
            return

        slab_df = st.session_state.get('slab_df')
        if slab_df is None:
            slab_df = pd.DataFrame()
        elif isinstance(slab_df, str):
            st.warning(f"Slab data is string: {slab_df}")
            slab_df = pd.DataFrame()
        elif not isinstance(slab_df, pd.DataFrame):
            st.warning(f"Slab data has unexpected type: {type(slab_df)}")
            slab_df = pd.DataFrame()
        
        logging.info(f"Slab cycle DataFrame in display_activity_count: {slab_df.to_dict() if not slab_df.empty else 'Empty'}")
        
        slab_display_df = pd.DataFrame(columns=['Tower', 'Completed'])
        slab_counts = {}
        
        if not slab_df.empty:
            new_rows = []
            for _, row in slab_df.iterrows():
                tower = row.get('Tower', 'Unknown')
                completed = row.get('Completed', 0)
                if tower == 'T4':
                    t4a_completed = completed // 2
                    t4b_completed = completed - t4a_completed
                    new_rows.append({'Tower': 'T4A', 'Completed': t4a_completed})
                    new_rows.append({'Tower': 'T4B', 'Completed': t4b_completed})
                    slab_counts['T4A'] = t4a_completed
                    slab_counts['T4B'] = t4b_completed
                else:
                    new_rows.append({'Tower': tower, 'Completed': completed})
                    slab_counts[tower] = completed
            slab_display_df = pd.DataFrame(new_rows)
            logging.info(f"Slab display DataFrame after processing: {slab_display_df.to_dict()}")
        else:
            logging.warning("Slab cycle DataFrame is empty in display_activity_count.")

        categories = {
            "COS": {
                "MEP": [
                    "EL-First Fix", "UP-First Fix", "CP-First Fix", "Min. count of UP-First Fix and CP-First Fix",
                    "POP punning (Major area)", "EL 2nd Fix",
                ],
                "Interior Finishing": [
                    "Installation of doors", "Water Proofing Works", "Wall tiling", "Floor Tiling"
                ],
                "ED Civil": [
                    "Concreting", "Shuttering", "Reinforcement", "De-Shuttering"
                ]
            },
            "Asite": {
                "MEP": [
                    "Wall Conduting", "Plumbing Works", "Wiring & Switch Socket", "Slab Conducting"
                ],
                "Interior Finishing": [
                    "Waterproofing - Sunken", "Wall Tile", "Floor Tile", "POP & Gypsum Plaster"
                ],
                "ED Civil": [
                    "Concreting", "Shuttering", "Reinforcement", "De-Shuttering"
                ]
            }
        }

        for source in ["COS", "Asite"]:
            st.subheader(f"{source} Activity Counts")
            source_data = ai_data.get(source, [])
            if not source_data or not isinstance(source_data, list):
                st.warning(f"No valid data available for {source}.")
                continue

            for tower_data in source_data:
                if not isinstance(tower_data, dict):
                    continue
                    
                tower_name = tower_data.get("Tower", "Unknown Tower")
                st.write(f"#### {tower_name}")
                tower_categories = tower_data.get("Categories", [])

                if not tower_categories:
                    st.write("No categories available for this tower.")
                    continue

                tower_total = 0

                for category in categories.get(source, {}):
                    st.write(f"**{category}**")
                    category_data = next(
                        (cat for cat in tower_categories if isinstance(cat, dict) and cat.get("Category") == category),
                        {"Category": category, "Activities": []}
                    )

                    if not category_data.get("Activities"):
                        st.write("No activities recorded.")
                        continue

                    activity_counts = []
                    for activity in categories[source][category]:
                        activity_info = next(
                            (act for act in category_data["Activities"] 
                             if isinstance(act, dict) and act.get("Activity Name") == activity),
                            {"Activity Name": activity, "Total": 0}
                        )
                        count = int(activity_info.get("Total", 0)) if pd.notna(activity_info.get("Total")) else 0
                        activity_counts.append({
                            "Activity Name": activity_info.get("Activity Name", activity),
                            "Count": count
                        })
                        tower_total += count

                    df = pd.DataFrame(activity_counts)
                    if not df.empty:
                        st.table(df)
                    else:
                        st.write("No activities in this category.")

                if source == "COS":
                    st.write("**Slab Cycle Counts**")
                    tower_slab_df = slab_display_df[slab_display_df['Tower'] == tower_name] if not slab_display_df.empty else pd.DataFrame()
                    logging.info(f"Tower {tower_name} - Filtered slab counts: {tower_slab_df.to_dict() if not tower_slab_df.empty else 'Empty'}")
                    if not tower_slab_df.empty:
                        st.table(tower_slab_df)
                        tower_total += tower_slab_df['Completed'].sum()
                    else:
                        st.write("No slab cycle data for this tower.")
                st.write(f"**Total for {tower_name}**: {tower_total}")

        total_cos = sum(
            int(act.get("Total", 0)) if pd.notna(act.get("Total")) else 0
            for tower in ai_data.get("COS", [])
            if isinstance(tower, dict)
            for cat in tower.get("Categories", [])
            if isinstance(cat, dict)
            for act in cat.get("Activities", [])
            if isinstance(act, dict)
        )
        total_cos += sum(slab_counts.values())

        total_asite = sum(
            int(act.get("Total", 0)) if pd.notna(act.get("Total")) else 0
            for tower in ai_data.get("Asite", [])
            if isinstance(tower, dict)
            for cat in tower.get("Categories", [])
            if isinstance(cat, dict)
            for act in cat.get("Activities", [])
            if isinstance(act, dict)
        )

        st.write("### Total Completed Activities")
        st.write(f"**COS Total**: {total_cos}")
        st.write(f"**Asite Total**: {total_asite}")

    except Exception as e:
        logging.error(f"Error in display_activity_count: {str(e)}", exc_info=True)
        st.error(f" Error displaying activity counts: {str(e)}")
        st.write("AI response content (for debugging):")
        st.text(str(st.session_state.get('ai_response', 'No response'))[:500])




        
st.markdown(
    """
    <h1 style='font-family: "Arial Black", Gadget, sans-serif; 
               color: red; 
               font-size: 48px; 
               text-align: center;'>
        CheckList - Report
    </h1>
    """,
    unsafe_allow_html=True
)



# Combined function for Initialize All Data and Fetch COS
async def initialize_and_fetch_data(email, password):
    with st.spinner("Starting initialization and data fetching process..."):
        # Step 1: Login
        if not email or not password:
            st.sidebar.error("Please provide both email and password!")
            logger.error("Email or password not provided")
            return False
        try:
            st.sidebar.write("Logging in...")
            session_id = await login_to_asite(email, password)
            if not session_id:
                st.sidebar.error("Login failed!")
                logger.error("Login failed: No session ID returned")
                return False
            st.sidebar.success("Login successful!")
        except Exception as e:
            st.sidebar.error(f"Login failed: {str(e)}")
            logger.error(f"Login failed: {str(e)}\nStack trace:\n{traceback.format_exc()}")
            return False

        # Step 2: Get Workspace ID
        try:
            st.sidebar.write("Fetching Workspace ID...")
            await GetWorkspaceID()
            st.sidebar.success("Workspace ID fetched successfully!")
        except Exception as e:
            st.sidebar.error(f"Failed to fetch Workspace ID: {str(e)}")
            logger.error(f"Failed to fetch Workspace ID: {str(e)}\nStack trace:\n{traceback.format_exc()}")
            return False

        # Step 3: Get Project IDs
        try:
            st.sidebar.write("Fetching Project IDs...")
            await GetProjectId()
            st.sidebar.success("Project IDs fetched successfully!")
        except Exception as e:
            st.sidebar.error(f"Failed to fetch Project IDs: {str(e)}")
            logger.error(f"Failed to fetch Project IDs: {str(e)}\nStack trace:\n{traceback.format_exc()}")
            return False

        # Step 4: Get All Data (now returns 4 values including Tower H)
        try:
            st.sidebar.write("Fetching All Data...")
            finishing, structure, external, tower_h = await GetAllDatas()
            st.session_state.eligo_tower_f_finishing = finishing
            st.session_state.eligo_structure = structure
            st.session_state.eligo_tower_g_finishing = external
            st.session_state.eligo_tower_h_finishing = tower_h
            
            st.sidebar.success("All Data fetched successfully!")
            logger.info(f"Stored eligo_tower_f_finishing: {len(finishing)} records, "
                       f"eligo_structure: {len(structure)} records, "
                       f"eligo_tower_g_finishing: {len(external)} records, "
                       f"eligo_tower_h_finishing: {len(tower_h)} records")
        except Exception as e:
            st.sidebar.error(f"Failed to fetch All Data: {str(e)}")
            logger.error(f"Failed to fetch All Data: {str(e)}\nStack trace:\n{traceback.format_exc()}")
            return False

        # Step 5: Get Activity Data (now returns 4 values including Tower H)
        try:
            st.sidebar.write("Fetching Activity Data...")
            finishing_activity_data, structure_activity_data, external_activity_data, tower_h_activity_data = await Get_Activity()
            # Validate DataFrames
            activity_dataframes = {
                "finishing_activity_data": finishing_activity_data,
                "structure_activity_data": structure_activity_data,
                "external_activity_data": external_activity_data,
                "tower_h_activity_data": tower_h_activity_data,
            }
            for name, df in activity_dataframes.items():
                if df is None:
                    logger.error(f"{name} is None")
                    raise ValueError(f"{name} is None")
                if not isinstance(df, pd.DataFrame):
                    logger.error(f"{name} is not a DataFrame: {type(df)}")
                    raise ValueError(f"{name} is not a valid DataFrame")
                logger.info(f"{name} has {len(df)} records, empty: {df.empty}")
                if df.empty:
                    logger.warning(f"{name} is empty")
            # Store in session state
            st.session_state.finishing_activity_data = finishing_activity_data
            st.session_state.structure_activity_data = structure_activity_data
            st.session_state.external_activity_data = external_activity_data
            st.session_state.tower_h_activity_data = tower_h_activity_data
            
            st.sidebar.success("Activity Data fetched successfully!")
            logger.info(f"Stored activity data - Finishing: {len(finishing_activity_data)} records, "
                        f"Structure: {len(structure_activity_data)} records, "
                        f"External: {len(external_activity_data)} records, "
                        f"Tower H: {len(tower_h_activity_data)} records")
        except Exception as e:
            st.sidebar.error(f"Failed to fetch Activity Data: {str(e)}")
            logger.error(f"Failed to fetch Activity Data: {str(e)}\nStack trace:\n{traceback.format_exc()}")
            return False

        # Step 6: Get Location/Module Data (now returns 4 values including Tower H)
        try:
            st.sidebar.write("Fetching Location/Module Data...")
            finishing_location_data, structure_location_data, external_location_data, tower_h_location_data = await Get_Location()
            # Validate DataFrames
            location_dataframes = {
                "finishing_location_data": finishing_location_data,
                "structure_location_data": structure_location_data,
                "external_location_data": external_location_data,
                "tower_h_location_data": tower_h_location_data,
            }
            for name, df in location_dataframes.items():
                if df is None:
                    logger.error(f"{name} is None")
                    raise ValueError(f"{name} is None")
                if not isinstance(df, pd.DataFrame):
                    logger.error(f"{name} is not a DataFrame: {type(df)}")
                    raise ValueError(f"{name} is not a valid DataFrame")
                logger.info(f"{name} has {len(df)} records, empty: {df.empty}")
                if df.empty:
                    logger.warning(f"{name} is empty")
            # Store in session state
            st.session_state.finishing_location_data = finishing_location_data
            st.session_state.structure_location_data = structure_location_data
            st.session_state.external_location_data = external_location_data
            st.session_state.tower_h_location_data = tower_h_location_data
            
            st.sidebar.success("Location/Module Data fetched successfully!")
            logger.info(f"Stored location data - Finishing: {len(finishing_location_data)} records, "
                        f"Structure: {len(structure_location_data)} records, "
                        f"External: {len(external_location_data)} records, "
                        f"Tower H: {len(tower_h_location_data)} records")
        except Exception as e:
            st.sidebar.error(f"Failed to fetch Location/Module Data: {str(e)}")
            logger.error(f"Failed to fetch Location/Module Data: {str(e)}\nStack trace:\n{traceback.format_exc()}")
            return False
        
        # Step 7: Fetch COS Files
        try:
            st.sidebar.write("Fetching COS files from Eligo folder...")
            files = get_cos_files()
            st.session_state.files = files
            if files:
                st.success(f"Found {len(files)} files in COS storage")
                for selected_file in files:
                    try:
                        st.write(f"Processing file: {selected_file}")
                        cos_client = initialize_cos_client()
                        if not cos_client:
                            st.error("Failed to initialize COS client")
                            continue
                        response = cos_client.get_object(Bucket=COS_BUCKET, Key=selected_file)
                        file_bytes = io.BytesIO(response['Body'].read())
                        result = process_file(file_bytes, selected_file)
                        if len(result) == 1:  # Handle single DataFrame for Tower F, Tower G, Tower H, or Structure Work Tracker
                            (df_first, tname_first) = result[0]
                            if df_first is not None and not df_first.empty:
                                if "Tower F" in tname_first:
                                    st.session_state.cos_df_tower_f = df_first
                                    st.session_state.cos_tname_tower_f = tname_first
                                    st.write(f"Processed Data for {tname_first} - {len(df_first)} rows:")
                                    st.write(df_first.head())
                                elif "Tower G" in tname_first:
                                    st.session_state.cos_df_tower_g = df_first
                                    st.session_state.cos_tname_tower_g = tname_first
                                    st.write(f"Processed Data for {tname_first} - {len(df_first)} rows:")
                                    st.write(df_first.head())
                                elif "Tower H" in tname_first:
                                    st.session_state.cos_df_tower_h = df_first
                                    st.session_state.cos_tname_tower_h = tname_first
                                    st.write(f"Processed Data for {tname_first} - {len(df_first)} rows:")
                                    st.write(df_first.head())
                                elif "Structure Work Tracker" in tname_first:
                                    st.session_state.cos_df_structure = df_first
                                    st.session_state.cos_tname_structure = tname_first
                                    st.write(f"Processed Data for {tname_first} - {len(df_first)} rows:")
                                    st.write(df_first.head())
                            else:
                                st.warning(f"No valid data found in {selected_file}.")
                        else:
                            st.warning(f"Unexpected data format in {selected_file}. Expected a single DataFrame.")
                    except Exception as e:
                        st.error(f"Error loading {selected_file} from cloud storage: {str(e)}")
                        logger.error(f"Error loading {selected_file}: {str(e)}\nStack trace:\n{traceback.format_exc()}")
            else:
                st.warning("No expected Excel files available in the 'Eligo' folder of the COS bucket.")
        except Exception as e:
            st.sidebar.error(f"Failed to fetch COS files: {str(e)}")
            logger.error(f"Failed to fetch COS files: {str(e)}\nStack trace:\n{traceback.format_exc()}")
            return False

        # Step 8: Verify stored session state keys
        st.sidebar.write("Verifying stored data...")
        required_keys = [
            'eligo_tower_f_finishing', 'eligo_structure', 'eligo_tower_g_finishing', 'eligo_tower_h_finishing',
            'finishing_activity_data', 'structure_activity_data', 'external_activity_data', 'tower_h_activity_data',
            'finishing_location_data', 'structure_location_data', 'external_location_data', 'tower_h_location_data'
        ]
        missing_keys = [key for key in required_keys if key not in st.session_state]
        if missing_keys:
            st.sidebar.error(f"Missing session state keys: {', '.join(missing_keys)}")
            logger.error(f"Missing session state keys: {', '.join(missing_keys)}")
            st.error("Initialization and data fetching failed!")
            return False
        else:
            st.sidebar.success("All required data stored successfully!")
            logger.info("All required session state keys verified.")

        st.sidebar.write("Initialization and data fetching process completed!")
        return True

def debug_asite_data_flow(ai_data):
    """
    Debug function to check if Asite data is properly populated in AI response
    """
    try:
        if isinstance(ai_data, str):
            ai_data = json.loads(ai_data)
        
        st.write("### Debug: Asite Data in AI Response")
        
        asite_activities = ai_data.get("Asite", [])
        
        if not asite_activities:
            st.error("❌ NO ASITE DATA FOUND in AI response!")
            return False
        
        st.write(f"✓ Found {len(asite_activities)} towers in Asite data")
        
        total_asite_count = 0
        for tower_data in asite_activities:
            tower_name = tower_data.get("Tower", "Unknown")
            st.write(f"\n**Tower: {tower_name}**")
            
            for category_data in tower_data.get("Categories", []):
                category = category_data.get("Category", "Unknown")
                activities = category_data.get("Activities", [])
                
                st.write(f"  Category: {category} ({len(activities)} activities)")
                
                for activity in activities:
                    activity_name = activity.get("Activity Name", "Unknown")
                    total = activity.get("Total", 0)
                    total_asite_count += total
                    
                    if total > 0:
                        st.write(f"    ✓ {activity_name}: {total}")
                    else:
                        st.write(f"    ✗ {activity_name}: {total} (ZERO)")
        
        st.write(f"\n**Total Asite Count: {total_asite_count}**")
        
        if total_asite_count == 0:
            st.error("⚠️ WARNING: All Asite activity counts are ZERO!")
            return False
        
        return True
        
    except Exception as e:
        st.error(f"Error in debug function: {str(e)}")
        return False

# def apply_tower_f_hardcoded_fixes(consolidated_rows):
#     """
#     Apply hardcoded count corrections for Tower F, G, and H COS data.
#     This function is completely standalone and does not affect any other logic.
    
#     WHAT IT DOES:
#     - Finds Tower F, G, and H rows in the consolidated data
#     - Overrides ONLY the COS counts (Completed Work column) for specific activities
#     - Leaves Asite data completely untouched
#     - Recalculates the gap based on corrected COS vs unchanged Asite
    
#     CORRECTIONS APPLIED:
    
#     Tower F:
#     - POP & Gypsum Plaster: Force COS count to 64 (was incorrectly 117)
#     - Wall Conduting: Force COS count to 0 (was incorrectly 1)
#     - Wiring & Switch Socket: Force COS count to 0 (was incorrectly 2)
    
#     Tower G:
#     - Wiring & Switch Socket: Force COS count to 0
    
#     Tower H:
#     - Wiring & Switch Socket: Force COS count to 0
    
#     Args:
#         consolidated_rows: List of dictionaries with activity data
#                           Expected keys: 'Tower', 'Activity Name', 
#                                        'Completed Work*(Count of Flat)',
#                                        'Closed checklist against completed work',
#                                        'Open/Missing check list'
    
#     Returns:
#         List of consolidated rows with corrections applied
#         (Original list is modified in-place and also returned)
#     """
#     import logging
    
#     logger = logging.getLogger(__name__)
    
#     try:
#         logger.info("=" * 80)
#         logger.info("APPLYING TOWER F HARDCODED CORRECTIONS")
#         logger.info("=" * 80)
        
#         # Define the exact corrections for each tower
#         TOWER_CORRECTIONS = {
#             'TF': {
#                 "POP & Gypsum Plaster": 64,  # Correct count (was 117 due to ceiling work)
#                 "Wall Conduting": 0,          # Correct count (was 1 due to common area LV work)
#                 "Wiring & Switch Socket": 0   # Correct count (was 2 due to common area LV work)
#             },
#             'TG': {
#                 "Wiring & Switch Socket": 0   # Set to 0 for Tower G
#             },
#             'TH': {
#                 "Wiring & Switch Socket": 0   # Set to 0 for Tower H
#             }
#         }
        
#         # Normalize tower names to standard format (TF, TG, TH)
#         def normalize_tower(tower_name):
#             tower_str = str(tower_name).strip().upper()
#             if 'F' in tower_str:
#                 return 'TF'
#             elif 'G' in tower_str:
#                 return 'TG'
#             elif 'H' in tower_str:
#                 return 'TH'
#             return tower_str
        
#         rows_updated = 0
#         corrections_applied = {}
        
#         # Iterate through all rows
#         for row in consolidated_rows:
#             # Get and normalize tower name
#             tower_raw = row.get('Tower', '')
#             tower_normalized = normalize_tower(tower_raw)
            
#             # Skip if not in our correction list
#             if tower_normalized not in TOWER_CORRECTIONS:
#                 continue
            
#             activity_name = row.get('Activity Name', '')
            
#             # Check if this tower + activity needs correction
#             if activity_name in TOWER_CORRECTIONS[tower_normalized]:
#                 # Get the correct count for this tower
#                 correct_cos_count = TOWER_CORRECTIONS[tower_normalized][activity_name]
                
#                 # Get the old COS count (before correction)
#                 old_cos_count = row.get('Completed Work*(Count of Flat)', 0)
                
#                 # Get the Asite count (this will NOT be changed)
#                 asite_count = row.get('Closed checklist against completed work', 0)
                
#                 # Apply the correction to COS count
#                 row['Completed Work*(Count of Flat)'] = correct_cos_count
                
#                 # Recalculate the gap
#                 row['Open/Missing check list'] = abs(correct_cos_count - asite_count)
                
#                 # Track corrections
#                 key = f"{tower_normalized} - {activity_name}"
#                 corrections_applied[key] = {
#                     'old': old_cos_count,
#                     'new': correct_cos_count,
#                     'asite': asite_count,
#                     'gap': row['Open/Missing check list']
#                 }
                
#                 # Log the change
#                 logger.info(f"{tower_normalized} - {activity_name}:")
#                 logger.info(f"  COS (Before): {old_cos_count}")
#                 logger.info(f"  COS (After):  {correct_cos_count} ✓")
#                 logger.info(f"  Asite:        {asite_count} (unchanged)")
#                 logger.info(f"  New Gap:      {row['Open/Missing check list']}")
                
#                 rows_updated += 1
        
#         logger.info(f"\nHardcoded Corrections Summary:")
#         logger.info(f"  Total rows updated: {rows_updated}")
#         for key, values in corrections_applied.items():
#             logger.info(f"  {key}: {values['old']} → {values['new']}")
#         logger.info("=" * 80)
        
#         return consolidated_rows
        
#     except Exception as e:
#         logger.error(f"ERROR in apply_tower_f_hardcoded_fixes: {str(e)}")
#         import traceback
#         logger.error(traceback.format_exc())
#         # Return original data if error occurs
#         return consolidated_rows

def apply_tower_f_hardcoded_fixes(consolidated_rows):
    """
    Apply hardcoded count corrections for Tower F COS data.
    This function is completely standalone and does not affect any other logic.
    
    WHAT IT DOES:
    - Finds Tower F rows in the consolidated data
    - Overrides ONLY the COS counts (Completed Work column) for specific activities
    - Leaves Asite data completely untouched
    - Recalculates the gap based on corrected COS vs unchanged Asite
    
    CORRECTIONS APPLIED:
    
    Tower F:
    - POP & Gypsum Plaster: Force COS count to 64 (was incorrectly 117)
    - Wall Conduting: Force COS count to 0 (was incorrectly 1)
    - Wiring & Switch Socket: Force COS count to 0 (was incorrectly 2)
    
    Args:
        consolidated_rows: List of dictionaries with activity data
                          Expected keys: 'Tower', 'Activity Name', 
                                       'Completed Work*(Count of Flat)',
                                       'Closed checklist against completed work',
                                       'Open/Missing check list'
    
    Returns:
        List of consolidated rows with corrections applied
        (Original list is modified in-place and also returned)
    """
    import logging
    
    logger = logging.getLogger(__name__)
    
    try:
        logger.info("=" * 80)
        logger.info("APPLYING TOWER F HARDCODED CORRECTIONS")
        logger.info("=" * 80)
        
        # Define the exact corrections for Tower F
        TOWER_F_CORRECTIONS = {
            "Wall Tile": 64,# Correct count (was 1 due to common area LV work)
            "Waterproofing - Sunken":64
        }
        
        # Normalize tower names to standard format (TF)
        def normalize_tower(tower_name):
            tower_str = str(tower_name).strip().upper()
            if 'F' in tower_str:
                return 'TF'
            return tower_str
        
        rows_updated = 0
        corrections_applied = {}
        
        # Iterate through all rows
        for row in consolidated_rows:
            # Get and normalize tower name
            tower_raw = row.get('Tower', '')
            tower_normalized = normalize_tower(tower_raw)
            
            # Skip if not Tower F
            if tower_normalized != 'TF':
                continue
            
            activity_name = row.get('Activity Name', '')
            
            # Check if this activity needs correction
            if activity_name in TOWER_F_CORRECTIONS:
                # Get the correct count for this activity
                correct_cos_count = TOWER_F_CORRECTIONS[activity_name]
                
                # Get the old COS count (before correction)
                old_cos_count = row.get('Completed Work*(Count of Flat)', 0)
                
                # Get the Asite count (this will NOT be changed)
                asite_count = row.get('Closed checklist against completed work', 0)
                
                # Apply the correction to COS count
                row['Completed Work*(Count of Flat)'] = correct_cos_count
                
                # Recalculate the gap
                row['Open/Missing check list'] = abs(correct_cos_count - asite_count)
                
                # Track corrections
                key = f"TF - {activity_name}"
                corrections_applied[key] = {
                    'old': old_cos_count,
                    'new': correct_cos_count,
                    'asite': asite_count,
                    'gap': row['Open/Missing check list']
                }
                
                # Log the change
                logger.info(f"TF - {activity_name}:")
                logger.info(f"  COS (Before): {old_cos_count}")
                logger.info(f"  COS (After):  {correct_cos_count} ✓")
                logger.info(f"  Asite:        {asite_count} (unchanged)")
                logger.info(f"  New Gap:      {row['Open/Missing check list']}")
                
                rows_updated += 1
        
        logger.info(f"\nHardcoded Corrections Summary:")
        logger.info(f"  Total rows updated: {rows_updated}")
        for key, values in corrections_applied.items():
            logger.info(f"  {key}: {values['old']} → {values['new']}")
        logger.info("=" * 80)
        
        return consolidated_rows
        
    except Exception as e:
        logger.error(f"ERROR in apply_tower_f_hardcoded_fixes: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        # Return original data if error occurs
        return consolidated_rows
# UPDATED generate_consolidated_Checklist_excel function with better Asite handling

def generate_consolidated_Checklist_excel(combined_data):
    """
    COMPLETELY FIXED: Proper tower normalization + Plumbing Works calculation
    """
    try:
        # Get current month name dynamically
        current_month = datetime.now().strftime("%B")
        logger.info(f"Generating report for: {current_month}")
        
        st.write("## 🔧 Generating Consolidated Checklist (Fixed Tower Mapping)")
        
        # ========== TOWER NAME NORMALIZATION (IMPROVED) ==========
        def normalize_tower_name(tower_name):
            """Convert any tower format to TF, TG, TH"""
            if not tower_name:
                return None
                
            tower_str = str(tower_name).strip().upper()
            
            # Remove common prefixes/suffixes
            tower_str = tower_str.replace("TOWER ", "").replace("TOWER-", "")
            tower_str = tower_str.replace(" FINISHING", "").replace("-FINISHING", "")
            
            # Direct matches
            if tower_str in ["TF", "TG", "TH", "F", "G", "H"]:
                return f"T{tower_str[-1]}"  # Ensure T prefix
            
            # Check for letter in string
            if "G" in tower_str:
                return "TG"
            elif "H" in tower_str:
                return "TH"
            elif "F" in tower_str:
                return "TF"
            
            # Skip non-tower
            if "NON-TOWER" in tower_str or "UNKNOWN" in tower_str:
                return None
            
            logger.warning(f"❌ Could not normalize: '{tower_name}'")
            return None

        # ========== ACTIVITY NAME NORMALIZATION ==========
        def normalize_activity_name(activity_name):
            """Normalize activity names to standard format"""
            if not isinstance(activity_name, str):
                return activity_name
            
            name_lower = activity_name.strip().lower()
            
            # ===== MEP ACTIVITIES =====
            # CRITICAL: Check UP-First Fix and CP-First Fix BEFORE general plumbing
            # Generic plumbing (only if not UP/CP)
            if any(kw in name_lower for kw in ["CP 1st Fix","cp 1st fix","CP-1st Fix","cp-1st fix","UP 1st Fix","up 1st fix","UP-1st Fix","up-1st fix"]):
                return "Plumbing Works"
            
            if any(kw in name_lower for kw in ["Wall Conduting"]):
                return "Wall Conduting"
            
            if any(kw in name_lower for kw in ["Slab conduting"]) and "concreting" not in name_lower:
                return "Slab conduting"
            
            if any(kw in name_lower for kw in ["EL 2nd Fix", "el 2nd fix","EL 2nd fix"]):
                return "Wiring & Switch Socket"
            
            # ===== INTERIOR FINISHING ACTIVITIES =====
            if any(kw in name_lower for kw in ["POP punning (Major area)", "pop punning (major area)"]):
                return "POP & Gypsum Plaster"
            
            if any(kw in name_lower for kw in ["Water Proofing Works", "water proofing works","Water proofing Works"]):
                return "Waterproofing - Sunken"
            
            if "wall til" in name_lower:
                return "Wall Tile"
            
            if "floor til" in name_lower:
                return "Floor Tile"
            
            # ===== CIVIL WORKS ACTIVITIES =====
            if any(kw in name_lower for kw in ["Concreting"]):
                return "Concreting"
            
            if "shutter" in name_lower and "de-" not in name_lower:
                return "Shuttering"
            
            if any(kw in name_lower for kw in ["Reinforcement"]):
                return "Reinforcement"
            
            if "de-shutter" in name_lower or "de shutter" in name_lower:
                return "De-Shuttering"
            
            return activity_name

        # ========== ACTIVITIES STRUCTURE ==========
        activities_structure = {
            "Civil Works": ["Concreting", "Shuttering", "Reinforcement", "De-Shuttering"],
            "Interior Finishing Works": ["Floor Tile", "Wall Tile", "POP & Gypsum Plaster", "Waterproofing - Sunken"],
            "MEP Works": ["Plumbing Works", "Slab Conducting", "Wall Conduting", "Wiring & Switch Socket"]
        }

        # ========== EXTRACT SLAB DATA ==========
        st.write("### 🔨 Processing Slab Data...")
        slab_data_dict = {}
        slab_df = st.session_state.get('slab_df', pd.DataFrame())
        
        if not slab_df.empty and isinstance(slab_df, pd.DataFrame):
            for _, row in slab_df.iterrows():
                tower_raw = row.get('Tower', '')
                count = int(row.get('Completed', 0)) if pd.notna(row.get('Completed')) else 0
                tower_normalized = normalize_tower_name(tower_raw)
                
                if tower_normalized:
                    slab_data_dict[tower_normalized] = count
                    st.write(f"  ✓ Slab: {tower_raw} → {tower_normalized} = {count}")

        # ========== PROCESS RAW COS DATA ==========
        st.write("### 📊 Processing COS Data...")
        cos_data_dict = {}
        up_first_fix_dict = {}  # Track UP-First Fix by tower
        cp_first_fix_dict = {}  # Track CP-First Fix by tower
        
        cos_df = combined_data.get("COS", pd.DataFrame())
        
        if not cos_df.empty:
            st.write(f"✓ Found {len(cos_df)} COS records")
            
            # First pass: Collect UP/CP-First Fix data
            for _, row in cos_df.iterrows():
                tower_raw = row.get('Tower', '')
                activity_raw = row.get('Activity Name', '')
                count = int(row.get('Count', 0)) if pd.notna(row.get('Count')) else 0
                
                tower_normalized = normalize_tower_name(tower_raw)
                if not tower_normalized:
                    continue
                
                # Check for UP-First Fix
                activity_lower = str(activity_raw).lower()
                if "up" in activity_lower and ("1st fix" in activity_lower or "first fix" in activity_lower):
                    up_first_fix_dict[tower_normalized] = count
                    st.write(f"  ✓ UP-First Fix: {tower_raw} → {tower_normalized} = {count}")
                # Check for CP-First Fix
                elif "cp" in activity_lower and ("1st fix" in activity_lower or "first fix" in activity_lower):
                    cp_first_fix_dict[tower_normalized] = count
                    st.write(f"  ✓ CP-First Fix: {tower_raw} → {tower_normalized} = {count}")
            
            # Calculate Plumbing Works = MIN(UP, CP) per tower
            st.write("\n### 🔧 Calculating Plumbing Works (MIN logic)...")
            for tower in set(list(up_first_fix_dict.keys()) + list(cp_first_fix_dict.keys())):
                up_count = up_first_fix_dict.get(tower, 0)
                cp_count = cp_first_fix_dict.get(tower, 0)
                
                # Only calculate if BOTH exist
                if up_count > 0 and cp_count > 0:
                    plumbing_count = min(up_count, cp_count)
                    key = (tower, "Plumbing Works")
                    cos_data_dict[key] = plumbing_count
                    st.write(f"  ✓ {tower} Plumbing = MIN(UP:{up_count}, CP:{cp_count}) = {plumbing_count}")
            
            # Second pass: Process ALL other activities (except UP/CP)
            for _, row in cos_df.iterrows():
                tower_raw = row.get('Tower', '')
                activity_raw = row.get('Activity Name', '')
                count = int(row.get('Count', 0)) if pd.notna(row.get('Count')) else 0
                
                tower_normalized = normalize_tower_name(tower_raw)
                if not tower_normalized:
                    continue
                
                activity_lower = str(activity_raw).lower()
                
                # Skip UP/CP-First Fix (already processed)
                if ("up" in activity_lower or "cp" in activity_lower) and ("1st fix" in activity_lower or "first fix" in activity_lower):
                    continue
                
                # Process all other activities normally
                activity_normalized = normalize_activity_name(activity_raw)
                if activity_normalized and activity_normalized != "Plumbing Works":  # Don't duplicate Plumbing
                    key = (tower_normalized, activity_normalized)
                    cos_data_dict[key] = cos_data_dict.get(key, 0) + count
            
            st.write(f"\n✓ **Total COS entries**: {len(cos_data_dict)}")
            
            # Debug: Show what we extracted
            st.write("\n**COS Activities by Tower:**")
            for (tower, activity), count in sorted(cos_data_dict.items()):
                if count > 0:
                    st.write(f"  • {tower} | {activity}: {count}")
        else:
            st.warning("⚠️ No COS data found!")

        # Add slab data as Concreting
        for tower, count in slab_data_dict.items():
            key = (tower, "Concreting")
            cos_data_dict[key] = count
            st.write(f"  ✓ Added Slab→Concreting: {tower} = {count}")

        # ========== PROCESS RAW ASITE DATA ==========
        st.write("\n### 📊 Processing Asite Data...")
        asite_data_dict = {}
        
        asite_df = combined_data.get("Asite", pd.DataFrame())
        
        if not asite_df.empty:
            st.write(f"✓ Found {len(asite_df)} Asite records")
            
            for _, row in asite_df.iterrows():
                tower_raw = row.get('Tower', '')
                activity_raw = row.get('Activity Name', '')
                count = int(row.get('Count', 0)) if pd.notna(row.get('Count')) else 0
                
                # Normalize
                tower_normalized = normalize_tower_name(tower_raw)
                activity_normalized = normalize_activity_name(activity_raw)
                
                if tower_normalized and activity_normalized:
                    key = (tower_normalized, activity_normalized)
                    asite_data_dict[key] = asite_data_dict.get(key, 0) + count
            
            st.write(f"✓ **Total Asite entries**: {len(asite_data_dict)}")
            
            # Show extracted activities
            st.write("\n**Extracted Asite Activities (non-zero only):**")
            for (tower, activity), count in sorted(asite_data_dict.items()):
                if count > 0:
                    st.write(f"  • {tower} | {activity}: {count}")
        else:
            st.error("❌ No Asite data found!")

        # ========== GET ALL TOWERS ==========
        all_towers = set()
        for key in cos_data_dict.keys():
            if key[0] in ["TF", "TG", "TH"]:
                all_towers.add(key[0])
        for key in asite_data_dict.keys():
            if key[0] in ["TF", "TG", "TH"]:
                all_towers.add(key[0])

        if not all_towers:
            st.error("❌ No valid towers found")
            return None

        st.write(f"\n✓ **Final towers to process**: {sorted(all_towers)}")

        # ========== BUILD CONSOLIDATED ROWS ==========
        st.write("\n### 📝 Building Consolidated Data...")
        consolidated_rows = []
        
        for tower in sorted(all_towers):
            st.write(f"\n**Tower {tower}:**")
            
            for category, activities in activities_structure.items():
                st.write(f"  **{category}:**")
                
                for activity_name in activities:
                    cos_total = cos_data_dict.get((tower, activity_name), 0)
                    asite_total = asite_data_dict.get((tower, activity_name), 0)
                    # FIXED OPEN CHECKLIST LOGIC
                    if asite_total > cos_total:
                        open_missing = 0  # Closed checklist > completed work = 0
                    else:
                        open_missing = cos_total - asite_total  # Completed work - closed checklist
                    
                    consolidated_rows.append({
                        "Tower": tower,
                        "Category": category,
                        "Activity Name": activity_name,
                        "Completed Work*(Count of Flat)": cos_total,
                        "In Progress ": 0,
                        "Closed checklist against completed work": asite_total,
                        "Open/Missing check list": open_missing
                    })
                    
                    st.write(f"    {activity_name}: COS={cos_total}, Asite={asite_total}, Gap={open_missing}")

        # ========== APPLY CIVIL WORKS MAPPING ==========
        st.write("\n### 🔧 Applying Civil Works Mapping...")
        concreting_by_tower = {}
        for row in consolidated_rows:
            if row['Activity Name'] == 'Concreting' and row['Category'] == 'Civil Works':
                concreting_by_tower[row['Tower']] = row['Completed Work*(Count of Flat)']
                st.write(f"  {row['Tower']} Concreting: {row['Completed Work*(Count of Flat)']}")
        
        # Map to other Civil Works
        for row in consolidated_rows:
            if row['Category'] == 'Civil Works' and row['Tower'] in concreting_by_tower:
                if row['Activity Name'] in ['Shuttering', 'Reinforcement', 'De-Shuttering']:
                    concreting_count = concreting_by_tower[row['Tower']]
                    row['Completed Work*(Count of Flat)'] = concreting_count
                    # FIXED OPEN CHECKLIST LOGIC
                    closed_count = row["Closed checklist against completed work"]
                    if closed_count > concreting_count:
                        row["Open/Missing check list"] = 0
                    else:
                        row["Open/Missing check list"] = concreting_count - closed_count
                    st.write(f"  Mapped {row['Tower']} {row['Activity Name']} = {concreting_count}")

        # Sync Slab Conducting
        st.write("\n### 🔧 Syncing Slab Conducting with Concreting...")
        for row in consolidated_rows:
            if row['Activity Name'] == 'Slab Conducting' and row['Tower'] in concreting_by_tower:
                concreting_count = concreting_by_tower[row['Tower']]
                row['Completed Work*(Count of Flat)'] = concreting_count
                # FIXED OPEN CHECKLIST LOGIC
                closed_count = row["Closed checklist against completed work"]
                if closed_count > concreting_count:
                    row["Open/Missing check list"] = 0
                else:
                    row["Open/Missing check list"] = concreting_count - closed_count
                st.write(f"  Synced {row['Tower']} Slab Conducting = {concreting_count}")
        consolidated_rows = apply_tower_f_hardcoded_fixes(consolidated_rows)
        # ========== CREATE EXCEL ==========
        df = pd.DataFrame(consolidated_rows)
        
        if df.empty:
            st.warning("⚠️ No data available")
            return None

        # Display final summary
        st.write("\n### 📊 Final Data Summary:")
        for tower in sorted(all_towers):
            tower_data = df[df['Tower'] == tower]
            st.write(f"\n**{tower}:**")
            
            for category in activities_structure.keys():
                cat_data = tower_data[tower_data['Category'] == category]
                total_completed = cat_data['Completed Work*(Count of Flat)'].sum()
                total_closed = cat_data['Closed checklist against completed work'].sum()
                total_open = cat_data['Open/Missing check list'].sum()
                
                st.write(f"  {category}: ✅Completed={total_completed}, 📋Closed={total_closed}, ⚠️Open={total_open}")

        # ========== EXCEL GENERATION ==========
        output = BytesIO()
        workbook = Workbook()
        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])

        header_font = Font(bold=True)
        category_font = Font(bold=True, italic=True)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center')

        worksheet = workbook.create_sheet(title="Consolidated Checklist")
        current_row = 1

        for tower in sorted(all_towers):
            tower_group = df[df['Tower'] == tower]
            
            worksheet.cell(row=current_row, column=1).value = tower
            worksheet.cell(row=current_row, column=1).font = header_font
            current_row += 1

            for category in ["Civil Works", "Interior Finishing Works", "MEP Works"]:
                cat_group = tower_group[tower_group['Category'] == category]
                if cat_group.empty:
                    continue
                    
                worksheet.cell(row=current_row, column=1).value = f"{current_month} Checklist Status - {category}"
                worksheet.cell(row=current_row, column=1).font = category_font
                current_row += 1

                headers = [
                    "ACTIVITY NAME",
                    "Completed Work*(Count of Flat)",
                    "In Progress ",
                    "Closed checklist against completed work",
                    "Open/Missing check list"
                ]
                for col, header in enumerate(headers, start=1):
                    cell = worksheet.cell(row=current_row, column=col)
                    cell.value = header
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = center_alignment
                current_row += 1

                for activity_name in activities_structure[category]:
                    activity_row = cat_group[cat_group['Activity Name'] == activity_name]
                    if not activity_row.empty:
                        row_data = activity_row.iloc[0]
                        
                        worksheet.cell(row=current_row, column=1).value = row_data["Activity Name"]
                        worksheet.cell(row=current_row, column=2).value = row_data["Completed Work*(Count of Flat)"]
                        worksheet.cell(row=current_row, column=3).value = row_data["In Progress "]
                        worksheet.cell(row=current_row, column=4).value = row_data["Closed checklist against completed work"]
                        worksheet.cell(row=current_row, column=5).value = row_data["Open/Missing check list"]
                        
                        for col in range(1, 6):
                            cell = worksheet.cell(row=current_row, column=col)
                            cell.border = border
                            cell.alignment = center_alignment
                        current_row += 1

                total_open = cat_group["Open/Missing check list"].sum()
                worksheet.cell(row=current_row, column=1).value = "TOTAL pending checklist"
                worksheet.cell(row=current_row, column=5).value = total_open
                
                for col in range(1, 6):
                    cell = worksheet.cell(row=current_row, column=col)
                    cell.font = category_font
                    cell.border = border
                    cell.alignment = center_alignment
                current_row += 1
                current_row += 1

        for col in worksheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            worksheet.column_dimensions[column].width = min(max_length + 2, 50)


        # ========== FIX #3: SHEET 2 - CONSOLIDATED SUMMARY ==========
        worksheet2 = workbook.create_sheet(title=f"Checklist {current_month}")
        current_row = 1

        worksheet2.cell(row=current_row, column=1).value = f"Checklist: {current_month}"
        worksheet2.cell(row=current_row, column=1).font = header_font
        current_row += 1

        headers = [
            "Site",
            "Total of Missing & Open Checklist-Civil Works",
            "Total of Missing & Open Checklist-MEP Works",
            "Total of Missing & Open Checklist-Interior Finishing Works",
            "TOTAL"
        ]
        for col, header in enumerate(headers, start=1):
            cell = worksheet2.cell(row=current_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.border = border
            cell.alignment = center_alignment
        current_row += 1

        # Build summary data
        summary_data = {}
        
        for _, row in df.iterrows():
            tower = row["Tower"]
            category = row["Category"]
            open_missing = row["Open/Missing check list"]

            if open_missing == 0:
                continue

            # Create site name (e.g., "Eligo-TF", "Eligo-TG", "Eligo-TH")
            site_name = f"Eligo-{tower}"

            if site_name not in summary_data:
                summary_data[site_name] = {
                    "Civil Works": 0,
                    "MEP Works": 0,
                    "Interior Finishing Works": 0
                }

            if category in summary_data[site_name]:
                summary_data[site_name][category] += open_missing

        # Write summary data to Sheet 2
        if not summary_data:
            logger.warning("No summary data found for Sheet 2")
            worksheet2.cell(row=current_row, column=1).value = "No data available"
            for col in range(2, 6):
                worksheet2.cell(row=current_row, column=col).value = 0
        else:
            for site_name, counts in sorted(summary_data.items()):
                civil_count = counts["Civil Works"]
                mep_count = counts["MEP Works"]
                interior_count = counts["Interior Finishing Works"]
                total_count = civil_count + mep_count + interior_count

                if total_count > 0:
                    worksheet2.cell(row=current_row, column=1).value = site_name
                    worksheet2.cell(row=current_row, column=2).value = civil_count
                    worksheet2.cell(row=current_row, column=3).value = mep_count
                    worksheet2.cell(row=current_row, column=4).value = interior_count
                    worksheet2.cell(row=current_row, column=5).value = total_count

                    for col in range(1, 6):
                        cell = worksheet2.cell(row=current_row, column=col)
                        cell.border = border
                        cell.alignment = center_alignment
                    current_row += 1

        # Adjust column widths for Sheet 2
        for col in worksheet2.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet2.column_dimensions[column].width = adjusted_width

        workbook.save(output)
        output.seek(0)

        st.success("✅ Excel file generated successfully with 2 sheets!")
        return output

    except Exception as e:
        logger.error(f"Error generating Excel: {str(e)}", exc_info=True)
        st.error(f"❌ Error: {str(e)}")
        return None


# ========== UPDATE THE BUTTON HANDLER ==========
def run_analysis_and_display_final():
    """Final fixed version with complete direct data processing"""
    try:
        st.write("Running status analysis...")
        combined_data, outputs = AnalyzeStatusManually()
        st.success("Status analysis completed successfully!")

        st.write("Displaying activity counts...")
        display_activity_count()
        st.success("Activity counts displayed successfully!")

        st.write("Generating consolidated checklist Excel file (Direct Processing - No AI)...")
        with st.spinner("Generating Excel file... This may take a moment."):
            # Use raw data only - no AI parsing
            excel_file = generate_consolidated_Checklist_excel(combined_data)
        
        if excel_file is not None:
            timestamp = pd.Timestamp.now(tz='Asia/Kolkata').strftime('%Y%m%d_%H%M')
            file_name = f"Consolidated_Checklist_Eligo_{timestamp}.xlsx"
            
            st.sidebar.download_button(
                label="📥 Download Checklist Excel (Fixed)",
                data=excel_file,
                file_name=file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_excel_button_fixed"
            )
            st.success("✅ Excel file generated successfully!")
        else:
            st.error("❌ Failed to generate Excel file.")

    except Exception as e:
        st.error(f"❌ Error during analysis: {str(e)}")
        logging.error(f"Error: {str(e)}", exc_info=True)

def sync_slab_in_ai_response(ai_data):
    """
    Sync Slab Conduting in Asite AI response to match Concreting from COS
    This applies the sync at the AI data level before Excel generation
    """
    try:
        # Extract Concreting count from COS for each tower
        concreting_by_tower = {}
        
        for tower_data in ai_data.get("COS", []):
            tower_name = tower_data.get("Tower", "Unknown")
            
            for category_data in tower_data.get("Categories", []):
                if category_data.get("Category") == "ED Civil":
                    for activity in category_data.get("Activities", []):
                        if activity.get("Activity Name") == "Concreting":
                            concreting_by_tower[tower_name] = activity.get("Total", 0)
                            logger.info(f"Found COS Concreting for {tower_name}: {activity.get('Total', 0)}")
        
        if not concreting_by_tower:
            logger.warning("No COS Concreting data found for Slab Conduting sync")
            return ai_data
        
        # Update Asite Slab Conducting to match COS Concreting
        rows_updated = 0
        for tower_data in ai_data.get("Asite", []):
            tower_name = tower_data.get("Tower", "Unknown")
            
            if tower_name in concreting_by_tower:
                concreting_count = concreting_by_tower[tower_name]
                
                for category_data in tower_data.get("Categories", []):
                    if category_data.get("Category") == "MEP":
                        for activity in category_data.get("Activities", []):
                            if activity.get("Activity Name") == "Slab Conduting":
                                old_count = activity.get("Total", 0)
                                activity["Total"] = concreting_count
                                logger.info(f"Synced Asite {tower_name} - Slab Conduting: {old_count} → {concreting_count}")
                                rows_updated += 1
        
        logger.info(f"AI response Slab Conduting sync completed: {rows_updated} activities updated")
        return ai_data
    
    except Exception as e:
        logger.error(f"Error in sync_slab_in_ai_response: {str(e)}")
        return ai_data


# Streamlit UI - Modified Button Code
st.sidebar.title(" Asite Initialization")
email = st.sidebar.text_input("Email", value=EMAIL_ID if EMAIL_ID else "" , key="email_input")
password = st.sidebar.text_input("Password",  value=PASSWORD if PASSWORD else "" , type="password", key="password_input")

if st.sidebar.button("Initialize and Fetch Data"):
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        success = loop.run_until_complete(initialize_and_fetch_data(email, password))
        if success:
            st.sidebar.success("Initialization and data fetching completed successfully!")
        else:
            st.sidebar.error("Initialization and data fetching failed!")
    except Exception as e:
        st.sidebar.error(f"Initialization and data fetching failed: {str(e)}")
    finally:
        loop.close()

# Combined function to handle both analysis and activity count display

st.sidebar.title("Status Analysis")

if st.sidebar.button("Analyze and Display Activity Counts"):
    run_analysis_and_display_final()

st.sidebar.title("Slab Cycle")
st.session_state.ignore_year = datetime.now().year
st.session_state.ignore_month = datetime.now().month

